from datetime import datetime, timedelta, date, time as datetime_time
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
import io
import openpyxl
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import TankStockLedger, LocationAccountingDaySetting, MaterialBalanceTemplate, MaterialBalanceTemplateColumn, OperationTransaction, OperationTransactionValue, User
from app.schemas import TankStockLedgerResponse, TankStockLedgerSummaryResponse, TankStockLedgerDailySummaryResponse, OutTurnReportResponse, MaterialBalanceDynamicReportResponse, FSOOTRReportResponse, FSOMaterialBalanceReportResponse, FSOOutturnReportResponse
from app.dependencies.auth import get_current_user_from_token
from app.dependencies.permissions import require_user_permission
from app.services.audit_service import create_audit_log
from app.utils.helpers import safe_float, clean_optional_text, get_transaction_ticket_number, get_location_by_code
from app.services.material_balance_helpers import (
    normalize_material_balance_code_value,
    get_active_material_balance_template_for_location,
    get_active_material_balance_template_columns,
    get_movement_value_for_unit,
    get_snapshot_value_for_unit,
    should_row_match_material_balance_column,
    get_global_internal_transfer_operation_codes,
    should_row_be_in_book_closing_formula,
    calculate_book_closing_from_eligible_ledger_rows,
)
from app.config import APPROVED_TRANSACTION_STATUS
from app.services.transaction_helpers import parse_date_filter

router = APIRouter(prefix="/tank-stock-ledger", tags=["Tank Stock Ledger"])


def build_tank_stock_ledger_response(
    ledger: TankStockLedger,
    db: Session,
):
    location = get_location_by_code(ledger.location_code, db)

    return {
        "id": ledger.id,
        "transaction_id": ledger.transaction_id,
        "ticket_number": ledger.ticket_number,
        "operation_number": ledger.operation_number,
        "location_code": ledger.location_code,
        "location_name": location.location_name if location else "",
        "tank_asset_code": ledger.tank_asset_code,
        "tank_asset_name": ledger.tank_asset_name,
        "operation_date": ledger.operation_date,
        "product_name": ledger.product_name,
        "accounting_date": ledger.accounting_date,
        "accounting_day_start": ledger.accounting_day_start,
        "accounting_day_end": ledger.accounting_day_end,
        "accounting_day_setting_id": ledger.accounting_day_setting_id,
        "tank_operation_code": ledger.tank_operation_code,
        "tank_operation_label": ledger.tank_operation_label,
        "tank_operation_category": ledger.tank_operation_category,
        "tank_operation_sign": ledger.tank_operation_sign,
        "movement_gsv_bbl": ledger.movement_gsv_bbl or 0,
        "movement_nsv_bbl": ledger.movement_nsv_bbl or 0,
        "movement_lt": ledger.movement_lt or 0,
        "movement_mt": ledger.movement_mt or 0,
        "stock_gsv_bbl": ledger.stock_gsv_bbl or 0,
        "stock_nsv_bbl": ledger.stock_nsv_bbl or 0,
        "stock_lt": ledger.stock_lt or 0,
        "stock_mt": ledger.stock_mt or 0,
        "previous_stock_gsv_bbl": ledger.previous_stock_gsv_bbl or 0,
        "previous_stock_nsv_bbl": ledger.previous_stock_nsv_bbl or 0,
        "previous_stock_lt": ledger.previous_stock_lt or 0,
        "previous_stock_mt": ledger.previous_stock_mt or 0,
        "running_balance_gsv_bbl": ledger.running_balance_gsv_bbl or 0,
        "running_balance_nsv_bbl": ledger.running_balance_nsv_bbl or 0,
        "running_balance_lt": ledger.running_balance_lt or 0,
        "running_balance_mt": ledger.running_balance_mt or 0,
        "status": ledger.status,
        "created_by": ledger.created_by,
        "remarks": ledger.remarks,
        "created_at": ledger.created_at,
        "updated_at": ledger.updated_at,
    }


def get_filtered_tank_stock_ledger_rows(
    db: Session,
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
):
    query = db.query(TankStockLedger)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)
    cleaned_status = clean_optional_text(status)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    if date_from:
        query = query.filter(TankStockLedger.accounting_date >= date_from)

    if date_to:
        query = query.filter(TankStockLedger.accounting_date <= date_to)

    if cleaned_status:
        query = query.filter(TankStockLedger.status == cleaned_status)

    return (
        query.order_by(
            TankStockLedger.location_code.asc(),
            TankStockLedger.tank_asset_code.asc(),
            TankStockLedger.accounting_date.asc(),
            TankStockLedger.operation_date.asc(),
            TankStockLedger.id.asc(),
        )
        .all()
    )


def build_date_range(start_date: date, end_date: date):
    if end_date < start_date:
        raise HTTPException(
            status_code=400,
            detail="Date To cannot be earlier than Date From",
        )

    dates = []
    current_date = start_date

    while current_date <= end_date:
        dates.append(current_date)
        current_date = current_date + timedelta(days=1)

    return dates


def get_active_location_day_setting(db: Session, location_code: str, on_date: date):
    return (
        db.query(LocationAccountingDaySetting)
        .filter(
            LocationAccountingDaySetting.location_code.ilike(location_code),
            LocationAccountingDaySetting.status == "Active",
            LocationAccountingDaySetting.effective_from <= on_date,
            or_(
                LocationAccountingDaySetting.effective_to.is_(None),
                LocationAccountingDaySetting.effective_to >= on_date,
            ),
        )
        .order_by(
            LocationAccountingDaySetting.effective_from.desc(),
            LocationAccountingDaySetting.id.desc(),
        )
        .first()
    )


def compute_accounting_date(
    op_date: date,
    event_time: str | None,
    day_start_time: datetime_time,
):
    if not event_time:
        return op_date

    try:
        hh, mm = event_time.split(":")
        t = datetime_time(int(hh), int(mm))
    except Exception:
        return op_date

    return op_date - timedelta(days=1) if t < day_start_time else op_date


def combine_operation_datetime(op_date: date, event_time: str | None, tz_name: str):
    try:
        if not event_time:
            return None
        hh, mm = event_time.split(":")
        dt = datetime(op_date.year, op_date.month, op_date.day, int(hh), int(mm))
        return dt.replace(tzinfo=ZoneInfo(tz_name))
    except Exception:
        return None


def get_tank_stock_rows_for_daily_summary(
    db: Session,
    location_code: str | None,
    tank_asset_code: str | None,
    product_name: str | None,
    date_to_value: date,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.accounting_date != None,
        TankStockLedger.accounting_date <= date_to_value,
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    return (
        query.order_by(
            TankStockLedger.location_code.asc(),
            TankStockLedger.tank_asset_code.asc(),
            TankStockLedger.product_name.asc(),
            TankStockLedger.accounting_date.asc(),
            TankStockLedger.operation_date.asc(),
            TankStockLedger.id.asc(),
        )
        .all()
    )


def get_stock_snapshot_values(row: TankStockLedger):
    stock_gsv = safe_float(row.stock_gsv_bbl)
    stock_nsv = safe_float(row.stock_nsv_bbl)
    stock_lt = safe_float(row.stock_lt)
    stock_mt = safe_float(row.stock_mt)

    if stock_gsv == 0 and stock_nsv == 0:
        stock_gsv = safe_float(row.running_balance_gsv_bbl)
        stock_nsv = safe_float(row.running_balance_nsv_bbl)
        stock_lt = safe_float(row.running_balance_lt)
        stock_mt = safe_float(row.running_balance_mt)

    return {
        "gsv": stock_gsv,
        "nsv": stock_nsv,
        "lt": stock_lt,
        "mt": stock_mt,
    }


def get_ledger_operation_datetime(row: TankStockLedger):
    try:
        payload = row.source_payload or {}
        inputs = payload.get("inputs") or {}

        gauging_date = clean_optional_text(inputs.get("gaugingDate"))
        gauging_time = clean_optional_text(inputs.get("gaugingTime"))

        if gauging_date and gauging_time:
            return datetime.fromisoformat(f"{gauging_date}T{gauging_time}")
    except Exception:
        pass

    if row.accounting_day_start is not None:
        return row.accounting_day_start

    if row.operation_date is not None:
        return datetime.combine(row.operation_date, datetime_time(0, 0))

    return None


def build_out_turn_report_response(
    row: TankStockLedger,
    db: Session,
):
    location = get_location_by_code(row.location_code, db)

    operation_datetime = get_ledger_operation_datetime(row)

    previous_gsv = safe_float(row.previous_stock_gsv_bbl)
    previous_nsv = safe_float(row.previous_stock_nsv_bbl)
    previous_lt = safe_float(row.previous_stock_lt)
    previous_mt = safe_float(row.previous_stock_mt)

    stock_snapshot = get_stock_snapshot_values(row)

    stock_after_gsv = stock_snapshot["gsv"]
    stock_after_nsv = stock_snapshot["nsv"]
    stock_after_lt = stock_snapshot["lt"]
    stock_after_mt = stock_snapshot["mt"]

    movement_gsv = safe_float(row.movement_gsv_bbl)
    movement_nsv = safe_float(row.movement_nsv_bbl)
    movement_lt = safe_float(row.movement_lt)
    movement_mt = safe_float(row.movement_mt)

    sign = str(row.tank_operation_sign or "").upper()

    net_receipt_gsv = 0
    net_receipt_nsv = 0
    net_receipt_lt = 0
    net_receipt_mt = 0

    net_dispatch_gsv = 0
    net_dispatch_nsv = 0
    net_dispatch_lt = 0
    net_dispatch_mt = 0

    signed_net_gsv = 0
    signed_net_nsv = 0
    signed_net_lt = 0
    signed_net_mt = 0

    if sign == "IN":
        net_receipt_gsv = movement_gsv
        net_receipt_nsv = movement_nsv
        net_receipt_lt = movement_lt
        net_receipt_mt = movement_mt

        signed_net_gsv = movement_gsv
        signed_net_nsv = movement_nsv
        signed_net_lt = movement_lt
        signed_net_mt = movement_mt

    elif sign == "OUT":
        net_dispatch_gsv = movement_gsv
        net_dispatch_nsv = movement_nsv
        net_dispatch_lt = movement_lt
        net_dispatch_mt = movement_mt

        signed_net_gsv = movement_gsv * -1
        signed_net_nsv = movement_nsv * -1
        signed_net_lt = movement_lt * -1
        signed_net_mt = movement_mt * -1

    elif sign == "SET":
        signed_net_gsv = 0
        signed_net_nsv = 0
        signed_net_lt = 0
        signed_net_mt = 0

    elif sign == "NEUTRAL":
        signed_net_gsv = 0
        signed_net_nsv = 0
        signed_net_lt = 0
        signed_net_mt = 0

    return {
        "ledger_id": row.id,
        "transaction_id": row.transaction_id,
        "ticket_number": row.ticket_number,
        "operation_number": row.operation_number,
        "accounting_date": row.accounting_date,
        "operation_datetime": operation_datetime,
        "location_code": row.location_code,
        "location_name": location.location_name if location else "",
        "tank_asset_code": row.tank_asset_code,
        "tank_asset_name": row.tank_asset_name,
        "product_name": row.product_name,
        "tank_operation_code": row.tank_operation_code,
        "tank_operation_label": row.tank_operation_label,
        "tank_operation_category": row.tank_operation_category,
        "tank_operation_sign": row.tank_operation_sign,
        "previous_stock_gsv_bbl": round(previous_gsv, 3),
        "previous_stock_nsv_bbl": round(previous_nsv, 3),
        "previous_stock_lt": round(previous_lt, 3),
        "previous_stock_mt": round(previous_mt, 3),
        "stock_after_gsv_bbl": round(stock_after_gsv, 3),
        "stock_after_nsv_bbl": round(stock_after_nsv, 3),
        "stock_after_lt": round(stock_after_lt, 3),
        "stock_after_mt": round(stock_after_mt, 3),
        "net_receipt_gsv_bbl": round(net_receipt_gsv, 3),
        "net_receipt_nsv_bbl": round(net_receipt_nsv, 3),
        "net_receipt_lt": round(net_receipt_lt, 3),
        "net_receipt_mt": round(net_receipt_mt, 3),
        "net_dispatch_gsv_bbl": round(net_dispatch_gsv, 3),
        "net_dispatch_nsv_bbl": round(net_dispatch_nsv, 3),
        "net_dispatch_lt": round(net_dispatch_lt, 3),
        "net_dispatch_mt": round(net_dispatch_mt, 3),
        "signed_net_movement_gsv_bbl": round(signed_net_gsv, 3),
        "signed_net_movement_nsv_bbl": round(signed_net_nsv, 3),
        "signed_net_movement_lt": round(signed_net_lt, 3),
        "signed_net_movement_mt": round(signed_net_mt, 3),
        "status": row.status,
        "remarks": row.remarks,
    }


def get_out_turn_report_rows(
    db: Session,
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
):
    query = db.query(TankStockLedger)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)
    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(TankStockLedger.status == cleaned_status)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value:
        query = query.filter(TankStockLedger.accounting_date >= date_from_value)

    if date_to_value:
        query = query.filter(TankStockLedger.accounting_date <= date_to_value)

    rows = query.all()

    rows = sorted(
        rows,
        key=lambda row: (
            row.accounting_date or date.min,
            get_ledger_operation_datetime(row) or datetime.min,
            row.location_code or "",
            row.tank_asset_code or "",
            row.product_name or "",
            row.id,
        ),
    )

    return rows


def build_tank_stock_daily_summary_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    date_from_value: date,
    date_to_value: date,
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    daily_summary_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key

        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                row.accounting_day_start or datetime.min,
                row.operation_date or date.min,
                row.id,
            ),
        )

        tank_asset_name = ""
        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_gsv = 0
        previous_closing_nsv = 0
        previous_closing_lt = 0
        previous_closing_mt = 0

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            last_before_period = rows_before_period[-1]
            previous_snapshot = get_stock_snapshot_values(last_before_period)

            previous_closing_gsv = previous_snapshot["gsv"]
            previous_closing_nsv = previous_snapshot["nsv"]
            previous_closing_lt = previous_snapshot["lt"]
            previous_closing_mt = previous_snapshot["mt"]

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    row.accounting_day_start or datetime.min,
                    row.operation_date or date.min,
                    row.id,
                ),
            )

            opening_gsv = previous_closing_gsv
            opening_nsv = previous_closing_nsv
            opening_lt = previous_closing_lt
            opening_mt = previous_closing_mt

            opening_rows = [
                row
                for row in day_rows
                if str(row.tank_operation_category or "").upper() == "OPENING"
            ]

            if opening_rows:
                opening_snapshot = get_stock_snapshot_values(opening_rows[-1])

                opening_gsv = opening_snapshot["gsv"]
                opening_nsv = opening_snapshot["nsv"]
                opening_lt = opening_snapshot["lt"]
                opening_mt = opening_snapshot["mt"]

            total_in_gsv = 0
            total_in_nsv = 0
            total_in_lt = 0
            total_in_mt = 0

            total_out_gsv = 0
            total_out_nsv = 0
            total_out_lt = 0
            total_out_mt = 0

            for row in day_rows:
                sign = str(row.tank_operation_sign or "").upper()

                if sign == "IN":
                    total_in_gsv += safe_float(row.movement_gsv_bbl)
                    total_in_nsv += safe_float(row.movement_nsv_bbl)
                    total_in_lt += safe_float(row.movement_lt)
                    total_in_mt += safe_float(row.movement_mt)

                elif sign == "OUT":
                    total_out_gsv += safe_float(row.movement_gsv_bbl)
                    total_out_nsv += safe_float(row.movement_nsv_bbl)
                    total_out_lt += safe_float(row.movement_lt)
                    total_out_mt += safe_float(row.movement_mt)

            book_closing_gsv = opening_gsv + total_in_gsv - total_out_gsv
            book_closing_nsv = opening_nsv + total_in_nsv - total_out_nsv
            book_closing_lt = opening_lt + total_in_lt - total_out_lt
            book_closing_mt = opening_mt + total_in_mt - total_out_mt

            actual_closing_gsv = book_closing_gsv
            actual_closing_nsv = book_closing_nsv
            actual_closing_lt = book_closing_lt
            actual_closing_mt = book_closing_mt

            last_ticket_number = None

            if day_rows:
                closing_rows = [
                    row
                    for row in day_rows
                    if str(row.tank_operation_category or "").upper()
                    == "CLOSING"
                ]

                if closing_rows:
                    closing_source_row = closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                closing_snapshot = get_stock_snapshot_values(closing_source_row)

                actual_closing_gsv = closing_snapshot["gsv"]
                actual_closing_nsv = closing_snapshot["nsv"]
                actual_closing_lt = closing_snapshot["lt"]
                actual_closing_mt = closing_snapshot["mt"]
                last_ticket_number = closing_source_row.ticket_number

            else:
                actual_closing_gsv = opening_gsv
                actual_closing_nsv = opening_nsv
                actual_closing_lt = opening_lt
                actual_closing_mt = opening_mt

                book_closing_gsv = opening_gsv
                book_closing_nsv = opening_nsv
                book_closing_lt = opening_lt
                book_closing_mt = opening_mt

            loss_gain_gsv = actual_closing_gsv - book_closing_gsv
            loss_gain_nsv = actual_closing_nsv - book_closing_nsv
            loss_gain_lt = actual_closing_lt - book_closing_lt
            loss_gain_mt = actual_closing_mt - book_closing_mt

            daily_summary_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "opening_gsv_bbl": round(opening_gsv, 3),
                    "opening_nsv_bbl": round(opening_nsv, 3),
                    "opening_lt": round(opening_lt, 3),
                    "opening_mt": round(opening_mt, 3),
                    "total_in_gsv_bbl": round(total_in_gsv, 3),
                    "total_in_nsv_bbl": round(total_in_nsv, 3),
                    "total_in_lt": round(total_in_lt, 3),
                    "total_in_mt": round(total_in_mt, 3),
                    "total_out_gsv_bbl": round(total_out_gsv, 3),
                    "total_out_nsv_bbl": round(total_out_nsv, 3),
                    "total_out_lt": round(total_out_lt, 3),
                    "total_out_mt": round(total_out_mt, 3),
                    "book_closing_gsv_bbl": round(book_closing_gsv, 3),
                    "book_closing_nsv_bbl": round(book_closing_nsv, 3),
                    "book_closing_lt": round(book_closing_lt, 3),
                    "book_closing_mt": round(book_closing_mt, 3),
                    "actual_closing_gsv_bbl": round(actual_closing_gsv, 3),
                    "actual_closing_nsv_bbl": round(actual_closing_nsv, 3),
                    "actual_closing_lt": round(actual_closing_lt, 3),
                    "actual_closing_mt": round(actual_closing_mt, 3),
                    "loss_gain_gsv_bbl": round(loss_gain_gsv, 3),
                    "loss_gain_nsv_bbl": round(loss_gain_nsv, 3),
                    "loss_gain_lt": round(loss_gain_lt, 3),
                    "loss_gain_mt": round(loss_gain_mt, 3),
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_gsv = actual_closing_gsv
            previous_closing_nsv = actual_closing_nsv
            previous_closing_lt = actual_closing_lt
            previous_closing_mt = actual_closing_mt

    return sorted(
        daily_summary_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"],
            row["product_name"] or "",
        ),
    )



def add_volume_values(target: dict, prefix: str, row: TankStockLedger):
    target[f"{prefix}_gsv"] += safe_float(row.movement_gsv_bbl)
    target[f"{prefix}_nsv"] += safe_float(row.movement_nsv_bbl)
    target[f"{prefix}_lt"] += safe_float(row.movement_lt)
    target[f"{prefix}_mt"] += safe_float(row.movement_mt)


def get_material_balance_rows_for_continuity(
    db: Session,
    location_code: str | None,
    tank_asset_code: str | None,
    product_name: str | None,
    date_to_value: date,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.accounting_date != None,
        TankStockLedger.accounting_date <= date_to_value,
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    rows = query.all()

    return sorted(
        rows,
        key=lambda row: (
            row.location_code or "",
            row.tank_asset_code or "",
            row.product_name or "",
            row.accounting_date or date.min,
            get_ledger_operation_datetime(row) or datetime.min,
            row.id,
        ),
    )




def build_dynamic_material_balance_columns_response(
    columns: list[MaterialBalanceTemplateColumn],
):
    return [
        {
            "column_key": column.column_key,
            "column_label": column.column_label,
            "column_order": column.column_order,
            "column_type": column.column_type,
            "movement_direction": column.movement_direction,
            "include_in_material_balance": column.include_in_material_balance,
            "include_in_book_closing": column.include_in_book_closing,
            "is_internal_transfer": column.is_internal_transfer,
        }
        for column in columns
    ]





def build_dynamic_material_balance_tank_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    columns: list[MaterialBalanceTemplateColumn],
    date_from_value: date,
    date_to_value: date,
    unit_key: str = "nsv",
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    report_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key
        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        tank_asset_name = ""

        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_snapshot = {
            "gsv": 0,
            "nsv": 0,
            "lt": 0,
            "mt": 0,
        }

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            previous_closing_snapshot = get_stock_snapshot_values(
                rows_before_period[-1]
            )

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    get_ledger_operation_datetime(row) or datetime.min,
                    row.id,
                ),
            )

            opening_value = get_snapshot_value_for_unit(
                previous_closing_snapshot,
                unit_key,
            )

            explicit_opening_rows = [
                row
                for row in day_rows
                if normalize_material_balance_code_value(
                    row.tank_operation_category
                )
                == "OPENING"
            ]

            if explicit_opening_rows:
                opening_snapshot = get_stock_snapshot_values(
                    explicit_opening_rows[-1]
                )
                opening_value = get_snapshot_value_for_unit(
                    opening_snapshot,
                    unit_key,
                )

            values = {}

            book_closing_value = opening_value
            actual_closing_value = opening_value
            last_ticket_number = None

            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "OPENING":
                    values[column_key] = round(opening_value, 3)
                    continue

                if column_type == "MOVEMENT":
                    movement_total = 0

                    for row in day_rows:
                        if should_row_match_material_balance_column(row, column):
                            movement_total += get_movement_value_for_unit(
                                row,
                                unit_key,
                            )

                    values[column_key] = round(movement_total, 3)
                    continue

                if column_type in ["INFO", "FORMULA"]:
                    values[column_key] = 0
                    continue

            book_closing_calculation = calculate_book_closing_from_eligible_ledger_rows(
                opening_value=opening_value,
                day_rows=day_rows,
                columns=columns,
                unit_key=unit_key,
            )

            book_closing_value = book_closing_calculation["book_closing_value"]

            if day_rows:
                explicit_closing_rows = [
                    row
                    for row in day_rows
                    if normalize_material_balance_code_value(
                        row.tank_operation_category
                    )
                    == "CLOSING"
                ]

                if explicit_closing_rows:
                    closing_source_row = explicit_closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                actual_closing_snapshot = get_stock_snapshot_values(
                    closing_source_row
                )

                actual_closing_value = get_snapshot_value_for_unit(
                    actual_closing_snapshot,
                    unit_key,
                )

                last_ticket_number = closing_source_row.ticket_number
            else:
                actual_closing_snapshot = previous_closing_snapshot
                actual_closing_value = opening_value

            loss_gain_value = actual_closing_value - book_closing_value

            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "BOOK_CLOSING":
                    values[column_key] = round(book_closing_value, 3)

                elif column_type == "ACTUAL_CLOSING":
                    values[column_key] = round(actual_closing_value, 3)

                elif column_type == "LOSS_GAIN":
                    values[column_key] = round(loss_gain_value, 3)

            report_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "values": values,
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_snapshot = {
                "gsv": actual_closing_snapshot.get("gsv", actual_closing_value),
                "nsv": actual_closing_snapshot.get("nsv", actual_closing_value),
                "lt": actual_closing_snapshot.get("lt", 0),
                "mt": actual_closing_snapshot.get("mt", 0),
            }

    return sorted(
        report_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"] or "",
            row["product_name"] or "",
        ),
    )


def consolidate_dynamic_material_balance_rows_by_location(
    tank_rows: list[dict],
    columns: list[MaterialBalanceTemplateColumn],
):
    consolidated_map = {}

    for row in tank_rows:
        key = (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        )

        if key not in consolidated_map:
            consolidated_map[key] = {
                "accounting_date": row["accounting_date"],
                "location_code": row["location_code"],
                "location_name": row["location_name"],
                "tank_asset_code": None,
                "tank_asset_name": "All Tanks",
                "product_name": row["product_name"],
                "values": {},
                "rows_count": 0,
                "last_ticket_number": None,
            }

            for column in columns:
                consolidated_map[key]["values"][column.column_key] = 0

        target = consolidated_map[key]

        for column in columns:
            column_key = column.column_key
            target["values"][column_key] = safe_float(
                target["values"].get(column_key)
            ) + safe_float(row["values"].get(column_key))

        target["rows_count"] += int(row.get("rows_count") or 0)

        if row.get("last_ticket_number"):
            target["last_ticket_number"] = row.get("last_ticket_number")

    consolidated_rows = []

    for row in consolidated_map.values():
        for column in columns:
            column_key = column.column_key
            row["values"][column_key] = round(
                safe_float(row["values"].get(column_key)),
                3,
            )

        consolidated_rows.append(row)

    return sorted(
        consolidated_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        ),
    )


def build_fso_otr_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
    shuttle_number: str | None = None,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)
    sn = clean_optional_text(shuttle_number)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )
    if sn:
        q = q.filter(OperationTransaction.convoy_number == sn)

    rows = []
    totals = {
        "receipt": 0.0,
        "export": 0.0,
        "movement": 0.0,
        "variance": 0.0,
        "compare_variance": 0.0,
    }

    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")
        op_label = str(meta.get("operation_label") or "").strip() or "FSO"
        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = float(safe_float(net.get("net_stock_bbl")))
        net_water = float(safe_float(net.get("net_water_bbl")))
        movement_qty = abs(net_stock) + abs(net_water)

        vessel_qty = float(safe_float(inputs.get("vessel_quantity_bbl")))
        variance = abs(net_stock + net_water) - vessel_qty

        src_discharge = float(safe_float(meta.get("source_shuttle_discharge_bbl")))
        compare_var = movement_qty - src_discharge if op_sign == "IN" and src_discharge > 0 else 0.0

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        row = {
            "transaction_id": tx.id,
            "ticket_number": get_transaction_ticket_number(tx),
            "accounting_date": acc_date,
            "operation_date": tx.operation_date,
            "event_time": event_time,
            "location_code": loc_code,
            "fso_asset_code": asset_code,
            "shuttle_number": inputs.get("shuttle_number") or meta.get("shuttle_number") or tx.convoy_number,
            "operation_label": op_label,
            "operation_sign": op_sign,
            "vessel_name": inputs.get("vessel_name"),
            "vessel_quantity_bbl": vessel_qty,
            "opening_stock_bbl": float(safe_float(inputs.get("opening_stock_bbl"))),
            "opening_water_bbl": float(safe_float(inputs.get("opening_water_bbl"))),
            "closing_stock_bbl": float(safe_float(inputs.get("closing_stock_bbl"))),
            "closing_water_bbl": float(safe_float(inputs.get("closing_water_bbl"))),
            "net_stock_bbl": net_stock,
            "net_water_bbl": net_water,
            "movement_qty_bbl": movement_qty,
            "variance_bbl": variance,
            "source_shuttle_discharge_bbl": src_discharge,
            "compare_variance_bbl": compare_var,
            "remarks": inputs.get("remarks"),
        }
        rows.append(row)

        totals["movement"] += movement_qty
        totals["variance"] += variance
        totals["compare_variance"] += compare_var
        if op_sign == "IN":
            totals["receipt"] += movement_qty
        elif op_sign == "OUT":
            totals["export"] += movement_qty

    return rows, totals


def build_fso_material_balance(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, inputs.get("event_time"), day_start)

        buckets.setdefault(acc_date, []).append((tx, meta, inputs, net))

    dates = sorted([d for d in buckets.keys() if d >= date_from and d <= date_to])
    rows = []
    prev_physical_close = None

    for acc_date in dates:
        items = buckets[acc_date]

        def sort_key(item):
            tx, meta, inputs, net = item
            t = inputs.get("event_time") or "00:00"
            return (str(tx.operation_date), str(t), tx.id)

        items = sorted(items, key=sort_key)

        if prev_physical_close is None:
            opening = float(safe_float(items[0][2].get("opening_stock_bbl")))
        else:
            opening = prev_physical_close

        receipt = 0.0
        export = 0.0

        for tx, meta, inputs, net in items:
            op_sign = str(meta.get("operation_sign") or "").strip().upper()
            net_stock = float(safe_float(net.get("net_stock_bbl")))
            net_water = float(safe_float(net.get("net_water_bbl")))
            qty = abs(net_stock) + abs(net_water)

            if op_sign == "IN":
                receipt += qty
            elif op_sign == "OUT":
                export += qty

        book_close = opening + receipt - export

        last_inputs = items[-1][2]
        physical_close = float(safe_float(last_inputs.get("closing_stock_bbl")))
        physical_close_water = float(safe_float(last_inputs.get("closing_water_bbl")))
        loss_gain = physical_close - book_close

        rows.append(
            {
                "accounting_date": acc_date,
                "opening_stock_bbl": opening,
                "receipt_bbl": receipt,
                "export_bbl": export,
                "book_closing_bbl": book_close,
                "physical_closing_bbl": physical_close,
                "physical_closing_water_bbl": physical_close_water,
                "loss_gain_bbl": loss_gain,
            }
        )

        prev_physical_close = physical_close

    return rows


def build_fso_outturn_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    from app.services.transaction_helpers import approved_transaction_not_on_correction_hold

    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            approved_transaction_not_on_correction_hold(db),
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    def _sf(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _abs_qty(net_stock, net_water):
        return abs(_sf(net_stock)) + abs(_sf(net_water))

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        shuttle_no = (
            inputs.get("shuttle_number")
            or meta.get("shuttle_number")
            or tx.convoy_number
            or ""
        ).strip()
        if shuttle_no == "":
            continue

        key = (acc_date, shuttle_no)
        buckets.setdefault(key, {"receipt": 0.0, "discharge": 0.0})

        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = net.get("net_stock_bbl")
        net_water = net.get("net_water_bbl")
        qty = _abs_qty(net_stock, net_water)

        if op_sign == "IN":
            buckets[key]["receipt"] += qty

        src = meta.get("source_shuttle_discharge_bbl")
        if src is not None:
            buckets[key]["discharge"] = max(buckets[key]["discharge"], _sf(src))

    rows = []
    totals = {"discharge": 0.0, "receipt": 0.0, "variance": 0.0}

    for (acc_date, shuttle_no) in sorted(buckets.keys()):
        discharge = float(buckets[(acc_date, shuttle_no)]["discharge"])
        receipt = float(buckets[(acc_date, shuttle_no)]["receipt"])
        variance = receipt - discharge
        pct = (variance / discharge * 100.0) if discharge != 0 else 0.0

        rows.append(
            {
                "accounting_date": acc_date,
                "shuttle_number": shuttle_no,
                "shuttle_discharge_bbl": discharge,
                "fso_receipt_bbl": receipt,
                "variance_bbl": variance,
                "variance_pct": pct,
            }
        )

        totals["discharge"] += discharge
        totals["receipt"] += receipt
        totals["variance"] += variance

    totals_pct = (totals["variance"] / totals["discharge"] * 100.0) if totals["discharge"] != 0 else 0.0
    return rows, totals, totals_pct


def _xlsx_autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


@router.get("", response_model=list[TankStockLedgerResponse])
def get_tank_stock_ledger(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )

    return [
        build_tank_stock_ledger_response(row, db)
        for row in ledger_rows
    ]


@router.get(
    "/summary",
    response_model=list[TankStockLedgerSummaryResponse],
)
def get_tank_stock_ledger_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status="Active",
    )

    summary_map = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in summary_map:
            location = get_location_by_code(row.location_code, db)

            summary_map[key] = {
                "location_code": row.location_code,
                "location_name": location.location_name if location else "",
                "tank_asset_code": row.tank_asset_code,
                "tank_asset_name": row.tank_asset_name,
                "product_name": row.product_name,
                "opening_nsv_bbl": 0,
                "total_in_nsv_bbl": 0,
                "total_out_nsv_bbl": 0,
                "closing_nsv_bbl": 0,
                "opening_lt": 0,
                "total_in_lt": 0,
                "total_out_lt": 0,
                "closing_lt": 0,
                "opening_mt": 0,
                "total_in_mt": 0,
                "total_out_mt": 0,
                "closing_mt": 0,
            }

        summary = summary_map[key]

        sign = row.tank_operation_sign
        category = row.tank_operation_category

        movement_nsv = row.movement_nsv_bbl or 0
        movement_lt = row.movement_lt or 0
        movement_mt = row.movement_mt or 0

        if category == "OPENING":
            summary["opening_nsv_bbl"] += movement_nsv
            summary["opening_lt"] += movement_lt
            summary["opening_mt"] += movement_mt

        if sign == "IN":
            summary["total_in_nsv_bbl"] += movement_nsv
            summary["total_in_lt"] += movement_lt
            summary["total_in_mt"] += movement_mt

        if sign == "OUT":
            summary["total_out_nsv_bbl"] += movement_nsv
            summary["total_out_lt"] += movement_lt
            summary["total_out_mt"] += movement_mt

        summary["closing_nsv_bbl"] = row.running_balance_nsv_bbl or 0
        summary["closing_lt"] = row.running_balance_lt or 0
        summary["closing_mt"] = row.running_balance_mt or 0

    return list(summary_map.values())


@router.get(
    "/daily-summary",
    response_model=list[TankStockLedgerDailySummaryResponse],
)
def get_tank_stock_ledger_daily_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for daily summary",
        )

    ledger_rows = get_tank_stock_rows_for_daily_summary(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
    )

    return build_tank_stock_daily_summary_rows(
        db=db,
        ledger_rows=ledger_rows,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
    )


@router.get(
    "/out-turn-report",
    response_model=list[OutTurnReportResponse],
)
def get_out_turn_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    rows = get_out_turn_report_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )

    return [
        build_out_turn_report_response(row, db)
        for row in rows
    ]


@router.get("/out-turn-report/validation")
def validate_out_turn_report_tank_sequence(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    continuity_rows = get_out_turn_report_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=None,
        date_to=date_to,
        status="Active",
    )

    visible_rows = []

    for row in continuity_rows:
        if date_from_value and row.accounting_date and row.accounting_date < date_from_value:
            continue

        if date_to_value and row.accounting_date and row.accounting_date > date_to_value:
            continue

        visible_rows.append(row)

    grouped_rows = {}

    for row in continuity_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    visible_ledger_ids = {row.id for row in visible_rows}

    issues = []

    for key, group_rows in grouped_rows.items():
        location, tank, product = key

        sorted_group_rows = sorted(
            group_rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        previous_row = None

        for row in sorted_group_rows:
            if previous_row is None:
                expected_previous_nsv = 0
            else:
                previous_snapshot = get_stock_snapshot_values(previous_row)
                expected_previous_nsv = previous_snapshot["nsv"]

            if row.id in visible_ledger_ids:
                actual_previous_nsv = safe_float(row.previous_stock_nsv_bbl)

                if round(actual_previous_nsv, 3) != round(expected_previous_nsv, 3):
                    issues.append(
                        {
                            "ledger_id": row.id,
                            "ticket_number": row.ticket_number,
                            "location_code": location,
                            "tank_asset_code": tank,
                            "product_name": product or None,
                            "expected_previous_nsv_bbl": round(
                                expected_previous_nsv,
                                3,
                            ),
                            "actual_previous_nsv_bbl": round(
                                actual_previous_nsv,
                                3,
                            ),
                            "message": (
                                "Previous stock does not match previous row "
                                "of the same tank/product sequence. Run ledger rebuild."
                            ),
                        }
                    )

            previous_row = row

    return {
        "rows_checked": len(visible_rows),
        "groups_checked": len(grouped_rows),
        "issues_count": len(issues),
        "issues": issues,
    }


@router.get(
    "/material-balance-report",
    response_model=MaterialBalanceDynamicReportResponse,
)
def get_material_balance_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    unit: str | None = "nsv",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Material Balance Report",
        db,
    )

    cleaned_location_code = clean_optional_text(location_code)

    if not cleaned_location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required for configurable Material Balance Report",
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for Material Balance Report",
        )

    unit_key = normalize_material_balance_code_value(unit).lower()

    if unit_key not in ["gsv", "nsv", "lt", "mt"]:
        raise HTTPException(
            status_code=400,
            detail="Unit must be one of: gsv, nsv, lt, mt",
        )

    template = get_active_material_balance_template_for_location(
        db=db,
        location_code=cleaned_location_code,
    )

    columns = get_active_material_balance_template_columns(
        db=db,
        template_id=template.id,
    )

    ledger_rows = get_material_balance_rows_for_continuity(
        db=db,
        location_code=cleaned_location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
    )

    tank_rows = build_dynamic_material_balance_tank_rows(
        db=db,
        ledger_rows=ledger_rows,
        columns=columns,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
        unit_key=unit_key,
    )

    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)

    if cleaned_tank_asset_code:
        report_rows = tank_rows
    else:
        report_rows = consolidate_dynamic_material_balance_rows_by_location(
            tank_rows=tank_rows,
            columns=columns,
        )

    return {
        "template": {
            "id": template.id,
            "location_code": template.location_code,
            "template_name": template.template_name,
        },
        "columns": build_dynamic_material_balance_columns_response(columns),
        "rows": report_rows,
    }


@router.get("/fso/reports/otr", response_model=FSOOTRReportResponse)
def get_fso_otr_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)
    return {
        "rows": rows,
        "total_receipt_bbl": totals["receipt"],
        "total_export_bbl": totals["export"],
        "total_movement_bbl": totals["movement"],
        "total_variance_bbl": totals["variance"],
        "total_compare_variance_bbl": totals["compare_variance"],
    }


@router.get("/fso/reports/material-balance", response_model=FSOMaterialBalanceReportResponse)
def get_fso_material_balance_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)
    return {"rows": rows}


@router.get("/fso/reports/outturn", response_model=FSOOutturnReportResponse)
def fso_report_outturn(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)
    return {
        "rows": rows,
        "total_shuttle_discharge_bbl": totals["discharge"],
        "total_fso_receipt_bbl": totals["receipt"],
        "total_variance_bbl": totals["variance"],
        "total_variance_pct": totals_pct,
    }


@router.get("/fso/reports/otr/export/xlsx")
def fso_report_otr_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO OTR"

    headers = [
        "Ticket", "Acc Date", "Op Date", "Time", "Operation", "Sign",
        "Shuttle", "Vessel", "Vessel Qty",
        "Open Stock", "Open Water", "Close Stock", "Close Water",
        "Net Stock", "Net Water", "Move Qty",
        "Variance", "Shuttle Discharge", "Compare Var", "Remarks",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["ticket_number"],
            str(r["accounting_date"]),
            str(r["operation_date"]),
            r.get("event_time") or "",
            r["operation_label"],
            r["operation_sign"],
            r.get("shuttle_number") or "",
            r.get("vessel_name") or "",
            round(float(r["vessel_quantity_bbl"]), 3),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["opening_water_bbl"]), 3),
            round(float(r["closing_stock_bbl"]), 3),
            round(float(r["closing_water_bbl"]), 3),
            round(float(r["net_stock_bbl"]), 3),
            round(float(r["net_water_bbl"]), 3),
            round(float(r["movement_qty_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["source_shuttle_discharge_bbl"]), 3),
            round(float(r["compare_variance_bbl"]), 3),
            r.get("remarks") or "",
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Receipt", totals["receipt"]])
    ws2.append(["Total Export", totals["export"]])
    ws2.append(["Total Movement", totals["movement"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Compare Variance", totals["compare_variance"]])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_otr_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/fso/reports/outturn/export/xlsx")
def fso_report_outturn_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Outturn"

    headers = ["Acc Date", "Shuttle Number", "Shuttle Discharge", "FSO Receipt", "Variance", "Variance %"]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            r["shuttle_number"],
            round(float(r["shuttle_discharge_bbl"]), 3),
            round(float(r["fso_receipt_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["variance_pct"]), 3),
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Shuttle Discharge", totals["discharge"]])
    ws2.append(["Total FSO Receipt", totals["receipt"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Variance %", totals_pct])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_outturn_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/fso/reports/material-balance/export/xlsx")
def fso_report_mb_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Material Balance"

    headers = [
        "Acc Date", "Opening", "Receipt", "Export",
        "Book Closing", "Physical Closing", "Closing Water", "Loss/Gain",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["receipt_bbl"]), 3),
            round(float(r["export_bbl"]), 3),
            round(float(r["book_closing_bbl"]), 3),
            round(float(r["physical_closing_bbl"]), 3),
            round(float(r["physical_closing_water_bbl"]), 3),
            round(float(r["loss_gain_bbl"]), 3),
        ])

    _xlsx_autofit(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_mb_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
