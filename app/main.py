from datetime import datetime, timedelta, date, time as datetime_time
from zoneinfo import ZoneInfo
from fastapi import Depends, FastAPI, HTTPException, Header
from fastapi.responses import StreamingResponse
import csv
import io
from jose import JWTError, jwt
import openpyxl
from openpyxl.utils import get_column_letter

from fastapi.middleware.cors import CORSMiddleware

allowed_origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "https://hydrocarbon-accounting-frontend.vercel.app",
]

app = FastAPI(
    title="Hydrocarbon Accounting API",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from passlib.context import CryptContext
from sqlalchemy import and_, func, inspect, literal, or_, text
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
from fastapi.encoders import jsonable_encoder
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.database import Base, engine, get_db
from app.models import (
    Asset,
    AssetAssignment,
    PrimeMoverTankerLink,
    AssetCalibrationData,
    AssetCalibrationTable,
    AssetType,
    CalibrationTemplate,
    CalibrationTemplateColumn,
    Location,
    LocationAccountingDaySetting,
    LocationOperationAvailability,
    OperationTemplate,
    OperationTemplateField,
    OperationTransaction,
    OperationTransactionValue,
    OperationTransactionStatusHistory,
    OperationType,
    TankOperation,
    VesselOperation,
    VesselStockLedger,
    MovementMapping,
    MovementMappingItem,
    MovementMappingComparison,
    TankStockLedger,
    TankerReceiptAcknowledgement,
    MaterialBalanceTemplate,
    MaterialBalanceTemplateColumn,
    Permission,
    Role,
    RolePermission,
    User,
    UserRole,
    Table11Factor,
    BargeSealMaster,
    CompanyReportProfile,
    DashboardConfig,
    DashboardVersion,
    DashboardDataSource,
    AuditLog,
    Trip,
    TripEvent,
    TripComparison,
    ShuttleVoyage,
    FSOVoyage,
)

from app.schemas import (
    AssetAssignmentCreate,
    AssetAssignmentResponse,
    PrimeMoverTankerLinkCreate,
    PrimeMoverTankerLinkResponse,
    CurrentPrimeMoverTankerLinkResponse,
    AssetCalibrationTableCreate,
    AssetCalibrationTableResponse,
    AssetCreate,
    AssetResponse,
    AssetTypeCreate,
    AssetTypeResponse,
    CalibrationTemplateCreate,
    CalibrationTemplateResponse,
    LocationCreate,
    LocationResponse,
    LocationAccountingDaySettingCreate,
    LocationAccountingDaySettingResponse,
    LocationOperationAvailabilityCreate,
    LocationOperationAvailabilityResponse,
    OperationEntryCreate,
    OperationEntryResponse,
    TankerTransactionReportResponse,
    TankerTrackingResponse,
    TankerTrackingTicketResponse,
    TankerReceiptAcknowledgementCreate,
    TankerReceiptAcknowledgementResponse,
    TankerTrackingClosureCreate,
    OperationTemplateCreate,
    OperationTemplateResponse,
    OperationTransactionCreate,
    OperationTransactionResponse,
    OperationTransactionRegisterPagedResponse,
    OperationTransactionStatusUpdate,
    OperationTypeCreate,
    OperationTypeResponse,
    TankOperationCreate,
    TankOperationResponse,
    VesselOperationCreate,
    VesselOperationResponse,
    VesselStockLedgerResponse,
    MovementMappingCreate,
    MovementMappingResponse,
    MovementMappingItemAddRequest,
    MovementMappingItemResponse,
    MovementMappingComparisonResponse,
    TankStockLedgerResponse,
    TankStockLedgerSummaryResponse,
    TankStockLedgerDailySummaryResponse,
    OutTurnReportResponse,
    MaterialBalanceReportResponse,
    MaterialBalanceDynamicReportResponse,
    FSOOTRReportResponse,
    FSOMaterialBalanceReportResponse,
    FSOOutturnReportResponse,
    MaterialBalanceTemplateCreate,
    MaterialBalanceTemplateUpdate,
    MaterialBalanceTemplateResponse,
    MaterialBalanceTemplateColumnCreate,
    MaterialBalanceTemplateColumnUpdate,
    MaterialBalanceTemplateColumnResponse,
    MaterialBalanceTemplateDetailResponse,
    PermissionCreate,
    PermissionResponse,
    RoleCreate,
    RolePermissionResponse,
    RolePermissionSaveRequest,
    RoleResponse,
    UserCreate,
    UserResponse,
    UserRoleResponse,
    UserRoleSaveRequest,
    UserUpdate,
    Table11FactorBulkCreate,
    Table11FactorCreate,
    Table11FactorResponse,
    Table11LookupResponse,
    BargeSealMasterBulkSaveRequest,
    BargeSealMasterResponse,
    CompanyReportProfileCreate,
    CompanyReportProfileResponse,
    DashboardConfigCreate,
    DashboardConfigUpdate,
    DashboardConfigResponse,
    DashboardVersionResponse,
    DashboardPublishRequest,
    DashboardRevertRequest,
    DashboardDataSourceCreate,
    DashboardDataSourceUpdate,
    DashboardDataSourceResponse,
    DashboardDataRequest,
    DashboardDataResponse,
    AuditLogResponse,
    TripCreate,
    TripResponse,
    TripEventCreate,
    TripEventResponse,
    TripComparisonCreate,
    TripComparisonResponse,
    ConvoyTrackerResponse,
    ShuttleTrackingResponse,
    ShuttleTrackingGroupResponse,
    ShuttleTrackingTicketResponse,
    FSOTrackingResponse,
    ShuttleVoyageCloseRequest,
    ShuttleVoyageReopenRequest,
    ShuttleVoyageResponse,
    FSOVoyageCloseRequest,
    FSOVoyageReopenRequest,
    FSOVoyageResponse,
)

password_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
JWT_SECRET_KEY = "hydrocarbon-development-secret-key-change-later"
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

class LoginRequest(BaseModel):
    username: str
    password: str


def hash_password(password: str):
    return password_context.hash(password)


def verify_password(plain_password: str, hashed_password: str):
    return password_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    token_data = data.copy()

    if expires_delta is None:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    else:
        expire = datetime.utcnow() + expires_delta

    token_data.update({"exp": expire})

    encoded_jwt = jwt.encode(
        token_data,
        JWT_SECRET_KEY,
        algorithm=JWT_ALGORITHM,
    )

    return encoded_jwt


def decode_access_token(token: str):
    try:
        payload = jwt.decode(
            token,
            JWT_SECRET_KEY,
            algorithms=[JWT_ALGORITHM],
        )

        return payload
    except JWTError:
        raise HTTPException(
            status_code=401,
            detail="Invalid or expired token",
        )

def get_current_user_from_token(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    if authorization is None:
        raise HTTPException(
            status_code=401,
            detail="Authorization header is missing",
        )

    if not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail="Invalid authorization header",
        )

    token = authorization.replace("Bearer ", "").strip()

    payload = decode_access_token(token)

    user_id = payload.get("user_id")

    if user_id is None:
        raise HTTPException(
            status_code=401,
            detail="Invalid token payload",
        )

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(
            status_code=401,
            detail="User not found",
        )

    if user.status != "Active":
        raise HTTPException(
            status_code=403,
            detail="User is not Active",
        )

    return user

def user_has_permission(
    user: User,
    permission_name: str,
    db: Session,
):
    user_role = (
        db.query(UserRole)
        .filter(UserRole.user_id == user.id)
        .first()
    )

    if not user_role:
        return False

    permission = (
        db.query(Permission)
        .join(RolePermission, RolePermission.permission_id == Permission.id)
        .filter(
            RolePermission.role_id == user_role.role_id,
            Permission.permission_name == permission_name,
            Permission.status == "Active",
        )
        .first()
    )

    return permission is not None


def require_user_permission(
    user: User,
    permission_name: str,
    db: Session,
):
    # --- Admin bypass (bootstrap) ---
    admin_role_names = {"admin"}  # compare lowercase
    user_role_names = {
        str(r.role_name or "").lower()
        for r in (
            db.query(Role)
            .join(UserRole, UserRole.role_id == Role.id)
            .filter(UserRole.user_id == user.id)
            .all()
        )
        if str(r.role_name or "").strip() != ""
    }

    if user_role_names.intersection(admin_role_names):
        return user
    # --- end Admin bypass ---

    if not user_has_permission(user, permission_name, db):
        raise HTTPException(
            status_code=403,
            detail=f"Permission required: {permission_name}",
        )

def get_required_permission_for_status_change(next_status: str):
    status_permission_map = {
        "Draft": "Submit Operation Transaction",
        "Submitted": "Submit Operation Transaction",
        "Approved": "Approve Operation Transaction",
        "Rejected": "Reject Operation Transaction",
        "Cancelled": "Cancel Operation Transaction",
    }

    return status_permission_map.get(next_status)

def build_logged_in_user_response(user: User, db: Session):
    user_role_assignment = (
        db.query(UserRole, Role)
        .join(Role, Role.id == UserRole.role_id)
        .filter(UserRole.user_id == user.id)
        .first()
    )

    role_data = None
    permissions_data = []

    if user_role_assignment:
        user_role, role = user_role_assignment

        role_data = {
            "id": role.id,
            "role_name": role.role_name,
            "description": role.description,
            "status": role.status,
        }

        permissions = (
            db.query(Permission)
            .join(RolePermission, RolePermission.permission_id == Permission.id)
            .filter(RolePermission.role_id == role.id)
            .order_by(Permission.module_name, Permission.permission_name)
            .all()
        )

        permissions_data = [
            {
                "id": permission.id,
                "permission_name": permission.permission_name,
                "module_name": permission.module_name,
                "description": permission.description,
                "status": permission.status,
            }
            for permission in permissions
        ]

    return {
        "id": user.id,
        "full_name": user.full_name,
        "username": user.username,
        "email": user.email,
        "phone": user.phone,
        "department": user.department,
        "designation": user.designation,
        "status": user.status,
        "role": role_data,
        "permissions": permissions_data,
    }


def clean_optional_text(value):
    if value is None:
        return None

    cleaned_value = str(value).strip()

    if cleaned_value == "":
        return None

    return cleaned_value


APPROVED_TRANSACTION_STATUS = "Approved"


def require_approved_transaction_for_tracking(
    transaction: OperationTransaction | None,
    action_label: str = "tracking",
):
    if not transaction:
        raise HTTPException(
            status_code=404,
            detail="Operation transaction not found",
        )

    if transaction.status != APPROVED_TRANSACTION_STATUS:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Only Approved transactions can be used for {action_label}. "
                f"Current status is {transaction.status}."
            ),
        )


def is_barge_transaction(transaction: OperationTransaction | None):
    if not transaction:
        return False

    return str(transaction.primary_asset_type_code or "").strip().upper() == "BARGE"


def resolve_barge_event_type(db: Session, transaction: OperationTransaction):
    code = str(transaction.operation_type_code or "").strip()
    code_upper = code.upper()

    operation_type_name = ""
    if code:
        op = (
            db.query(OperationType)
            .filter(OperationType.operation_type_code.ilike(code))
            .first()
        )
        if op and op.operation_type_name:
            operation_type_name = str(op.operation_type_name)

    text = f"{code_upper} {operation_type_name}".upper()

    if (
        "UNLOAD" in text
        or "DISCHARGE" in text
        or "RECEIPT" in text
        or "RECEIVE" in text
    ):
        return "UNLOAD"

    return "LOAD"

def ensure_operation_ticket_number_column():
    with engine.connect() as connection:
        connection.execute(
            text(
                """
                ALTER TABLE operation_transactions
                ADD COLUMN IF NOT EXISTS operation_ticket_number VARCHAR(100);
                """
            )
        )

        connection.execute(
            text(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS
                ix_operation_transactions_operation_ticket_number
                ON operation_transactions(operation_ticket_number)
                WHERE operation_ticket_number IS NOT NULL;
                """
            )
        )

        connection.commit()


def ensure_operation_template_layout_columns():
    with engine.connect() as connection:
        connection.execute(
            text(
                """
                ALTER TABLE operation_templates
                ADD COLUMN IF NOT EXISTS entry_layout_type VARCHAR(80);
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE operation_templates
                ADD COLUMN IF NOT EXISTS calculation_engine VARCHAR(100);
                """
            )
        )

        connection.execute(
            text(
                """
                UPDATE operation_templates
                SET entry_layout_type = 'Standard Form'
                WHERE entry_layout_type IS NULL OR TRIM(entry_layout_type) = '';
                """
            )
        )

        connection.execute(
            text(
                """
                UPDATE operation_templates
                SET calculation_engine = 'None'
                WHERE calculation_engine IS NULL OR TRIM(calculation_engine) = '';
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE operation_templates
                ALTER COLUMN entry_layout_type SET DEFAULT 'Standard Form';
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE operation_templates
                ALTER COLUMN calculation_engine SET DEFAULT 'None';
                """
            )
        )

        connection.commit()


def ensure_tank_stock_ledger_accounting_columns():
    with engine.connect() as connection:
        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS accounting_date DATE;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS accounting_day_start TIMESTAMP;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS accounting_day_end TIMESTAMP;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS accounting_day_setting_id INTEGER;
                """
            )
        )

        connection.commit()

def ensure_tank_stock_ledger_stock_snapshot_columns():
    with engine.connect() as connection:
        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS stock_gsv_bbl DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS stock_nsv_bbl DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS stock_lt DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS stock_mt DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS previous_stock_gsv_bbl DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS previous_stock_nsv_bbl DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS previous_stock_lt DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.execute(
            text(
                """
                ALTER TABLE tank_stock_ledger
                ADD COLUMN IF NOT EXISTS previous_stock_mt DOUBLE PRECISION DEFAULT 0;
                """
            )
        )

        connection.commit()


def ensure_tanker_acknowledgement_closure_columns():
    inspector = inspect(engine)

    if "tanker_receipt_acknowledgements" not in inspector.get_table_names():
        return

    columns = {
        column["name"]
        for column in inspector.get_columns("tanker_receipt_acknowledgements")
    }

    with engine.begin() as connection:
        if "closed_by" not in columns:
            connection.execute(
                text(
                    "ALTER TABLE tanker_receipt_acknowledgements "
                    "ADD COLUMN closed_by VARCHAR(150)"
                )
            )

        if "closed_at" not in columns:
            connection.execute(
                text(
                    "ALTER TABLE tanker_receipt_acknowledgements "
                    "ADD COLUMN closed_at TIMESTAMP"
                )
            )

        if "closure_remarks" not in columns:
            connection.execute(
                text(
                    "ALTER TABLE tanker_receipt_acknowledgements "
                    "ADD COLUMN closure_remarks TEXT"
                )
            )

def ensure_vessel_operation_show_in_column():
    inspector = inspect(engine)

    if "vessel_operations" not in inspector.get_table_names():
        return

    cols = {c["name"] for c in inspector.get_columns("vessel_operations")}
    if "show_in" in cols:
        return

    with engine.begin() as conn:
        conn.execute(
            text(
                "ALTER TABLE vessel_operations ADD COLUMN show_in VARCHAR(20) DEFAULT 'Both';"
            )
        )
        conn.execute(
            text(
                "UPDATE vessel_operations SET show_in = 'Both' "
                "WHERE show_in IS NULL OR TRIM(show_in) = '';"
            )
        )

def ensure_barge_event_type_template_field():
    """
    Ensures Operation Templates used for BARGE Multi-Tank Before/After have
    a manual field_code = 'barge_event_type' to store:
    LOAD_1 / LOAD_2_TOPUP / UNLOAD / STS
    """
    db = Session(bind=engine)
    try:
        barge_op_codes = [
            x.operation_type_code
            for x in db.query(OperationType)
            .filter(
                OperationType.applicable_asset_type_code == "BARGE",
                OperationType.status == "Active",
            )
            .all()
        ]

        if not barge_op_codes:
            return

        barge_templates = (
            db.query(OperationTemplate)
            .filter(
                OperationTemplate.entry_layout_type == "Multi-Tank Before/After",
                OperationTemplate.status == "Active",
                OperationTemplate.operation_type_code.in_(barge_op_codes),
            )
            .all()
        )

        for t in barge_templates:
            exists = (
                db.query(OperationTemplateField)
                .filter(
                    OperationTemplateField.template_id == t.id,
                    OperationTemplateField.field_code == "barge_event_type",
                )
                .first()
            )

            if exists:
                continue

            db.add(
                OperationTemplateField(
                    template_id=t.id,
                    field_name="Barge Movement Stage",
                    field_code="barge_event_type",
                    field_group="Barge Tracking",
                    data_type="Text",
                    unit=None,
                    is_required="No",
                    input_mode="Manual",
                    calculation_role="Input",
                    sort_order=0,
                    status="Active",
                )
            )

        db.commit()
    finally:
        db.close()

Base.metadata.create_all(bind=engine)
ensure_operation_ticket_number_column()
ensure_operation_template_layout_columns()
ensure_tank_stock_ledger_accounting_columns()
ensure_tank_stock_ledger_stock_snapshot_columns()
ensure_vessel_operation_show_in_column()
ensure_barge_event_type_template_field()


@app.get("/")
def root():
    return {
        "message": "Hydrocarbon Accounting API is running"
    }


@app.get("/health")
def health_check():
    return {
        "status": "ok"
    }


@app.get("/db-test")
def database_test(db: Session = Depends(get_db)):
    db.execute(text("SELECT 1"))

    return {
        "database": "connected"
    }

# -------------------------
# Authentication APIs
# -------------------------

@app.post("/auth/login")
def login_user(
    login_request: LoginRequest,
    db: Session = Depends(get_db),
):
    username = login_request.username.strip()

    if username == "":
        raise HTTPException(
            status_code=400,
            detail="Username is required",
        )

    if login_request.password.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Password is required",
        )

    user = (
        db.query(User)
        .filter(User.username.ilike(username))
        .first()
    )

    if not user:
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password",
        )

    if user.status != "Active":
        raise HTTPException(
            status_code=403,
            detail="User is not Active",
        )

    if not verify_password(login_request.password, user.password_hash):
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password",
        )

    logged_in_user = build_logged_in_user_response(user, db)

    access_token = create_access_token(
        data={
            "user_id": user.id,
            "username": user.username,
        }
    )

    return {
        "message": "Login successful",
        "access_token": access_token,
        "token_type": "bearer",
        "user": logged_in_user,
        "role": logged_in_user["role"],
        "permissions": logged_in_user["permissions"],
    }

@app.get("/auth/me")
def get_logged_in_user(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    logged_in_user = build_logged_in_user_response(current_user, db)

    return {
        "user": logged_in_user,
        "role": logged_in_user["role"],
        "permissions": logged_in_user["permissions"],
    }


class DevResetPasswordRequest(BaseModel):
    username: str
    new_password: str


@app.post("/auth/dev-reset-password")
def dev_reset_password(
    reset_request: DevResetPasswordRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    # Dev endpoint: restrict it
    require_user_permission(
        current_user,
        "Manage User",
        db,
    )

    username = reset_request.username.strip()

    if username == "":
        raise HTTPException(
            status_code=400,
            detail="Username is required",
        )

    if reset_request.new_password.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="New password is required",
        )

    user = (
        db.query(User)
        .filter(User.username.ilike(username))
        .first()
    )

    if not user:
        raise HTTPException(
            status_code=404,
            detail="User not found",
        )

    # Snapshot (do NOT log password)
    before_data = {
        "user_id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "status": user.status,
    }

    user.password_hash = hash_password(reset_request.new_password)
    db.flush()

    create_audit_log(
        db=db,
        module_name="User Master",
        action="Dev Reset Password",
        current_user=current_user,
        entity_type="User",
        entity_id=user.id,
        entity_label=f"{user.full_name} ({user.username})",
        remarks="Password reset via dev endpoint",
        request_path="/auth/dev-reset-password",
        details={
            "target_user": before_data,
            "password_reset": True,
        },
    )

    db.commit()

    return {
        "message": "Password reset successfully",
        "username": user.username,
    }


# -------------------------
# User APIs
# -------------------------

@app.get("/users", response_model=list[UserResponse])
def get_users(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View User",
        db,
    )

    users = db.query(User).order_by(User.id).all()
    return users


@app.post("/users", response_model=UserResponse)
def create_user(
    user: UserCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage User",
        db,
    )

    existing_user = (
        db.query(User)
        .filter(User.username.ilike(user.username))
        .first()
    )

    if existing_user:
        raise HTTPException(
            status_code=400,
            detail="Username already exists",
        )

    if user.password.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Password is required",
        )

    new_user = User(
        full_name=user.full_name.strip(),
        username=user.username.strip(),
        email=user.email.strip(),
        phone=clean_optional_text(user.phone),
        department=clean_optional_text(user.department),
        designation=clean_optional_text(user.designation),
        password_hash=hash_password(user.password),
        status=user.status,
    )

    db.add(new_user)
    db.flush()  # get id for audit log

    after_data = {
        "full_name": new_user.full_name,
        "username": new_user.username,
        "email": new_user.email,
        "phone": new_user.phone,
        "department": new_user.department,
        "designation": new_user.designation,
        "status": new_user.status,
    }

    create_audit_log(
        db=db,
        module_name="User Master",
        action="Create User",
        current_user=current_user,
        entity_type="User",
        entity_id=new_user.id,
        entity_label=f"{new_user.full_name} ({new_user.username})",
        remarks="User created",
        request_path="/users",
        details={
            "after": after_data,
            "password_set": True,
        },
    )

    db.commit()
    db.refresh(new_user)

    return new_user


@app.put("/users/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    user: UserUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage User",
        db,
    )

    existing_user = db.query(User).filter(User.id == user_id).first()

    if not existing_user:
        raise HTTPException(
            status_code=404,
            detail="User not found",
        )

    duplicate_user = (
        db.query(User)
        .filter(
            User.username.ilike(user.username),
            User.id != user_id,
        )
        .first()
    )

    if duplicate_user:
        raise HTTPException(
            status_code=400,
            detail="Username already exists",
        )

    before_data = {
        "full_name": existing_user.full_name,
        "username": existing_user.username,
        "email": existing_user.email,
        "phone": existing_user.phone,
        "department": existing_user.department,
        "designation": existing_user.designation,
        "status": existing_user.status,
    }

    password_changed = False

    existing_user.full_name = user.full_name.strip()
    existing_user.username = user.username.strip()
    existing_user.email = user.email.strip()
    existing_user.phone = clean_optional_text(user.phone)
    existing_user.department = clean_optional_text(user.department)
    existing_user.designation = clean_optional_text(user.designation)
    existing_user.status = user.status

    if user.password is not None and user.password.strip() != "":
        existing_user.password_hash = hash_password(user.password)
        password_changed = True

    after_data = {
        "full_name": existing_user.full_name,
        "username": existing_user.username,
        "email": existing_user.email,
        "phone": existing_user.phone,
        "department": existing_user.department,
        "designation": existing_user.designation,
        "status": existing_user.status,
    }

    create_audit_log(
        db=db,
        module_name="User Master",
        action="Update User",
        current_user=current_user,
        entity_type="User",
        entity_id=existing_user.id,
        entity_label=f"{existing_user.full_name} ({existing_user.username})",
        remarks="User updated",
        request_path=f"/users/{user_id}",
        details={
            "before": before_data,
            "after": after_data,
            "password_changed": password_changed,
        },
    )

    db.commit()
    db.refresh(existing_user)

    return existing_user


@app.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage User",
        db,
    )

    existing_user = db.query(User).filter(User.id == user_id).first()

    if not existing_user:
        raise HTTPException(
            status_code=404,
            detail="User not found",
        )

    if existing_user.id == current_user.id:
        raise HTTPException(
            status_code=400,
            detail="You cannot delete your own logged-in user account",
        )

    assigned_role = (
        db.query(UserRole)
        .filter(UserRole.user_id == user_id)
        .first()
    )

    if assigned_role:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete user because a role is assigned. Remove user role assignment first.",
        )

    deleted_data = {
        "full_name": existing_user.full_name,
        "username": existing_user.username,
        "email": existing_user.email,
        "phone": existing_user.phone,
        "department": existing_user.department,
        "designation": existing_user.designation,
        "status": existing_user.status,
    }

    create_audit_log(
        db=db,
        module_name="User Master",
        action="Delete User",
        current_user=current_user,
        entity_type="User",
        entity_id=existing_user.id,
        entity_label=f"{existing_user.full_name} ({existing_user.username})",
        remarks="User deleted",
        request_path=f"/users/{user_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.delete(existing_user)
    db.commit()

    return {
        "message": "User deleted successfully"
    }


# -------------------------
# Role APIs
# -------------------------

@app.get("/roles", response_model=list[RoleResponse])
def get_roles(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Role",
        db,
    )

    roles = db.query(Role).order_by(Role.id).all()
    return roles


@app.post("/roles", response_model=RoleResponse)
def create_role(
    role: RoleCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Role",
        db,
    )

    existing_role = (
        db.query(Role)
        .filter(Role.role_name.ilike(role.role_name))
        .first()
    )

    if existing_role:
        raise HTTPException(
            status_code=400,
            detail="Role name already exists",
        )

    new_role = Role(
        role_name=role.role_name.strip(),
        description=clean_optional_text(role.description),
        status=role.status,
    )

    db.add(new_role)
    db.flush()  # get new_role.id before audit log

    role_data = {
        "role_name": new_role.role_name,
        "description": new_role.description,
        "status": new_role.status,
    }

    create_audit_log(
        db=db,
        module_name="Role Master",
        action="Create Role",
        current_user=current_user,
        entity_type="Role",
        entity_id=new_role.id,
        entity_label=new_role.role_name,
        remarks="Role created",
        request_path="/roles",
        details={
            "after": role_data,
        },
    )

    db.commit()
    db.refresh(new_role)

    return new_role


@app.put("/roles/{role_id}", response_model=RoleResponse)
def update_role(
    role_id: int,
    role: RoleCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Role",
        db,
    )

    existing_role = db.query(Role).filter(Role.id == role_id).first()

    if not existing_role:
        raise HTTPException(
            status_code=404,
            detail="Role not found",
        )

    duplicate_role = (
        db.query(Role)
        .filter(
            Role.role_name.ilike(role.role_name),
            Role.id != role_id,
        )
        .first()
    )

    if duplicate_role:
        raise HTTPException(
            status_code=400,
            detail="Role name already exists",
        )

    old_role_data = {
        "role_name": existing_role.role_name,
        "description": existing_role.description,
        "status": existing_role.status,
    }

    existing_role.role_name = role.role_name.strip()
    existing_role.description = clean_optional_text(role.description)
    existing_role.status = role.status

    new_role_data = {
        "role_name": existing_role.role_name,
        "description": existing_role.description,
        "status": existing_role.status,
    }

    create_audit_log(
        db=db,
        module_name="Role Master",
        action="Update Role",
        current_user=current_user,
        entity_type="Role",
        entity_id=existing_role.id,
        entity_label=existing_role.role_name,
        remarks="Role updated",
        request_path=f"/roles/{role_id}",
        details={
            "before": old_role_data,
            "after": new_role_data,
        },
    )

    db.commit()
    db.refresh(existing_role)

    return existing_role


@app.delete("/roles/{role_id}")
def delete_role(
    role_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Role",
        db,
    )

    existing_role = db.query(Role).filter(Role.id == role_id).first()

    if not existing_role:
        raise HTTPException(
            status_code=404,
            detail="Role not found",
        )

    user_role = db.query(UserRole).filter(UserRole.role_id == role_id).first()

    if user_role:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete role because it is assigned to users",
        )

    role_permission = (
        db.query(RolePermission)
        .filter(RolePermission.role_id == role_id)
        .first()
    )

    if role_permission:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete role because permissions are assigned to it",
        )

    deleted_role_data = {
        "role_name": existing_role.role_name,
        "description": existing_role.description,
        "status": existing_role.status,
    }

    create_audit_log(
        db=db,
        module_name="Role Master",
        action="Delete Role",
        current_user=current_user,
        entity_type="Role",
        entity_id=existing_role.id,
        entity_label=existing_role.role_name,
        remarks="Role deleted",
        request_path=f"/roles/{role_id}",
        details={
            "deleted": deleted_role_data,
        },
    )

    db.delete(existing_role)
    db.commit()

    return {
        "message": "Role deleted successfully"
    }

# -------------------------
# Permission APIs
# -------------------------

@app.get("/permissions", response_model=list[PermissionResponse])
def get_permissions(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Permission",
        db,
    )

    permissions = db.query(Permission).order_by(Permission.id).all()
    return permissions


@app.post("/permissions", response_model=PermissionResponse)
def create_permission(
    permission: PermissionCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Permission",
        db,
    )

    existing_permission = (
        db.query(Permission)
        .filter(
            Permission.permission_name.ilike(permission.permission_name),
            Permission.module_name.ilike(permission.module_name),
        )
        .first()
    )

    if existing_permission:
        raise HTTPException(
            status_code=400,
            detail="Permission already exists for this module",
        )

    new_permission = Permission(
        permission_name=permission.permission_name.strip(),
        module_name=permission.module_name.strip(),
        description=clean_optional_text(permission.description),
        status=permission.status,
    )

    db.add(new_permission)
    db.flush()  # get new_permission.id before audit log

    after_data = {
        "permission_name": new_permission.permission_name,
        "module_name": new_permission.module_name,
        "description": new_permission.description,
        "status": new_permission.status,
    }

    create_audit_log(
        db=db,
        module_name="Permission Master",
        action="Create Permission",
        current_user=current_user,
        entity_type="Permission",
        entity_id=new_permission.id,
        entity_label=f"{new_permission.module_name} - {new_permission.permission_name}",
        remarks="Permission created",
        request_path="/permissions",
        details={
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(new_permission)

    return new_permission

@app.put("/permissions/{permission_id}", response_model=PermissionResponse)
def update_permission(
    permission_id: int,
    permission: PermissionCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Permission",
        db,
    )

    existing_permission = (
        db.query(Permission)
        .filter(Permission.id == permission_id)
        .first()
    )

    if not existing_permission:
        raise HTTPException(
            status_code=404,
            detail="Permission not found",
        )

    duplicate_permission = (
        db.query(Permission)
        .filter(
            Permission.permission_name.ilike(permission.permission_name),
            Permission.module_name.ilike(permission.module_name),
            Permission.id != permission_id,
        )
        .first()
    )

    if duplicate_permission:
        raise HTTPException(
            status_code=400,
            detail="Permission already exists for this module",
        )

    before_data = {
        "permission_name": existing_permission.permission_name,
        "module_name": existing_permission.module_name,
        "description": existing_permission.description,
        "status": existing_permission.status,
    }

    existing_permission.permission_name = permission.permission_name.strip()
    existing_permission.module_name = permission.module_name.strip()
    existing_permission.description = clean_optional_text(permission.description)
    existing_permission.status = permission.status

    after_data = {
        "permission_name": existing_permission.permission_name,
        "module_name": existing_permission.module_name,
        "description": existing_permission.description,
        "status": existing_permission.status,
    }

    create_audit_log(
        db=db,
        module_name="Permission Master",
        action="Update Permission",
        current_user=current_user,
        entity_type="Permission",
        entity_id=existing_permission.id,
        entity_label=f"{existing_permission.module_name} - {existing_permission.permission_name}",
        remarks="Permission updated",
        request_path=f"/permissions/{permission_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_permission)

    return existing_permission


@app.delete("/permissions/{permission_id}")
def delete_permission(
    permission_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Permission",
        db,
    )

    existing_permission = (
        db.query(Permission)
        .filter(Permission.id == permission_id)
        .first()
    )

    if not existing_permission:
        raise HTTPException(
            status_code=404,
            detail="Permission not found",
        )

    role_permission = (
        db.query(RolePermission)
        .filter(RolePermission.permission_id == permission_id)
        .first()
    )

    if role_permission:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete permission because it is assigned to roles",
        )

    deleted_data = {
        "permission_name": existing_permission.permission_name,
        "module_name": existing_permission.module_name,
        "description": existing_permission.description,
        "status": existing_permission.status,
    }

    create_audit_log(
        db=db,
        module_name="Permission Master",
        action="Delete Permission",
        current_user=current_user,
        entity_type="Permission",
        entity_id=existing_permission.id,
        entity_label=f"{existing_permission.module_name} - {existing_permission.permission_name}",
        remarks="Permission deleted",
        request_path=f"/permissions/{permission_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.delete(existing_permission)
    db.commit()

    return {
        "message": "Permission deleted successfully"
    }

# -------------------------
# Permission Seed API
# -------------------------

@app.post("/permissions/seed-standard")
def seed_standard_permissions(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Permission",
        db,
    )

    standard_permissions = [
        # (keep your existing list unchanged)
        # User Management
        {
            "permission_name": "View User",
            "module_name": "User Master",
            "description": "Can view users",
        },
        {
            "permission_name": "Manage User",
            "module_name": "User Master",
            "description": "Can create, update, and delete users",
        },
        {
            "permission_name": "View Role",
            "module_name": "Role Master",
            "description": "Can view roles",
        },
        {
            "permission_name": "Manage Role",
            "module_name": "Role Master",
            "description": "Can create, update, and delete roles",
        },
        {
            "permission_name": "View Permission",
            "module_name": "Permission Master",
            "description": "Can view permissions",
        },
        {
            "permission_name": "Manage Permission",
            "module_name": "Permission Master",
            "description": "Can create, update, and delete permissions",
        },
        {
            "permission_name": "View Role Permission Assignment",
            "module_name": "Role Permission Assignment",
            "description": "Can view role permission assignments",
        },
        {
            "permission_name": "Manage Role Permission Assignment",
            "module_name": "Role Permission Assignment",
            "description": "Can assign permissions to roles",
        },
        {
            "permission_name": "View User Role Assignment",
            "module_name": "User Role Assignment",
            "description": "Can view user role assignments",
        },
        {
            "permission_name": "Manage User Role Assignment",
            "module_name": "User Role Assignment",
            "description": "Can assign roles to users",
        },
        {
            "permission_name": "View Access Summary",
            "module_name": "Access Summary",
            "description": "Can view final RBAC access summary",
        },
        {
            "permission_name": "View Dashboard",
            "module_name": "Dashboard",
            "description": "Can view dashboard configurations",
        },
        {
            "permission_name": "Manage Dashboard",
            "module_name": "Dashboard",
            "description": "Can create, update, publish, and revert dashboards",
        },

        # Master Data
        {
            "permission_name": "View Location",
            "module_name": "Location Master",
            "description": "Can view locations",
        },
        {
            "permission_name": "Manage Location",
            "module_name": "Location Master",
            "description": "Can create, update, and delete locations",
        },
        {
            "permission_name": "View Location Accounting Day Setting",
            "module_name": "Location Accounting Day Setting",
            "description": "Can view location-wise accounting day settings",
        },
        {
            "permission_name": "Manage Location Accounting Day Setting",
            "module_name": "Location Accounting Day Setting",
            "description": "Can create, update, and delete location-wise accounting day settings",
        },
        {
            "permission_name": "View Asset Type",
            "module_name": "Asset Type Master",
            "description": "Can view asset types",
        },
        {
            "permission_name": "Manage Asset Type",
            "module_name": "Asset Type Master",
            "description": "Can create, update, and delete asset types",
        },
        {
            "permission_name": "View Asset",
            "module_name": "Asset Master",
            "description": "Can view assets",
        },
        {
            "permission_name": "Manage Asset",
            "module_name": "Asset Master",
            "description": "Can create, update, and delete assets",
        },
        {
            "permission_name": "View Calibration Template",
            "module_name": "Calibration Template Master",
            "description": "Can view calibration templates",
        },
        {
            "permission_name": "Manage Calibration Template",
            "module_name": "Calibration Template Master",
            "description": "Can create, update, and delete calibration templates",
        },
        {
            "permission_name": "View Asset Calibration",
            "module_name": "Asset Calibration Table",
            "description": "Can view asset calibration tables",
        },
        {
            "permission_name": "Manage Asset Calibration",
            "module_name": "Asset Calibration Table",
            "description": "Can create, upload, update, and delete calibration data",
        },
        {
            "permission_name": "View Asset Assignment",
            "module_name": "Asset Assignment",
            "description": "Can view asset assignments",
        },
        {
            "permission_name": "Manage Asset Assignment",
            "module_name": "Asset Assignment",
            "description": "Can create, update, and delete asset assignments",
        },
        {
            "permission_name": "View Asset Assignment Summary",
            "module_name": "Asset Assignment Summary",
            "description": "Can view asset assignment summary",
        },

        # Operations
        {
            "permission_name": "View Operation Type",
            "module_name": "Operations",
            "description": "Can view operation type master",
        },
        {
            "permission_name": "Manage Operation Type",
            "module_name": "Operations",
            "description": "Can create, update, and delete operation types",
        },
        {
            "permission_name": "View Tank Operation",
            "module_name": "Operations",
            "description": "Can view location-wise tank operation master",
        },
        {
            "permission_name": "Manage Tank Operation",
            "module_name": "Operations",
            "description": "Can create, update, and delete location-wise tank operations",
        },
        # Vessel / FSO soft-coded operations
        {
            "permission_name": "View Vessel Operation",
            "module_name": "Operations",
            "description": "Can view Vessel Operation Master (soft-coded Loading/Unloading/STS/Decanting etc.)",
        },
        {
            "permission_name": "Manage Vessel Operation",
            "module_name": "Operations",
            "description": "Can create, update, and delete Vessel Operation Master entries",
        },

        # Vessel Stock Ledger (Shuttle / FSO)
        {
            "permission_name": "View Vessel Stock Ledger",
            "module_name": "Operations",
            "description": "Can view Vessel Stock Ledger (approved-only derived ledger for Shuttle/FSO)",
        },

        # Movement Mapping (Barge ↔ Shuttle ↔ FSO reconciliation)
        {
            "permission_name": "View Movement Mapping",
            "module_name": "Operations",
            "description": "Can view Movement Mapping and reconciliation comparisons",
        },
        {
            "permission_name": "Manage Movement Mapping",
            "module_name": "Operations",
            "description": "Can create/update/close Movement Mapping and attach approved tickets for reconciliation",
        },
        {
            "permission_name": "View Shuttle Tracking",
            "module_name": "Operations",
            "description": "Can view Shuttle Tracking (Approved-only voyage tracking within location)",
        },
        {
            "permission_name": "Manage Shuttle Tracking",
            "module_name": "Operations",
            "description": "Can close/reopen Shuttle voyages",
        },
        {
            "permission_name": "View FSO Tracking",
            "module_name": "Operations",
            "description": "Can view FSO Tracking (Approved-only, shuttle-number based)",
        },
        {
            "permission_name": "Manage FSO Tracking",
            "module_name": "Operations",
            "description": "Can close/reopen FSO voyages",
        },
        {
            "permission_name": "View Tank Stock Ledger",
            "module_name": "Operations",
            "description": "Can view tank stock ledger and stock movement summary",
        },
        {
            "permission_name": "Manage Tank Stock Ledger",
            "module_name": "Operations",
            "description": "Can rebuild or manage tank stock ledger entries",
        },
        {
            "permission_name": "View Location Operation Availability",
            "module_name": "Operations",
            "description": "Can view location operation availability",
        },
        {
            "permission_name": "Manage Location Operation Availability",
            "module_name": "Operations",
            "description": "Can configure operation availability by location",
        },
        {
            "permission_name": "View Operation Template",
            "module_name": "Operations",
            "description": "Can view operation templates",
        },
        {
            "permission_name": "Manage Operation Template",
            "module_name": "Operations",
            "description": "Can create, update, and delete operation templates",
        },
        {
            "permission_name": "Create Operation Entry",
            "module_name": "Operations",
            "description": "Can create new operation tickets from Operation Entry",
        },
        {
            "permission_name": "View Operation Transaction",
            "module_name": "Operations",
            "description": "Can view operation transaction register and detail",
        },
        {
            "permission_name": "Submit Operation Transaction",
            "module_name": "Operations",
            "description": "Can submit draft operation tickets",
        },
        {
            "permission_name": "Approve Operation Transaction",
            "module_name": "Operations",
            "description": "Can approve submitted operation tickets",
        },
        {
            "permission_name": "Reject Operation Transaction",
            "module_name": "Operations",
            "description": "Can reject submitted operation tickets",
        },
        {
            "permission_name": "Cancel Operation Transaction",
            "module_name": "Operations",
            "description": "Can cancel draft or rejected operation tickets",
        },
        # Barge Seal Master
        {
            "permission_name": "View Barge Seal Master",
            "module_name": "Barge Seal Master",
            "description": "Can view barge seal master",
        },
        {
            "permission_name": "Manage Barge Seal Master",
            "module_name": "Barge Seal Master",
            "description": "Can create/update barge seal master",
        },
        # Company / Report Profiles
        {
            "permission_name": "View Company Report Profile",
            "module_name": "Company Report Profile",
            "description": "Can view company report profiles used for printable reports",
        },
        {
            "permission_name": "Manage Company Report Profile",
            "module_name": "Company Report Profile",
            "description": "Can create, update, and delete company report profiles",
        },

        # Audit Logs
        {
            "permission_name": "View Audit Log",
            "module_name": "Audit Log",
            "description": "Can view system audit logs",
        },

        # Reports / Admin Future
        {
            "permission_name": "View Reports",
            "module_name": "Reports",
            "description": "Can view reports",
        },
        {
            "permission_name": "Export Reports",
            "module_name": "Reports",
            "description": "Can export reports",
        },
        {
            "permission_name": "View Admin Settings",
            "module_name": "Admin",
            "description": "Can view admin settings",
        },
        {
            "permission_name": "Manage Admin Settings",
            "module_name": "Admin",
            "description": "Can manage admin settings",
        },
        {
            "permission_name": "View Out-Turn Report",
            "module_name": "Reports",
            "description": "Can view Out-Turn Report from approved tank stock ledger rows",
        },
        {
            "permission_name": "View Material Balance Report",
            "module_name": "Reports",
            "description": "Can view Material Balance Report from tank stock ledger",
        },
        {
            "permission_name": "View Material Balance Template",
            "module_name": "Configuration",
            "description": "Can view Material Balance template configuration",
        },
        {
            "permission_name": "Manage Material Balance Template",
            "module_name": "Configuration",
            "description": "Can create, edit, and delete Material Balance template configuration",
        },
    ]

    created_count = 0
    existing_count = 0

    for permission_data in standard_permissions:
        existing_permission = (
            db.query(Permission)
            .filter(
                Permission.permission_name.ilike(permission_data["permission_name"]),
                Permission.module_name.ilike(permission_data["module_name"]),
            )
            .first()
        )

        if existing_permission:
            existing_count += 1
            continue

        new_permission = Permission(
            permission_name=permission_data["permission_name"],
            module_name=permission_data["module_name"],
            description=permission_data["description"],
            status="Active",
        )

        db.add(new_permission)
        created_count += 1

    create_audit_log(
        db=db,
        module_name="Permission Master",
        action="Seed Standard Permissions",
        current_user=current_user,
        entity_type="Permission",
        entity_id=None,
        entity_label="Standard Permission Seed",
        remarks="Seeded standard permissions",
        request_path="/permissions/seed-standard",
        details={
            "created_count": created_count,
            "existing_count": existing_count,
            "total_standard_permissions": len(standard_permissions),
        },
    )

    db.commit()

    return {
        "message": "Standard permissions seed completed",
        "created_count": created_count,
        "existing_count": existing_count,
        "total_standard_permissions": len(standard_permissions),
    }

# -------------------------
# Role Permission Assignment APIs
# -------------------------

def build_role_permission_response(
    role: Role,
    db: Session,
):
    assigned_permissions = (
        db.query(Permission)
        .join(RolePermission, RolePermission.permission_id == Permission.id)
        .filter(RolePermission.role_id == role.id)
        .order_by(Permission.module_name, Permission.permission_name)
        .all()
    )

    return {
        "role_id": role.id,
        "role_name": role.role_name,
        "permissions": [
            {
                "id": permission.id,
                "permission_id": permission.id,
                "permission_name": permission.permission_name,
                "module_name": permission.module_name,
                "description": permission.description,
                "status": permission.status,
            }
            for permission in assigned_permissions
        ],
    }


@app.get("/role-permissions", response_model=list[RolePermissionResponse])
def get_all_role_permissions(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Role Permission Assignment",
        db,
    )

    roles = db.query(Role).order_by(Role.id).all()

    return [
        build_role_permission_response(role, db)
        for role in roles
    ]


@app.get("/role-permissions/{role_id}", response_model=RolePermissionResponse)
def get_role_permissions(
    role_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Role Permission Assignment",
        db,
    )

    role = db.query(Role).filter(Role.id == role_id).first()

    if not role:
        raise HTTPException(
            status_code=404,
            detail="Role not found",
        )

    return build_role_permission_response(role, db)


@app.post("/role-permissions/{role_id}", response_model=RolePermissionResponse)
def save_role_permissions(
    role_id: int,
    request: RolePermissionSaveRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Role Permission Assignment",
        db,
    )

    role = db.query(Role).filter(Role.id == role_id).first()

    if not role:
        raise HTTPException(
            status_code=404,
            detail="Role not found",
        )

    # --- BEFORE snapshot (current permissions for this role) ---
    before_assigned_permissions = (
        db.query(Permission)
        .join(RolePermission, RolePermission.permission_id == Permission.id)
        .filter(RolePermission.role_id == role_id)
        .order_by(Permission.module_name, Permission.permission_name)
        .all()
    )

    before_permission_ids = [p.id for p in before_assigned_permissions]

    before_permissions_info = [
        {
            "id": p.id,
            "permission_name": p.permission_name,
            "module_name": p.module_name,
            "status": p.status,
        }
        for p in before_assigned_permissions
    ]

    # Validate request: no duplicates
    if len(request.permission_ids) != len(set(request.permission_ids)):
        raise HTTPException(
            status_code=400,
            detail="Duplicate permission IDs are not allowed",
        )

    # Validate request: all IDs exist
    permissions = (
        db.query(Permission)
        .filter(Permission.id.in_(request.permission_ids))
        .order_by(Permission.module_name, Permission.permission_name)
        .all()
    )

    if len(permissions) != len(request.permission_ids):
        raise HTTPException(
            status_code=400,
            detail="One or more permission IDs are invalid",
        )

    after_permission_ids = sorted(request.permission_ids)

    after_permissions_info = [
        {
            "id": p.id,
            "permission_name": p.permission_name,
            "module_name": p.module_name,
            "status": p.status,
        }
        for p in permissions
    ]

    before_set = set(before_permission_ids)
    after_set = set(after_permission_ids)

    added_permission_ids = sorted(list(after_set - before_set))
    removed_permission_ids = sorted(list(before_set - after_set))

    changed = (len(added_permission_ids) > 0) or (len(removed_permission_ids) > 0)

    # --- Apply change: replace all assignments ---
    db.query(RolePermission).filter(
        RolePermission.role_id == role_id
    ).delete()

    for permission_id in after_permission_ids:
        db.add(
            RolePermission(
                role_id=role_id,
                permission_id=permission_id,
            )
        )

    # --- Audit log (same transaction) ---
    create_audit_log(
        db=db,
        module_name="Role Permission Assignment",
        action="Update Role Permission Assignment",
        current_user=current_user,
        entity_type="Role",
        entity_id=role.id,
        entity_label=role.role_name,
        remarks=(
            "Role permissions updated"
            if changed
            else "Role permissions saved (no change)"
        ),
        request_path=f"/role-permissions/{role_id}",
        details={
            "role": {
                "id": role.id,
                "role_name": role.role_name,
            },
            "changed": changed,
            "before_permission_ids": sorted(before_permission_ids),
            "after_permission_ids": after_permission_ids,
            "added_permission_ids": added_permission_ids,
            "removed_permission_ids": removed_permission_ids,
            "before_permissions": before_permissions_info,
            "after_permissions": after_permissions_info,
            "counts": {
                "before": len(before_permission_ids),
                "after": len(after_permission_ids),
                "added": len(added_permission_ids),
                "removed": len(removed_permission_ids),
            },
        },
    )

    db.commit()

    return build_role_permission_response(role, db)


# -------------------------
# User Role Assignment APIs
# -------------------------

@app.get("/user-roles", response_model=list[UserRoleResponse])
def get_user_roles(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View User Role Assignment",
        db,
    )

    assignments = (
        db.query(UserRole, User, Role)
        .join(User, User.id == UserRole.user_id)
        .join(Role, Role.id == UserRole.role_id)
        .order_by(User.full_name, User.username)
        .all()
    )

    return [
        {
            "id": assignment.id,
            "user_id": user.id,
            "full_name": user.full_name,
            "username": user.username,
            "role_id": role.id,
            "role_name": role.role_name,
        }
        for assignment, user, role in assignments
    ]


@app.post("/user-roles", response_model=UserRoleResponse)
def save_user_role(
    request: UserRoleSaveRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage User Role Assignment",
        db,
    )

    user = db.query(User).filter(User.id == request.user_id).first()

    if not user:
        raise HTTPException(
            status_code=404,
            detail="User not found",
        )

    if user.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active users can be assigned roles",
        )

    role = db.query(Role).filter(Role.id == request.role_id).first()

    if not role:
        raise HTTPException(
            status_code=404,
            detail="Role not found",
        )

    if role.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active roles can be assigned to users",
        )

    existing_assignment = (
        db.query(UserRole)
        .filter(UserRole.user_id == request.user_id)
        .first()
    )

    # -------------------------
    # UPDATE existing assignment
    # -------------------------
    if existing_assignment:
        old_role = db.query(Role).filter(Role.id == existing_assignment.role_id).first()

        before_role = {
            "role_id": existing_assignment.role_id,
            "role_name": old_role.role_name if old_role else None,
        }

        after_role = {
            "role_id": role.id,
            "role_name": role.role_name,
        }

        changed = (before_role["role_id"] != after_role["role_id"])

        existing_assignment.role_id = request.role_id

        create_audit_log(
            db=db,
            module_name="User Role Assignment",
            action="Update User Role Assignment",
            current_user=current_user,
            entity_type="User",
            entity_id=user.id,
            entity_label=f"{user.full_name} ({user.username})",
            remarks="User role updated" if changed else "User role saved (no change)",
            request_path="/user-roles",
            details={
                "changed": changed,
                "assignment_id": existing_assignment.id,
                "user": {
                    "user_id": user.id,
                    "full_name": user.full_name,
                    "username": user.username,
                    "status": user.status,
                },
                "before_role": before_role,
                "after_role": after_role,
            },
        )

        db.commit()
        db.refresh(existing_assignment)

        return {
            "id": existing_assignment.id,
            "user_id": user.id,
            "full_name": user.full_name,
            "username": user.username,
            "role_id": role.id,
            "role_name": role.role_name,
        }

    # -------------------------
    # CREATE new assignment
    # -------------------------
    new_assignment = UserRole(
        user_id=request.user_id,
        role_id=request.role_id,
    )

    db.add(new_assignment)
    db.flush()  # get id before audit log

    create_audit_log(
        db=db,
        module_name="User Role Assignment",
        action="Create User Role Assignment",
        current_user=current_user,
        entity_type="User",
        entity_id=user.id,
        entity_label=f"{user.full_name} ({user.username})",
        remarks="User role assigned",
        request_path="/user-roles",
        details={
            "assignment_id": new_assignment.id,
            "user": {
                "user_id": user.id,
                "full_name": user.full_name,
                "username": user.username,
                "status": user.status,
            },
            "assigned_role": {
                "role_id": role.id,
                "role_name": role.role_name,
            },
        },
    )

    db.commit()
    db.refresh(new_assignment)

    return {
        "id": new_assignment.id,
        "user_id": user.id,
        "full_name": user.full_name,
        "username": user.username,
        "role_id": role.id,
        "role_name": role.role_name,
    }


@app.delete("/user-roles/{assignment_id}")
def delete_user_role(
    assignment_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage User Role Assignment",
        db,
    )

    assignment = db.query(UserRole).filter(UserRole.id == assignment_id).first()

    if not assignment:
        raise HTTPException(
            status_code=404,
            detail="User role assignment not found",
        )

    user = db.query(User).filter(User.id == assignment.user_id).first()
    role = db.query(Role).filter(Role.id == assignment.role_id).first()

    # Audit BEFORE delete
    create_audit_log(
        db=db,
        module_name="User Role Assignment",
        action="Delete User Role Assignment",
        current_user=current_user,
        entity_type="User",
        entity_id=assignment.user_id,
        entity_label=(
            f"{user.full_name} ({user.username})" if user else f"UserId={assignment.user_id}"
        ),
        remarks="User role assignment deleted",
        request_path=f"/user-roles/{assignment_id}",
        details={
            "assignment_id": assignment.id,
            "user": {
                "user_id": user.id if user else assignment.user_id,
                "full_name": user.full_name if user else None,
                "username": user.username if user else None,
                "status": user.status if user else None,
            },
            "removed_role": {
                "role_id": assignment.role_id,
                "role_name": role.role_name if role else None,
            },
        },
    )

    db.delete(assignment)
    db.commit()

    return {
        "message": "User role assignment deleted successfully"
    }

# -------------------------
# Location APIs
# -------------------------

@app.get("/locations", response_model=list[LocationResponse])
def get_locations(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Location",
        db,
    )

    locations = db.query(Location).order_by(Location.id).all()
    return locations


@app.post("/locations", response_model=LocationResponse)
def create_location(
    location: LocationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Location", db)

    existing_location = (
        db.query(Location)
        .filter(Location.location_code.ilike(location.location_code))
        .first()
    )

    if existing_location:
        raise HTTPException(status_code=400, detail="Location code already exists")

    if location.parent_location_code:
        parent_location = (
            db.query(Location)
            .filter(Location.location_code.ilike(location.parent_location_code))
            .first()
        )

        if not parent_location:
            raise HTTPException(status_code=400, detail="Parent location not found")

    new_location = Location(
        location_name=location.location_name,
        location_code=location.location_code,
        location_type=location.location_type,
        parent_location_code=location.parent_location_code,
        description=location.description,
        status=location.status,
    )

    db.add(new_location)
    db.flush()  # get id before audit

    after_data = {
        "location_name": new_location.location_name,
        "location_code": new_location.location_code,
        "location_type": new_location.location_type,
        "parent_location_code": new_location.parent_location_code,
        "description": new_location.description,
        "status": new_location.status,
    }

    create_audit_log(
        db=db,
        module_name="Location Master",
        action="Create Location",
        current_user=current_user,
        entity_type="Location",
        entity_id=new_location.id,
        entity_label=f"{new_location.location_name} ({new_location.location_code})",
        remarks="Location created",
        request_path="/locations",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_location)
    return new_location


@app.put("/locations/{location_id}", response_model=LocationResponse)
def update_location(
    location_id: int,
    location: LocationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Location", db)

    existing_location = db.query(Location).filter(Location.id == location_id).first()

    if not existing_location:
        raise HTTPException(status_code=404, detail="Location not found")

    duplicate_location = (
        db.query(Location)
        .filter(
            Location.location_code.ilike(location.location_code),
            Location.id != location_id,
        )
        .first()
    )

    if duplicate_location:
        raise HTTPException(status_code=400, detail="Location code already exists")

    if (
        location.parent_location_code
        and location.parent_location_code.lower() == location.location_code.lower()
    ):
        raise HTTPException(status_code=400, detail="Location cannot be its own parent")

    if location.parent_location_code:
        parent_location = (
            db.query(Location)
            .filter(Location.location_code.ilike(location.parent_location_code))
            .first()
        )

        if not parent_location:
            raise HTTPException(status_code=400, detail="Parent location not found")

    before_data = {
        "location_name": existing_location.location_name,
        "location_code": existing_location.location_code,
        "location_type": existing_location.location_type,
        "parent_location_code": existing_location.parent_location_code,
        "description": existing_location.description,
        "status": existing_location.status,
    }

    existing_location.location_name = location.location_name
    existing_location.location_code = location.location_code
    existing_location.location_type = location.location_type
    existing_location.parent_location_code = location.parent_location_code
    existing_location.description = location.description
    existing_location.status = location.status

    after_data = {
        "location_name": existing_location.location_name,
        "location_code": existing_location.location_code,
        "location_type": existing_location.location_type,
        "parent_location_code": existing_location.parent_location_code,
        "description": existing_location.description,
        "status": existing_location.status,
    }

    create_audit_log(
        db=db,
        module_name="Location Master",
        action="Update Location",
        current_user=current_user,
        entity_type="Location",
        entity_id=existing_location.id,
        entity_label=f"{existing_location.location_name} ({existing_location.location_code})",
        remarks="Location updated",
        request_path=f"/locations/{location_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_location)
    return existing_location


@app.delete("/locations/{location_id}")
def delete_location(
    location_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Location", db)

    existing_location = db.query(Location).filter(Location.id == location_id).first()

    if not existing_location:
        raise HTTPException(status_code=404, detail="Location not found")

    child_location = (
        db.query(Location)
        .filter(Location.parent_location_code.ilike(existing_location.location_code))
        .first()
    )

    if child_location:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete location because it is used as a parent location",
        )

    deleted_data = {
        "location_name": existing_location.location_name,
        "location_code": existing_location.location_code,
        "location_type": existing_location.location_type,
        "parent_location_code": existing_location.parent_location_code,
        "description": existing_location.description,
        "status": existing_location.status,
    }

    create_audit_log(
        db=db,
        module_name="Location Master",
        action="Delete Location",
        current_user=current_user,
        entity_type="Location",
        entity_id=existing_location.id,
        entity_label=f"{existing_location.location_name} ({existing_location.location_code})",
        remarks="Location deleted",
        request_path=f"/locations/{location_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_location)
    db.commit()

    return {"message": "Location deleted successfully"}


# -------------------------
# Location Accounting Day Setting APIs
# -------------------------

def build_location_accounting_day_setting_response(
    setting: LocationAccountingDaySetting,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(setting.location_code))
        .first()
    )

    return {
        "id": setting.id,
        "location_code": setting.location_code,
        "location_name": location.location_name if location else "",
        "day_start_time": setting.day_start_time,
        "day_end_time": setting.day_end_time,
        "effective_from": setting.effective_from,
        "effective_to": setting.effective_to,
        "timezone_name": setting.timezone_name,
        "description": setting.description,
        "status": setting.status,
        "created_at": setting.created_at,
        "updated_at": setting.updated_at,
    }


def build_location_accounting_day_setting_audit_snapshot(
    setting: LocationAccountingDaySetting,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(setting.location_code))
        .first()
    )

    return {
        "id": setting.id,
        "location_code": setting.location_code,
        "location_name": location.location_name if location else "",
        "day_start_time": setting.day_start_time.strftime("%H:%M:%S")
        if setting.day_start_time
        else None,
        "day_end_time": setting.day_end_time.strftime("%H:%M:%S")
        if setting.day_end_time
        else None,
        "effective_from": str(setting.effective_from)
        if setting.effective_from
        else None,
        "effective_to": str(setting.effective_to)
        if setting.effective_to
        else None,
        "timezone_name": setting.timezone_name,
        "description": setting.description,
        "status": setting.status,
    }


def validate_location_accounting_day_setting(
    setting: LocationAccountingDaySettingCreate,
    db: Session,
    setting_id: int | None = None,
):
    location_code = str(setting.location_code or "").strip().upper()

    if location_code == "":
        raise HTTPException(
            status_code=400,
            detail="Location is required",
        )

    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(location_code))
        .first()
    )

    if not location:
        raise HTTPException(
            status_code=400,
            detail="Location not found",
        )

    if location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active locations can be configured",
        )

    if setting.effective_to is not None:
        if setting.effective_to < setting.effective_from:
            raise HTTPException(
                status_code=400,
                detail="Effective To cannot be earlier than Effective From",
            )

    timezone_name = str(setting.timezone_name or "").strip()

    if timezone_name == "":
        raise HTTPException(
            status_code=400,
            detail="Timezone is required",
        )

    if setting.day_start_time == setting.day_end_time:
        raise HTTPException(
            status_code=400,
            detail="Day Start Time and Day End Time cannot be same",
        )

    # Prevent overlapping active settings for the same location.
    # Treat NULL effective_to as open-ended far future.
    if setting.status == "Active":
        new_from = setting.effective_from
        new_to = setting.effective_to or date(9999, 12, 31)

        active_settings_query = db.query(LocationAccountingDaySetting).filter(
            LocationAccountingDaySetting.location_code.ilike(location_code),
            LocationAccountingDaySetting.status == "Active",
        )

        if setting_id is not None:
            active_settings_query = active_settings_query.filter(
                LocationAccountingDaySetting.id != setting_id
            )

        active_settings = active_settings_query.all()

        for existing in active_settings:
            existing_from = existing.effective_from
            existing_to = existing.effective_to or date(9999, 12, 31)

            overlaps = new_from <= existing_to and new_to >= existing_from

            if overlaps:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Another Active accounting day setting already exists "
                        "for this location within the selected effective period"
                    ),
                )

    return {
        "location_code": location_code,
        "timezone_name": timezone_name,
    }


@app.get(
    "/location-accounting-day-settings",
    response_model=list[LocationAccountingDaySettingResponse],
)
def get_location_accounting_day_settings(
    location_code: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Location Accounting Day Setting",
        db,
    )

    query = db.query(LocationAccountingDaySetting)

    cleaned_location_code = clean_optional_text(location_code)

    if cleaned_location_code:
        query = query.filter(
            LocationAccountingDaySetting.location_code.ilike(
                cleaned_location_code
            )
        )

    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(LocationAccountingDaySetting.status == cleaned_status)

    settings = (
        query.order_by(
            LocationAccountingDaySetting.location_code.asc(),
            LocationAccountingDaySetting.effective_from.desc(),
            LocationAccountingDaySetting.id.desc(),
        )
        .all()
    )

    return [
        build_location_accounting_day_setting_response(setting, db)
        for setting in settings
    ]


@app.post(
    "/location-accounting-day-settings",
    response_model=LocationAccountingDaySettingResponse,
)
def create_location_accounting_day_setting(
    setting: LocationAccountingDaySettingCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Accounting Day Setting",
        db,
    )

    validated_data = validate_location_accounting_day_setting(
        setting=setting,
        db=db,
    )

    new_setting = LocationAccountingDaySetting(
        location_code=validated_data["location_code"],
        day_start_time=setting.day_start_time,
        day_end_time=setting.day_end_time,
        effective_from=setting.effective_from,
        effective_to=setting.effective_to,
        timezone_name=validated_data["timezone_name"],
        description=clean_optional_text(setting.description),
        status=setting.status,
    )

    db.add(new_setting)
    db.flush()

    after_data = build_location_accounting_day_setting_audit_snapshot(
        new_setting,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Location Accounting Day Setting",
        action="Create Location Accounting Day Setting",
        current_user=current_user,
        entity_type="LocationAccountingDaySetting",
        entity_id=new_setting.id,
        entity_label=(
            f"{new_setting.location_code} "
            f"{new_setting.day_start_time.strftime('%H:%M')} - "
            f"{new_setting.day_end_time.strftime('%H:%M')}"
        ),
        remarks="Location accounting day setting created",
        request_path="/location-accounting-day-settings",
        details={
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(new_setting)

    return build_location_accounting_day_setting_response(new_setting, db)


@app.put(
    "/location-accounting-day-settings/{setting_id}",
    response_model=LocationAccountingDaySettingResponse,
)
def update_location_accounting_day_setting(
    setting_id: int,
    setting: LocationAccountingDaySettingCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Accounting Day Setting",
        db,
    )

    existing_setting = (
        db.query(LocationAccountingDaySetting)
        .filter(LocationAccountingDaySetting.id == setting_id)
        .first()
    )

    if not existing_setting:
        raise HTTPException(
            status_code=404,
            detail="Location Accounting Day Setting not found",
        )

    before_data = build_location_accounting_day_setting_audit_snapshot(
        existing_setting,
        db,
    )

    validated_data = validate_location_accounting_day_setting(
        setting=setting,
        db=db,
        setting_id=setting_id,
    )

    existing_setting.location_code = validated_data["location_code"]
    existing_setting.day_start_time = setting.day_start_time
    existing_setting.day_end_time = setting.day_end_time
    existing_setting.effective_from = setting.effective_from
    existing_setting.effective_to = setting.effective_to
    existing_setting.timezone_name = validated_data["timezone_name"]
    existing_setting.description = clean_optional_text(setting.description)
    existing_setting.status = setting.status
    existing_setting.updated_at = datetime.now()

    db.flush()

    after_data = build_location_accounting_day_setting_audit_snapshot(
        existing_setting,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Location Accounting Day Setting",
        action="Update Location Accounting Day Setting",
        current_user=current_user,
        entity_type="LocationAccountingDaySetting",
        entity_id=existing_setting.id,
        entity_label=(
            f"{existing_setting.location_code} "
            f"{existing_setting.day_start_time.strftime('%H:%M')} - "
            f"{existing_setting.day_end_time.strftime('%H:%M')}"
        ),
        remarks="Location accounting day setting updated",
        request_path=f"/location-accounting-day-settings/{setting_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_setting)

    return build_location_accounting_day_setting_response(existing_setting, db)


@app.delete("/location-accounting-day-settings/{setting_id}")
def delete_location_accounting_day_setting(
    setting_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Accounting Day Setting",
        db,
    )

    existing_setting = (
        db.query(LocationAccountingDaySetting)
        .filter(LocationAccountingDaySetting.id == setting_id)
        .first()
    )

    if not existing_setting:
        raise HTTPException(
            status_code=404,
            detail="Location Accounting Day Setting not found",
        )

    deleted_data = build_location_accounting_day_setting_audit_snapshot(
        existing_setting,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Location Accounting Day Setting",
        action="Delete Location Accounting Day Setting",
        current_user=current_user,
        entity_type="LocationAccountingDaySetting",
        entity_id=existing_setting.id,
        entity_label=(
            f"{existing_setting.location_code} "
            f"{existing_setting.day_start_time.strftime('%H:%M')} - "
            f"{existing_setting.day_end_time.strftime('%H:%M')}"
        ),
        remarks="Location accounting day setting deleted",
        request_path=f"/location-accounting-day-settings/{setting_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.delete(existing_setting)
    db.commit()

    return {
        "message": "Location Accounting Day Setting deleted successfully"
    }

# -------------------------
# Asset Type APIs
# -------------------------

@app.get("/asset-types", response_model=list[AssetTypeResponse])
def get_asset_types(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset Type",
        db,
    )

    asset_types = db.query(AssetType).order_by(AssetType.id).all()
    return asset_types


@app.post("/asset-types", response_model=AssetTypeResponse)
def create_asset_type(
    asset_type: AssetTypeCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Asset Type", db)

    existing_asset_type = db.query(AssetType).filter(
        AssetType.asset_type_code.ilike(asset_type.asset_type_code)
    ).first()

    if existing_asset_type:
        raise HTTPException(status_code=400, detail="Asset Type Code already exists")

    new_asset_type = AssetType(
        asset_type_name=asset_type.asset_type_name,
        asset_type_code=asset_type.asset_type_code,
        description=asset_type.description,
        status=asset_type.status,
    )

    db.add(new_asset_type)
    db.flush()

    after_data = {
        "asset_type_name": new_asset_type.asset_type_name,
        "asset_type_code": new_asset_type.asset_type_code,
        "description": new_asset_type.description,
        "status": new_asset_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Type Master",
        action="Create Asset Type",
        current_user=current_user,
        entity_type="AssetType",
        entity_id=new_asset_type.id,
        entity_label=f"{new_asset_type.asset_type_name} ({new_asset_type.asset_type_code})",
        remarks="Asset type created",
        request_path="/asset-types",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_asset_type)
    return new_asset_type


@app.put("/asset-types/{asset_type_id}", response_model=AssetTypeResponse)
def update_asset_type(
    asset_type_id: int,
    asset_type: AssetTypeCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Asset Type", db)

    existing_asset_type = db.query(AssetType).filter(
        AssetType.id == asset_type_id
    ).first()

    if not existing_asset_type:
        raise HTTPException(status_code=404, detail="Asset Type not found")

    duplicate_asset_type = db.query(AssetType).filter(
        AssetType.asset_type_code.ilike(asset_type.asset_type_code),
        AssetType.id != asset_type_id,
    ).first()

    if duplicate_asset_type:
        raise HTTPException(status_code=400, detail="Asset Type Code already exists")

    before_data = {
        "asset_type_name": existing_asset_type.asset_type_name,
        "asset_type_code": existing_asset_type.asset_type_code,
        "description": existing_asset_type.description,
        "status": existing_asset_type.status,
    }

    existing_asset_type.asset_type_name = asset_type.asset_type_name
    existing_asset_type.asset_type_code = asset_type.asset_type_code
    existing_asset_type.description = asset_type.description
    existing_asset_type.status = asset_type.status

    after_data = {
        "asset_type_name": existing_asset_type.asset_type_name,
        "asset_type_code": existing_asset_type.asset_type_code,
        "description": existing_asset_type.description,
        "status": existing_asset_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Type Master",
        action="Update Asset Type",
        current_user=current_user,
        entity_type="AssetType",
        entity_id=existing_asset_type.id,
        entity_label=f"{existing_asset_type.asset_type_name} ({existing_asset_type.asset_type_code})",
        remarks="Asset type updated",
        request_path=f"/asset-types/{asset_type_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_asset_type)
    return existing_asset_type


@app.delete("/asset-types/{asset_type_id}")
def delete_asset_type(
    asset_type_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Asset Type", db)

    existing_asset_type = db.query(AssetType).filter(
        AssetType.id == asset_type_id
    ).first()

    if not existing_asset_type:
        raise HTTPException(status_code=404, detail="Asset Type not found")

    used_asset = db.query(Asset).filter(
        Asset.asset_type_code.ilike(existing_asset_type.asset_type_code)
    ).first()

    if used_asset:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete asset type because it is already used by assets",
        )

    deleted_data = {
        "asset_type_name": existing_asset_type.asset_type_name,
        "asset_type_code": existing_asset_type.asset_type_code,
        "description": existing_asset_type.description,
        "status": existing_asset_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Type Master",
        action="Delete Asset Type",
        current_user=current_user,
        entity_type="AssetType",
        entity_id=existing_asset_type.id,
        entity_label=f"{existing_asset_type.asset_type_name} ({existing_asset_type.asset_type_code})",
        remarks="Asset type deleted",
        request_path=f"/asset-types/{asset_type_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_asset_type)
    db.commit()

    return {"message": "Asset Type deleted successfully"}

# -------------------------
# Asset APIs
# -------------------------

@app.get("/assets", response_model=list[AssetResponse])
def get_assets(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset",
        db,
    )

    assets = db.query(Asset).order_by(Asset.id).all()
    return assets


@app.post("/assets", response_model=AssetResponse)
def create_asset(
    asset: AssetCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_asset = db.query(Asset).filter(
        Asset.asset_code.ilike(asset.asset_code)
    ).first()

    if existing_asset:
        raise HTTPException(
            status_code=400,
            detail="Asset code already exists",
        )

    asset_type = db.query(AssetType).filter(
        AssetType.asset_type_code.ilike(asset.asset_type_code)
    ).first()

    if not asset_type:
        raise HTTPException(
            status_code=400,
            detail="Asset type not found",
        )

    if asset.asset_scope not in ["Local", "Global"]:
        raise HTTPException(
            status_code=400,
            detail="Asset scope must be Local or Global",
        )

    location_code = clean_optional_text(asset.location_code)

    if asset.asset_scope == "Local" and location_code is None:
        raise HTTPException(
            status_code=400,
            detail="Location is required for Local assets",
        )

    if asset.asset_scope == "Local":
        location = db.query(Location).filter(
            Location.location_code.ilike(location_code)
        ).first()

        if not location:
            raise HTTPException(
                status_code=400,
                detail="Location not found",
            )

        if location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active location can be used for Local assets",
            )

    new_asset = Asset(
        asset_name=asset.asset_name.strip(),
        asset_code=asset.asset_code.strip(),
        asset_scope=asset.asset_scope,
        asset_type_code=asset.asset_type_code.strip(),
        location_code=location_code if asset.asset_scope == "Local" else None,
        serial_number=clean_optional_text(asset.serial_number),
        manufacturer=clean_optional_text(asset.manufacturer),
        model=clean_optional_text(asset.model),
        commission_date=asset.commission_date,
        description=clean_optional_text(asset.description),
        status=asset.status,
    )

    db.add(new_asset)
    db.flush()  # IMPORTANT: get id before audit

    after_data = {
        "asset_name": new_asset.asset_name,
        "asset_code": new_asset.asset_code,
        "asset_scope": new_asset.asset_scope,
        "asset_type_code": new_asset.asset_type_code,
        "location_code": new_asset.location_code,
        "serial_number": new_asset.serial_number,
        "manufacturer": new_asset.manufacturer,
        "model": new_asset.model,
        "commission_date": str(new_asset.commission_date) if new_asset.commission_date else None,
        "description": new_asset.description,
        "status": new_asset.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Master",
        action="Create Asset",
        current_user=current_user,
        entity_type="Asset",
        entity_id=new_asset.id,
        entity_label=f"{new_asset.asset_name} ({new_asset.asset_code})",
        remarks="Asset created",
        request_path="/assets",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_asset)

    return new_asset


@app.put("/assets/{asset_id}", response_model=AssetResponse)
def update_asset(
    asset_id: int,
    asset: AssetCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_asset = db.query(Asset).filter(
        Asset.id == asset_id
    ).first()

    if not existing_asset:
        raise HTTPException(
            status_code=404,
            detail="Asset not found",
        )

    duplicate_asset = db.query(Asset).filter(
        Asset.asset_code.ilike(asset.asset_code),
        Asset.id != asset_id,
    ).first()

    if duplicate_asset:
        raise HTTPException(
            status_code=400,
            detail="Asset code already exists",
        )

    asset_type = db.query(AssetType).filter(
        AssetType.asset_type_code.ilike(asset.asset_type_code)
    ).first()

    if not asset_type:
        raise HTTPException(
            status_code=400,
            detail="Asset type not found",
        )

    if asset.asset_scope not in ["Local", "Global"]:
        raise HTTPException(
            status_code=400,
            detail="Asset scope must be Local or Global",
        )

    location_code = clean_optional_text(asset.location_code)

    if asset.asset_scope == "Local" and location_code is None:
        raise HTTPException(
            status_code=400,
            detail="Location is required for Local assets",
        )

    if asset.asset_scope == "Local":
        location = db.query(Location).filter(
            Location.location_code.ilike(location_code)
        ).first()

        if not location:
            raise HTTPException(
                status_code=400,
                detail="Location not found",
            )

        if location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active location can be used for Local assets",
            )

    before_data = {
        "asset_name": existing_asset.asset_name,
        "asset_code": existing_asset.asset_code,
        "asset_scope": existing_asset.asset_scope,
        "asset_type_code": existing_asset.asset_type_code,
        "location_code": existing_asset.location_code,
        "serial_number": existing_asset.serial_number,
        "manufacturer": existing_asset.manufacturer,
        "model": existing_asset.model,
        "commission_date": str(existing_asset.commission_date) if existing_asset.commission_date else None,
        "description": existing_asset.description,
        "status": existing_asset.status,
    }

    existing_asset.asset_name = asset.asset_name.strip()
    existing_asset.asset_code = asset.asset_code.strip()
    existing_asset.asset_scope = asset.asset_scope
    existing_asset.asset_type_code = asset.asset_type_code.strip()
    existing_asset.location_code = (
        location_code if asset.asset_scope == "Local" else None
    )
    existing_asset.serial_number = clean_optional_text(asset.serial_number)
    existing_asset.manufacturer = clean_optional_text(asset.manufacturer)
    existing_asset.model = clean_optional_text(asset.model)
    existing_asset.commission_date = asset.commission_date
    existing_asset.description = clean_optional_text(asset.description)
    existing_asset.status = asset.status

    after_data = {
        "asset_name": existing_asset.asset_name,
        "asset_code": existing_asset.asset_code,
        "asset_scope": existing_asset.asset_scope,
        "asset_type_code": existing_asset.asset_type_code,
        "location_code": existing_asset.location_code,
        "serial_number": existing_asset.serial_number,
        "manufacturer": existing_asset.manufacturer,
        "model": existing_asset.model,
        "commission_date": str(existing_asset.commission_date) if existing_asset.commission_date else None,
        "description": existing_asset.description,
        "status": existing_asset.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Master",
        action="Update Asset",
        current_user=current_user,
        entity_type="Asset",
        entity_id=existing_asset.id,
        entity_label=f"{existing_asset.asset_name} ({existing_asset.asset_code})",
        remarks="Asset updated",
        request_path=f"/assets/{asset_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_asset)

    return existing_asset


@app.delete("/assets/{asset_id}")
def delete_asset(
    asset_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_asset = db.query(Asset).filter(
        Asset.id == asset_id
    ).first()

    if not existing_asset:
        raise HTTPException(
            status_code=404,
            detail="Asset not found",
        )

    calibration_table = db.query(AssetCalibrationTable).filter(
        AssetCalibrationTable.asset_code.ilike(existing_asset.asset_code)
    ).first()

    if calibration_table:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete asset because calibration table exists for this asset",
        )

    assignment = db.query(AssetAssignment).filter(
        AssetAssignment.asset_code.ilike(existing_asset.asset_code)
    ).first()

    if assignment:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete asset because assignment history exists for this asset",
        )

    deleted_data = {
        "asset_name": existing_asset.asset_name,
        "asset_code": existing_asset.asset_code,
        "asset_scope": existing_asset.asset_scope,
        "asset_type_code": existing_asset.asset_type_code,
        "location_code": existing_asset.location_code,
        "serial_number": existing_asset.serial_number,
        "manufacturer": existing_asset.manufacturer,
        "model": existing_asset.model,
        "commission_date": str(existing_asset.commission_date) if existing_asset.commission_date else None,
        "description": existing_asset.description,
        "status": existing_asset.status,
    }

    create_audit_log(
        db=db,
        module_name="Asset Master",
        action="Delete Asset",
        current_user=current_user,
        entity_type="Asset",
        entity_id=existing_asset.id,
        entity_label=f"{existing_asset.asset_name} ({existing_asset.asset_code})",
        remarks="Asset deleted",
        request_path=f"/assets/{asset_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_asset)
    db.commit()

    return {"message": "Asset deleted successfully"}

# -------------------------
# Calibration Template APIs
# -------------------------

def build_calibration_template_response(
    template: CalibrationTemplate,
    db: Session,
):
    template_columns = (
        db.query(CalibrationTemplateColumn)
        .filter(CalibrationTemplateColumn.template_id == template.id)
        .order_by(
            CalibrationTemplateColumn.sort_order,
            CalibrationTemplateColumn.id,
        )
        .all()
    )

    return {
        "id": template.id,
        "template_name": template.template_name,
        "asset_type_code": template.asset_type_code,
        "calibration_type": template.calibration_type,
        "description": template.description,
        "status": template.status,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
        "columns": [
            {
                "id": column.id,
                "column_name": column.column_name,
                "data_type": column.data_type,
                "unit": column.unit,
                "is_required": column.is_required,
                "interpolation_role": column.interpolation_role,
                "sort_order": column.sort_order,
            }
            for column in template_columns
        ],
    }


def validate_calibration_template(
    template: CalibrationTemplateCreate,
    db: Session,
):
    asset_type = db.query(AssetType).filter(
        AssetType.asset_type_code.ilike(template.asset_type_code)
    ).first()

    if not asset_type:
        raise HTTPException(
            status_code=400,
            detail="Asset type not found",
        )

    if len(template.columns) == 0:
        raise HTTPException(
            status_code=400,
            detail="Please add at least one template column",
        )

    column_names = [
        column.column_name.strip().lower()
        for column in template.columns
    ]

    if len(column_names) != len(set(column_names)):
        raise HTTPException(
            status_code=400,
            detail="Duplicate column names are not allowed in the same template",
        )

    input_x_exists = any(
        column.interpolation_role == "Input X"
        for column in template.columns
    )

    output_exists = any(
        column.interpolation_role == "Output"
        for column in template.columns
    )

    if not input_x_exists:
        raise HTTPException(
            status_code=400,
            detail="At least one column must have Interpolation Role as Input X",
        )

    if not output_exists:
        raise HTTPException(
            status_code=400,
            detail="At least one column must have Interpolation Role as Output",
        )


@app.get(
    "/calibration-templates",
    response_model=list[CalibrationTemplateResponse],
)
def get_calibration_templates(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Calibration Template",
        db,
    )

    templates = (
        db.query(CalibrationTemplate)
        .order_by(CalibrationTemplate.id)
        .all()
    )

    return [
        build_calibration_template_response(template, db)
        for template in templates
    ]


@app.post(
    "/calibration-templates",
    response_model=CalibrationTemplateResponse,
)
def create_calibration_template(
    template: CalibrationTemplateCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Calibration Template",
        db,
    )

    existing_template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.template_name.ilike(template.template_name)
    ).first()

    if existing_template:
        raise HTTPException(
            status_code=400,
            detail="Template name already exists",
        )

    validate_calibration_template(template, db)

    # Create template header (no commit yet)
    new_template = CalibrationTemplate(
        template_name=template.template_name.strip(),
        asset_type_code=template.asset_type_code.strip(),
        calibration_type=template.calibration_type.strip(),
        description=clean_optional_text(template.description),
        status=template.status,
    )

    db.add(new_template)
    db.flush()  # get new_template.id

    # Create template columns
    for index, column in enumerate(template.columns):
        new_column = CalibrationTemplateColumn(
            template_id=new_template.id,
            column_name=column.column_name.strip(),
            data_type=column.data_type,
            unit=clean_optional_text(column.unit),
            is_required=column.is_required,
            interpolation_role=column.interpolation_role,
            sort_order=column.sort_order or index + 1,
        )
        db.add(new_column)

    db.flush()

    after_data = build_calibration_template_response(new_template, db)

    create_audit_log(
        db=db,
        module_name="Calibration Template Master",
        action="Create Calibration Template",
        current_user=current_user,
        entity_type="CalibrationTemplate",
        entity_id=new_template.id,
        entity_label=new_template.template_name,
        remarks="Calibration template created",
        request_path="/calibration-templates",
        details={
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(new_template)

    return build_calibration_template_response(new_template, db)


@app.put(
    "/calibration-templates/{template_id}",
    response_model=CalibrationTemplateResponse,
)
def update_calibration_template(
    template_id: int,
    template: CalibrationTemplateCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Calibration Template",
        db,
    )

    existing_template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.id == template_id
    ).first()

    if not existing_template:
        raise HTTPException(
            status_code=404,
            detail="Calibration template not found",
        )

    duplicate_template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.template_name.ilike(template.template_name),
        CalibrationTemplate.id != template_id,
    ).first()

    if duplicate_template:
        raise HTTPException(
            status_code=400,
            detail="Template name already exists",
        )

    validate_calibration_template(template, db)

    before_data = build_calibration_template_response(existing_template, db)

    # Update header
    existing_template.template_name = template.template_name.strip()
    existing_template.asset_type_code = template.asset_type_code.strip()
    existing_template.calibration_type = template.calibration_type.strip()
    existing_template.description = clean_optional_text(template.description)
    existing_template.status = template.status

    # Replace columns
    db.query(CalibrationTemplateColumn).filter(
        CalibrationTemplateColumn.template_id == template_id
    ).delete()

    for index, column in enumerate(template.columns):
        new_column = CalibrationTemplateColumn(
            template_id=template_id,
            column_name=column.column_name.strip(),
            data_type=column.data_type,
            unit=clean_optional_text(column.unit),
            is_required=column.is_required,
            interpolation_role=column.interpolation_role,
            sort_order=column.sort_order or index + 1,
        )
        db.add(new_column)

    db.flush()

    after_data = build_calibration_template_response(existing_template, db)

    create_audit_log(
        db=db,
        module_name="Calibration Template Master",
        action="Update Calibration Template",
        current_user=current_user,
        entity_type="CalibrationTemplate",
        entity_id=existing_template.id,
        entity_label=existing_template.template_name,
        remarks="Calibration template updated",
        request_path=f"/calibration-templates/{template_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_template)

    return build_calibration_template_response(existing_template, db)


@app.delete("/calibration-templates/{template_id}")
def delete_calibration_template(
    template_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Calibration Template",
        db,
    )

    existing_template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.id == template_id
    ).first()

    if not existing_template:
        raise HTTPException(
            status_code=404,
            detail="Calibration template not found",
        )

    used_calibration_table = (
        db.query(AssetCalibrationTable)
        .filter(AssetCalibrationTable.template_id == template_id)
        .first()
    )

    if used_calibration_table:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete calibration template because it is used by asset calibration tables",
        )

    deleted_data = build_calibration_template_response(existing_template, db)

    create_audit_log(
        db=db,
        module_name="Calibration Template Master",
        action="Delete Calibration Template",
        current_user=current_user,
        entity_type="CalibrationTemplate",
        entity_id=existing_template.id,
        entity_label=existing_template.template_name,
        remarks="Calibration template deleted",
        request_path=f"/calibration-templates/{template_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.query(CalibrationTemplateColumn).filter(
        CalibrationTemplateColumn.template_id == template_id
    ).delete()

    db.delete(existing_template)
    db.commit()

    return {
        "message": "Calibration template deleted successfully"
    }

# -------------------------
# Asset Calibration Table APIs
# -------------------------

def build_asset_calibration_table_response(
    calibration_table: AssetCalibrationTable,
    db: Session,
):
    asset = db.query(Asset).filter(
        Asset.asset_code == calibration_table.asset_code
    ).first()

    template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.id == calibration_table.template_id
    ).first()

    rows = (
        db.query(AssetCalibrationData)
        .filter(
            AssetCalibrationData.calibration_table_id == calibration_table.id
        )
        .order_by(AssetCalibrationData.row_number)
        .all()
    )

    return {
        "id": calibration_table.id,
        "calibration_name": calibration_table.calibration_name,
        "asset_code": calibration_table.asset_code,
        "asset_name": asset.asset_name if asset else "",
        "template_id": calibration_table.template_id,
        "template_name": template.template_name if template else "",
        "effective_date": calibration_table.effective_date,
        "remarks": calibration_table.remarks,
        "status": calibration_table.status,
        "created_at": calibration_table.created_at,
        "updated_at": calibration_table.updated_at,
        "rows": [
            {
                "id": row.id,
                "row_number": row.row_number,
                "row_data": row.row_data,
            }
            for row in rows
        ],
    }


def validate_asset_calibration_table(
    calibration_table: AssetCalibrationTableCreate,
    db: Session,
):
    asset = db.query(Asset).filter(
        Asset.asset_code.ilike(calibration_table.asset_code)
    ).first()

    if not asset:
        raise HTTPException(
            status_code=400,
            detail="Asset not found",
        )

    template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.id == calibration_table.template_id
    ).first()

    if not template:
        raise HTTPException(
            status_code=400,
            detail="Calibration template not found",
        )

    if asset.asset_type_code.lower() != template.asset_type_code.lower():
        raise HTTPException(
            status_code=400,
            detail="Selected template does not belong to this asset type",
        )

    if len(calibration_table.rows) == 0:
        raise HTTPException(
            status_code=400,
            detail="Please add at least one calibration data row",
        )

    template_columns = (
        db.query(CalibrationTemplateColumn)
        .filter(CalibrationTemplateColumn.template_id == template.id)
        .order_by(CalibrationTemplateColumn.sort_order)
        .all()
    )

    required_columns = [
        column.column_name
        for column in template_columns
        if column.is_required == "Yes"
    ]

    def _norm_col(name: str) -> str:
        return str(name or "").strip().lower()

    # normalized template column -> exact template column name
    template_col_map = {
        _norm_col(col.column_name): col.column_name
        for col in template_columns
    }

    for row in calibration_table.rows:
        original = row.row_data or {}
        original_keys = list(original.keys())

        # Normalize uploaded headers to match template columns
        normalized_row_data = {}
        for k, v in original.items():
            nk = _norm_col(k)
            if nk in template_col_map:
                normalized_row_data[template_col_map[nk]] = v
            else:
                normalized_row_data[k] = v  # keep extra columns

        row.row_data = normalized_row_data

        normalized_keys = list(row.row_data.keys())
        row_keys_norm = {_norm_col(k) for k in row.row_data.keys()}

        for required_column in required_columns:
            rn = _norm_col(required_column)

            if rn not in row_keys_norm:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Required column missing: {required_column}. "
                        f"RowNumber={getattr(row, 'row_number', None)}. "
                        f"IncomingKeys={original_keys}. "
                        f"NormalizedKeys={normalized_keys}."
                    ),
                )

            template_key = template_col_map.get(rn, required_column)
            value = row.row_data.get(template_key)

            if value is None or str(value).strip() == "":
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Required column cannot be blank: {required_column}. "
                        f"RowNumber={getattr(row, 'row_number', None)}."
                    ),
                )


@app.get(
    "/asset-calibration-tables",
    response_model=list[AssetCalibrationTableResponse],
)
def get_asset_calibration_tables(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset Calibration",
        db,
    )

    calibration_tables = (
        db.query(AssetCalibrationTable)
        .order_by(AssetCalibrationTable.id)
        .all()
    )

    return [
        build_asset_calibration_table_response(calibration_table, db)
        for calibration_table in calibration_tables
    ]

def build_asset_calibration_table_audit_snapshot(
    calibration_table: AssetCalibrationTable,
    db: Session,
    max_rows: int = 50,
):
    asset = db.query(Asset).filter(
        Asset.asset_code == calibration_table.asset_code
    ).first()

    template = db.query(CalibrationTemplate).filter(
        CalibrationTemplate.id == calibration_table.template_id
    ).first()

    rows = (
        db.query(AssetCalibrationData)
        .filter(AssetCalibrationData.calibration_table_id == calibration_table.id)
        .order_by(AssetCalibrationData.row_number.asc())
        .all()
    )

    row_count = len(rows)

    preview_rows = rows[: max_rows]

    return {
        "id": calibration_table.id,
        "calibration_name": calibration_table.calibration_name,
        "asset_code": calibration_table.asset_code,
        "asset_name": asset.asset_name if asset else "",
        "template_id": calibration_table.template_id,
        "template_name": template.template_name if template else "",
        "effective_date": str(calibration_table.effective_date)
        if calibration_table.effective_date
        else None,
        "remarks": calibration_table.remarks,
        "status": calibration_table.status,
        "row_count": row_count,
        "row_numbers": [r.row_number for r in rows],
        "rows_preview_limit": max_rows,
        "rows_preview": [
            {
                "row_number": r.row_number,
                "row_data": r.row_data,
            }
            for r in preview_rows
        ],
    }

@app.post(
    "/asset-calibration-tables",
    response_model=AssetCalibrationTableResponse,
)
def create_asset_calibration_table(
    calibration_table: AssetCalibrationTableCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Calibration",
        db,
    )

    validate_asset_calibration_table(calibration_table, db)

    new_calibration_table = AssetCalibrationTable(
        calibration_name=calibration_table.calibration_name.strip(),
        asset_code=calibration_table.asset_code.strip(),
        template_id=calibration_table.template_id,
        effective_date=calibration_table.effective_date,
        remarks=clean_optional_text(calibration_table.remarks),
        status=calibration_table.status,
    )

    db.add(new_calibration_table)
    db.flush()

    for index, row in enumerate(calibration_table.rows):
        new_row = AssetCalibrationData(
            calibration_table_id=new_calibration_table.id,
            row_number=row.row_number or index + 1,
            row_data=row.row_data,
        )
        db.add(new_row)

    db.flush()

    after_data = build_asset_calibration_table_audit_snapshot(
        new_calibration_table, db
    )

    create_audit_log(
        db=db,
        module_name="Asset Calibration Table",
        action="Create Asset Calibration Table",
        current_user=current_user,
        entity_type="AssetCalibrationTable",
        entity_id=new_calibration_table.id,
        entity_label=new_calibration_table.calibration_name,
        remarks="Asset calibration table created",
        request_path="/asset-calibration-tables",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_calibration_table)

    return build_asset_calibration_table_response(new_calibration_table, db)


@app.put(
    "/asset-calibration-tables/{calibration_table_id}",
    response_model=AssetCalibrationTableResponse,
)
def update_asset_calibration_table(
    calibration_table_id: int,
    calibration_table: AssetCalibrationTableCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Calibration",
        db,
    )

    existing_calibration_table = db.query(AssetCalibrationTable).filter(
        AssetCalibrationTable.id == calibration_table_id
    ).first()

    if not existing_calibration_table:
        raise HTTPException(
            status_code=404,
            detail="Asset calibration table not found",
        )

    before_data = build_asset_calibration_table_audit_snapshot(
        existing_calibration_table, db
    )

    validate_asset_calibration_table(calibration_table, db)

    existing_calibration_table.calibration_name = (
        calibration_table.calibration_name.strip()
    )
    existing_calibration_table.asset_code = calibration_table.asset_code.strip()
    existing_calibration_table.template_id = calibration_table.template_id
    existing_calibration_table.effective_date = calibration_table.effective_date
    existing_calibration_table.remarks = clean_optional_text(
        calibration_table.remarks
    )
    existing_calibration_table.status = calibration_table.status

    db.query(AssetCalibrationData).filter(
        AssetCalibrationData.calibration_table_id == calibration_table_id
    ).delete()

    for index, row in enumerate(calibration_table.rows):
        new_row = AssetCalibrationData(
            calibration_table_id=calibration_table_id,
            row_number=row.row_number or index + 1,
            row_data=row.row_data,
        )
        db.add(new_row)

    db.flush()

    after_data = build_asset_calibration_table_audit_snapshot(
        existing_calibration_table, db
    )

    create_audit_log(
        db=db,
        module_name="Asset Calibration Table",
        action="Update Asset Calibration Table",
        current_user=current_user,
        entity_type="AssetCalibrationTable",
        entity_id=existing_calibration_table.id,
        entity_label=existing_calibration_table.calibration_name,
        remarks="Asset calibration table updated",
        request_path=f"/asset-calibration-tables/{calibration_table_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_calibration_table)

    return build_asset_calibration_table_response(existing_calibration_table, db)


@app.delete("/asset-calibration-tables/{calibration_table_id}")
def delete_asset_calibration_table(
    calibration_table_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Calibration",
        db,
    )

    existing_calibration_table = db.query(AssetCalibrationTable).filter(
        AssetCalibrationTable.id == calibration_table_id
    ).first()

    if not existing_calibration_table:
        raise HTTPException(
            status_code=404,
            detail="Asset calibration table not found",
        )

    deleted_data = build_asset_calibration_table_audit_snapshot(
        existing_calibration_table, db
    )

    create_audit_log(
        db=db,
        module_name="Asset Calibration Table",
        action="Delete Asset Calibration Table",
        current_user=current_user,
        entity_type="AssetCalibrationTable",
        entity_id=existing_calibration_table.id,
        entity_label=existing_calibration_table.calibration_name,
        remarks="Asset calibration table deleted",
        request_path=f"/asset-calibration-tables/{calibration_table_id}",
        details={"deleted": deleted_data},
    )

    db.query(AssetCalibrationData).filter(
        AssetCalibrationData.calibration_table_id == calibration_table_id
    ).delete()

    db.delete(existing_calibration_table)
    db.commit()

    return {
        "message": "Asset calibration table deleted successfully"
    }


# -------------------------
# Asset Assignment APIs
# -------------------------

def build_asset_assignment_response(
    assignment: AssetAssignment,
    db: Session,
):
    asset = (
        db.query(Asset)
        .filter(Asset.asset_code.ilike(assignment.asset_code))
        .first()
    )

    location = (
        db.query(Location)
        .filter(
            Location.location_code.ilike(
                assignment.assignment_location_code
            )
        )
        .first()
    )

    assigned_to_display = assignment.assigned_to

    if assignment.assigned_to_type == "User":
        assigned_user = (
            db.query(User)
            .filter(User.username.ilike(assignment.assigned_to))
            .first()
        )

        if assigned_user:
            assigned_to_display = (
                f"{assigned_user.full_name} ({assigned_user.username})"
            )

    return {
        "id": assignment.id,
        "asset_code": assignment.asset_code,
        "asset_name": asset.asset_name if asset else "",
        "asset_scope": assignment.asset_scope,
        "assignment_location_code": assignment.assignment_location_code,
        "assignment_location_name": location.location_name if location else "",
        "assigned_to_type": assignment.assigned_to_type,
        "assigned_to": assignment.assigned_to,
        "assigned_to_display": assigned_to_display,
        "assignment_date": assignment.assignment_date,
        "return_date": assignment.return_date,
        "remarks": assignment.remarks,
        "status": assignment.status,
        "created_at": assignment.created_at,
        "updated_at": assignment.updated_at,
    }


def validate_asset_assignment(
    assignment: AssetAssignmentCreate,
    db: Session,
    assignment_id: int | None = None,
):
    asset = (
        db.query(Asset)
        .filter(Asset.asset_code.ilike(assignment.asset_code))
        .first()
    )

    if not asset:
        raise HTTPException(
            status_code=400,
            detail="Asset not found",
        )

    if asset.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active assets can be assigned",
        )

    if asset.asset_scope != assignment.asset_scope:
        raise HTTPException(
            status_code=400,
            detail="Selected asset scope does not match Asset Master",
        )

    location = (
        db.query(Location)
        .filter(
            Location.location_code.ilike(
                assignment.assignment_location_code
            )
        )
        .first()
    )

    if not location:
        raise HTTPException(
            status_code=400,
            detail="Assignment location not found",
        )

    if location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active locations can be used for assignment",
        )

    if assignment.assigned_to_type not in ["User", "Location", "External"]:
        raise HTTPException(
            status_code=400,
            detail="Assigned To Type must be User, Location, or External",
        )

    if assignment.assigned_to.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Assigned To is required",
        )

    if assignment.assigned_to_type == "User":
        assigned_user = (
            db.query(User)
            .filter(User.username.ilike(assignment.assigned_to))
            .first()
        )

        if not assigned_user:
            raise HTTPException(
                status_code=400,
                detail="Assigned user not found",
            )

        if assigned_user.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active users can be assigned",
            )

    if assignment.assigned_to_type == "Location":
        assigned_location = (
            db.query(Location)
            .filter(Location.location_code.ilike(assignment.assigned_to))
            .first()
        )

        if not assigned_location:
            raise HTTPException(
                status_code=400,
                detail="Assigned location not found",
            )

        if assigned_location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active locations can be assigned",
            )

    active_assignment_query = db.query(AssetAssignment).filter(
        AssetAssignment.asset_code.ilike(assignment.asset_code),
        AssetAssignment.status == "Active",
    )

    if assignment_id is not None:
        active_assignment_query = active_assignment_query.filter(
            AssetAssignment.id != assignment_id
        )

    active_assignment = active_assignment_query.first()

    if active_assignment and assignment.status == "Active":
        raise HTTPException(
            status_code=400,
            detail="This asset already has an active assignment",
        )


@app.get(
    "/asset-assignments",
    response_model=list[AssetAssignmentResponse],
)
def get_asset_assignments(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset Assignment",
        db,
    )

    assignments = (
        db.query(AssetAssignment)
        .order_by(AssetAssignment.id)
        .all()
    )

    return [
        build_asset_assignment_response(assignment, db)
        for assignment in assignments
    ]

def build_asset_assignment_audit_snapshot(
    assignment: AssetAssignment,
    db: Session,
):
    asset = db.query(Asset).filter(
        Asset.asset_code.ilike(assignment.asset_code)
    ).first()

    location = db.query(Location).filter(
        Location.location_code.ilike(assignment.assignment_location_code)
    ).first()

    assigned_user_display = None

    if assignment.assigned_to_type == "User":
        assigned_user = db.query(User).filter(
            User.username.ilike(assignment.assigned_to)
        ).first()

        if assigned_user:
            assigned_user_display = f"{assigned_user.full_name} ({assigned_user.username})"

    return {
        "id": assignment.id,
        "asset_code": assignment.asset_code,
        "asset_name": asset.asset_name if asset else "",
        "asset_scope": assignment.asset_scope,
        "assignment_location_code": assignment.assignment_location_code,
        "assignment_location_name": location.location_name if location else "",
        "assigned_to_type": assignment.assigned_to_type,
        "assigned_to": assignment.assigned_to,
        "assigned_to_display": assigned_user_display or assignment.assigned_to,
        "assignment_date": str(assignment.assignment_date) if assignment.assignment_date else None,
        "return_date": str(assignment.return_date) if assignment.return_date else None,
        "remarks": assignment.remarks,
        "status": assignment.status,
    }

@app.post(
    "/asset-assignments",
    response_model=AssetAssignmentResponse,
)
def create_asset_assignment(
    assignment: AssetAssignmentCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Assignment",
        db,
    )

    validate_asset_assignment(assignment, db)

    new_assignment = AssetAssignment(
        asset_code=assignment.asset_code.strip(),
        asset_scope=assignment.asset_scope,
        assignment_location_code=assignment.assignment_location_code.strip(),
        assigned_to_type=assignment.assigned_to_type,
        assigned_to=assignment.assigned_to.strip(),
        assignment_date=assignment.assignment_date,
        return_date=assignment.return_date,
        remarks=clean_optional_text(assignment.remarks),
        status=assignment.status,
    )

    db.add(new_assignment)
    db.flush()

    after_data = build_asset_assignment_audit_snapshot(new_assignment, db)

    create_audit_log(
        db=db,
        module_name="Asset Assignment",
        action="Create Asset Assignment",
        current_user=current_user,
        entity_type="AssetAssignment",
        entity_id=new_assignment.id,
        entity_label=f"{after_data.get('asset_name','')} ({new_assignment.asset_code})",
        remarks="Asset assignment created",
        request_path="/asset-assignments",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_assignment)

    return build_asset_assignment_response(new_assignment, db)


@app.put(
    "/asset-assignments/{assignment_id}",
    response_model=AssetAssignmentResponse,
)
def update_asset_assignment(
    assignment_id: int,
    assignment: AssetAssignmentCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Assignment",
        db,
    )

    existing_assignment = (
        db.query(AssetAssignment)
        .filter(AssetAssignment.id == assignment_id)
        .first()
    )

    if not existing_assignment:
        raise HTTPException(
            status_code=404,
            detail="Asset assignment not found",
        )

    before_data = build_asset_assignment_audit_snapshot(existing_assignment, db)

    validate_asset_assignment(assignment, db, assignment_id)

    existing_assignment.asset_code = assignment.asset_code.strip()
    existing_assignment.asset_scope = assignment.asset_scope
    existing_assignment.assignment_location_code = (
        assignment.assignment_location_code.strip()
    )
    existing_assignment.assigned_to_type = assignment.assigned_to_type
    existing_assignment.assigned_to = assignment.assigned_to.strip()
    existing_assignment.assignment_date = assignment.assignment_date
    existing_assignment.return_date = assignment.return_date
    existing_assignment.remarks = clean_optional_text(assignment.remarks)
    existing_assignment.status = assignment.status

    db.flush()

    after_data = build_asset_assignment_audit_snapshot(existing_assignment, db)

    create_audit_log(
        db=db,
        module_name="Asset Assignment",
        action="Update Asset Assignment",
        current_user=current_user,
        entity_type="AssetAssignment",
        entity_id=existing_assignment.id,
        entity_label=f"{after_data.get('asset_name','')} ({existing_assignment.asset_code})",
        remarks="Asset assignment updated",
        request_path=f"/asset-assignments/{assignment_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_assignment)

    return build_asset_assignment_response(existing_assignment, db)


@app.delete("/asset-assignments/{assignment_id}")
def delete_asset_assignment(
    assignment_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Assignment",
        db,
    )

    existing_assignment = (
        db.query(AssetAssignment)
        .filter(AssetAssignment.id == assignment_id)
        .first()
    )

    if not existing_assignment:
        raise HTTPException(
            status_code=404,
            detail="Asset assignment not found",
        )

    deleted_data = build_asset_assignment_audit_snapshot(existing_assignment, db)

    create_audit_log(
        db=db,
        module_name="Asset Assignment",
        action="Delete Asset Assignment",
        current_user=current_user,
        entity_type="AssetAssignment",
        entity_id=existing_assignment.id,
        entity_label=f"{deleted_data.get('asset_name','')} ({existing_assignment.asset_code})",
        remarks="Asset assignment deleted",
        request_path=f"/asset-assignments/{assignment_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_assignment)
    db.commit()

    return {
        "message": "Asset assignment deleted successfully"
    }


# -------------------------
# Prime Mover - Tanker Link APIs
# -------------------------

def get_asset_by_code_case_insensitive(
    asset_code: str | None,
    db: Session,
):
    cleaned_asset_code = clean_optional_text(asset_code)

    if not cleaned_asset_code:
        return None

    return (
        db.query(Asset)
        .filter(Asset.asset_code.ilike(cleaned_asset_code))
        .first()
    )


def is_prime_mover_asset(asset: Asset | None):
    if not asset:
        return False

    asset_type_code = str(asset.asset_type_code or "").strip().upper()
    asset_name = str(asset.asset_name or "").strip().upper()
    asset_code = str(asset.asset_code or "").strip().upper()

    return (
        "PRIME" in asset_type_code
        or "MOVER" in asset_type_code
        or "PRIME" in asset_name
        or "MOVER" in asset_name
        or asset_code.startswith("PM")
    )


def is_tanker_trailer_asset(asset: Asset | None):
    if not asset:
        return False

    asset_type_code = str(asset.asset_type_code or "").strip().upper()
    asset_name = str(asset.asset_name or "").strip().upper()

    return (
        "TANKER" in asset_type_code
        or "TRAILER" in asset_type_code
        or "TRUCK" in asset_type_code
        or "TANKER" in asset_name
        or "TRAILER" in asset_name
    )


def build_prime_mover_tanker_link_response(
    link: PrimeMoverTankerLink,
    db: Session,
):
    prime_mover_asset = get_asset_by_code_case_insensitive(
        link.prime_mover_asset_code,
        db,
    )

    tanker_asset = get_asset_by_code_case_insensitive(
        link.tanker_asset_code,
        db,
    )

    return {
        "id": link.id,
        "prime_mover_asset_code": link.prime_mover_asset_code,
        "prime_mover_asset_name": prime_mover_asset.asset_name
        if prime_mover_asset
        else "",
        "prime_mover_asset_type_code": prime_mover_asset.asset_type_code
        if prime_mover_asset
        else "",
        "tanker_asset_code": link.tanker_asset_code,
        "tanker_asset_name": tanker_asset.asset_name if tanker_asset else "",
        "tanker_asset_type_code": tanker_asset.asset_type_code
        if tanker_asset
        else "",
        "tanker_chassis_number": tanker_asset.serial_number
        if tanker_asset
        else "",
        "linked_from": link.linked_from,
        "linked_to": link.linked_to,
        "remarks": link.remarks,
        "status": link.status,
        "created_by": link.created_by,
        "created_at": link.created_at,
        "updated_at": link.updated_at,
    }


def build_prime_mover_tanker_link_audit_snapshot(
    link: PrimeMoverTankerLink,
    db: Session,
):
    return build_prime_mover_tanker_link_response(link, db)


def validate_prime_mover_tanker_link(
    link_request: PrimeMoverTankerLinkCreate,
    db: Session,
    link_id: int | None = None,
):
    prime_mover_asset_code = str(
        link_request.prime_mover_asset_code or ""
    ).strip()

    tanker_asset_code = str(
        link_request.tanker_asset_code or ""
    ).strip()

    if prime_mover_asset_code == "":
        raise HTTPException(
            status_code=400,
            detail="Prime Mover asset is required",
        )

    if tanker_asset_code == "":
        raise HTTPException(
            status_code=400,
            detail="Tanker Trailer asset is required",
        )

    if prime_mover_asset_code.lower() == tanker_asset_code.lower():
        raise HTTPException(
            status_code=400,
            detail="Prime Mover and Tanker Trailer cannot be the same asset",
        )

    prime_mover_asset = get_asset_by_code_case_insensitive(
        prime_mover_asset_code,
        db,
    )

    if not prime_mover_asset:
        raise HTTPException(
            status_code=400,
            detail="Prime Mover asset not found",
        )

    if prime_mover_asset.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active Prime Mover asset can be linked",
        )

    if not is_prime_mover_asset(prime_mover_asset):
        raise HTTPException(
            status_code=400,
            detail=(
                "Selected Prime Mover asset does not look like a Prime Mover. "
                "Use asset type code such as PRIME_MOVER."
            ),
        )

    tanker_asset = get_asset_by_code_case_insensitive(
        tanker_asset_code,
        db,
    )

    if not tanker_asset:
        raise HTTPException(
            status_code=400,
            detail="Tanker Trailer asset not found",
        )

    if tanker_asset.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active Tanker Trailer asset can be linked",
        )

    if not is_tanker_trailer_asset(tanker_asset):
        raise HTTPException(
            status_code=400,
            detail=(
                "Selected Tanker asset does not look like a Tanker Trailer. "
                "Use asset type code such as TANKER_TRAILER."
            ),
        )

    if link_request.linked_to is not None:
        if link_request.linked_to < link_request.linked_from:
            raise HTTPException(
                status_code=400,
                detail="Linked To cannot be earlier than Linked From",
            )

    cleaned_status = str(link_request.status or "").strip()

    if cleaned_status not in ["Active", "Inactive"]:
        raise HTTPException(
            status_code=400,
            detail="Status must be Active or Inactive",
        )

    # Active link rule:
    # 1 Prime Mover can have only 1 Active Tanker.
    # 1 Tanker Trailer can have only 1 Active Prime Mover.
    if cleaned_status == "Active":
        active_prime_mover_link_query = (
            db.query(PrimeMoverTankerLink)
            .filter(
                PrimeMoverTankerLink.prime_mover_asset_code.ilike(
                    prime_mover_asset_code
                ),
                PrimeMoverTankerLink.status == "Active",
            )
        )

        active_tanker_link_query = (
            db.query(PrimeMoverTankerLink)
            .filter(
                PrimeMoverTankerLink.tanker_asset_code.ilike(
                    tanker_asset_code
                ),
                PrimeMoverTankerLink.status == "Active",
            )
        )

        if link_id is not None:
            active_prime_mover_link_query = active_prime_mover_link_query.filter(
                PrimeMoverTankerLink.id != link_id
            )

            active_tanker_link_query = active_tanker_link_query.filter(
                PrimeMoverTankerLink.id != link_id
            )

        active_prime_mover_link = active_prime_mover_link_query.first()

        if active_prime_mover_link:
            raise HTTPException(
                status_code=400,
                detail=(
                    "This Prime Mover already has an Active Tanker link. "
                    "Close or deactivate the old link before creating a new one."
                ),
            )

        active_tanker_link = active_tanker_link_query.first()

        if active_tanker_link:
            raise HTTPException(
                status_code=400,
                detail=(
                    "This Tanker Trailer is already linked to another Active "
                    "Prime Mover. Close or deactivate the old link first."
                ),
            )

    return {
        "prime_mover_asset_code": prime_mover_asset.asset_code,
        "tanker_asset_code": tanker_asset.asset_code,
        "linked_from": link_request.linked_from,
        "linked_to": link_request.linked_to,
        "remarks": clean_optional_text(link_request.remarks),
        "status": cleaned_status,
    }


@app.get(
    "/prime-mover-tanker-links",
    response_model=list[PrimeMoverTankerLinkResponse],
)
def get_prime_mover_tanker_links(
    status: str | None = None,
    prime_mover_asset_code: str | None = None,
    tanker_asset_code: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset",
        db,
    )

    query = db.query(PrimeMoverTankerLink)

    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(PrimeMoverTankerLink.status == cleaned_status)

    cleaned_prime_mover_asset_code = clean_optional_text(prime_mover_asset_code)

    if cleaned_prime_mover_asset_code:
        query = query.filter(
            PrimeMoverTankerLink.prime_mover_asset_code.ilike(
                cleaned_prime_mover_asset_code
            )
        )

    cleaned_tanker_asset_code = clean_optional_text(tanker_asset_code)

    if cleaned_tanker_asset_code:
        query = query.filter(
            PrimeMoverTankerLink.tanker_asset_code.ilike(
                cleaned_tanker_asset_code
            )
        )

    links = (
        query.order_by(
            PrimeMoverTankerLink.status.asc(),
            PrimeMoverTankerLink.linked_from.desc(),
            PrimeMoverTankerLink.id.desc(),
        )
        .all()
    )

    return [
        build_prime_mover_tanker_link_response(link, db)
        for link in links
    ]


@app.get(
    "/prime-mover-tanker-links/current-by-prime-mover/{prime_mover_asset_code}",
    response_model=CurrentPrimeMoverTankerLinkResponse,
)
def get_current_prime_mover_tanker_link(
    prime_mover_asset_code: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset",
        db,
    )

    link = (
        db.query(PrimeMoverTankerLink)
        .filter(
            PrimeMoverTankerLink.prime_mover_asset_code.ilike(
                prime_mover_asset_code
            ),
            PrimeMoverTankerLink.status == "Active",
        )
        .order_by(
            PrimeMoverTankerLink.linked_from.desc(),
            PrimeMoverTankerLink.id.desc(),
        )
        .first()
    )

    if not link:
        return {
            "has_active_link": False,
            "link": None,
        }

    return {
        "has_active_link": True,
        "link": build_prime_mover_tanker_link_response(link, db),
    }


@app.post(
    "/prime-mover-tanker-links",
    response_model=PrimeMoverTankerLinkResponse,
)
def create_prime_mover_tanker_link(
    link_request: PrimeMoverTankerLinkCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    validated_data = validate_prime_mover_tanker_link(
        link_request,
        db,
    )

    new_link = PrimeMoverTankerLink(
        prime_mover_asset_code=validated_data["prime_mover_asset_code"],
        tanker_asset_code=validated_data["tanker_asset_code"],
        linked_from=validated_data["linked_from"],
        linked_to=validated_data["linked_to"],
        remarks=validated_data["remarks"],
        status=validated_data["status"],
        created_by=get_current_user_display_name(current_user),
    )

    db.add(new_link)
    db.flush()

    after_data = build_prime_mover_tanker_link_audit_snapshot(
        new_link,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Prime Mover Tanker Link",
        action="Create Prime Mover Tanker Link",
        current_user=current_user,
        entity_type="PrimeMoverTankerLink",
        entity_id=new_link.id,
        entity_label=(
            f"{new_link.prime_mover_asset_code} -> "
            f"{new_link.tanker_asset_code}"
        ),
        remarks="Prime mover tanker link created",
        request_path="/prime-mover-tanker-links",
        details={
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(new_link)

    return build_prime_mover_tanker_link_response(new_link, db)


@app.put(
    "/prime-mover-tanker-links/{link_id}",
    response_model=PrimeMoverTankerLinkResponse,
)
def update_prime_mover_tanker_link(
    link_id: int,
    link_request: PrimeMoverTankerLinkCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_link = (
        db.query(PrimeMoverTankerLink)
        .filter(PrimeMoverTankerLink.id == link_id)
        .first()
    )

    if not existing_link:
        raise HTTPException(
            status_code=404,
            detail="Prime Mover Tanker link not found",
        )

    before_data = build_prime_mover_tanker_link_audit_snapshot(
        existing_link,
        db,
    )

    validated_data = validate_prime_mover_tanker_link(
        link_request,
        db,
        link_id=link_id,
    )

    existing_link.prime_mover_asset_code = validated_data[
        "prime_mover_asset_code"
    ]
    existing_link.tanker_asset_code = validated_data["tanker_asset_code"]
    existing_link.linked_from = validated_data["linked_from"]
    existing_link.linked_to = validated_data["linked_to"]
    existing_link.remarks = validated_data["remarks"]
    existing_link.status = validated_data["status"]
    existing_link.updated_at = datetime.now()

    db.flush()

    after_data = build_prime_mover_tanker_link_audit_snapshot(
        existing_link,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Prime Mover Tanker Link",
        action="Update Prime Mover Tanker Link",
        current_user=current_user,
        entity_type="PrimeMoverTankerLink",
        entity_id=existing_link.id,
        entity_label=(
            f"{existing_link.prime_mover_asset_code} -> "
            f"{existing_link.tanker_asset_code}"
        ),
        remarks="Prime mover tanker link updated",
        request_path=f"/prime-mover-tanker-links/{link_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_link)

    return build_prime_mover_tanker_link_response(existing_link, db)


@app.post(
    "/prime-mover-tanker-links/{link_id}/close",
    response_model=PrimeMoverTankerLinkResponse,
)
def close_prime_mover_tanker_link(
    link_id: int,
    linked_to: date | None = None,
    remarks: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_link = (
        db.query(PrimeMoverTankerLink)
        .filter(PrimeMoverTankerLink.id == link_id)
        .first()
    )

    if not existing_link:
        raise HTTPException(
            status_code=404,
            detail="Prime Mover Tanker link not found",
        )

    close_date = linked_to or date.today()

    if close_date < existing_link.linked_from:
        raise HTTPException(
            status_code=400,
            detail="Close date cannot be earlier than Linked From",
        )

    before_data = build_prime_mover_tanker_link_audit_snapshot(
        existing_link,
        db,
    )

    existing_link.linked_to = close_date
    existing_link.status = "Inactive"

    cleaned_remarks = clean_optional_text(remarks)

    if cleaned_remarks:
        if existing_link.remarks:
            existing_link.remarks = (
                f"{existing_link.remarks}\nClose Remarks: {cleaned_remarks}"
            )
        else:
            existing_link.remarks = f"Close Remarks: {cleaned_remarks}"

    existing_link.updated_at = datetime.now()

    db.flush()

    after_data = build_prime_mover_tanker_link_audit_snapshot(
        existing_link,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Prime Mover Tanker Link",
        action="Close Prime Mover Tanker Link",
        current_user=current_user,
        entity_type="PrimeMoverTankerLink",
        entity_id=existing_link.id,
        entity_label=(
            f"{existing_link.prime_mover_asset_code} -> "
            f"{existing_link.tanker_asset_code}"
        ),
        remarks="Prime mover tanker link closed",
        request_path=f"/prime-mover-tanker-links/{link_id}/close",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_link)

    return build_prime_mover_tanker_link_response(existing_link, db)


@app.delete("/prime-mover-tanker-links/{link_id}")
def delete_prime_mover_tanker_link(
    link_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset",
        db,
    )

    existing_link = (
        db.query(PrimeMoverTankerLink)
        .filter(PrimeMoverTankerLink.id == link_id)
        .first()
    )

    if not existing_link:
        raise HTTPException(
            status_code=404,
            detail="Prime Mover Tanker link not found",
        )

    deleted_data = build_prime_mover_tanker_link_audit_snapshot(
        existing_link,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Prime Mover Tanker Link",
        action="Delete Prime Mover Tanker Link",
        current_user=current_user,
        entity_type="PrimeMoverTankerLink",
        entity_id=existing_link.id,
        entity_label=(
            f"{existing_link.prime_mover_asset_code} -> "
            f"{existing_link.tanker_asset_code}"
        ),
        remarks="Prime mover tanker link deleted",
        request_path=f"/prime-mover-tanker-links/{link_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.delete(existing_link)
    db.commit()

    return {
        "message": "Prime Mover Tanker link deleted successfully",
    }

# -------------------------
# Operation Type APIs
# -------------------------

@app.get("/operation-types", response_model=list[OperationTypeResponse])
def get_operation_types(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Operation Type", db)

    operation_types = db.query(OperationType).order_by(OperationType.id).all()
    return operation_types


@app.post("/operation-types", response_model=OperationTypeResponse)
def create_operation_type(
    operation_type: OperationTypeCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Type", db)

    existing_operation_type = (
        db.query(OperationType)
        .filter(OperationType.operation_type_code.ilike(operation_type.operation_type_code))
        .first()
    )
    if existing_operation_type:
        raise HTTPException(status_code=400, detail="Operation type code already exists")

    asset_type = (
        db.query(AssetType)
        .filter(AssetType.asset_type_code.ilike(operation_type.applicable_asset_type_code))
        .first()
    )
    if not asset_type:
        raise HTTPException(status_code=400, detail="Applicable asset type not found")

    new_operation_type = OperationType(
        operation_type_name=operation_type.operation_type_name.strip(),
        operation_type_code=operation_type.operation_type_code.strip(),
        operation_category=operation_type.operation_category,
        applicable_asset_type_code=operation_type.applicable_asset_type_code.strip(),
        requires_sender_location=operation_type.requires_sender_location,
        requires_receiver_location=operation_type.requires_receiver_location,
        requires_comparison=operation_type.requires_comparison,
        requires_approval=operation_type.requires_approval,
        description=clean_optional_text(operation_type.description),
        status=operation_type.status,
    )

    db.add(new_operation_type)
    db.flush()

    after_data = {
        "operation_type_name": new_operation_type.operation_type_name,
        "operation_type_code": new_operation_type.operation_type_code,
        "operation_category": new_operation_type.operation_category,
        "applicable_asset_type_code": new_operation_type.applicable_asset_type_code,
        "requires_sender_location": new_operation_type.requires_sender_location,
        "requires_receiver_location": new_operation_type.requires_receiver_location,
        "requires_comparison": new_operation_type.requires_comparison,
        "requires_approval": new_operation_type.requires_approval,
        "description": new_operation_type.description,
        "status": new_operation_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Create Operation Type",
        current_user=current_user,
        entity_type="OperationType",
        entity_id=new_operation_type.id,
        entity_label=f"{new_operation_type.operation_type_name} ({new_operation_type.operation_type_code})",
        remarks="Operation type created",
        request_path="/operation-types",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_operation_type)
    return new_operation_type


@app.put("/operation-types/{operation_type_id}", response_model=OperationTypeResponse)
def update_operation_type(
    operation_type_id: int,
    operation_type: OperationTypeCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Type", db)

    existing_operation_type = (
        db.query(OperationType)
        .filter(OperationType.id == operation_type_id)
        .first()
    )
    if not existing_operation_type:
        raise HTTPException(status_code=404, detail="Operation type not found")

    duplicate_operation_type = (
        db.query(OperationType)
        .filter(
            OperationType.operation_type_code.ilike(operation_type.operation_type_code),
            OperationType.id != operation_type_id,
        )
        .first()
    )
    if duplicate_operation_type:
        raise HTTPException(status_code=400, detail="Operation type code already exists")

    asset_type = (
        db.query(AssetType)
        .filter(AssetType.asset_type_code.ilike(operation_type.applicable_asset_type_code))
        .first()
    )
    if not asset_type:
        raise HTTPException(status_code=400, detail="Applicable asset type not found")

    before_data = {
        "operation_type_name": existing_operation_type.operation_type_name,
        "operation_type_code": existing_operation_type.operation_type_code,
        "operation_category": existing_operation_type.operation_category,
        "applicable_asset_type_code": existing_operation_type.applicable_asset_type_code,
        "requires_sender_location": existing_operation_type.requires_sender_location,
        "requires_receiver_location": existing_operation_type.requires_receiver_location,
        "requires_comparison": existing_operation_type.requires_comparison,
        "requires_approval": existing_operation_type.requires_approval,
        "description": existing_operation_type.description,
        "status": existing_operation_type.status,
    }

    existing_operation_type.operation_type_name = operation_type.operation_type_name.strip()
    existing_operation_type.operation_type_code = operation_type.operation_type_code.strip()
    existing_operation_type.operation_category = operation_type.operation_category
    existing_operation_type.applicable_asset_type_code = operation_type.applicable_asset_type_code.strip()
    existing_operation_type.requires_sender_location = operation_type.requires_sender_location
    existing_operation_type.requires_receiver_location = operation_type.requires_receiver_location
    existing_operation_type.requires_comparison = operation_type.requires_comparison
    existing_operation_type.requires_approval = operation_type.requires_approval
    existing_operation_type.description = clean_optional_text(operation_type.description)
    existing_operation_type.status = operation_type.status

    after_data = {
        "operation_type_name": existing_operation_type.operation_type_name,
        "operation_type_code": existing_operation_type.operation_type_code,
        "operation_category": existing_operation_type.operation_category,
        "applicable_asset_type_code": existing_operation_type.applicable_asset_type_code,
        "requires_sender_location": existing_operation_type.requires_sender_location,
        "requires_receiver_location": existing_operation_type.requires_receiver_location,
        "requires_comparison": existing_operation_type.requires_comparison,
        "requires_approval": existing_operation_type.requires_approval,
        "description": existing_operation_type.description,
        "status": existing_operation_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Update Operation Type",
        current_user=current_user,
        entity_type="OperationType",
        entity_id=existing_operation_type.id,
        entity_label=f"{existing_operation_type.operation_type_name} ({existing_operation_type.operation_type_code})",
        remarks="Operation type updated",
        request_path=f"/operation-types/{operation_type_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_operation_type)
    return existing_operation_type


@app.delete("/operation-types/{operation_type_id}")
def delete_operation_type(
    operation_type_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Type", db)

    existing_operation_type = (
        db.query(OperationType)
        .filter(OperationType.id == operation_type_id)
        .first()
    )
    if not existing_operation_type:
        raise HTTPException(status_code=404, detail="Operation type not found")

    operation_template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.operation_type_code.ilike(existing_operation_type.operation_type_code))
        .first()
    )
    if operation_template:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete operation type because operation templates exist for it",
        )

    operation_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.operation_type_code.ilike(existing_operation_type.operation_type_code))
        .first()
    )
    if operation_transaction:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete operation type because transactions exist for it",
        )

    deleted_data = {
        "operation_type_name": existing_operation_type.operation_type_name,
        "operation_type_code": existing_operation_type.operation_type_code,
        "operation_category": existing_operation_type.operation_category,
        "applicable_asset_type_code": existing_operation_type.applicable_asset_type_code,
        "requires_sender_location": existing_operation_type.requires_sender_location,
        "requires_receiver_location": existing_operation_type.requires_receiver_location,
        "requires_comparison": existing_operation_type.requires_comparison,
        "requires_approval": existing_operation_type.requires_approval,
        "description": existing_operation_type.description,
        "status": existing_operation_type.status,
    }

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Delete Operation Type",
        current_user=current_user,
        entity_type="OperationType",
        entity_id=existing_operation_type.id,
        entity_label=f"{existing_operation_type.operation_type_name} ({existing_operation_type.operation_type_code})",
        remarks="Operation type deleted",
        request_path=f"/operation-types/{operation_type_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_operation_type)
    db.commit()

    return {"message": "Operation type deleted successfully"}


# -------------------------
# Tank Operation Master APIs
# -------------------------

VALID_TANK_OPERATION_CATEGORIES = [
    "OPENING",
    "RECEIPT",
    "PRODUCTION",
    "DISPATCH",
    "DRAINING",
    "CLOSING",
    "ADJUSTMENT",
]

VALID_TANK_OPERATION_SIGNS = [
    "SET",
    "IN",
    "OUT",
    "NEUTRAL",
]


def normalize_code(value: str):
    return str(value or "").strip().upper()


def build_tank_operation_response(
    tank_operation: TankOperation,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(tank_operation.location_code))
        .first()
    )

    return {
        "id": tank_operation.id,
        "location_code": tank_operation.location_code,
        "location_name": location.location_name if location else "",
        "operation_code": tank_operation.operation_code,
        "operation_label": tank_operation.operation_label,
        "operation_category": tank_operation.operation_category,
        "operation_sign": tank_operation.operation_sign,
        "sort_order": tank_operation.sort_order,
        "description": tank_operation.description,
        "status": tank_operation.status,
        "created_at": tank_operation.created_at,
        "updated_at": tank_operation.updated_at,
    }

def build_vessel_operation_response(
    vessel_operation: VesselOperation,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(vessel_operation.location_code))
        .first()
    )

    return {
        "id": vessel_operation.id,
        "location_code": vessel_operation.location_code,
        "location_name": location.location_name if location else "",
        "applicable_asset_type_code": vessel_operation.applicable_asset_type_code,
        "operation_code": vessel_operation.operation_code,
        "operation_label": vessel_operation.operation_label,
        "operation_category": vessel_operation.operation_category,
        "operation_sign": vessel_operation.operation_sign,
        "sort_order": vessel_operation.sort_order,
        "description": vessel_operation.description,
        "status": vessel_operation.status,
        "created_at": vessel_operation.created_at,
        "updated_at": vessel_operation.updated_at,
    }

def build_tank_operation_audit_snapshot(
    tank_operation: TankOperation,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(tank_operation.location_code))
        .first()
    )

    return {
        "id": tank_operation.id,
        "location_code": tank_operation.location_code,
        "location_name": location.location_name if location else "",
        "operation_code": tank_operation.operation_code,
        "operation_label": tank_operation.operation_label,
        "operation_category": tank_operation.operation_category,
        "operation_sign": tank_operation.operation_sign,
        "sort_order": tank_operation.sort_order,
        "description": tank_operation.description,
        "status": tank_operation.status,
    }


def validate_tank_operation(
    tank_operation: TankOperationCreate,
    db: Session,
    tank_operation_id: int | None = None,
):
    location_code = normalize_code(tank_operation.location_code)
    operation_code = normalize_code(tank_operation.operation_code)
    operation_label = str(tank_operation.operation_label or "").strip()
    operation_category = normalize_code(tank_operation.operation_category)
    operation_sign = normalize_code(tank_operation.operation_sign)

    if location_code == "":
        raise HTTPException(
            status_code=400,
            detail="Location is required",
        )

    if operation_code == "":
        raise HTTPException(
            status_code=400,
            detail="Operation Code is required",
        )

    if operation_label == "":
        raise HTTPException(
            status_code=400,
            detail="Operation Label is required",
        )

    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(location_code))
        .first()
    )

    if not location:
        raise HTTPException(
            status_code=400,
            detail="Location not found",
        )

    if location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active locations can be used for Tank Operations",
        )

    if operation_category not in VALID_TANK_OPERATION_CATEGORIES:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid Operation Category. Allowed values are: "
                + ", ".join(VALID_TANK_OPERATION_CATEGORIES)
            ),
        )

    if operation_sign not in VALID_TANK_OPERATION_SIGNS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid Operation Sign. Allowed values are: "
                + ", ".join(VALID_TANK_OPERATION_SIGNS)
            ),
        )

    duplicate_code_query = db.query(TankOperation).filter(
        TankOperation.location_code.ilike(location_code),
        TankOperation.operation_code.ilike(operation_code),
    )

    duplicate_label_query = db.query(TankOperation).filter(
        TankOperation.location_code.ilike(location_code),
        TankOperation.operation_label.ilike(operation_label),
    )

    if tank_operation_id is not None:
        duplicate_code_query = duplicate_code_query.filter(
            TankOperation.id != tank_operation_id
        )

        duplicate_label_query = duplicate_label_query.filter(
            TankOperation.id != tank_operation_id
        )

    duplicate_code = duplicate_code_query.first()

    if duplicate_code:
        raise HTTPException(
            status_code=400,
            detail="Operation Code already exists for this location",
        )

    duplicate_label = duplicate_label_query.first()

    if duplicate_label:
        raise HTTPException(
            status_code=400,
            detail="Operation Label already exists for this location",
        )

    return {
        "location_code": location_code,
        "operation_code": operation_code,
        "operation_label": operation_label,
        "operation_category": operation_category,
        "operation_sign": operation_sign,
    }

def validate_vessel_operation(
    vessel_operation: VesselOperationCreate,
    db: Session,
    vessel_operation_id: int | None = None,
):
    location_code = normalize_code(vessel_operation.location_code)
    asset_type_code = normalize_code(vessel_operation.applicable_asset_type_code)

    operation_code = normalize_code(vessel_operation.operation_code)
    operation_label = str(vessel_operation.operation_label or "").strip()

    operation_category = normalize_code(vessel_operation.operation_category)
    operation_sign = normalize_code(vessel_operation.operation_sign)

    if location_code == "":
        raise HTTPException(status_code=400, detail="Location is required")

    if asset_type_code == "":
        raise HTTPException(status_code=400, detail="Applicable Asset Type is required")

    if operation_code == "":
        raise HTTPException(status_code=400, detail="Operation Code is required")

    if operation_label == "":
        raise HTTPException(status_code=400, detail="Operation Label is required")

    if operation_category == "":
        raise HTTPException(status_code=400, detail="Operation Category is required")

    if operation_sign not in ["IN", "OUT", "NEUTRAL", "SET"]:
        raise HTTPException(
            status_code=400,
            detail="Operation Sign must be IN / OUT / NEUTRAL / SET",
        )

    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(location_code))
        .first()
    )
    if not location:
        raise HTTPException(status_code=400, detail="Location not found")

    asset_type = (
        db.query(AssetType)
        .filter(AssetType.asset_type_code.ilike(asset_type_code))
        .first()
    )
    if not asset_type:
        raise HTTPException(status_code=400, detail="Asset Type not found")

    # Unique checks (code + label)
    code_q = db.query(VesselOperation).filter(
        VesselOperation.location_code.ilike(location_code),
        VesselOperation.applicable_asset_type_code.ilike(asset_type_code),
        VesselOperation.operation_code.ilike(operation_code),
    )
    label_q = db.query(VesselOperation).filter(
        VesselOperation.location_code.ilike(location_code),
        VesselOperation.applicable_asset_type_code.ilike(asset_type_code),
        VesselOperation.operation_label.ilike(operation_label),
    )

    if vessel_operation_id:
        code_q = code_q.filter(VesselOperation.id != vessel_operation_id)
        label_q = label_q.filter(VesselOperation.id != vessel_operation_id)

    if code_q.first():
        raise HTTPException(status_code=400, detail="Operation Code already exists")

    if label_q.first():
        raise HTTPException(status_code=400, detail="Operation Label already exists")

    return {
        "location_code": location_code,
        "asset_type_code": asset_type_code,
        "operation_code": operation_code,
        "operation_label": operation_label,
        "operation_category": operation_category,
        "operation_sign": operation_sign,
    }

@app.get(
    "/tank-operations",
    response_model=list[TankOperationResponse],
)
def get_tank_operations(
    location_code: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Operation",
        db,
    )

    query = db.query(TankOperation)

    cleaned_location_code = clean_optional_text(location_code)

    if cleaned_location_code:
        query = query.filter(
            TankOperation.location_code.ilike(cleaned_location_code)
        )

    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(TankOperation.status == cleaned_status)

    tank_operations = (
        query.order_by(
            TankOperation.location_code.asc(),
            TankOperation.sort_order.asc(),
            TankOperation.operation_label.asc(),
        )
        .all()
    )

    return [
        build_tank_operation_response(tank_operation, db)
        for tank_operation in tank_operations
    ]


@app.post(
    "/tank-operations",
    response_model=TankOperationResponse,
)
def create_tank_operation(
    tank_operation: TankOperationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Tank Operation",
        db,
    )

    validated_data = validate_tank_operation(tank_operation, db)

    new_tank_operation = TankOperation(
        location_code=validated_data["location_code"],
        operation_code=validated_data["operation_code"],
        operation_label=validated_data["operation_label"],
        operation_category=validated_data["operation_category"],
        operation_sign=validated_data["operation_sign"],
        sort_order=tank_operation.sort_order or 1,
        description=clean_optional_text(tank_operation.description),
        status=tank_operation.status,
    )

    db.add(new_tank_operation)
    db.flush()

    after_data = build_tank_operation_audit_snapshot(new_tank_operation, db)

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Create Tank Operation",
        current_user=current_user,
        entity_type="TankOperation",
        entity_id=new_tank_operation.id,
        entity_label=(
            f"{new_tank_operation.location_code} - "
            f"{new_tank_operation.operation_label}"
        ),
        remarks="Tank operation created",
        request_path="/tank-operations",
        details={
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(new_tank_operation)

    return build_tank_operation_response(new_tank_operation, db)


@app.put(
    "/tank-operations/{tank_operation_id}",
    response_model=TankOperationResponse,
)
def update_tank_operation(
    tank_operation_id: int,
    tank_operation: TankOperationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Tank Operation",
        db,
    )

    existing_tank_operation = (
        db.query(TankOperation)
        .filter(TankOperation.id == tank_operation_id)
        .first()
    )

    if not existing_tank_operation:
        raise HTTPException(
            status_code=404,
            detail="Tank Operation not found",
        )

    before_data = build_tank_operation_audit_snapshot(
        existing_tank_operation,
        db,
    )

    validated_data = validate_tank_operation(
        tank_operation,
        db,
        tank_operation_id,
    )

    existing_tank_operation.location_code = validated_data["location_code"]
    existing_tank_operation.operation_code = validated_data["operation_code"]
    existing_tank_operation.operation_label = validated_data["operation_label"]
    existing_tank_operation.operation_category = validated_data[
        "operation_category"
    ]
    existing_tank_operation.operation_sign = validated_data["operation_sign"]
    existing_tank_operation.sort_order = tank_operation.sort_order or 1
    existing_tank_operation.description = clean_optional_text(
        tank_operation.description
    )
    existing_tank_operation.status = tank_operation.status
    existing_tank_operation.updated_at = datetime.now()

    db.flush()

    after_data = build_tank_operation_audit_snapshot(
        existing_tank_operation,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Update Tank Operation",
        current_user=current_user,
        entity_type="TankOperation",
        entity_id=existing_tank_operation.id,
        entity_label=(
            f"{existing_tank_operation.location_code} - "
            f"{existing_tank_operation.operation_label}"
        ),
        remarks="Tank operation updated",
        request_path=f"/tank-operations/{tank_operation_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_tank_operation)

    return build_tank_operation_response(existing_tank_operation, db)


@app.delete("/tank-operations/{tank_operation_id}")
def delete_tank_operation(
    tank_operation_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Tank Operation",
        db,
    )

    existing_tank_operation = (
        db.query(TankOperation)
        .filter(TankOperation.id == tank_operation_id)
        .first()
    )

    if not existing_tank_operation:
        raise HTTPException(
            status_code=404,
            detail="Tank Operation not found",
        )

    deleted_data = build_tank_operation_audit_snapshot(
        existing_tank_operation,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Delete Tank Operation",
        current_user=current_user,
        entity_type="TankOperation",
        entity_id=existing_tank_operation.id,
        entity_label=(
            f"{existing_tank_operation.location_code} - "
            f"{existing_tank_operation.operation_label}"
        ),
        remarks="Tank operation deleted",
        request_path=f"/tank-operations/{tank_operation_id}",
        details={
            "deleted": deleted_data,
        },
    )

    db.delete(existing_tank_operation)
    db.commit()

    return {
        "message": "Tank operation deleted successfully"
    }

# -------------------------
# Vessel Operation Master APIs
# -------------------------

VALID_VESSEL_OPERATION_SIGNS = ["SET", "IN", "OUT", "NEUTRAL"]

def build_vessel_operation_response(vessel_operation: VesselOperation, db: Session):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(vessel_operation.location_code))
        .first()
    )

    return {
        "id": vessel_operation.id,
        "location_code": vessel_operation.location_code,
        "location_name": location.location_name if location else "",
        "applicable_asset_type_code": vessel_operation.applicable_asset_type_code,
        "operation_code": vessel_operation.operation_code,
        "operation_label": vessel_operation.operation_label,
        "operation_category": vessel_operation.operation_category,
        "operation_sign": vessel_operation.operation_sign,
        "show_in": vessel_operation.show_in,
        "sort_order": vessel_operation.sort_order,
        "description": vessel_operation.description,
        "status": vessel_operation.status,
        "created_at": vessel_operation.created_at,
        "updated_at": vessel_operation.updated_at,
    }

def validate_vessel_operation(v: VesselOperationCreate, db: Session, vessel_operation_id: int | None = None):
    location_code = normalize_code(v.location_code)
    asset_type_code = normalize_code(v.applicable_asset_type_code)
    operation_code = normalize_code(v.operation_code)

    operation_label = str(v.operation_label or "").strip()
    operation_category = normalize_code(v.operation_category)
    operation_sign = normalize_code(v.operation_sign)

    if not location_code:
        raise HTTPException(status_code=400, detail="Location is required")
    if not asset_type_code:
        raise HTTPException(status_code=400, detail="Applicable Asset Type is required")
    if not operation_code:
        raise HTTPException(status_code=400, detail="Operation Code is required")
    if not operation_label:
        raise HTTPException(status_code=400, detail="Operation Label is required")
    if not operation_category:
        raise HTTPException(status_code=400, detail="Operation Category is required")
    if operation_sign not in VALID_VESSEL_OPERATION_SIGNS:
        raise HTTPException(status_code=400, detail="Operation Sign must be SET / IN / OUT / NEUTRAL")

    show_in_raw = str(getattr(v, "show_in", "") or "").strip()
    if show_in_raw == "":
        show_in_raw = "Both"

    show_in = show_in_raw[:1].upper() + show_in_raw[1:].lower()
    if show_in not in ["Entry", "Tracking", "Both"]:
        raise HTTPException(status_code=400, detail="Show In must be Entry / Tracking / Both")

    if not db.query(Location).filter(Location.location_code.ilike(location_code)).first():
        raise HTTPException(status_code=400, detail="Location not found")

    if not db.query(AssetType).filter(AssetType.asset_type_code.ilike(asset_type_code)).first():
        raise HTTPException(status_code=400, detail="Asset Type not found")

    code_q = db.query(VesselOperation).filter(
        VesselOperation.location_code.ilike(location_code),
        VesselOperation.applicable_asset_type_code.ilike(asset_type_code),
        VesselOperation.operation_code.ilike(operation_code),
    )
    label_q = db.query(VesselOperation).filter(
        VesselOperation.location_code.ilike(location_code),
        VesselOperation.applicable_asset_type_code.ilike(asset_type_code),
        VesselOperation.operation_label.ilike(operation_label),
    )

    if vessel_operation_id:
        code_q = code_q.filter(VesselOperation.id != vessel_operation_id)
        label_q = label_q.filter(VesselOperation.id != vessel_operation_id)

    if code_q.first():
        raise HTTPException(status_code=400, detail="Operation Code already exists")
    if label_q.first():
        raise HTTPException(status_code=400, detail="Operation Label already exists")

    return {
        "location_code": location_code,
        "asset_type_code": asset_type_code,
        "operation_code": operation_code,
        "operation_label": operation_label,
        "operation_category": operation_category,
        "operation_sign": operation_sign,
        "show_in": show_in,
    }

@app.get("/vessel-operations", response_model=list[VesselOperationResponse])
def get_vessel_operations(
    location_code: str | None = None,
    applicable_asset_type_code: str | None = None,
    status: str | None = None,
    show_in: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Vessel Operation", db)

    q = db.query(VesselOperation)

    lc = clean_optional_text(location_code)
    if lc:
        q = q.filter(VesselOperation.location_code.ilike(lc))

    at = clean_optional_text(applicable_asset_type_code)
    if at:
        q = q.filter(VesselOperation.applicable_asset_type_code.ilike(at))

    st = clean_optional_text(status)
    if st:
        q = q.filter(VesselOperation.status == st)

    si = clean_optional_text(show_in)
    if si:
        normalized = si[:1].upper() + si[1:].lower()
        if normalized == "Entry":
            q = q.filter(VesselOperation.show_in.in_(["Entry", "Both"]))
        elif normalized == "Tracking":
            q = q.filter(VesselOperation.show_in.in_(["Tracking", "Both"]))
        elif normalized == "Both":
            q = q.filter(VesselOperation.show_in == "Both")

    rows = q.order_by(
        VesselOperation.location_code.asc(),
        VesselOperation.applicable_asset_type_code.asc(),
        VesselOperation.sort_order.asc(),
        VesselOperation.operation_label.asc(),
    ).all()

    return [build_vessel_operation_response(r, db) for r in rows]

@app.post("/vessel-operations", response_model=VesselOperationResponse)
def create_vessel_operation(
    vessel_operation: VesselOperationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Vessel Operation", db)

    d = validate_vessel_operation(vessel_operation, db)

    row = VesselOperation(
        location_code=d["location_code"],
        applicable_asset_type_code=d["asset_type_code"],
        operation_code=d["operation_code"],
        operation_label=d["operation_label"],
        operation_category=d["operation_category"],
        operation_sign=d["operation_sign"],
        show_in=d["show_in"],
        sort_order=vessel_operation.sort_order or 1,
        description=clean_optional_text(vessel_operation.description),
        status=vessel_operation.status or "Active",
    )

    db.add(row)
    db.commit()
    db.refresh(row)
    return build_vessel_operation_response(row, db)

@app.put("/vessel-operations/{vessel_operation_id}", response_model=VesselOperationResponse)
def update_vessel_operation(
    vessel_operation_id: int,
    vessel_operation: VesselOperationCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Vessel Operation", db)

    existing = db.query(VesselOperation).filter(VesselOperation.id == vessel_operation_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Vessel operation not found")

    d = validate_vessel_operation(vessel_operation, db, vessel_operation_id)

    existing.location_code = d["location_code"]
    existing.applicable_asset_type_code = d["asset_type_code"]
    existing.operation_code = d["operation_code"]
    existing.operation_label = d["operation_label"]
    existing.operation_category = d["operation_category"]
    existing.operation_sign = d["operation_sign"]
    existing.show_in = d["show_in"]
    existing.sort_order = vessel_operation.sort_order or 1
    existing.description = clean_optional_text(vessel_operation.description)
    existing.status = vessel_operation.status or "Active"
    existing.updated_at = datetime.now()

    db.commit()
    db.refresh(existing)
    return build_vessel_operation_response(existing, db)

@app.delete("/vessel-operations/{vessel_operation_id}")
def delete_vessel_operation(
    vessel_operation_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Vessel Operation", db)

    existing = db.query(VesselOperation).filter(VesselOperation.id == vessel_operation_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Vessel operation not found")

    db.delete(existing)
    db.commit()
    return {"message": "Vessel operation deleted successfully"}

# -------------------------
# Vessel Stock Ledger (Approved-only) APIs
# -------------------------

def get_value_text(db: Session, transaction_id: int, field_code: str):
    row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == field_code,
        )
        .first()
    )
    if not row:
        return None
    if row.field_value is None:
        return None
    if isinstance(row.field_value, (dict, list)):
        return row.field_value
    return str(row.field_value).strip()


def extract_multitank_net(payload: dict):
    net = (((payload or {}).get("calculated") or {}).get("net") or {})
    qty = safe_float(net.get("TOV"))
    water = safe_float(net.get("FW"))
    nsv = safe_float(net.get("NSV"))
    # mapping/ledger should use positive magnitudes
    return abs(qty), abs(water), abs(nsv)


def create_or_update_vessel_stock_ledger_from_approved_transaction(
    db: Session,
    transaction: OperationTransaction,
    current_user: User,
):
    if transaction.status != "Approved":
        return None

    template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.id == transaction.operation_template_id)
        .first()
    )
    if not template:
        return None

    layout = str(template.entry_layout_type or "").strip()

    if layout not in ["Vessel Cycle", "Stock Movement"]:
        return None

    asset = (
        db.query(Asset)
        .filter(Asset.asset_code.ilike(transaction.primary_asset_code))
        .first()
    )
    if not asset:
        return None

    created_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )

    # Resolve operation code from values
    vessel_operation_code = get_value_text(db, transaction.id, "vessel_operation_code")
    vessel_operation_code = clean_optional_text(vessel_operation_code if isinstance(vessel_operation_code, str) else "")

    # Reference
    movement_reference = (
        get_value_text(db, transaction.id, "movement_reference")
        or get_value_text(db, transaction.id, "shuttle_number")
        or get_value_text(db, transaction.id, "reference_number")
    )
    movement_reference = clean_optional_text(movement_reference if isinstance(movement_reference, str) else "")

    # Quantities
    qty_bbl = water_bbl = nsv_bbl = 0
    opening_stock = opening_water = closing_stock = closing_water = 0
    net_stock = net_water = net_nsv = 0

    if layout == "Vessel Cycle":
        qty_bbl = safe_float(get_value_text(db, transaction.id, "quantity_bbl") or get_value_text(db, transaction.id, "gross_qty_bbl"))
        water_bbl = safe_float(get_value_text(db, transaction.id, "water_bbl"))
        nsv_bbl = safe_float(get_value_text(db, transaction.id, "nsv_bbl"))
    else:
        opening_stock = safe_float(get_value_text(db, transaction.id, "opening_stock"))
        opening_water = safe_float(get_value_text(db, transaction.id, "opening_water"))
        closing_stock = safe_float(get_value_text(db, transaction.id, "closing_stock"))
        closing_water = safe_float(get_value_text(db, transaction.id, "closing_water"))

        net_stock = safe_float(get_value_text(db, transaction.id, "net_stock")) or (closing_stock - opening_stock)
        net_water = safe_float(get_value_text(db, transaction.id, "net_water")) or (closing_water - opening_water)
        net_nsv = safe_float(get_value_text(db, transaction.id, "net_nsv")) or (net_stock - net_water)

        qty_bbl = net_stock
        water_bbl = net_water
        nsv_bbl = net_nsv

    # Lookup soft-coded vessel operation details (optional but recommended)
    vessel_op = None
    if vessel_operation_code:
        vessel_op = (
            db.query(VesselOperation)
            .filter(
                VesselOperation.location_code.ilike(transaction.origin_location_code),
                VesselOperation.applicable_asset_type_code.ilike(asset.asset_type_code),
                VesselOperation.operation_code.ilike(vessel_operation_code),
                VesselOperation.status == "Active",
            )
            .first()
        )

    ledger = (
        db.query(VesselStockLedger)
        .filter(VesselStockLedger.transaction_id == transaction.id)
        .first()
    )

    if not ledger:
        ledger = VesselStockLedger(transaction_id=transaction.id)
        db.add(ledger)

    ledger.ticket_number = get_transaction_ticket_number(transaction)
    ledger.operation_number = transaction.operation_number
    ledger.status = transaction.status

    ledger.location_code = transaction.origin_location_code

    ledger.vessel_asset_code = asset.asset_code
    ledger.vessel_asset_name = asset.asset_name
    ledger.vessel_asset_type_code = asset.asset_type_code

    ledger.operation_date = transaction.operation_date
    ledger.product_name = transaction.product_name

    ledger.movement_reference = movement_reference

    ledger.vessel_operation_code = vessel_operation_code
    ledger.vessel_operation_label = vessel_op.operation_label if vessel_op else vessel_operation_code
    ledger.vessel_operation_category = vessel_op.operation_category if vessel_op else None
    ledger.vessel_operation_sign = vessel_op.operation_sign if vessel_op else None

    ledger.qty_bbl = qty_bbl
    ledger.water_bbl = water_bbl
    ledger.nsv_bbl = nsv_bbl

    ledger.opening_stock = opening_stock
    ledger.opening_water = opening_water
    ledger.closing_stock = closing_stock
    ledger.closing_water = closing_water
    ledger.net_stock = net_stock
    ledger.net_water = net_water
    ledger.net_nsv = net_nsv

    ledger.created_by = ledger.created_by or created_by
    ledger.updated_at = datetime.now()

    db.flush()

    return ledger


@app.get("/vessel-stock-ledger", response_model=list[VesselStockLedgerResponse])
def get_vessel_stock_ledger(
    location_code: str | None = None,
    vessel_asset_code: str | None = None,
    reference_number: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Operation Transaction", db)

    query = db.query(VesselStockLedger)

    lc = clean_optional_text(location_code)
    if lc:
        query = query.filter(VesselStockLedger.location_code.ilike(lc))

    ac = clean_optional_text(vessel_asset_code)
    if ac:
        query = query.filter(VesselStockLedger.vessel_asset_code.ilike(ac))

    ref = clean_optional_text(reference_number)
    if ref:
        query = query.filter(VesselStockLedger.movement_reference.ilike(ref))

    if clean_optional_text(date_from):
        query = query.filter(VesselStockLedger.operation_date >= date.fromisoformat(date_from))

    if clean_optional_text(date_to):
        query = query.filter(VesselStockLedger.operation_date <= date.fromisoformat(date_to))

    rows = query.order_by(VesselStockLedger.operation_date.desc(), VesselStockLedger.id.desc()).all()

    # attach location_name
    loc_map = {l.location_code: l.location_name for l in db.query(Location).all()}

    results = []
    for r in rows:
        item = r.__dict__.copy()
        item["location_name"] = loc_map.get(r.location_code, "")
        results.append(item)

    return results

# -------------------------
# Movement Mapping + Comparison APIs
# -------------------------

def extract_transaction_quantities(db: Session, transaction: OperationTransaction):
    # Multi-tank payload (barges)
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction.id,
            OperationTransactionValue.field_code == "multi_tank_payload",
        )
        .first()
    )

    if payload_row and isinstance(payload_row.field_value, dict):
        net = (((payload_row.field_value or {}).get("calculated") or {}).get("net") or {})
        qty = abs(safe_float(net.get("TOV")))
        water = abs(safe_float(net.get("FW")))
        nsv = abs(safe_float(net.get("NSV")))
        return qty, water, nsv

    # ✅ Shuttle payload (Shuttle Tracking)
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction.id,
            OperationTransactionValue.field_code == "shuttle_payload",
        )
        .first()
    )

    if payload_row and isinstance(payload_row.field_value, dict):
        net = (((payload_row.field_value or {}).get("calculated") or {}).get("net") or {})
        qty = abs(safe_float(net.get("TOV")))
        water = abs(safe_float(net.get("FW")))
        nsv = abs(safe_float(net.get("NSV")))
        return qty, water, nsv

    # Vessel Cycle manual
    def get_val(code):
        row = (
            db.query(OperationTransactionValue)
            .filter(
                OperationTransactionValue.transaction_id == transaction.id,
                OperationTransactionValue.field_code == code,
            )
            .first()
        )
        if not row:
            return None
        return row.field_value

    qty = safe_float(get_val("quantity_bbl") or get_val("gross_qty_bbl"))
    water = safe_float(get_val("water_bbl"))
    nsv = safe_float(get_val("nsv_bbl"))
    if qty or water or nsv:
        return abs(qty), abs(water), abs(nsv)

    # Stock Movement manual (net)
    net_stock = safe_float(get_val("net_stock"))
    net_water = safe_float(get_val("net_water"))
    net_nsv = safe_float(get_val("net_nsv"))
    if net_stock or net_water or net_nsv:
        return abs(net_stock), abs(net_water), abs(net_nsv)

    return 0, 0, 0


def recompute_mapping_comparison(db: Session, mapping_id: int):
    items = db.query(MovementMappingItem).filter(MovementMappingItem.mapping_id == mapping_id).all()

    source = [i for i in items if str(i.role).upper() == "SOURCE"]
    target = [i for i in items if str(i.role).upper() == "TARGET"]

    source_qty = sum(safe_float(i.qty_bbl) for i in source)
    source_water = sum(safe_float(i.water_bbl) for i in source)
    source_nsv = sum(safe_float(i.nsv_bbl) for i in source)

    target_qty = sum(safe_float(i.qty_bbl) for i in target)
    target_water = sum(safe_float(i.water_bbl) for i in target)
    target_nsv = sum(safe_float(i.nsv_bbl) for i in target)

    diff_nsv = target_nsv - source_nsv
    diff_pct = (diff_nsv / source_nsv * 100) if source_nsv else 0

    summary = {
        "source": {"qty_bbl": source_qty, "water_bbl": source_water, "nsv_bbl": source_nsv},
        "target": {"qty_bbl": target_qty, "water_bbl": target_water, "nsv_bbl": target_nsv},
        "diff": {"nsv_bbl": diff_nsv, "nsv_percent": diff_pct},
    }

    cmp_row = db.query(MovementMappingComparison).filter(MovementMappingComparison.mapping_id == mapping_id).first()
    if not cmp_row:
        cmp_row = MovementMappingComparison(mapping_id=mapping_id)
        db.add(cmp_row)

    cmp_row.source_qty_bbl = source_qty
    cmp_row.source_water_bbl = source_water
    cmp_row.source_nsv_bbl = source_nsv

    cmp_row.target_qty_bbl = target_qty
    cmp_row.target_water_bbl = target_water
    cmp_row.target_nsv_bbl = target_nsv

    cmp_row.diff_nsv_bbl = diff_nsv
    cmp_row.diff_nsv_percent = diff_pct

    cmp_row.summary_json = summary
    cmp_row.updated_at = datetime.now()

    db.flush()
    return cmp_row


def build_mapping_response(db: Session, mapping: MovementMapping):
    items = db.query(MovementMappingItem).filter(MovementMappingItem.mapping_id == mapping.id).order_by(MovementMappingItem.id.asc()).all()
    cmp_row = db.query(MovementMappingComparison).filter(MovementMappingComparison.mapping_id == mapping.id).first()

    return {
        "id": mapping.id,
        "mapping_type": mapping.mapping_type,
        "location_code": mapping.location_code,
        "reference_number": mapping.reference_number,
        "product_name": mapping.product_name,
        "status": mapping.status,
        "remarks": mapping.remarks,
        "created_by": mapping.created_by,
        "closed_by": mapping.closed_by,
        "closed_at": mapping.closed_at,
        "created_at": mapping.created_at,
        "updated_at": mapping.updated_at,
        "items": items,
        "comparison": cmp_row,
    }


@app.get("/movement-mappings", response_model=list[MovementMappingResponse])
def list_movement_mappings(
    mapping_type: str | None = None,
    location_code: str | None = None,
    reference_number: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Movement Mapping", db)

    q = db.query(MovementMapping)

    if clean_optional_text(mapping_type):
        q = q.filter(MovementMapping.mapping_type.ilike(mapping_type))

    if clean_optional_text(location_code):
        q = q.filter(MovementMapping.location_code.ilike(location_code))

    if clean_optional_text(reference_number):
        q = q.filter(MovementMapping.reference_number.ilike(reference_number))

    if clean_optional_text(status):
        q = q.filter(MovementMapping.status.ilike(status))

    rows = q.order_by(MovementMapping.created_at.desc(), MovementMapping.id.desc()).all()
    return [build_mapping_response(db, r) for r in rows]


@app.get("/movement-mappings/{mapping_id}", response_model=MovementMappingResponse)
def get_movement_mapping(
    mapping_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Movement Mapping", db)

    m = db.query(MovementMapping).filter(MovementMapping.id == mapping_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Mapping not found")

    return build_mapping_response(db, m)


@app.post("/movement-mappings", response_model=MovementMappingResponse)
def create_movement_mapping(
    request: MovementMappingCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Movement Mapping", db)

    created_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )

    m = MovementMapping(
        mapping_type=normalize_code(request.mapping_type),
        location_code=normalize_code(request.location_code),
        reference_number=str(request.reference_number or "").strip(),
        product_name=clean_optional_text(request.product_name),
        remarks=clean_optional_text(request.remarks),
        status="OPEN",
        created_by=created_by,
        updated_at=datetime.now(),
    )

    db.add(m)
    db.commit()
    db.refresh(m)

    return build_mapping_response(db, m)


@app.post("/movement-mappings/{mapping_id}/items", response_model=MovementMappingResponse)
def add_mapping_items(
    mapping_id: int,
    request: MovementMappingItemAddRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Movement Mapping", db)

    mapping = db.query(MovementMapping).filter(MovementMapping.id == mapping_id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")

    if str(mapping.status).upper() == "CLOSED":
        raise HTTPException(status_code=400, detail="Mapping is CLOSED")

    role = normalize_code(request.role)
    if role not in ["SOURCE", "TARGET"]:
        raise HTTPException(status_code=400, detail="role must be SOURCE or TARGET")

    for tid in request.transaction_ids:
        tx = db.query(OperationTransaction).filter(OperationTransaction.id == tid).first()
        if not tx:
            raise HTTPException(status_code=404, detail=f"Transaction {tid} not found")

        if tx.status != "Approved":
            raise HTTPException(status_code=400, detail=f"Only Approved transactions can be mapped (ticket {tid})")

        exists = db.query(MovementMappingItem).filter(
            MovementMappingItem.mapping_id == mapping_id,
            MovementMappingItem.transaction_id == tid,
        ).first()
        if exists:
            continue

        qty, water, nsv = extract_transaction_quantities(db, tx)

        asset_type_code = None
        asset = db.query(Asset).filter(Asset.asset_code.ilike(tx.primary_asset_code)).first()
        if asset:
            asset_type_code = asset.asset_type_code

        item = MovementMappingItem(
            mapping_id=mapping_id,
            transaction_id=tid,
            role=role,
            asset_code=tx.primary_asset_code,
            asset_type_code=asset_type_code,
            ticket_number=get_transaction_ticket_number(tx),
            operation_date=tx.operation_date,
            qty_bbl=qty,
            water_bbl=water,
            nsv_bbl=nsv,
        )
        db.add(item)

    db.flush()
    recompute_mapping_comparison(db, mapping_id)

    mapping.updated_at = datetime.now()
    db.commit()
    db.refresh(mapping)

    return build_mapping_response(db, mapping)


@app.delete("/movement-mappings/{mapping_id}/items/{item_id}", response_model=MovementMappingResponse)
def remove_mapping_item(
    mapping_id: int,
    item_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Movement Mapping", db)

    mapping = db.query(MovementMapping).filter(MovementMapping.id == mapping_id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")

    if str(mapping.status).upper() == "CLOSED":
        raise HTTPException(status_code=400, detail="Mapping is CLOSED")

    item = db.query(MovementMappingItem).filter(
        MovementMappingItem.id == item_id,
        MovementMappingItem.mapping_id == mapping_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    db.delete(item)
    db.flush()

    recompute_mapping_comparison(db, mapping_id)

    mapping.updated_at = datetime.now()
    db.commit()
    db.refresh(mapping)

    return build_mapping_response(db, mapping)


@app.post("/movement-mappings/{mapping_id}/close", response_model=MovementMappingResponse)
def close_mapping(
    mapping_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Movement Mapping", db)

    mapping = db.query(MovementMapping).filter(MovementMapping.id == mapping_id).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")

    if str(mapping.status).upper() == "CLOSED":
        return build_mapping_response(db, mapping)

    mapping.status = "CLOSED"
    mapping.closed_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )
    mapping.closed_at = datetime.now()
    mapping.updated_at = datetime.now()

    db.commit()
    db.refresh(mapping)

    return build_mapping_response(db, mapping)

# -------------------------
# Tank Stock Ledger APIs
# -------------------------

def build_tank_stock_ledger_response(
    ledger: TankStockLedger,
    db: Session,
):
    location = get_location_by_code(ledger.location_code, db)

    return {
        "id": ledger.id,
        "transaction_id": ledger.transaction_id,
        "ticket_number": ledger.ticket_number,
        "operation_number": ledger.operation_number,
        "location_code": ledger.location_code,
        "location_name": location.location_name if location else "",
        "tank_asset_code": ledger.tank_asset_code,
        "tank_asset_name": ledger.tank_asset_name,
        "operation_date": ledger.operation_date,
        "product_name": ledger.product_name,
        "accounting_date": ledger.accounting_date,
        "accounting_day_start": ledger.accounting_day_start,
        "accounting_day_end": ledger.accounting_day_end,
        "accounting_day_setting_id": ledger.accounting_day_setting_id,
        "tank_operation_code": ledger.tank_operation_code,
        "tank_operation_label": ledger.tank_operation_label,
        "tank_operation_category": ledger.tank_operation_category,
        "tank_operation_sign": ledger.tank_operation_sign,
        "movement_gsv_bbl": ledger.movement_gsv_bbl or 0,
        "movement_nsv_bbl": ledger.movement_nsv_bbl or 0,
        "movement_lt": ledger.movement_lt or 0,
        "movement_mt": ledger.movement_mt or 0,
        "stock_gsv_bbl": ledger.stock_gsv_bbl or 0,
        "stock_nsv_bbl": ledger.stock_nsv_bbl or 0,
        "stock_lt": ledger.stock_lt or 0,
        "stock_mt": ledger.stock_mt or 0,
        "previous_stock_gsv_bbl": ledger.previous_stock_gsv_bbl or 0,
        "previous_stock_nsv_bbl": ledger.previous_stock_nsv_bbl or 0,
        "previous_stock_lt": ledger.previous_stock_lt or 0,
        "previous_stock_mt": ledger.previous_stock_mt or 0,
        "running_balance_gsv_bbl": ledger.running_balance_gsv_bbl or 0,
        "running_balance_nsv_bbl": ledger.running_balance_nsv_bbl or 0,
        "running_balance_lt": ledger.running_balance_lt or 0,
        "running_balance_mt": ledger.running_balance_mt or 0,
        "status": ledger.status,
        "created_by": ledger.created_by,
        "remarks": ledger.remarks,
        "created_at": ledger.created_at,
        "updated_at": ledger.updated_at,
    }


def get_filtered_tank_stock_ledger_rows(
    db: Session,
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
):
    query = db.query(TankStockLedger)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)
    cleaned_status = clean_optional_text(status)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    if date_from:
        query = query.filter(TankStockLedger.accounting_date >= date_from)

    if date_to:
        query = query.filter(TankStockLedger.accounting_date <= date_to)

    if cleaned_status:
        query = query.filter(TankStockLedger.status == cleaned_status)

    return (
        query.order_by(
            TankStockLedger.location_code.asc(),
            TankStockLedger.tank_asset_code.asc(),
            TankStockLedger.accounting_date.asc(),
            TankStockLedger.operation_date.asc(),
            TankStockLedger.id.asc(),
        )
        .all()
    )


def parse_date_filter(value: str | None, field_name: str):
    cleaned_value = clean_optional_text(value)

    if not cleaned_value:
        return None

    try:
        return date.fromisoformat(cleaned_value)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"{field_name} must be in YYYY-MM-DD format",
        )


def build_date_range(start_date: date, end_date: date):
    if end_date < start_date:
        raise HTTPException(
            status_code=400,
            detail="Date To cannot be earlier than Date From",
        )

    dates = []
    current_date = start_date

    while current_date <= end_date:
        dates.append(current_date)
        current_date = current_date + timedelta(days=1)

    return dates


def get_active_location_day_setting(db: Session, location_code: str, on_date: date):
    return (
        db.query(LocationAccountingDaySetting)
        .filter(
            LocationAccountingDaySetting.location_code.ilike(location_code),
            LocationAccountingDaySetting.status == "Active",
            LocationAccountingDaySetting.effective_from <= on_date,
            or_(
                LocationAccountingDaySetting.effective_to.is_(None),
                LocationAccountingDaySetting.effective_to >= on_date,
            ),
        )
        .order_by(
            LocationAccountingDaySetting.effective_from.desc(),
            LocationAccountingDaySetting.id.desc(),
        )
        .first()
    )


def compute_accounting_date(
    op_date: date,
    event_time: str | None,
    day_start_time: datetime_time,
):
    if not event_time:
        return op_date

    try:
        hh, mm = event_time.split(":")
        t = datetime_time(int(hh), int(mm))
    except Exception:
        return op_date

    return op_date - timedelta(days=1) if t < day_start_time else op_date


def combine_operation_datetime(op_date: date, event_time: str | None, tz_name: str):
    try:
        if not event_time:
            return None
        hh, mm = event_time.split(":")
        dt = datetime(op_date.year, op_date.month, op_date.day, int(hh), int(mm))
        return dt.replace(tzinfo=ZoneInfo(tz_name))
    except Exception:
        return None


def get_tank_stock_rows_for_daily_summary(
    db: Session,
    location_code: str | None,
    tank_asset_code: str | None,
    product_name: str | None,
    date_to_value: date,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.accounting_date != None,
        TankStockLedger.accounting_date <= date_to_value,
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    return (
        query.order_by(
            TankStockLedger.location_code.asc(),
            TankStockLedger.tank_asset_code.asc(),
            TankStockLedger.product_name.asc(),
            TankStockLedger.accounting_date.asc(),
            TankStockLedger.operation_date.asc(),
            TankStockLedger.id.asc(),
        )
        .all()
    )


def get_stock_snapshot_values(row: TankStockLedger):
    stock_gsv = safe_float(row.stock_gsv_bbl)
    stock_nsv = safe_float(row.stock_nsv_bbl)
    stock_lt = safe_float(row.stock_lt)
    stock_mt = safe_float(row.stock_mt)

    # Backward compatibility for old rows created before stock_* existed.
    if stock_gsv == 0 and stock_nsv == 0:
        stock_gsv = safe_float(row.running_balance_gsv_bbl)
        stock_nsv = safe_float(row.running_balance_nsv_bbl)
        stock_lt = safe_float(row.running_balance_lt)
        stock_mt = safe_float(row.running_balance_mt)

    return {
        "gsv": stock_gsv,
        "nsv": stock_nsv,
        "lt": stock_lt,
        "mt": stock_mt,
    }

# -------------------------
# Out-Turn Report Helpers
# -------------------------

def get_ledger_operation_datetime(row: TankStockLedger):
    try:
        payload = row.source_payload or {}
        inputs = payload.get("inputs") or {}

        gauging_date = clean_optional_text(inputs.get("gaugingDate"))
        gauging_time = clean_optional_text(inputs.get("gaugingTime"))

        if gauging_date and gauging_time:
            return datetime.fromisoformat(f"{gauging_date}T{gauging_time}")
    except Exception:
        pass

    if row.accounting_day_start is not None:
        return row.accounting_day_start

    if row.operation_date is not None:
        return datetime.combine(row.operation_date, datetime_time(0, 0))

    return None


def build_out_turn_report_response(
    row: TankStockLedger,
    db: Session,
):
    location = get_location_by_code(row.location_code, db)

    operation_datetime = get_ledger_operation_datetime(row)

    previous_gsv = safe_float(row.previous_stock_gsv_bbl)
    previous_nsv = safe_float(row.previous_stock_nsv_bbl)
    previous_lt = safe_float(row.previous_stock_lt)
    previous_mt = safe_float(row.previous_stock_mt)

    stock_snapshot = get_stock_snapshot_values(row)

    stock_after_gsv = stock_snapshot["gsv"]
    stock_after_nsv = stock_snapshot["nsv"]
    stock_after_lt = stock_snapshot["lt"]
    stock_after_mt = stock_snapshot["mt"]

    movement_gsv = safe_float(row.movement_gsv_bbl)
    movement_nsv = safe_float(row.movement_nsv_bbl)
    movement_lt = safe_float(row.movement_lt)
    movement_mt = safe_float(row.movement_mt)

    sign = str(row.tank_operation_sign or "").upper()

    net_receipt_gsv = 0
    net_receipt_nsv = 0
    net_receipt_lt = 0
    net_receipt_mt = 0

    net_dispatch_gsv = 0
    net_dispatch_nsv = 0
    net_dispatch_lt = 0
    net_dispatch_mt = 0

    signed_net_gsv = 0
    signed_net_nsv = 0
    signed_net_lt = 0
    signed_net_mt = 0

    if sign == "IN":
        net_receipt_gsv = movement_gsv
        net_receipt_nsv = movement_nsv
        net_receipt_lt = movement_lt
        net_receipt_mt = movement_mt

        signed_net_gsv = movement_gsv
        signed_net_nsv = movement_nsv
        signed_net_lt = movement_lt
        signed_net_mt = movement_mt

    elif sign == "OUT":
        net_dispatch_gsv = movement_gsv
        net_dispatch_nsv = movement_nsv
        net_dispatch_lt = movement_lt
        net_dispatch_mt = movement_mt

        signed_net_gsv = movement_gsv * -1
        signed_net_nsv = movement_nsv * -1
        signed_net_lt = movement_lt * -1
        signed_net_mt = movement_mt * -1

    elif sign == "SET":
        # SET rows are stock declaration rows, not receipt/dispatch movement rows.
        signed_net_gsv = 0
        signed_net_nsv = 0
        signed_net_lt = 0
        signed_net_mt = 0

    elif sign == "NEUTRAL":
        signed_net_gsv = 0
        signed_net_nsv = 0
        signed_net_lt = 0
        signed_net_mt = 0

    return {
        "ledger_id": row.id,
        "transaction_id": row.transaction_id,
        "ticket_number": row.ticket_number,
        "operation_number": row.operation_number,
        "accounting_date": row.accounting_date,
        "operation_datetime": operation_datetime,
        "location_code": row.location_code,
        "location_name": location.location_name if location else "",
        "tank_asset_code": row.tank_asset_code,
        "tank_asset_name": row.tank_asset_name,
        "product_name": row.product_name,
        "tank_operation_code": row.tank_operation_code,
        "tank_operation_label": row.tank_operation_label,
        "tank_operation_category": row.tank_operation_category,
        "tank_operation_sign": row.tank_operation_sign,
        "previous_stock_gsv_bbl": round(previous_gsv, 3),
        "previous_stock_nsv_bbl": round(previous_nsv, 3),
        "previous_stock_lt": round(previous_lt, 3),
        "previous_stock_mt": round(previous_mt, 3),
        "stock_after_gsv_bbl": round(stock_after_gsv, 3),
        "stock_after_nsv_bbl": round(stock_after_nsv, 3),
        "stock_after_lt": round(stock_after_lt, 3),
        "stock_after_mt": round(stock_after_mt, 3),
        "net_receipt_gsv_bbl": round(net_receipt_gsv, 3),
        "net_receipt_nsv_bbl": round(net_receipt_nsv, 3),
        "net_receipt_lt": round(net_receipt_lt, 3),
        "net_receipt_mt": round(net_receipt_mt, 3),
        "net_dispatch_gsv_bbl": round(net_dispatch_gsv, 3),
        "net_dispatch_nsv_bbl": round(net_dispatch_nsv, 3),
        "net_dispatch_lt": round(net_dispatch_lt, 3),
        "net_dispatch_mt": round(net_dispatch_mt, 3),
        "signed_net_movement_gsv_bbl": round(signed_net_gsv, 3),
        "signed_net_movement_nsv_bbl": round(signed_net_nsv, 3),
        "signed_net_movement_lt": round(signed_net_lt, 3),
        "signed_net_movement_mt": round(signed_net_mt, 3),
        "status": row.status,
        "remarks": row.remarks,
    }


def get_out_turn_report_rows(
    db: Session,
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
):
    query = db.query(TankStockLedger)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)
    cleaned_status = clean_optional_text(status)

    if cleaned_status:
        query = query.filter(TankStockLedger.status == cleaned_status)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value:
        query = query.filter(TankStockLedger.accounting_date >= date_from_value)

    if date_to_value:
        query = query.filter(TankStockLedger.accounting_date <= date_to_value)

    rows = query.all()

    # Display order is global chronological order.
    # Net movement itself must NOT depend on this display order.
    # Net movement comes from ledger.previous_stock_* and movement_*,
    # which are calculated per location + tank + product group.
    rows = sorted(
        rows,
        key=lambda row: (
            row.accounting_date or date.min,
            get_ledger_operation_datetime(row) or datetime.min,
            row.location_code or "",
            row.tank_asset_code or "",
            row.product_name or "",
            row.id,
        ),
    )

    return rows

def build_tank_stock_daily_summary_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    date_from_value: date,
    date_to_value: date,
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    daily_summary_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key

        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                row.accounting_day_start or datetime.min,
                row.operation_date or date.min,
                row.id,
            ),
        )

        tank_asset_name = ""
        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_gsv = 0
        previous_closing_nsv = 0
        previous_closing_lt = 0
        previous_closing_mt = 0

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            last_before_period = rows_before_period[-1]
            previous_snapshot = get_stock_snapshot_values(last_before_period)

            previous_closing_gsv = previous_snapshot["gsv"]
            previous_closing_nsv = previous_snapshot["nsv"]
            previous_closing_lt = previous_snapshot["lt"]
            previous_closing_mt = previous_snapshot["mt"]

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    row.accounting_day_start or datetime.min,
                    row.operation_date or date.min,
                    row.id,
                ),
            )

            # Default opening is previous accounting day's actual closing.
            opening_gsv = previous_closing_gsv
            opening_nsv = previous_closing_nsv
            opening_lt = previous_closing_lt
            opening_mt = previous_closing_mt

            # If user entered an explicit Opening Stock, use its stock snapshot as opening.
            opening_rows = [
                row
                for row in day_rows
                if str(row.tank_operation_category or "").upper() == "OPENING"
            ]

            if opening_rows:
                opening_snapshot = get_stock_snapshot_values(opening_rows[-1])

                opening_gsv = opening_snapshot["gsv"]
                opening_nsv = opening_snapshot["nsv"]
                opening_lt = opening_snapshot["lt"]
                opening_mt = opening_snapshot["mt"]

            total_in_gsv = 0
            total_in_nsv = 0
            total_in_lt = 0
            total_in_mt = 0

            total_out_gsv = 0
            total_out_nsv = 0
            total_out_lt = 0
            total_out_mt = 0

            for row in day_rows:
                sign = str(row.tank_operation_sign or "").upper()

                if sign == "IN":
                    total_in_gsv += safe_float(row.movement_gsv_bbl)
                    total_in_nsv += safe_float(row.movement_nsv_bbl)
                    total_in_lt += safe_float(row.movement_lt)
                    total_in_mt += safe_float(row.movement_mt)

                elif sign == "OUT":
                    total_out_gsv += safe_float(row.movement_gsv_bbl)
                    total_out_nsv += safe_float(row.movement_nsv_bbl)
                    total_out_lt += safe_float(row.movement_lt)
                    total_out_mt += safe_float(row.movement_mt)

            book_closing_gsv = opening_gsv + total_in_gsv - total_out_gsv
            book_closing_nsv = opening_nsv + total_in_nsv - total_out_nsv
            book_closing_lt = opening_lt + total_in_lt - total_out_lt
            book_closing_mt = opening_mt + total_in_mt - total_out_mt

            actual_closing_gsv = book_closing_gsv
            actual_closing_nsv = book_closing_nsv
            actual_closing_lt = book_closing_lt
            actual_closing_mt = book_closing_mt

            last_ticket_number = None

            if day_rows:
                # Correct rule:
                # Closing stock is the latest stock snapshot in the accounting day.
                # User does not need to create an explicit Closing Stock ticket.
                closing_rows = [
                    row
                    for row in day_rows
                    if str(row.tank_operation_category or "").upper()
                    == "CLOSING"
                ]

                if closing_rows:
                    closing_source_row = closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                closing_snapshot = get_stock_snapshot_values(closing_source_row)

                actual_closing_gsv = closing_snapshot["gsv"]
                actual_closing_nsv = closing_snapshot["nsv"]
                actual_closing_lt = closing_snapshot["lt"]
                actual_closing_mt = closing_snapshot["mt"]
                last_ticket_number = closing_source_row.ticket_number

            else:
                # No entry on this accounting day:
                # carry forward previous day closing.
                actual_closing_gsv = opening_gsv
                actual_closing_nsv = opening_nsv
                actual_closing_lt = opening_lt
                actual_closing_mt = opening_mt

                book_closing_gsv = opening_gsv
                book_closing_nsv = opening_nsv
                book_closing_lt = opening_lt
                book_closing_mt = opening_mt

            loss_gain_gsv = actual_closing_gsv - book_closing_gsv
            loss_gain_nsv = actual_closing_nsv - book_closing_nsv
            loss_gain_lt = actual_closing_lt - book_closing_lt
            loss_gain_mt = actual_closing_mt - book_closing_mt

            daily_summary_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "opening_gsv_bbl": round(opening_gsv, 3),
                    "opening_nsv_bbl": round(opening_nsv, 3),
                    "opening_lt": round(opening_lt, 3),
                    "opening_mt": round(opening_mt, 3),
                    "total_in_gsv_bbl": round(total_in_gsv, 3),
                    "total_in_nsv_bbl": round(total_in_nsv, 3),
                    "total_in_lt": round(total_in_lt, 3),
                    "total_in_mt": round(total_in_mt, 3),
                    "total_out_gsv_bbl": round(total_out_gsv, 3),
                    "total_out_nsv_bbl": round(total_out_nsv, 3),
                    "total_out_lt": round(total_out_lt, 3),
                    "total_out_mt": round(total_out_mt, 3),
                    "book_closing_gsv_bbl": round(book_closing_gsv, 3),
                    "book_closing_nsv_bbl": round(book_closing_nsv, 3),
                    "book_closing_lt": round(book_closing_lt, 3),
                    "book_closing_mt": round(book_closing_mt, 3),
                    "actual_closing_gsv_bbl": round(actual_closing_gsv, 3),
                    "actual_closing_nsv_bbl": round(actual_closing_nsv, 3),
                    "actual_closing_lt": round(actual_closing_lt, 3),
                    "actual_closing_mt": round(actual_closing_mt, 3),
                    "loss_gain_gsv_bbl": round(loss_gain_gsv, 3),
                    "loss_gain_nsv_bbl": round(loss_gain_nsv, 3),
                    "loss_gain_lt": round(loss_gain_lt, 3),
                    "loss_gain_mt": round(loss_gain_mt, 3),
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_gsv = actual_closing_gsv
            previous_closing_nsv = actual_closing_nsv
            previous_closing_lt = actual_closing_lt
            previous_closing_mt = actual_closing_mt

    return sorted(
        daily_summary_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"],
            row["product_name"] or "",
        ),
    )

# -------------------------
# Material Balance Report Helpers
# -------------------------

def normalize_material_balance_category(value: str | None):
    text = str(value or "").strip().upper()

    if "RECEIPT" in text:
        return "RECEIPT"

    if "PRODUCTION" in text:
        return "PRODUCTION"

    if "DISPATCH" in text:
        return "DISPATCH"

    if "DRAIN" in text:
        return "DRAINING"

    if "OPENING" in text:
        return "OPENING"

    if "CLOSING" in text:
        return "CLOSING"

    return text or "OTHER"


def add_volume_values(target: dict, prefix: str, row: TankStockLedger):
    target[f"{prefix}_gsv"] += safe_float(row.movement_gsv_bbl)
    target[f"{prefix}_nsv"] += safe_float(row.movement_nsv_bbl)
    target[f"{prefix}_lt"] += safe_float(row.movement_lt)
    target[f"{prefix}_mt"] += safe_float(row.movement_mt)


def get_material_balance_rows_for_continuity(
    db: Session,
    location_code: str | None,
    tank_asset_code: str | None,
    product_name: str | None,
    date_to_value: date,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.accounting_date != None,
        TankStockLedger.accounting_date <= date_to_value,
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(
            TankStockLedger.location_code.ilike(cleaned_location_code)
        )

    if cleaned_tank_asset_code:
        query = query.filter(
            TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code)
        )

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )

    rows = query.all()

    return sorted(
        rows,
        key=lambda row: (
            row.location_code or "",
            row.tank_asset_code or "",
            row.product_name or "",
            row.accounting_date or date.min,
            get_ledger_operation_datetime(row) or datetime.min,
            row.id,
        ),
    )


def build_material_balance_report_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    date_from_value: date,
    date_to_value: date,
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    material_balance_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key

        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        tank_asset_name = ""
        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_gsv = 0
        previous_closing_nsv = 0
        previous_closing_lt = 0
        previous_closing_mt = 0

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            last_before_period = rows_before_period[-1]
            previous_snapshot = get_stock_snapshot_values(last_before_period)

            previous_closing_gsv = previous_snapshot["gsv"]
            previous_closing_nsv = previous_snapshot["nsv"]
            previous_closing_lt = previous_snapshot["lt"]
            previous_closing_mt = previous_snapshot["mt"]

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    get_ledger_operation_datetime(row) or datetime.min,
                    row.id,
                ),
            )

            opening_gsv = previous_closing_gsv
            opening_nsv = previous_closing_nsv
            opening_lt = previous_closing_lt
            opening_mt = previous_closing_mt

            opening_rows = [
                row
                for row in day_rows
                if normalize_material_balance_category(
                    row.tank_operation_category
                )
                == "OPENING"
            ]

            if opening_rows:
                opening_snapshot = get_stock_snapshot_values(opening_rows[-1])

                opening_gsv = opening_snapshot["gsv"]
                opening_nsv = opening_snapshot["nsv"]
                opening_lt = opening_snapshot["lt"]
                opening_mt = opening_snapshot["mt"]

            buckets = {
                "receipt_gsv": 0,
                "receipt_nsv": 0,
                "receipt_lt": 0,
                "receipt_mt": 0,
                "production_gsv": 0,
                "production_nsv": 0,
                "production_lt": 0,
                "production_mt": 0,
                "dispatch_gsv": 0,
                "dispatch_nsv": 0,
                "dispatch_lt": 0,
                "dispatch_mt": 0,
                "draining_gsv": 0,
                "draining_nsv": 0,
                "draining_lt": 0,
                "draining_mt": 0,
                "other_in_gsv": 0,
                "other_in_nsv": 0,
                "other_in_lt": 0,
                "other_in_mt": 0,
                "other_out_gsv": 0,
                "other_out_nsv": 0,
                "other_out_lt": 0,
                "other_out_mt": 0,
            }

            for row in day_rows:
                sign = str(row.tank_operation_sign or "").upper()
                category = normalize_material_balance_category(
                    row.tank_operation_category
                )

                if sign == "IN":
                    if category == "RECEIPT":
                        add_volume_values(buckets, "receipt", row)
                    elif category == "PRODUCTION":
                        add_volume_values(buckets, "production", row)
                    else:
                        add_volume_values(buckets, "other_in", row)

                elif sign == "OUT":
                    if category == "DISPATCH":
                        add_volume_values(buckets, "dispatch", row)
                    elif category == "DRAINING":
                        add_volume_values(buckets, "draining", row)
                    else:
                        add_volume_values(buckets, "other_out", row)

            total_in_gsv = (
                buckets["receipt_gsv"]
                + buckets["production_gsv"]
                + buckets["other_in_gsv"]
            )
            total_in_nsv = (
                buckets["receipt_nsv"]
                + buckets["production_nsv"]
                + buckets["other_in_nsv"]
            )
            total_in_lt = (
                buckets["receipt_lt"]
                + buckets["production_lt"]
                + buckets["other_in_lt"]
            )
            total_in_mt = (
                buckets["receipt_mt"]
                + buckets["production_mt"]
                + buckets["other_in_mt"]
            )

            total_out_gsv = (
                buckets["dispatch_gsv"]
                + buckets["draining_gsv"]
                + buckets["other_out_gsv"]
            )
            total_out_nsv = (
                buckets["dispatch_nsv"]
                + buckets["draining_nsv"]
                + buckets["other_out_nsv"]
            )
            total_out_lt = (
                buckets["dispatch_lt"]
                + buckets["draining_lt"]
                + buckets["other_out_lt"]
            )
            total_out_mt = (
                buckets["dispatch_mt"]
                + buckets["draining_mt"]
                + buckets["other_out_mt"]
            )

            book_closing_gsv = opening_gsv + total_in_gsv - total_out_gsv
            book_closing_nsv = opening_nsv + total_in_nsv - total_out_nsv
            book_closing_lt = opening_lt + total_in_lt - total_out_lt
            book_closing_mt = opening_mt + total_in_mt - total_out_mt

            actual_closing_gsv = book_closing_gsv
            actual_closing_nsv = book_closing_nsv
            actual_closing_lt = book_closing_lt
            actual_closing_mt = book_closing_mt

            last_ticket_number = None

            if day_rows:
                closing_rows = [
                    row
                    for row in day_rows
                    if normalize_material_balance_category(
                        row.tank_operation_category
                    )
                    == "CLOSING"
                ]

                if closing_rows:
                    closing_source_row = closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                closing_snapshot = get_stock_snapshot_values(closing_source_row)

                actual_closing_gsv = closing_snapshot["gsv"]
                actual_closing_nsv = closing_snapshot["nsv"]
                actual_closing_lt = closing_snapshot["lt"]
                actual_closing_mt = closing_snapshot["mt"]

                last_ticket_number = closing_source_row.ticket_number

            else:
                # No entry in this accounting day: carry forward previous closing.
                actual_closing_gsv = opening_gsv
                actual_closing_nsv = opening_nsv
                actual_closing_lt = opening_lt
                actual_closing_mt = opening_mt

                book_closing_gsv = opening_gsv
                book_closing_nsv = opening_nsv
                book_closing_lt = opening_lt
                book_closing_mt = opening_mt

            loss_gain_gsv = actual_closing_gsv - book_closing_gsv
            loss_gain_nsv = actual_closing_nsv - book_closing_nsv
            loss_gain_lt = actual_closing_lt - book_closing_lt
            loss_gain_mt = actual_closing_mt - book_closing_mt

            material_balance_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "opening_gsv_bbl": round(opening_gsv, 3),
                    "opening_nsv_bbl": round(opening_nsv, 3),
                    "opening_lt": round(opening_lt, 3),
                    "opening_mt": round(opening_mt, 3),
                    "receipt_gsv_bbl": round(buckets["receipt_gsv"], 3),
                    "receipt_nsv_bbl": round(buckets["receipt_nsv"], 3),
                    "receipt_lt": round(buckets["receipt_lt"], 3),
                    "receipt_mt": round(buckets["receipt_mt"], 3),
                    "production_gsv_bbl": round(buckets["production_gsv"], 3),
                    "production_nsv_bbl": round(buckets["production_nsv"], 3),
                    "production_lt": round(buckets["production_lt"], 3),
                    "production_mt": round(buckets["production_mt"], 3),
                    "dispatch_gsv_bbl": round(buckets["dispatch_gsv"], 3),
                    "dispatch_nsv_bbl": round(buckets["dispatch_nsv"], 3),
                    "dispatch_lt": round(buckets["dispatch_lt"], 3),
                    "dispatch_mt": round(buckets["dispatch_mt"], 3),
                    "draining_gsv_bbl": round(buckets["draining_gsv"], 3),
                    "draining_nsv_bbl": round(buckets["draining_nsv"], 3),
                    "draining_lt": round(buckets["draining_lt"], 3),
                    "draining_mt": round(buckets["draining_mt"], 3),
                    "other_in_gsv_bbl": round(buckets["other_in_gsv"], 3),
                    "other_in_nsv_bbl": round(buckets["other_in_nsv"], 3),
                    "other_in_lt": round(buckets["other_in_lt"], 3),
                    "other_in_mt": round(buckets["other_in_mt"], 3),
                    "other_out_gsv_bbl": round(buckets["other_out_gsv"], 3),
                    "other_out_nsv_bbl": round(buckets["other_out_nsv"], 3),
                    "other_out_lt": round(buckets["other_out_lt"], 3),
                    "other_out_mt": round(buckets["other_out_mt"], 3),
                    "total_in_gsv_bbl": round(total_in_gsv, 3),
                    "total_in_nsv_bbl": round(total_in_nsv, 3),
                    "total_in_lt": round(total_in_lt, 3),
                    "total_in_mt": round(total_in_mt, 3),
                    "total_out_gsv_bbl": round(total_out_gsv, 3),
                    "total_out_nsv_bbl": round(total_out_nsv, 3),
                    "total_out_lt": round(total_out_lt, 3),
                    "total_out_mt": round(total_out_mt, 3),
                    "book_closing_gsv_bbl": round(book_closing_gsv, 3),
                    "book_closing_nsv_bbl": round(book_closing_nsv, 3),
                    "book_closing_lt": round(book_closing_lt, 3),
                    "book_closing_mt": round(book_closing_mt, 3),
                    "actual_closing_gsv_bbl": round(actual_closing_gsv, 3),
                    "actual_closing_nsv_bbl": round(actual_closing_nsv, 3),
                    "actual_closing_lt": round(actual_closing_lt, 3),
                    "actual_closing_mt": round(actual_closing_mt, 3),
                    "loss_gain_gsv_bbl": round(loss_gain_gsv, 3),
                    "loss_gain_nsv_bbl": round(loss_gain_nsv, 3),
                    "loss_gain_lt": round(loss_gain_lt, 3),
                    "loss_gain_mt": round(loss_gain_mt, 3),
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_gsv = actual_closing_gsv
            previous_closing_nsv = actual_closing_nsv
            previous_closing_lt = actual_closing_lt
            previous_closing_mt = actual_closing_mt

    return sorted(
        material_balance_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"] or "",
            row["product_name"] or "",
        ),
    )

def consolidate_material_balance_rows_by_location(
    tank_wise_rows: list[dict],
):
    consolidated_map = {}

    for row in tank_wise_rows:
        key = (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        )

        if key not in consolidated_map:
            consolidated_map[key] = {
                "accounting_date": row["accounting_date"],
                "location_code": row["location_code"],
                "location_name": row["location_name"],
                "tank_asset_code": None,
                "tank_asset_name": "All Tanks",
                "product_name": row["product_name"],
                "opening_gsv_bbl": 0,
                "opening_nsv_bbl": 0,
                "opening_lt": 0,
                "opening_mt": 0,
                "receipt_gsv_bbl": 0,
                "receipt_nsv_bbl": 0,
                "receipt_lt": 0,
                "receipt_mt": 0,
                "production_gsv_bbl": 0,
                "production_nsv_bbl": 0,
                "production_lt": 0,
                "production_mt": 0,
                "dispatch_gsv_bbl": 0,
                "dispatch_nsv_bbl": 0,
                "dispatch_lt": 0,
                "dispatch_mt": 0,
                "draining_gsv_bbl": 0,
                "draining_nsv_bbl": 0,
                "draining_lt": 0,
                "draining_mt": 0,
                "other_in_gsv_bbl": 0,
                "other_in_nsv_bbl": 0,
                "other_in_lt": 0,
                "other_in_mt": 0,
                "other_out_gsv_bbl": 0,
                "other_out_nsv_bbl": 0,
                "other_out_lt": 0,
                "other_out_mt": 0,
                "total_in_gsv_bbl": 0,
                "total_in_nsv_bbl": 0,
                "total_in_lt": 0,
                "total_in_mt": 0,
                "total_out_gsv_bbl": 0,
                "total_out_nsv_bbl": 0,
                "total_out_lt": 0,
                "total_out_mt": 0,
                "book_closing_gsv_bbl": 0,
                "book_closing_nsv_bbl": 0,
                "book_closing_lt": 0,
                "book_closing_mt": 0,
                "actual_closing_gsv_bbl": 0,
                "actual_closing_nsv_bbl": 0,
                "actual_closing_lt": 0,
                "actual_closing_mt": 0,
                "loss_gain_gsv_bbl": 0,
                "loss_gain_nsv_bbl": 0,
                "loss_gain_lt": 0,
                "loss_gain_mt": 0,
                "rows_count": 0,
                "last_ticket_number": None,
            }

        target = consolidated_map[key]

        numeric_fields = [
            "opening_gsv_bbl",
            "opening_nsv_bbl",
            "opening_lt",
            "opening_mt",
            "receipt_gsv_bbl",
            "receipt_nsv_bbl",
            "receipt_lt",
            "receipt_mt",
            "production_gsv_bbl",
            "production_nsv_bbl",
            "production_lt",
            "production_mt",
            "dispatch_gsv_bbl",
            "dispatch_nsv_bbl",
            "dispatch_lt",
            "dispatch_mt",
            "draining_gsv_bbl",
            "draining_nsv_bbl",
            "draining_lt",
            "draining_mt",
            "other_in_gsv_bbl",
            "other_in_nsv_bbl",
            "other_in_lt",
            "other_in_mt",
            "other_out_gsv_bbl",
            "other_out_nsv_bbl",
            "other_out_lt",
            "other_out_mt",
            "total_in_gsv_bbl",
            "total_in_nsv_bbl",
            "total_in_lt",
            "total_in_mt",
            "total_out_gsv_bbl",
            "total_out_nsv_bbl",
            "total_out_lt",
            "total_out_mt",
            "book_closing_gsv_bbl",
            "book_closing_nsv_bbl",
            "book_closing_lt",
            "book_closing_mt",
            "actual_closing_gsv_bbl",
            "actual_closing_nsv_bbl",
            "actual_closing_lt",
            "actual_closing_mt",
            "loss_gain_gsv_bbl",
            "loss_gain_nsv_bbl",
            "loss_gain_lt",
            "loss_gain_mt",
        ]

        for field in numeric_fields:
            target[field] += safe_float(row.get(field))

        target["rows_count"] += int(row.get("rows_count") or 0)

        if row.get("last_ticket_number"):
            target["last_ticket_number"] = row.get("last_ticket_number")

    consolidated_rows = []

    for row in consolidated_map.values():
        for key, value in list(row.items()):
            if isinstance(value, float):
                row[key] = round(value, 3)

        consolidated_rows.append(row)

    return sorted(
        consolidated_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        ),
    )

# -------------------------
# Dynamic Material Balance Report Helpers
# -------------------------

def normalize_material_balance_code_value(value):
    return str(value or "").strip().upper()


def get_active_material_balance_template_for_location(
    db: Session,
    location_code: str,
):
    template = (
        db.query(MaterialBalanceTemplate)
        .filter(
            MaterialBalanceTemplate.location_code.ilike(location_code),
            MaterialBalanceTemplate.status == "Active",
        )
        .order_by(MaterialBalanceTemplate.id.desc())
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=400,
            detail=(
                "No Active Material Balance Template found for this location. "
                "Please configure Material Balance Template first."
            ),
        )

    return template


def get_active_material_balance_template_columns(
    db: Session,
    template_id: int,
):
    columns = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(
            MaterialBalanceTemplateColumn.template_id == template_id,
            MaterialBalanceTemplateColumn.status == "Active",
        )
        .order_by(
            MaterialBalanceTemplateColumn.column_order.asc(),
            MaterialBalanceTemplateColumn.id.asc(),
        )
        .all()
    )

    if not columns:
        raise HTTPException(
            status_code=400,
            detail="No Active columns configured for this Material Balance Template.",
        )

    return columns


def build_dynamic_material_balance_columns_response(
    columns: list[MaterialBalanceTemplateColumn],
):
    return [
        {
            "column_key": column.column_key,
            "column_label": column.column_label,
            "column_order": column.column_order,
            "column_type": column.column_type,
            "movement_direction": column.movement_direction,
            "include_in_material_balance": column.include_in_material_balance,
            "include_in_book_closing": column.include_in_book_closing,
            "is_internal_transfer": column.is_internal_transfer,
        }
        for column in columns
    ]


def get_movement_value_for_unit(row: TankStockLedger, unit_key: str = "nsv"):
    if unit_key == "gsv":
        return safe_float(row.movement_gsv_bbl)

    if unit_key == "lt":
        return safe_float(row.movement_lt)

    if unit_key == "mt":
        return safe_float(row.movement_mt)

    return safe_float(row.movement_nsv_bbl)


def get_snapshot_value_for_unit(snapshot: dict, unit_key: str = "nsv"):
    unit_key = str(unit_key or "nsv").strip().lower()

    if unit_key == "gsv":
        return safe_float(snapshot.get("gsv"))

    if unit_key == "lt":
        return safe_float(snapshot.get("lt"))

    if unit_key == "mt":
        return safe_float(snapshot.get("mt"))

    # default nsv
    return safe_float(snapshot.get("nsv"))


def should_row_match_material_balance_column(
    row: TankStockLedger,
    column: MaterialBalanceTemplateColumn,
):
    if normalize_material_balance_code_value(column.column_type) != "MOVEMENT":
        return False

    row_operation_code = normalize_material_balance_code_value(
        row.tank_operation_code
    )

    row_sign = normalize_material_balance_code_value(row.tank_operation_sign)
    column_direction = normalize_material_balance_code_value(
        column.movement_direction
    )

    mapped_operation_codes = {
        normalize_material_balance_code_value(code)
        for code in (column.mapped_operation_codes or [])
    }

    excluded_operation_codes = {
        normalize_material_balance_code_value(code)
        for code in (column.excluded_operation_codes or [])
    }

    if row_operation_code in excluded_operation_codes:
        return False

    if mapped_operation_codes and row_operation_code not in mapped_operation_codes:
        return False

    if column_direction and row_sign != column_direction:
        return False

    return True

def get_global_internal_transfer_operation_codes(
    columns: list[MaterialBalanceTemplateColumn],
):
    internal_codes = set()

    for column in columns:
        is_internal_transfer = (
            normalize_material_balance_code_value(column.is_internal_transfer)
            == "YES"
        )

        include_in_material_balance = (
            normalize_material_balance_code_value(column.include_in_material_balance)
            == "YES"
        )

        include_in_book_closing = (
            normalize_material_balance_code_value(column.include_in_book_closing)
            == "YES"
        )

        # Any operation mapped to an internal transfer column must be globally
        # excluded from Book Closing, even if accidentally mapped elsewhere.
        if is_internal_transfer:
            for code in column.mapped_operation_codes or []:
                internal_codes.add(normalize_material_balance_code_value(code))

        # Extra safety:
        # If a movement column is explicitly excluded from both MB and Book Closing,
        # treat its mapped operation codes as non-book-closing operations.
        if (
            normalize_material_balance_code_value(column.column_type) == "MOVEMENT"
            and not include_in_material_balance
            and not include_in_book_closing
        ):
            for code in column.mapped_operation_codes or []:
                internal_codes.add(normalize_material_balance_code_value(code))

    return internal_codes


def should_row_be_in_book_closing_formula(
    row: TankStockLedger,
    columns: list[MaterialBalanceTemplateColumn],
    global_internal_transfer_codes: set[str],
):
    row_operation_code = normalize_material_balance_code_value(
        row.tank_operation_code
    )

    row_sign = normalize_material_balance_code_value(row.tank_operation_sign)

    if row_sign not in ["IN", "OUT"]:
        return False

    if row_operation_code in global_internal_transfer_codes:
        return False

    for column in columns:
        column_type = normalize_material_balance_code_value(column.column_type)

        if column_type != "MOVEMENT":
            continue

        include_in_material_balance = (
            normalize_material_balance_code_value(column.include_in_material_balance)
            == "YES"
        )

        include_in_book_closing = (
            normalize_material_balance_code_value(column.include_in_book_closing)
            == "YES"
        )

        is_internal_transfer = (
            normalize_material_balance_code_value(column.is_internal_transfer)
            == "YES"
        )

        if not include_in_material_balance:
            continue

        if not include_in_book_closing:
            continue

        if is_internal_transfer:
            continue

        if should_row_match_material_balance_column(row, column):
            return True

    return False


def calculate_book_closing_from_eligible_ledger_rows(
    opening_value: float,
    day_rows: list[TankStockLedger],
    columns: list[MaterialBalanceTemplateColumn],
    unit_key: str,
):
    global_internal_transfer_codes = get_global_internal_transfer_operation_codes(
        columns
    )

    eligible_in_total = 0
    eligible_out_total = 0
    included_ledger_ids = set()

    for row in day_rows:
        if row.id in included_ledger_ids:
            continue

        if not should_row_be_in_book_closing_formula(
            row=row,
            columns=columns,
            global_internal_transfer_codes=global_internal_transfer_codes,
        ):
            continue

        movement_value = get_movement_value_for_unit(row, unit_key)
        row_sign = normalize_material_balance_code_value(row.tank_operation_sign)

        if row_sign == "IN":
            eligible_in_total += movement_value
            included_ledger_ids.add(row.id)

        elif row_sign == "OUT":
            eligible_out_total += movement_value
            included_ledger_ids.add(row.id)

    book_closing_value = opening_value + eligible_in_total - eligible_out_total

    return {
        "book_closing_value": book_closing_value,
        "eligible_in_total": eligible_in_total,
        "eligible_out_total": eligible_out_total,
        "included_ledger_ids": sorted(list(included_ledger_ids)),
    }

def build_dynamic_material_balance_tank_rows(
    db: Session,
    ledger_rows: list[TankStockLedger],
    columns: list[MaterialBalanceTemplateColumn],
    date_from_value: date,
    date_to_value: date,
    unit_key: str = "nsv",
):
    date_range = build_date_range(date_from_value, date_to_value)

    grouped_rows = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    report_rows = []

    for key, rows in grouped_rows.items():
        location_code, tank_asset_code, product_name_value = key
        location = get_location_by_code(location_code, db)

        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        tank_asset_name = ""

        if sorted_rows:
            tank_asset_name = sorted_rows[-1].tank_asset_name or ""

        previous_closing_snapshot = {
            "gsv": 0,
            "nsv": 0,
            "lt": 0,
            "mt": 0,
        }

        rows_before_period = [
            row
            for row in sorted_rows
            if row.accounting_date is not None
            and row.accounting_date < date_from_value
        ]

        if rows_before_period:
            previous_closing_snapshot = get_stock_snapshot_values(
                rows_before_period[-1]
            )

        for accounting_date_value in date_range:
            day_rows = [
                row
                for row in sorted_rows
                if row.accounting_date == accounting_date_value
            ]

            day_rows = sorted(
                day_rows,
                key=lambda row: (
                    get_ledger_operation_datetime(row) or datetime.min,
                    row.id,
                ),
            )

            opening_value = get_snapshot_value_for_unit(
                previous_closing_snapshot,
                unit_key,
            )

            explicit_opening_rows = [
                row
                for row in day_rows
                if normalize_material_balance_code_value(
                    row.tank_operation_category
                )
                == "OPENING"
            ]

            if explicit_opening_rows:
                opening_snapshot = get_stock_snapshot_values(
                    explicit_opening_rows[-1]
                )
                opening_value = get_snapshot_value_for_unit(
                    opening_snapshot,
                    unit_key,
                )

            values = {}

            book_closing_value = opening_value
            actual_closing_value = opening_value
            last_ticket_number = None

            # First pass: calculate configured columns except computed closing/loss.
            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "OPENING":
                    values[column_key] = round(opening_value, 3)
                    continue

                if column_type == "MOVEMENT":
                    movement_total = 0

                    for row in day_rows:
                        if should_row_match_material_balance_column(row, column):
                            movement_total += get_movement_value_for_unit(
                                row,
                                unit_key,
                            )

                    values[column_key] = round(movement_total, 3)
                    continue

                if column_type in ["INFO", "FORMULA"]:
                    values[column_key] = 0
                    continue

            book_closing_calculation = calculate_book_closing_from_eligible_ledger_rows(
                opening_value=opening_value,
                day_rows=day_rows,
                columns=columns,
                unit_key=unit_key,
            )

            book_closing_value = book_closing_calculation["book_closing_value"]

            if day_rows:
                explicit_closing_rows = [
                    row
                    for row in day_rows
                    if normalize_material_balance_code_value(
                        row.tank_operation_category
                    )
                    == "CLOSING"
                ]

                if explicit_closing_rows:
                    closing_source_row = explicit_closing_rows[-1]
                else:
                    closing_source_row = day_rows[-1]

                actual_closing_snapshot = get_stock_snapshot_values(
                    closing_source_row
                )

                actual_closing_value = get_snapshot_value_for_unit(
                    actual_closing_snapshot,
                    unit_key,
                )

                last_ticket_number = closing_source_row.ticket_number
            else:
                actual_closing_snapshot = previous_closing_snapshot
                actual_closing_value = opening_value

            loss_gain_value = actual_closing_value - book_closing_value

            # Second pass: fill computed columns.
            for column in columns:
                column_key = column.column_key
                column_type = normalize_material_balance_code_value(
                    column.column_type
                )

                if column_type == "BOOK_CLOSING":
                    values[column_key] = round(book_closing_value, 3)

                elif column_type == "ACTUAL_CLOSING":
                    values[column_key] = round(actual_closing_value, 3)

                elif column_type == "LOSS_GAIN":
                    values[column_key] = round(loss_gain_value, 3)

            report_rows.append(
                {
                    "accounting_date": accounting_date_value,
                    "location_code": location_code,
                    "location_name": location.location_name if location else "",
                    "tank_asset_code": tank_asset_code,
                    "tank_asset_name": tank_asset_name,
                    "product_name": product_name_value or None,
                    "values": values,
                    "rows_count": len(day_rows),
                    "last_ticket_number": last_ticket_number,
                }
            )

            previous_closing_snapshot = {
                "gsv": actual_closing_snapshot.get("gsv", actual_closing_value),
                "nsv": actual_closing_snapshot.get("nsv", actual_closing_value),
                "lt": actual_closing_snapshot.get("lt", 0),
                "mt": actual_closing_snapshot.get("mt", 0),
            }

    return sorted(
        report_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["tank_asset_code"] or "",
            row["product_name"] or "",
        ),
    )


def consolidate_dynamic_material_balance_rows_by_location(
    tank_rows: list[dict],
    columns: list[MaterialBalanceTemplateColumn],
):
    consolidated_map = {}

    for row in tank_rows:
        key = (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        )

        if key not in consolidated_map:
            consolidated_map[key] = {
                "accounting_date": row["accounting_date"],
                "location_code": row["location_code"],
                "location_name": row["location_name"],
                "tank_asset_code": None,
                "tank_asset_name": "All Tanks",
                "product_name": row["product_name"],
                "values": {},
                "rows_count": 0,
                "last_ticket_number": None,
            }

            for column in columns:
                consolidated_map[key]["values"][column.column_key] = 0

        target = consolidated_map[key]

        for column in columns:
            column_key = column.column_key
            target["values"][column_key] = safe_float(
                target["values"].get(column_key)
            ) + safe_float(row["values"].get(column_key))

        target["rows_count"] += int(row.get("rows_count") or 0)

        if row.get("last_ticket_number"):
            target["last_ticket_number"] = row.get("last_ticket_number")

    consolidated_rows = []

    for row in consolidated_map.values():
        for column in columns:
            column_key = column.column_key
            row["values"][column_key] = round(
                safe_float(row["values"].get(column_key)),
                3,
            )

        consolidated_rows.append(row)

    return sorted(
        consolidated_rows,
        key=lambda row: (
            row["accounting_date"],
            row["location_code"],
            row["product_name"] or "",
        ),
    )

@app.get(
    "/tank-stock-ledger",
    response_model=list[TankStockLedgerResponse],
)
def get_tank_stock_ledger(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )

    return [
        build_tank_stock_ledger_response(row, db)
        for row in ledger_rows
    ]


@app.get(
    "/tank-stock-ledger/summary",
    response_model=list[TankStockLedgerSummaryResponse],
)
def get_tank_stock_ledger_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    ledger_rows = get_filtered_tank_stock_ledger_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status="Active",
    )

    summary_map = {}

    for row in ledger_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in summary_map:
            location = get_location_by_code(row.location_code, db)

            summary_map[key] = {
                "location_code": row.location_code,
                "location_name": location.location_name if location else "",
                "tank_asset_code": row.tank_asset_code,
                "tank_asset_name": row.tank_asset_name,
                "product_name": row.product_name,
                "opening_nsv_bbl": 0,
                "total_in_nsv_bbl": 0,
                "total_out_nsv_bbl": 0,
                "closing_nsv_bbl": 0,
                "opening_lt": 0,
                "total_in_lt": 0,
                "total_out_lt": 0,
                "closing_lt": 0,
                "opening_mt": 0,
                "total_in_mt": 0,
                "total_out_mt": 0,
                "closing_mt": 0,
            }

        summary = summary_map[key]

        sign = row.tank_operation_sign
        category = row.tank_operation_category

        movement_nsv = row.movement_nsv_bbl or 0
        movement_lt = row.movement_lt or 0
        movement_mt = row.movement_mt or 0

        if category == "OPENING":
            summary["opening_nsv_bbl"] += movement_nsv
            summary["opening_lt"] += movement_lt
            summary["opening_mt"] += movement_mt

        if sign == "IN":
            summary["total_in_nsv_bbl"] += movement_nsv
            summary["total_in_lt"] += movement_lt
            summary["total_in_mt"] += movement_mt

        if sign == "OUT":
            summary["total_out_nsv_bbl"] += movement_nsv
            summary["total_out_lt"] += movement_lt
            summary["total_out_mt"] += movement_mt

        # Closing values are based on last running balance in the selected period.
        summary["closing_nsv_bbl"] = row.running_balance_nsv_bbl or 0
        summary["closing_lt"] = row.running_balance_lt or 0
        summary["closing_mt"] = row.running_balance_mt or 0

    return list(summary_map.values())


@app.get(
    "/tank-stock-ledger/daily-summary",
    response_model=list[TankStockLedgerDailySummaryResponse],
)
def get_tank_stock_ledger_daily_summary(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Tank Stock Ledger",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for daily summary",
        )

    ledger_rows = get_tank_stock_rows_for_daily_summary(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
    )

    return build_tank_stock_daily_summary_rows(
        db=db,
        ledger_rows=ledger_rows,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
    )

@app.get(
    "/out-turn-report",
    response_model=list[OutTurnReportResponse],
)
def get_out_turn_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = "Active",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    rows = get_out_turn_report_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=date_from,
        date_to=date_to,
        status=status,
    )

    return [
        build_out_turn_report_response(row, db)
        for row in rows
    ]

@app.get("/out-turn-report/validation")
def validate_out_turn_report_tank_sequence(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Out-Turn Report",
        db,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    # Important:
    # For validation, fetch all rows up to date_to so that first visible row
    # can still be validated against earlier stock continuity.
    continuity_rows = get_out_turn_report_rows(
        db=db,
        location_code=location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_from=None,
        date_to=date_to,
        status="Active",
    )

    visible_rows = []

    for row in continuity_rows:
        if date_from_value and row.accounting_date and row.accounting_date < date_from_value:
            continue

        if date_to_value and row.accounting_date and row.accounting_date > date_to_value:
            continue

        visible_rows.append(row)

    grouped_rows = {}

    for row in continuity_rows:
        key = (
            row.location_code,
            row.tank_asset_code,
            row.product_name or "",
        )

        if key not in grouped_rows:
            grouped_rows[key] = []

        grouped_rows[key].append(row)

    visible_ledger_ids = {row.id for row in visible_rows}

    issues = []

    for key, group_rows in grouped_rows.items():
        location, tank, product = key

        sorted_group_rows = sorted(
            group_rows,
            key=lambda row: (
                row.accounting_date or date.min,
                get_ledger_operation_datetime(row) or datetime.min,
                row.id,
            ),
        )

        previous_row = None

        for row in sorted_group_rows:
            if previous_row is None:
                expected_previous_nsv = 0
            else:
                previous_snapshot = get_stock_snapshot_values(previous_row)
                expected_previous_nsv = previous_snapshot["nsv"]

            # Only report issues for rows inside the requested visible date range.
            if row.id in visible_ledger_ids:
                actual_previous_nsv = safe_float(row.previous_stock_nsv_bbl)

                if round(actual_previous_nsv, 3) != round(expected_previous_nsv, 3):
                    issues.append(
                        {
                            "ledger_id": row.id,
                            "ticket_number": row.ticket_number,
                            "location_code": location,
                            "tank_asset_code": tank,
                            "product_name": product or None,
                            "expected_previous_nsv_bbl": round(
                                expected_previous_nsv,
                                3,
                            ),
                            "actual_previous_nsv_bbl": round(
                                actual_previous_nsv,
                                3,
                            ),
                            "message": (
                                "Previous stock does not match previous row "
                                "of the same tank/product sequence. Run ledger rebuild."
                            ),
                        }
                    )

            previous_row = row

    return {
        "rows_checked": len(visible_rows),
        "groups_checked": len(grouped_rows),
        "issues_count": len(issues),
        "issues": issues,
    }

@app.get(
    "/material-balance-report",
    response_model=MaterialBalanceDynamicReportResponse,
)
def get_material_balance_report(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    unit: str | None = "nsv",
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Material Balance Report",
        db,
    )

    cleaned_location_code = clean_optional_text(location_code)

    if not cleaned_location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required for configurable Material Balance Report",
        )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value is None or date_to_value is None:
        raise HTTPException(
            status_code=400,
            detail="Date From and Date To are required for Material Balance Report",
        )

    unit_key = normalize_material_balance_code_value(unit).lower()

    if unit_key not in ["gsv", "nsv", "lt", "mt"]:
        raise HTTPException(
            status_code=400,
            detail="Unit must be one of: gsv, nsv, lt, mt",
        )

    template = get_active_material_balance_template_for_location(
        db=db,
        location_code=cleaned_location_code,
    )

    columns = get_active_material_balance_template_columns(
        db=db,
        template_id=template.id,
    )

    ledger_rows = get_material_balance_rows_for_continuity(
        db=db,
        location_code=cleaned_location_code,
        tank_asset_code=tank_asset_code,
        product_name=product_name,
        date_to_value=date_to_value,
    )

    tank_rows = build_dynamic_material_balance_tank_rows(
        db=db,
        ledger_rows=ledger_rows,
        columns=columns,
        date_from_value=date_from_value,
        date_to_value=date_to_value,
        unit_key=unit_key,
    )

    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)

    if cleaned_tank_asset_code:
        report_rows = tank_rows
    else:
        report_rows = consolidate_dynamic_material_balance_rows_by_location(
            tank_rows=tank_rows,
            columns=columns,
        )

    return {
        "template": {
            "id": template.id,
            "location_code": template.location_code,
            "template_name": template.template_name,
        },
        "columns": build_dynamic_material_balance_columns_response(columns),
        "rows": report_rows,
    }

def build_fso_otr_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
    shuttle_number: str | None = None,
):
    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)
    sn = clean_optional_text(shuttle_number)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )
    if sn:
        q = q.filter(OperationTransaction.convoy_number == sn)

    rows = []
    totals = {
        "receipt": 0.0,
        "export": 0.0,
        "movement": 0.0,
        "variance": 0.0,
        "compare_variance": 0.0,
    }

    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")
        op_label = str(meta.get("operation_label") or "").strip() or "FSO"
        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = float(safe_float(net.get("net_stock_bbl")))
        net_water = float(safe_float(net.get("net_water_bbl")))
        movement_qty = abs(net_stock) + abs(net_water)

        vessel_qty = float(safe_float(inputs.get("vessel_quantity_bbl")))
        # ✅ your final variance rule
        variance = abs(net_stock + net_water) - vessel_qty

        src_discharge = float(safe_float(meta.get("source_shuttle_discharge_bbl")))
        compare_var = movement_qty - src_discharge if op_sign == "IN" and src_discharge > 0 else 0.0

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        row = {
            "transaction_id": tx.id,
            "ticket_number": get_transaction_ticket_number(tx),
            "accounting_date": acc_date,
            "operation_date": tx.operation_date,
            "event_time": event_time,
            "location_code": loc_code,
            "fso_asset_code": asset_code,
            "shuttle_number": inputs.get("shuttle_number") or meta.get("shuttle_number") or tx.convoy_number,
            "operation_label": op_label,
            "operation_sign": op_sign,
            "vessel_name": inputs.get("vessel_name"),
            "vessel_quantity_bbl": vessel_qty,
            "opening_stock_bbl": float(safe_float(inputs.get("opening_stock_bbl"))),
            "opening_water_bbl": float(safe_float(inputs.get("opening_water_bbl"))),
            "closing_stock_bbl": float(safe_float(inputs.get("closing_stock_bbl"))),
            "closing_water_bbl": float(safe_float(inputs.get("closing_water_bbl"))),
            "net_stock_bbl": net_stock,
            "net_water_bbl": net_water,
            "movement_qty_bbl": movement_qty,
            "variance_bbl": variance,
            "source_shuttle_discharge_bbl": src_discharge,
            "compare_variance_bbl": compare_var,
            "remarks": inputs.get("remarks"),
        }
        rows.append(row)

        totals["movement"] += movement_qty
        totals["variance"] += variance
        totals["compare_variance"] += compare_var
        if op_sign == "IN":
            totals["receipt"] += movement_qty
        elif op_sign == "OUT":
            totals["export"] += movement_qty

    return rows, totals


def build_fso_material_balance(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, inputs.get("event_time"), day_start)

        buckets.setdefault(acc_date, []).append((tx, meta, inputs, net))

    dates = sorted([d for d in buckets.keys() if d >= date_from and d <= date_to])
    rows = []
    prev_physical_close = None

    for acc_date in dates:
        items = buckets[acc_date]

        def sort_key(item):
            tx, meta, inputs, net = item
            t = inputs.get("event_time") or "00:00"
            return (str(tx.operation_date), str(t), tx.id)

        items = sorted(items, key=sort_key)

        if prev_physical_close is None:
            opening = float(safe_float(items[0][2].get("opening_stock_bbl")))
        else:
            opening = prev_physical_close

        receipt = 0.0
        export = 0.0

        for tx, meta, inputs, net in items:
            op_sign = str(meta.get("operation_sign") or "").strip().upper()
            net_stock = float(safe_float(net.get("net_stock_bbl")))
            net_water = float(safe_float(net.get("net_water_bbl")))
            qty = abs(net_stock) + abs(net_water)

            if op_sign == "IN":
                receipt += qty
            elif op_sign == "OUT":
                export += qty

        book_close = opening + receipt - export

        last_inputs = items[-1][2]
        physical_close = float(safe_float(last_inputs.get("closing_stock_bbl")))
        physical_close_water = float(safe_float(last_inputs.get("closing_water_bbl")))
        loss_gain = physical_close - book_close

        rows.append(
            {
                "accounting_date": acc_date,
                "opening_stock_bbl": opening,
                "receipt_bbl": receipt,
                "export_bbl": export,
                "book_closing_bbl": book_close,
                "physical_closing_bbl": physical_close,
                "physical_closing_water_bbl": physical_close_water,
                "loss_gain_bbl": loss_gain,
            }
        )

        prev_physical_close = physical_close

    return rows


def build_fso_outturn_report(
    db: Session,
    location_code: str,
    fso_asset_code: str,
    date_from: date,
    date_to: date,
):
    loc_code = clean_optional_text(location_code)
    asset_code = clean_optional_text(fso_asset_code)

    q = (
        db.query(OperationTransaction, OperationTransactionValue)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.primary_asset_code == asset_code,
            OperationTransaction.operation_date >= date_from,
            OperationTransaction.operation_date <= date_to,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
    )

    def _sf(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _abs_qty(net_stock, net_water):
        return abs(_sf(net_stock)) + abs(_sf(net_water))

    buckets = {}
    for tx, val in q.all():
        payload = val.field_value if isinstance(val.field_value, dict) else {}
        meta = payload.get("meta") or {}
        inputs = payload.get("inputs") or {}
        net = ((payload.get("calculated") or {}).get("net") or {})

        event_time = inputs.get("event_time")

        setting = get_active_location_day_setting(db, loc_code, tx.operation_date)
        day_start = setting.day_start_time if setting else datetime_time(0, 0)
        acc_date = compute_accounting_date(tx.operation_date, event_time, day_start)

        shuttle_no = (
            inputs.get("shuttle_number")
            or meta.get("shuttle_number")
            or tx.convoy_number
            or ""
        ).strip()
        if shuttle_no == "":
            continue

        key = (acc_date, shuttle_no)
        buckets.setdefault(key, {"receipt": 0.0, "discharge": 0.0})

        op_sign = str(meta.get("operation_sign") or "").strip().upper()

        net_stock = net.get("net_stock_bbl")
        net_water = net.get("net_water_bbl")
        qty = _abs_qty(net_stock, net_water)

        if op_sign == "IN":
            buckets[key]["receipt"] += qty

        src = meta.get("source_shuttle_discharge_bbl")
        if src is not None:
            buckets[key]["discharge"] = max(buckets[key]["discharge"], _sf(src))

    rows = []
    totals = {"discharge": 0.0, "receipt": 0.0, "variance": 0.0}

    for (acc_date, shuttle_no) in sorted(buckets.keys()):
        discharge = float(buckets[(acc_date, shuttle_no)]["discharge"])
        receipt = float(buckets[(acc_date, shuttle_no)]["receipt"])
        variance = receipt - discharge
        pct = (variance / discharge * 100.0) if discharge != 0 else 0.0

        rows.append(
            {
                "accounting_date": acc_date,
                "shuttle_number": shuttle_no,
                "shuttle_discharge_bbl": discharge,
                "fso_receipt_bbl": receipt,
                "variance_bbl": variance,
                "variance_pct": pct,
            }
        )

        totals["discharge"] += discharge
        totals["receipt"] += receipt
        totals["variance"] += variance

    totals_pct = (totals["variance"] / totals["discharge"] * 100.0) if totals["discharge"] != 0 else 0.0
    return rows, totals, totals_pct


def _xlsx_autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


@app.get("/fso/reports/otr", response_model=FSOOTRReportResponse)
def get_fso_otr_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)
    return {
        "rows": rows,
        "total_receipt_bbl": totals["receipt"],
        "total_export_bbl": totals["export"],
        "total_movement_bbl": totals["movement"],
        "total_variance_bbl": totals["variance"],
        "total_compare_variance_bbl": totals["compare_variance"],
    }


@app.get("/fso/reports/material-balance", response_model=FSOMaterialBalanceReportResponse)
def get_fso_material_balance_report(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)
    return {"rows": rows}


@app.get("/fso/reports/outturn", response_model=FSOOutturnReportResponse)
def fso_report_outturn(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)
    return {
        "rows": rows,
        "total_shuttle_discharge_bbl": totals["discharge"],
        "total_fso_receipt_bbl": totals["receipt"],
        "total_variance_bbl": totals["variance"],
        "total_variance_pct": totals_pct,
    }


@app.get("/fso/reports/otr/export/xlsx")
def fso_report_otr_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    shuttle_number: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows, totals = build_fso_otr_report(db, location_code, fso_asset_code, df, dt, shuttle_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO OTR"

    headers = [
        "Ticket", "Acc Date", "Op Date", "Time", "Operation", "Sign",
        "Shuttle", "Vessel", "Vessel Qty",
        "Open Stock", "Open Water", "Close Stock", "Close Water",
        "Net Stock", "Net Water", "Move Qty",
        "Variance", "Shuttle Discharge", "Compare Var", "Remarks",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["ticket_number"],
            str(r["accounting_date"]),
            str(r["operation_date"]),
            r.get("event_time") or "",
            r["operation_label"],
            r["operation_sign"],
            r.get("shuttle_number") or "",
            r.get("vessel_name") or "",
            round(float(r["vessel_quantity_bbl"]), 3),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["opening_water_bbl"]), 3),
            round(float(r["closing_stock_bbl"]), 3),
            round(float(r["closing_water_bbl"]), 3),
            round(float(r["net_stock_bbl"]), 3),
            round(float(r["net_water_bbl"]), 3),
            round(float(r["movement_qty_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["source_shuttle_discharge_bbl"]), 3),
            round(float(r["compare_variance_bbl"]), 3),
            r.get("remarks") or "",
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Receipt", totals["receipt"]])
    ws2.append(["Total Export", totals["export"]])
    ws2.append(["Total Movement", totals["movement"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Compare Variance", totals["compare_variance"]])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_otr_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/fso/reports/outturn/export/xlsx")
def fso_report_outturn_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    rows, totals, totals_pct = build_fso_outturn_report(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Outturn"

    headers = ["Acc Date", "Shuttle Number", "Shuttle Discharge", "FSO Receipt", "Variance", "Variance %"]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            r["shuttle_number"],
            round(float(r["shuttle_discharge_bbl"]), 3),
            round(float(r["fso_receipt_bbl"]), 3),
            round(float(r["variance_bbl"]), 3),
            round(float(r["variance_pct"]), 3),
        ])

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Total Shuttle Discharge", totals["discharge"]])
    ws2.append(["Total FSO Receipt", totals["receipt"]])
    ws2.append(["Total Variance", totals["variance"]])
    ws2.append(["Total Variance %", totals_pct])

    _xlsx_autofit(ws)
    _xlsx_autofit(ws2)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_outturn_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/fso/reports/material-balance/export/xlsx")
def fso_report_mb_xlsx(
    location_code: str,
    fso_asset_code: str,
    date_from: str,
    date_to: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)
    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")
    rows = build_fso_material_balance(db, location_code, fso_asset_code, df, dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSO Material Balance"

    headers = [
        "Acc Date", "Opening", "Receipt", "Export",
        "Book Closing", "Physical Closing", "Closing Water", "Loss/Gain",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            str(r["accounting_date"]),
            round(float(r["opening_stock_bbl"]), 3),
            round(float(r["receipt_bbl"]), 3),
            round(float(r["export_bbl"]), 3),
            round(float(r["book_closing_bbl"]), 3),
            round(float(r["physical_closing_bbl"]), 3),
            round(float(r["physical_closing_water_bbl"]), 3),
            round(float(r["loss_gain_bbl"]), 3),
        ])

    _xlsx_autofit(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"fso_mb_{location_code}_{fso_asset_code}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/tank-stock-ledger/rebuild")
def rebuild_tank_stock_ledger(
    location_code: str | None = None,
    tank_asset_code: str | None = None,
    product_name: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Tank Stock Ledger",
        db,
    )

    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
    )

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_tank_asset_code = clean_optional_text(tank_asset_code)
    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_location_code:
        query = query.filter(TankStockLedger.location_code.ilike(cleaned_location_code))

    if cleaned_tank_asset_code:
        query = query.filter(TankStockLedger.tank_asset_code.ilike(cleaned_tank_asset_code))

    if cleaned_product_name:
        query = query.filter(TankStockLedger.product_name.ilike(cleaned_product_name))

    rows = query.all()

    group_keys = set()

    for row in rows:
        group_keys.add(
            (
                row.location_code,
                row.tank_asset_code,
                row.product_name,
            )
        )

    for location, tank_asset, product in group_keys:
        rebuild_tank_stock_running_balances(
            db=db,
            location_code=location,
            tank_asset_code=tank_asset,
            product_name=product,
        )

    create_audit_log(
        db=db,
        module_name="Tank Stock Ledger",
        action="Rebuild Tank Stock Ledger",
        current_user=current_user,
        entity_type="TankStockLedger",
        entity_id=None,
        entity_label="Tank Stock Ledger Rebuild",
        remarks="Rebuilt stock movements from chronological tank stock snapshots",
        request_path="/tank-stock-ledger/rebuild",
        details={
            "location_code": cleaned_location_code,
            "tank_asset_code": cleaned_tank_asset_code,
            "product_name": cleaned_product_name,
            "groups_rebuilt": len(group_keys),
            "rows_scanned": len(rows),
        },
    )

    db.commit()

    return {
        "message": "Tank Stock Ledger rebuilt successfully",
        "groups_rebuilt": len(group_keys),
        "rows_scanned": len(rows),
    }
# -------------------------
# Operation Transaction APIs
# -------------------------

def generate_operation_number(db: Session):
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"OP-{today}"

    existing_count = db.query(OperationTransaction).filter(
        OperationTransaction.operation_number.ilike(f"{prefix}%")
    ).count()

    next_number = existing_count + 1

    return f"{prefix}-{next_number:04d}"


def get_transaction_ticket_number(transaction: OperationTransaction):
    return transaction.operation_ticket_number or transaction.operation_number or ""

def get_current_user_display_name(current_user: User):
    if current_user.full_name:
        return f"{current_user.full_name} ({current_user.username})"

    return current_user.username


def create_audit_log(
    db: Session,
    module_name: str,
    action: str,
    current_user: User | None = None,
    entity_type: str | None = None,
    entity_id: int | None = None,
    entity_label: str | None = None,
    ticket_number: str | None = None,
    operation_number: str | None = None,
    old_status: str | None = None,
    new_status: str | None = None,
    remarks: str | None = None,
    request_path: str | None = None,
    details: dict | None = None,
):
    performed_by = None

    if current_user:
        performed_by = get_current_user_display_name(current_user)

    # ✅ Convert datetime/date/Decimal/etc into JSON-safe values
    safe_details = jsonable_encoder(details) if details is not None else None

    audit_log = AuditLog(
        module_name=module_name,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_label=entity_label,
        ticket_number=ticket_number,
        operation_number=operation_number,
        old_status=old_status,
        new_status=new_status,
        performed_by=performed_by,
        remarks=remarks,
        request_path=request_path,
        details=safe_details,
    )

    db.add(audit_log)
    return audit_log

def build_audit_log_response(audit_log: AuditLog):
    return {
        "id": audit_log.id,
        "module_name": audit_log.module_name,
        "action": audit_log.action,
        "entity_type": audit_log.entity_type,
        "entity_id": audit_log.entity_id,
        "entity_label": audit_log.entity_label,
        "ticket_number": audit_log.ticket_number,
        "operation_number": audit_log.operation_number,
        "old_status": audit_log.old_status,
        "new_status": audit_log.new_status,
        "performed_by": audit_log.performed_by,
        "remarks": audit_log.remarks,
        "request_path": audit_log.request_path,
        "details": audit_log.details,
        "created_at": audit_log.created_at,
    }

def get_location_name_by_code(location_code: str | None, db: Session):
    if not location_code:
        return None

    location = db.query(Location).filter(
        Location.location_code == location_code
    ).first()

    if not location:
        return None

    return location.location_name


def get_location_by_code(location_code: str | None, db: Session):
    if not location_code:
        return None

    return db.query(Location).filter(
        Location.location_code == location_code
    ).first()


def get_asset_by_code(asset_code: str | None, db: Session):
    if not asset_code:
        return None

    return db.query(Asset).filter(
        Asset.asset_code == asset_code
    ).first()


def get_operation_type_by_code(operation_type_code: str | None, db: Session):
    if not operation_type_code:
        return None

    return db.query(OperationType).filter(
        OperationType.operation_type_code == operation_type_code
    ).first()


def build_operation_transaction_response(
    transaction: OperationTransaction,
    db: Session,
):
    operation_type = get_operation_type_by_code(
        transaction.operation_type_code,
        db,
    )

    asset = get_asset_by_code(transaction.primary_asset_code, db)

    return {
        "id": transaction.id,
        "operation_number": transaction.operation_number,
        "operation_ticket_number": get_transaction_ticket_number(transaction),
        "ticket_number": get_transaction_ticket_number(transaction),
        "operation_type_code": transaction.operation_type_code,
        "operation_type_name": (
            operation_type.operation_type_name if operation_type else ""
        ),
        "primary_asset_code": transaction.primary_asset_code,
        "primary_asset_name": asset.asset_name if asset else "",
        "primary_asset_type_code": transaction.primary_asset_type_code,
        "convoy_number": transaction.convoy_number,
        "origin_location_code": transaction.origin_location_code,
        "origin_location_name": get_location_name_by_code(
            transaction.origin_location_code,
            db,
        ),
        "destination_location_code": transaction.destination_location_code,
        "destination_location_name": get_location_name_by_code(
            transaction.destination_location_code,
            db,
        ),
        "sender_location_code": transaction.sender_location_code,
        "sender_location_name": get_location_name_by_code(
            transaction.sender_location_code,
            db,
        ),
        "receiver_location_code": transaction.receiver_location_code,
        "receiver_location_name": get_location_name_by_code(
            transaction.receiver_location_code,
            db,
        ),
        "operation_date": transaction.operation_date,
        "operation_start_datetime": transaction.operation_start_datetime,
        "operation_end_datetime": transaction.operation_end_datetime,
        "product_name": transaction.product_name,
        "created_by": transaction.created_by,
        "remarks": transaction.remarks,
        "status": transaction.status,
        "created_at": transaction.created_at,
        "updated_at": transaction.updated_at,
    }


def build_operation_transaction_register_row(
    transaction: OperationTransaction,
    db: Session,
):
    operation_type = get_operation_type_by_code(
        transaction.operation_type_code,
        db,
    )

    location = get_location_by_code(transaction.origin_location_code, db)
    primary_asset = get_asset_by_code(transaction.primary_asset_code, db)

    field_count = (
        db.query(OperationTransactionValue)
        .filter(OperationTransactionValue.transaction_id == transaction.id)
        .count()
    )

    return {
        "id": transaction.id,
        "operation_number": transaction.operation_number,
        "operation_ticket_number": get_transaction_ticket_number(transaction),
        "ticket_number": get_transaction_ticket_number(transaction),
        "operation_date": transaction.operation_date,
        "operation_type_id": operation_type.id if operation_type else None,
        "operation_type_code": transaction.operation_type_code,
        "operation_type_name": operation_type.operation_type_name
        if operation_type
        else "",
        "location_id": location.id if location else None,
        "location_name": location.location_name if location else "",
        "location_code": transaction.origin_location_code,
        "primary_asset_id": primary_asset.id if primary_asset else None,
        "primary_asset_name": primary_asset.asset_name
        if primary_asset
        else "",
        "primary_asset_code": transaction.primary_asset_code,
        "convoy_number": transaction.convoy_number,
        "status": transaction.status,
        "field_count": field_count,
        "created_at": transaction.created_at,
    }


def validate_operation_transaction(
    transaction: OperationTransactionCreate,
    db: Session,
):
    if not transaction.operation_type_code:
        raise HTTPException(
            status_code=400,
            detail="Operation type is missing in operation entry request",
        )

    if not transaction.primary_asset_code:
        raise HTTPException(
            status_code=400,
            detail="Primary asset is missing in operation entry request",
        )

    if not transaction.origin_location_code:
        raise HTTPException(
            status_code=400,
            detail="Origin location is missing in operation entry request",
        )

    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code.ilike(transaction.operation_type_code)
    ).first()

    if not operation_type:
        raise HTTPException(
            status_code=400,
            detail="Operation type not found",
        )

    if operation_type.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active operation types can be used",
        )

    asset = db.query(Asset).filter(
        Asset.asset_code.ilike(transaction.primary_asset_code)
    ).first()

    if not asset:
        raise HTTPException(
            status_code=400,
            detail="Asset not found",
        )

    if asset.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active assets can be used for operation",
        )

    if (
        asset.asset_type_code.lower()
        != operation_type.applicable_asset_type_code.lower()
    ):
        raise HTTPException(
            status_code=400,
            detail="Selected operation type is not applicable for this asset type",
        )

    origin_location = db.query(Location).filter(
        Location.location_code.ilike(transaction.origin_location_code)
    ).first()

    if not origin_location:
        raise HTTPException(
            status_code=400,
            detail="Origin location not found",
        )

    if origin_location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active origin location can be used",
        )

    if transaction.destination_location_code:
        destination_location = db.query(Location).filter(
            Location.location_code.ilike(transaction.destination_location_code)
        ).first()

        if not destination_location:
            raise HTTPException(
                status_code=400,
                detail="Destination location not found",
            )

        if destination_location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active destination location can be used",
            )

    if operation_type.requires_sender_location == "Yes":
        if not transaction.sender_location_code:
            raise HTTPException(
                status_code=400,
                detail="Sender location is required for this operation type",
            )

    if operation_type.requires_receiver_location == "Yes":
        if not transaction.receiver_location_code:
            raise HTTPException(
                status_code=400,
                detail="Receiver location is required for this operation type",
            )

    return operation_type, asset

def get_filtered_operation_transaction_rows(
    db: Session,
    date_from: str | None = None,
    date_to: str | None = None,
    operation_type_id: int | None = None,
    operation_type_code: str | None = None,
    location_id: int | None = None,
    location_code: str | None = None,
    asset_id: int | None = None,
    asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
):
    query = db.query(OperationTransaction)

    if date_from:
        query = query.filter(OperationTransaction.operation_date >= date_from)

    if date_to:
        query = query.filter(OperationTransaction.operation_date <= date_to)

    resolved_operation_type_code = clean_optional_text(operation_type_code)

    if operation_type_id:
        operation_type = (
            db.query(OperationType)
            .filter(OperationType.id == operation_type_id)
            .first()
        )

        if operation_type:
            resolved_operation_type_code = operation_type.operation_type_code

    if resolved_operation_type_code:
        query = query.filter(
            OperationTransaction.operation_type_code.ilike(
                resolved_operation_type_code
            )
        )

    resolved_location_code = clean_optional_text(location_code)

    if location_id:
        location = (
            db.query(Location)
            .filter(Location.id == location_id)
            .first()
        )

        if location:
            resolved_location_code = location.location_code

    if resolved_location_code:
        query = query.filter(
            OperationTransaction.origin_location_code.ilike(
                resolved_location_code
            )
        )

    resolved_asset_code = clean_optional_text(asset_code)

    if asset_id:
        asset = (
            db.query(Asset)
            .filter(Asset.id == asset_id)
            .first()
        )

        if asset:
            resolved_asset_code = asset.asset_code

    if resolved_asset_code:
        query = query.filter(
            OperationTransaction.primary_asset_code.ilike(
                resolved_asset_code
            )
        )

    if status:
        query = query.filter(OperationTransaction.status == status)

    transactions = query.order_by(OperationTransaction.id.desc()).all()

    result = []

    for transaction in transactions:
        row = build_operation_transaction_register_row(transaction, db)

        if search:
            search_value = search.lower().strip()

            searchable_text = " ".join(
                [
                    str(row["ticket_number"] or ""),
                    str(row["operation_number"] or ""),
                    str(row["operation_type_code"] or ""),
                    str(row["operation_type_name"] or ""),
                    str(row["location_name"] or ""),
                    str(row["location_code"] or ""),
                    str(row["primary_asset_name"] or ""),
                    str(row["primary_asset_code"] or ""),
                    str(row["status"] or ""),
                ]
            ).lower()

            if search_value not in searchable_text:
                continue

        result.append(row)

    return result

@app.get("/operation-transactions")
def get_operation_transactions(
    date_from: str | None = None,
    date_to: str | None = None,
    operation_type_id: int | None = None,
    operation_type_code: str | None = None,
    location_id: int | None = None,
    location_code: str | None = None,
    asset_id: int | None = None,
    asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    return get_filtered_operation_transaction_rows(
        db=db,
        date_from=date_from,
        date_to=date_to,
        operation_type_id=operation_type_id,
        operation_type_code=operation_type_code,
        location_id=location_id,
        location_code=location_code,
        asset_id=asset_id,
        asset_code=asset_code,
        status=status,
        search=search,
    )
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )
    query = db.query(OperationTransaction)

    if date_from:
        query = query.filter(OperationTransaction.operation_date >= date_from)

    if date_to:
        query = query.filter(OperationTransaction.operation_date <= date_to)

    resolved_operation_type_code = clean_optional_text(operation_type_code)

    if operation_type_id:
        operation_type = db.query(OperationType).filter(
            OperationType.id == operation_type_id
        ).first()

        if operation_type:
            resolved_operation_type_code = operation_type.operation_type_code

    if resolved_operation_type_code:
        query = query.filter(
            OperationTransaction.operation_type_code.ilike(
                resolved_operation_type_code
            )
        )

    resolved_location_code = clean_optional_text(location_code)

    if location_id:
        location = db.query(Location).filter(Location.id == location_id).first()

        if location:
            resolved_location_code = location.location_code

    if resolved_location_code:
        query = query.filter(
            OperationTransaction.origin_location_code.ilike(
                resolved_location_code
            )
        )

    resolved_asset_code = clean_optional_text(asset_code)

    if asset_id:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()

        if asset:
            resolved_asset_code = asset.asset_code

    if resolved_asset_code:
        query = query.filter(
            OperationTransaction.primary_asset_code.ilike(
                resolved_asset_code
            )
        )

    if status:
        query = query.filter(OperationTransaction.status == status)

    transactions = query.order_by(OperationTransaction.id.desc()).all()

    result = []

    for transaction in transactions:
        row = build_operation_transaction_register_row(transaction, db)

        if search:
            search_value = search.lower().strip()

            searchable_text = " ".join(
                [
                    str(row["ticket_number"] or ""),
                    str(row["operation_number"] or ""),
                    str(row["operation_type_code"] or ""),
                    str(row["operation_type_name"] or ""),
                    str(row["location_name"] or ""),
                    str(row["location_code"] or ""),
                    str(row["primary_asset_name"] or ""),
                    str(row["primary_asset_code"] or ""),
                    str(row["status"] or ""),
                ]
            ).lower()

            if search_value not in searchable_text:
                continue

        result.append(row)

    return result

from sqlalchemy.orm import aliased

@app.get(
    "/operation-transactions/paged",
    response_model=OperationTransactionRegisterPagedResponse,
)
def get_operation_transactions_paged(
    date_from: str | None = None,
    date_to: str | None = None,
    operation_type_id: int | None = None,
    operation_type_code: str | None = None,
    location_id: int | None = None,
    location_code: str | None = None,
    asset_id: int | None = None,
    asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Operation Transaction", db)

    # Guardrails
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 1
    if page_size > 200:
        page_size = 200

    resolved_operation_type_code = clean_optional_text(operation_type_code)
    resolved_location_code = clean_optional_text(location_code)
    resolved_asset_code = clean_optional_text(asset_code)
    resolved_search = clean_optional_text(search)
    resolved_status = clean_optional_text(status)

    # Resolve IDs -> codes (same behavior as existing filter helper)
    if operation_type_id:
        op = (
            db.query(OperationType)
            .filter(OperationType.id == operation_type_id)
            .first()
        )
        if op:
            resolved_operation_type_code = op.operation_type_code

    if location_id:
        loc = db.query(Location).filter(Location.id == location_id).first()
        if loc:
            resolved_location_code = loc.location_code

    if asset_id:
        ast = db.query(Asset).filter(Asset.id == asset_id).first()
        if ast:
            resolved_asset_code = ast.asset_code

    # Aliases (avoid any ambiguous join issues later)
    OT = aliased(OperationType)
    LOC = aliased(Location)
    AST = aliased(Asset)

    # Field count subquery
    value_count_subq = (
        db.query(
            OperationTransactionValue.transaction_id.label("tx_id"),
            func.count(OperationTransactionValue.id).label("field_count"),
        )
        .group_by(OperationTransactionValue.transaction_id)
        .subquery()
    )

    base_query = (
        db.query(
            OperationTransaction.id.label("id"),
            OperationTransaction.operation_number.label("operation_number"),
            OperationTransaction.operation_ticket_number.label("operation_ticket_number"),
            OperationTransaction.convoy_number.label("convoy_number"),
            OperationTransaction.operation_date.label("operation_date"),
            OperationTransaction.operation_type_code.label("operation_type_code"),
            OperationTransaction.origin_location_code.label("origin_location_code"),
            OperationTransaction.primary_asset_code.label("primary_asset_code"),
            OperationTransaction.status.label("status"),
            OperationTransaction.created_at.label("created_at"),
            OT.id.label("operation_type_id"),
            OT.operation_type_name.label("operation_type_name"),
            LOC.id.label("location_id"),
            LOC.location_name.label("location_name"),
            AST.id.label("asset_id"),
            AST.asset_name.label("asset_name"),
            func.coalesce(value_count_subq.c.field_count, 0).label("field_count"),
        )
        .outerjoin(OT, OT.operation_type_code == OperationTransaction.operation_type_code)
        .outerjoin(LOC, LOC.location_code == OperationTransaction.origin_location_code)
        .outerjoin(AST, AST.asset_code == OperationTransaction.primary_asset_code)
        .outerjoin(value_count_subq, value_count_subq.c.tx_id == OperationTransaction.id)
    )

    # Filters
    if date_from:
        base_query = base_query.filter(OperationTransaction.operation_date >= date_from)

    if date_to:
        base_query = base_query.filter(OperationTransaction.operation_date <= date_to)

    if resolved_operation_type_code:
        base_query = base_query.filter(
            OperationTransaction.operation_type_code.ilike(resolved_operation_type_code)
        )

    if resolved_location_code:
        base_query = base_query.filter(
            OperationTransaction.origin_location_code.ilike(resolved_location_code)
        )

    if resolved_asset_code:
        base_query = base_query.filter(
            OperationTransaction.primary_asset_code.ilike(resolved_asset_code)
        )

    if resolved_status:
        base_query = base_query.filter(OperationTransaction.status == resolved_status)

    if resolved_search:
        s = f"%{resolved_search.lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(func.coalesce(OperationTransaction.operation_ticket_number, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.operation_number, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.operation_type_code, "")).ilike(s),
                func.lower(func.coalesce(OT.operation_type_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.origin_location_code, "")).ilike(s),
                func.lower(func.coalesce(LOC.location_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.primary_asset_code, "")).ilike(s),
                func.lower(func.coalesce(AST.asset_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.status, "")).ilike(s),
            )
        )

    # Total rows count
    total_rows = base_query.count()

    # Status counts (same filters, but ignore status filter so tabs can show totals)
    count_query = (
        db.query(OperationTransaction.status, func.count(OperationTransaction.id))
        .outerjoin(OT, OT.operation_type_code == OperationTransaction.operation_type_code)
        .outerjoin(LOC, LOC.location_code == OperationTransaction.origin_location_code)
        .outerjoin(AST, AST.asset_code == OperationTransaction.primary_asset_code)
    )

    if date_from:
        count_query = count_query.filter(OperationTransaction.operation_date >= date_from)
    if date_to:
        count_query = count_query.filter(OperationTransaction.operation_date <= date_to)
    if resolved_operation_type_code:
        count_query = count_query.filter(OperationTransaction.operation_type_code.ilike(resolved_operation_type_code))
    if resolved_location_code:
        count_query = count_query.filter(OperationTransaction.origin_location_code.ilike(resolved_location_code))
    if resolved_asset_code:
        count_query = count_query.filter(OperationTransaction.primary_asset_code.ilike(resolved_asset_code))
    if resolved_search:
        s = f"%{resolved_search.lower()}%"
        count_query = count_query.filter(
            or_(
                func.lower(func.coalesce(OperationTransaction.operation_ticket_number, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.operation_number, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.operation_type_code, "")).ilike(s),
                func.lower(func.coalesce(OT.operation_type_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.origin_location_code, "")).ilike(s),
                func.lower(func.coalesce(LOC.location_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.primary_asset_code, "")).ilike(s),
                func.lower(func.coalesce(AST.asset_name, "")).ilike(s),
                func.lower(func.coalesce(OperationTransaction.status, "")).ilike(s),
            )
        )

    status_counts_raw = (
        count_query.group_by(OperationTransaction.status)
        .all()
    )

    status_counts = [
        {"status": (row[0] or ""), "count": int(row[1] or 0)}
        for row in status_counts_raw
        if (row[0] or "").strip() != ""
    ]

    # Paging
    offset = (page - 1) * page_size

    rows_raw = (
        base_query.order_by(OperationTransaction.id.desc())
        .offset(offset)
        .limit(page_size + 1)
        .all()
    )

    has_more = len(rows_raw) > page_size
    rows_raw = rows_raw[:page_size]

    rows = []
    for r in rows_raw:
        ticket_number = r.operation_ticket_number or r.operation_number or ""
        rows.append(
            {
                "id": r.id,
                "ticket_number": ticket_number,
                "operation_number": r.operation_number,
                "convoy_number": r.convoy_number,
                "operation_date": r.operation_date,
                "operation_type_id": r.operation_type_id,
                "operation_type_code": r.operation_type_code,
                "operation_type_name": r.operation_type_name or "",
                "location_id": r.location_id,
                "location_code": r.origin_location_code,
                "location_name": r.location_name or "",
                "primary_asset_id": r.asset_id,
                "primary_asset_code": r.primary_asset_code,
                "primary_asset_name": r.asset_name or "",
                "status": r.status or "",
                "field_count": int(r.field_count or 0),
                "created_at": r.created_at,
            }
        )

    return {
        "rows": rows,
        "total_rows": total_rows,
        "page": page,
        "page_size": page_size,
        "has_more": has_more,
        "status_counts": status_counts,
    }

@app.get("/operation-transactions/export/csv")
def export_operation_transactions_csv(
    date_from: str | None = None,
    date_to: str | None = None,
    operation_type_id: int | None = None,
    operation_type_code: str | None = None,
    location_id: int | None = None,
    location_code: str | None = None,
    asset_id: int | None = None,
    asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    rows = get_filtered_operation_transaction_rows(
        db=db,
        date_from=date_from,
        date_to=date_to,
        operation_type_id=operation_type_id,
        operation_type_code=operation_type_code,
        location_id=location_id,
        location_code=location_code,
        asset_id=asset_id,
        asset_code=asset_code,
        status=status,
        search=search,
    )

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Operation Transaction Register"])
    writer.writerow(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow(["Record Count", len(rows)])
    writer.writerow([])

    writer.writerow(["Applied Filters"])
    writer.writerow(["Date From", date_from or "All"])
    writer.writerow(["Date To", date_to or "All"])
    writer.writerow(["Operation Type ID", operation_type_id or "All"])
    writer.writerow(["Operation Type Code", operation_type_code or "All"])
    writer.writerow(["Location ID", location_id or "All"])
    writer.writerow(["Location Code", location_code or "All"])
    writer.writerow(["Asset ID", asset_id or "All"])
    writer.writerow(["Asset Code", asset_code or "All"])
    writer.writerow(["Status", status or "All"])
    writer.writerow(["Search", search or ""])
    writer.writerow([])

    writer.writerow(
        [
            "Ticket Number",
            "Operation Number",
            "Operation Date",
            "Operation Type Code",
            "Operation Type Name",
            "Location Code",
            "Location Name",
            "Primary Asset Code",
            "Primary Asset Name",
            "Field Count",
            "Status",
            "Created At",
        ]
    )

    for row in rows:
        writer.writerow(
            [
                row.get("ticket_number", ""),
                row.get("operation_number", ""),
                row.get("operation_date", ""),
                row.get("operation_type_code", ""),
                row.get("operation_type_name", ""),
                row.get("location_code", ""),
                row.get("location_name", ""),
                row.get("primary_asset_code", ""),
                row.get("primary_asset_name", ""),
                row.get("field_count", ""),
                row.get("status", ""),
                row.get("created_at", ""),
            ]
        )

    output.seek(0)

    filename = f"operation-transaction-register-{datetime.now().strftime('%Y%m%d-%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )

@app.post(
    "/operation-transactions",
    response_model=OperationTransactionResponse,
)
def create_operation_transaction(
    transaction: OperationTransactionCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    # This is a legacy/direct endpoint. Keep it secure.
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    operation_type, asset = validate_operation_transaction(transaction, db)

    created_by_display = get_current_user_display_name(current_user)

    # Always control created_by server-side (prevent spoofing)
    new_transaction = OperationTransaction(
        operation_number=generate_operation_number(db),
        operation_type_code=operation_type.operation_type_code,
        primary_asset_code=asset.asset_code,
        primary_asset_type_code=asset.asset_type_code,
        convoy_number=clean_optional_text(transaction.convoy_number),
        origin_location_code=transaction.origin_location_code.strip(),
        destination_location_code=clean_optional_text(
            transaction.destination_location_code
        ),
        sender_location_code=clean_optional_text(transaction.sender_location_code),
        receiver_location_code=clean_optional_text(transaction.receiver_location_code),
        operation_date=transaction.operation_date,
        operation_start_datetime=transaction.operation_start_datetime,
        operation_end_datetime=transaction.operation_end_datetime,
        product_name=clean_optional_text(transaction.product_name),
        created_by=created_by_display,
        remarks=clean_optional_text(transaction.remarks),
        status=transaction.status or "Draft",
    )

    db.add(new_transaction)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Create Operation Transaction",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=new_transaction.id,
        entity_label=get_transaction_ticket_number(new_transaction),
        ticket_number=get_transaction_ticket_number(new_transaction),
        operation_number=new_transaction.operation_number,
        new_status=new_transaction.status,
        remarks="Created via /operation-transactions",
        request_path="/operation-transactions",
        details={
            "operation_type_code": new_transaction.operation_type_code,
            "primary_asset_code": new_transaction.primary_asset_code,
            "origin_location_code": new_transaction.origin_location_code,
            "destination_location_code": new_transaction.destination_location_code,
            "sender_location_code": new_transaction.sender_location_code,
            "receiver_location_code": new_transaction.receiver_location_code,
            "operation_date": str(new_transaction.operation_date),
        },
    )

    db.commit()
    db.refresh(new_transaction)

    return build_operation_transaction_response(new_transaction, db)


@app.put(
    "/operation-transactions/{transaction_id}",
    response_model=OperationTransactionResponse,
)
def update_operation_transaction(
    transaction_id: int,
    transaction: OperationTransactionCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    # This is a legacy/direct endpoint. Keep it secure.
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    existing_transaction = db.query(OperationTransaction).filter(
        OperationTransaction.id == transaction_id
    ).first()

    if not existing_transaction:
        raise HTTPException(
            status_code=404,
            detail="Operation transaction not found",
        )

    # Match Operation Entry edit rule
    if existing_transaction.status not in ["Draft", "Rejected"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Rejected operation transactions can be edited."
            ),
        )

    before_data = {
        "operation_type_code": existing_transaction.operation_type_code,
        "primary_asset_code": existing_transaction.primary_asset_code,
        "convoy_number": existing_transaction.convoy_number,
        "origin_location_code": existing_transaction.origin_location_code,
        "destination_location_code": existing_transaction.destination_location_code,
        "sender_location_code": existing_transaction.sender_location_code,
        "receiver_location_code": existing_transaction.receiver_location_code,
        "operation_date": str(existing_transaction.operation_date),
        "product_name": existing_transaction.product_name,
        "remarks": existing_transaction.remarks,
        "status": existing_transaction.status,
        "created_by": existing_transaction.created_by,
    }

    operation_type, asset = validate_operation_transaction(transaction, db)

    existing_transaction.operation_type_code = operation_type.operation_type_code
    existing_transaction.primary_asset_code = asset.asset_code
    existing_transaction.primary_asset_type_code = asset.asset_type_code
    existing_transaction.convoy_number = clean_optional_text(transaction.convoy_number)
    existing_transaction.origin_location_code = transaction.origin_location_code.strip()
    existing_transaction.destination_location_code = clean_optional_text(
        transaction.destination_location_code
    )
    existing_transaction.sender_location_code = clean_optional_text(
        transaction.sender_location_code
    )
    existing_transaction.receiver_location_code = clean_optional_text(
        transaction.receiver_location_code
    )
    existing_transaction.operation_date = transaction.operation_date
    existing_transaction.operation_start_datetime = transaction.operation_start_datetime
    existing_transaction.operation_end_datetime = transaction.operation_end_datetime
    existing_transaction.product_name = clean_optional_text(transaction.product_name)

    # IMPORTANT: do NOT allow client to change created_by
    existing_transaction.remarks = clean_optional_text(transaction.remarks)
    existing_transaction.updated_at = datetime.now()

    after_data = {
        "operation_type_code": existing_transaction.operation_type_code,
        "primary_asset_code": existing_transaction.primary_asset_code,
        "convoy_number": existing_transaction.convoy_number,
        "origin_location_code": existing_transaction.origin_location_code,
        "destination_location_code": existing_transaction.destination_location_code,
        "sender_location_code": existing_transaction.sender_location_code,
        "receiver_location_code": existing_transaction.receiver_location_code,
        "operation_date": str(existing_transaction.operation_date),
        "product_name": existing_transaction.product_name,
        "remarks": existing_transaction.remarks,
        "status": existing_transaction.status,
        "created_by": existing_transaction.created_by,
    }

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Update Operation Transaction",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=existing_transaction.id,
        entity_label=get_transaction_ticket_number(existing_transaction),
        ticket_number=get_transaction_ticket_number(existing_transaction),
        operation_number=existing_transaction.operation_number,
        old_status=existing_transaction.status,
        new_status=existing_transaction.status,
        remarks="Updated via /operation-transactions",
        request_path=f"/operation-transactions/{transaction_id}",
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(existing_transaction)

    return build_operation_transaction_response(existing_transaction, db)


@app.delete("/operation-transactions/{transaction_id}")
def delete_operation_transaction(
    transaction_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Cancel Operation Transaction",
        db,
    )

    existing_transaction = db.query(OperationTransaction).filter(
        OperationTransaction.id == transaction_id
    ).first()

    if not existing_transaction:
        raise HTTPException(
            status_code=404,
            detail="Operation transaction not found",
        )

    if existing_transaction.status not in ["Draft", "Rejected"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Rejected operation transactions can be cancelled. "
                "Submitted tickets must be recalled to Draft before cancelling. "
                "Approved and Cancelled tickets are locked."
            ),
        )

    old_status = existing_transaction.status

    changed_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )

    existing_transaction.status = "Cancelled"
    existing_transaction.updated_at = datetime.now()

    existing_remarks = existing_transaction.remarks or ""

    existing_transaction.remarks = (
        f"{existing_remarks}\n"
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
        f"Cancelled by {changed_by}"
    ).strip()

    history = OperationTransactionStatusHistory(
        transaction_id=existing_transaction.id,
        old_status=old_status,
        new_status="Cancelled",
        changed_by=changed_by,
        remarks="Cancelled from Operation Transaction Register",
        changed_at=datetime.now(),
    )

    db.add(history)

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Cancel Operation Transaction",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=existing_transaction.id,
        entity_label=get_transaction_ticket_number(existing_transaction),
        ticket_number=get_transaction_ticket_number(existing_transaction),
        operation_number=existing_transaction.operation_number,
        old_status=old_status,
        new_status="Cancelled",
        remarks="Cancelled from Operation Transaction Register",
        request_path=f"/operation-transactions/{transaction_id}",
        details={
            "operation_type_code": existing_transaction.operation_type_code,
            "operation_template_id": existing_transaction.operation_template_id,
            "primary_asset_code": existing_transaction.primary_asset_code,
            "origin_location_code": existing_transaction.origin_location_code,
            "operation_date": str(existing_transaction.operation_date),
        },
    )

    db.commit()
    db.refresh(existing_transaction)

    return {
        "message": "Operation transaction cancelled successfully"
    }

# -------------------------
# Location Operation Availability APIs
# -------------------------

def build_location_operation_availability_response(
    availability: LocationOperationAvailability,
    db: Session,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(availability.location_code))
        .first()
    )

    operation_type = (
        db.query(OperationType)
        .filter(
            OperationType.operation_type_code.ilike(
                availability.operation_type_code
            )
        )
        .first()
    )

    return {
        "id": availability.id,
        "location_code": availability.location_code,
        "location_name": location.location_name if location else "",
        "operation_type_code": availability.operation_type_code,
        "operation_type_name": (
            operation_type.operation_type_name if operation_type else ""
        ),
        "status": availability.status,
        "remarks": availability.remarks,
        "created_at": availability.created_at,
        "updated_at": availability.updated_at,
    }


def validate_location_operation_availability(
    availability: LocationOperationAvailabilityCreate,
    db: Session,
    availability_id: int | None = None,
):
    location = (
        db.query(Location)
        .filter(Location.location_code.ilike(availability.location_code))
        .first()
    )

    if not location:
        raise HTTPException(
            status_code=400,
            detail="Location not found",
        )

    if location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active locations can be configured",
        )

    operation_type = (
        db.query(OperationType)
        .filter(
            OperationType.operation_type_code.ilike(
                availability.operation_type_code
            )
        )
        .first()
    )

    if not operation_type:
        raise HTTPException(
            status_code=400,
            detail="Operation type not found",
        )

    if operation_type.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active operation types can be configured",
        )

    duplicate_query = db.query(LocationOperationAvailability).filter(
        LocationOperationAvailability.location_code.ilike(
            availability.location_code
        ),
        LocationOperationAvailability.operation_type_code.ilike(
            availability.operation_type_code
        ),
    )

    if availability_id is not None:
        duplicate_query = duplicate_query.filter(
            LocationOperationAvailability.id != availability_id
        )

    duplicate = duplicate_query.first()

    if duplicate:
        raise HTTPException(
            status_code=400,
            detail="This operation type is already configured for this location",
        )


@app.get(
    "/location-operation-availability",
    response_model=list[LocationOperationAvailabilityResponse],
)
def get_location_operation_availability(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Location Operation Availability",
        db,
    )

    availability_records = (
        db.query(LocationOperationAvailability)
        .order_by(LocationOperationAvailability.id)
        .all()
    )

    return [
        build_location_operation_availability_response(record, db)
        for record in availability_records
    ]

def build_location_operation_availability_audit_snapshot(
    availability: LocationOperationAvailability,
    db: Session,
):
    location = db.query(Location).filter(
        Location.location_code.ilike(availability.location_code)
    ).first()

    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code.ilike(availability.operation_type_code)
    ).first()

    return {
        "id": availability.id,
        "location_code": availability.location_code,
        "location_name": location.location_name if location else "",
        "operation_type_code": availability.operation_type_code,
        "operation_type_name": operation_type.operation_type_name if operation_type else "",
        "status": availability.status,
        "remarks": availability.remarks,
    }

@app.post(
    "/location-operation-availability",
    response_model=LocationOperationAvailabilityResponse,
)
def create_location_operation_availability(
    availability: LocationOperationAvailabilityCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Operation Availability",
        db,
    )

    validate_location_operation_availability(availability, db)

    new_record = LocationOperationAvailability(
        location_code=availability.location_code.strip(),
        operation_type_code=availability.operation_type_code.strip(),
        status=availability.status,
        remarks=clean_optional_text(availability.remarks),
    )

    db.add(new_record)
    db.flush()

    after_data = build_location_operation_availability_audit_snapshot(new_record, db)

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Create Location Operation Availability",
        current_user=current_user,
        entity_type="LocationOperationAvailability",
        entity_id=new_record.id,
        entity_label=f"{after_data.get('location_code')} - {after_data.get('operation_type_code')}",
        remarks="Location operation availability created",
        request_path="/location-operation-availability",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_record)

    return build_location_operation_availability_response(new_record, db)


@app.put(
    "/location-operation-availability/{availability_id}",
    response_model=LocationOperationAvailabilityResponse,
)
def update_location_operation_availability(
    availability_id: int,
    availability: LocationOperationAvailabilityCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Operation Availability",
        db,
    )

    existing_record = (
        db.query(LocationOperationAvailability)
        .filter(LocationOperationAvailability.id == availability_id)
        .first()
    )

    if not existing_record:
        raise HTTPException(
            status_code=404,
            detail="Location operation availability not found",
        )

    before_data = build_location_operation_availability_audit_snapshot(
        existing_record, db
    )

    validate_location_operation_availability(
        availability,
        db,
        availability_id,
    )

    existing_record.location_code = availability.location_code.strip()
    existing_record.operation_type_code = availability.operation_type_code.strip()
    existing_record.status = availability.status
    existing_record.remarks = clean_optional_text(availability.remarks)

    db.flush()

    after_data = build_location_operation_availability_audit_snapshot(
        existing_record, db
    )

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Update Location Operation Availability",
        current_user=current_user,
        entity_type="LocationOperationAvailability",
        entity_id=existing_record.id,
        entity_label=f"{after_data.get('location_code')} - {after_data.get('operation_type_code')}",
        remarks="Location operation availability updated",
        request_path=f"/location-operation-availability/{availability_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_record)

    return build_location_operation_availability_response(existing_record, db)


@app.delete("/location-operation-availability/{availability_id}")
def delete_location_operation_availability(
    availability_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Location Operation Availability",
        db,
    )

    existing_record = (
        db.query(LocationOperationAvailability)
        .filter(LocationOperationAvailability.id == availability_id)
        .first()
    )

    if not existing_record:
        raise HTTPException(
            status_code=404,
            detail="Location operation availability not found",
        )

    deleted_data = build_location_operation_availability_audit_snapshot(
        existing_record, db
    )

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Delete Location Operation Availability",
        current_user=current_user,
        entity_type="LocationOperationAvailability",
        entity_id=existing_record.id,
        entity_label=f"{deleted_data.get('location_code')} - {deleted_data.get('operation_type_code')}",
        remarks="Location operation availability deleted",
        request_path=f"/location-operation-availability/{availability_id}",
        details={"deleted": deleted_data},
    )

    db.delete(existing_record)
    db.commit()

    return {"message": "Location operation availability deleted successfully"}

# -------------------------
# Operation Template APIs
# -------------------------

VALID_ENTRY_LAYOUT_TYPES = [
    "Standard Form",
    "Stock Movement",
    "Tank Gauging",
    "Multi-Tank Before/After",
    "Vessel Cycle",
    "Tanker Loading",
    "Meter Reading",
    "Shuttle Tracking",
    "FSO Tracking",   # ✅ NEW
]

VALID_CALCULATION_ENGINES = [
    "None",
    "Stock Movement Net/Variance",
    "Tank Quantity",
    "Barge Before/After Quantity",
    "Vessel Cycle Quantity",
    "Tanker Quantity",
    "Meter Reading Quantity",
]

def build_operation_template_response(
    template: OperationTemplate,
    db: Session,
):
    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code == template.operation_type_code
    ).first()

    fields = (
        db.query(OperationTemplateField)
        .filter(OperationTemplateField.template_id == template.id)
        .order_by(OperationTemplateField.sort_order, OperationTemplateField.id)
        .all()
    )

    return {
        "id": template.id,
        "template_name": template.template_name,
        "operation_type_code": template.operation_type_code,
        "operation_type_name": (
            operation_type.operation_type_name if operation_type else ""
        ),
        "entry_layout_type": template.entry_layout_type or "Standard Form",
        "calculation_engine": template.calculation_engine or "None",
        "description": template.description,
        "status": template.status,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
        "fields": [
            {
                "id": field.id,
                "field_name": field.field_name,
                "field_code": field.field_code,
                "field_group": field.field_group,
                "data_type": field.data_type,
                "unit": field.unit,
                "is_required": field.is_required,
                "input_mode": field.input_mode,
                "calculation_role": field.calculation_role,
                "sort_order": field.sort_order,
                "status": field.status,
            }
            for field in fields
        ],
    }


def validate_operation_template(
    template: OperationTemplateCreate,
    db: Session,
):
    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code.ilike(template.operation_type_code)
    ).first()

    if not operation_type:
        raise HTTPException(
            status_code=400,
            detail="Operation type not found",
        )

    if operation_type.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active operation types can be used",
        )

    if template.entry_layout_type not in VALID_ENTRY_LAYOUT_TYPES:
        raise HTTPException(
            status_code=400,
            detail="Invalid entry layout type",
        )

    if template.calculation_engine not in VALID_CALCULATION_ENGINES:
        raise HTTPException(
            status_code=400,
            detail="Invalid calculation engine",
        )

    if len(template.fields) == 0:
        raise HTTPException(
            status_code=400,
            detail="Please add at least one operation template field",
        )

    field_codes = [
        field.field_code.strip().lower()
        for field in template.fields
    ]

    if len(field_codes) != len(set(field_codes)):
        raise HTTPException(
            status_code=400,
            detail="Duplicate field codes are not allowed in the same template",
        )

    field_names = [
        field.field_name.strip().lower()
        for field in template.fields
    ]

    if len(field_names) != len(set(field_names)):
        raise HTTPException(
            status_code=400,
            detail="Duplicate field names are not allowed in the same template",
        )

    return operation_type


@app.get(
    "/operation-templates",
    response_model=list[OperationTemplateResponse],
)
def get_operation_templates(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Template",
        db,
    )

    templates = (
        db.query(OperationTemplate)
        .order_by(OperationTemplate.id)
        .all()
    )

    return [
        build_operation_template_response(template, db)
        for template in templates
    ]

def build_operation_template_audit_snapshot(
    template: OperationTemplate,
    db: Session,
):
    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code == template.operation_type_code
    ).first()

    fields = (
        db.query(OperationTemplateField)
        .filter(OperationTemplateField.template_id == template.id)
        .order_by(OperationTemplateField.sort_order, OperationTemplateField.id)
        .all()
    )

    return {
        "id": template.id,
        "template_name": template.template_name,
        "operation_type_code": template.operation_type_code,
        "operation_type_name": operation_type.operation_type_name if operation_type else "",
        "entry_layout_type": template.entry_layout_type or "Standard Form",
        "calculation_engine": template.calculation_engine or "None",
        "description": template.description,
        "status": template.status,
        "field_count": len(fields),
        "fields": [
            {
                "id": field.id,
                "field_name": field.field_name,
                "field_code": field.field_code,
                "field_group": field.field_group,
                "data_type": field.data_type,
                "unit": field.unit,
                "is_required": field.is_required,
                "input_mode": field.input_mode,
                "calculation_role": field.calculation_role,
                "sort_order": field.sort_order,
                "status": field.status,
            }
            for field in fields
        ],
    }

@app.post(
    "/operation-templates",
    response_model=OperationTemplateResponse,
)
def create_operation_template(
    template: OperationTemplateCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Template", db)

    existing_template = db.query(OperationTemplate).filter(
        OperationTemplate.template_name.ilike(template.template_name)
    ).first()

    if existing_template:
        raise HTTPException(
            status_code=400,
            detail="Operation template name already exists",
        )

    operation_type = validate_operation_template(template, db)

    new_template = OperationTemplate(
        template_name=template.template_name.strip(),
        operation_type_code=operation_type.operation_type_code,
        entry_layout_type=template.entry_layout_type,
        calculation_engine=template.calculation_engine,
        description=clean_optional_text(template.description),
        status=template.status,
    )

    db.add(new_template)
    db.flush()

    for index, field in enumerate(template.fields):
        new_field = OperationTemplateField(
            template_id=new_template.id,
            field_name=field.field_name.strip(),
            field_code=field.field_code.strip(),
            field_group=field.field_group,
            data_type=field.data_type,
            unit=clean_optional_text(field.unit),
            is_required=field.is_required,
            input_mode=field.input_mode,
            calculation_role=field.calculation_role,
            sort_order=field.sort_order or index + 1,
            status=field.status,
        )
        db.add(new_field)

    db.flush()

    after_data = build_operation_template_audit_snapshot(new_template, db)

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Create Operation Template",
        current_user=current_user,
        entity_type="OperationTemplate",
        entity_id=new_template.id,
        entity_label=new_template.template_name,
        remarks="Operation template created",
        request_path="/operation-templates",
        details={"after": after_data},
    )

    db.commit()
    db.refresh(new_template)

    return build_operation_template_response(new_template, db)


@app.put(
    "/operation-templates/{template_id}",
    response_model=OperationTemplateResponse,
)
def update_operation_template(
    template_id: int,
    template: OperationTemplateCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Template", db)

    existing_template = db.query(OperationTemplate).filter(
        OperationTemplate.id == template_id
    ).first()

    if not existing_template:
        raise HTTPException(
            status_code=404,
            detail="Operation template not found",
        )

    duplicate_template = db.query(OperationTemplate).filter(
        OperationTemplate.template_name.ilike(template.template_name),
        OperationTemplate.id != template_id,
    ).first()

    if duplicate_template:
        raise HTTPException(
            status_code=400,
            detail="Operation template name already exists",
        )

    before_data = build_operation_template_audit_snapshot(existing_template, db)

    operation_type = validate_operation_template(template, db)

    existing_template.template_name = template.template_name.strip()
    existing_template.operation_type_code = operation_type.operation_type_code
    existing_template.entry_layout_type = template.entry_layout_type
    existing_template.calculation_engine = template.calculation_engine
    existing_template.description = clean_optional_text(template.description)
    existing_template.status = template.status

    db.query(OperationTemplateField).filter(
        OperationTemplateField.template_id == template_id
    ).delete()

    for index, field in enumerate(template.fields):
        new_field = OperationTemplateField(
            template_id=template_id,
            field_name=field.field_name.strip(),
            field_code=field.field_code.strip(),
            field_group=field.field_group,
            data_type=field.data_type,
            unit=clean_optional_text(field.unit),
            is_required=field.is_required,
            input_mode=field.input_mode,
            calculation_role=field.calculation_role,
            sort_order=field.sort_order or index + 1,
            status=field.status,
        )
        db.add(new_field)

    db.flush()

    after_data = build_operation_template_audit_snapshot(existing_template, db)

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Update Operation Template",
        current_user=current_user,
        entity_type="OperationTemplate",
        entity_id=existing_template.id,
        entity_label=existing_template.template_name,
        remarks="Operation template updated",
        request_path=f"/operation-templates/{template_id}",
        details={"before": before_data, "after": after_data},
    )

    db.commit()
    db.refresh(existing_template)

    return build_operation_template_response(existing_template, db)


@app.delete("/operation-templates/{template_id}")
def delete_operation_template(
    template_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Operation Template", db)

    existing_template = db.query(OperationTemplate).filter(
        OperationTemplate.id == template_id
    ).first()

    if not existing_template:
        raise HTTPException(
            status_code=404,
            detail="Operation template not found",
        )

    existing_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.operation_template_id == template_id)
        .first()
    )

    if existing_transaction:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete operation template because transactions exist for it",
        )

    deleted_data = build_operation_template_audit_snapshot(existing_template, db)

    create_audit_log(
        db=db,
        module_name="Operations",
        action="Delete Operation Template",
        current_user=current_user,
        entity_type="OperationTemplate",
        entity_id=existing_template.id,
        entity_label=existing_template.template_name,
        remarks="Operation template deleted",
        request_path=f"/operation-templates/{template_id}",
        details={"deleted": deleted_data},
    )

    db.query(OperationTemplateField).filter(
        OperationTemplateField.template_id == template_id
    ).delete()

    db.delete(existing_template)
    db.commit()

    return {"message": "Operation template deleted successfully"}

# -------------------------
# Operation Entry APIs
# -------------------------

def build_operation_entry_response(
    transaction: OperationTransaction,
    db: Session,
):
    template = None

    if transaction.operation_template_id:
        template = db.query(OperationTemplate).filter(
            OperationTemplate.id == transaction.operation_template_id
        ).first()

    values = (
        db.query(OperationTransactionValue)
        .filter(OperationTransactionValue.transaction_id == transaction.id)
        .order_by(
            OperationTransactionValue.sort_order,
            OperationTransactionValue.id,
        )
        .all()
    )

    return {
        "transaction": build_operation_transaction_response(transaction, db),
        "operation_template_id": transaction.operation_template_id,
        "operation_template_name": template.template_name if template else "",
        "values": [
            {
                "id": value.id,
                "field_code": value.field_code,
                "field_name": value.field_name,
                "field_group": value.field_group,
                "data_type": value.data_type,
                "unit": value.unit,
                "input_mode": value.input_mode,
                "calculation_role": value.calculation_role,
                "field_value": value.field_value,
                "sort_order": value.sort_order,
            }
            for value in values
        ],
    }


def validate_operation_entry(
    entry: OperationEntryCreate,
    db: Session,
):
    template = db.query(OperationTemplate).filter(
        OperationTemplate.id == entry.operation_template_id
    ).first()

    if not template:
        raise HTTPException(
            status_code=400,
            detail="Operation template not found",
        )

    if template.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active operation templates can be used",
        )

    transaction_operation_type_code = clean_optional_text(
        getattr(entry.transaction, "operation_type_code", None)
    )

    if transaction_operation_type_code is None:
        transaction_operation_type_code = template.operation_type_code

    if transaction_operation_type_code is None:
        raise HTTPException(
            status_code=400,
            detail="Operation type is missing in operation entry request",
        )

    if template.operation_type_code.lower() != transaction_operation_type_code.lower():
        raise HTTPException(
            status_code=400,
            detail="Selected template does not belong to selected operation type",
        )

    operation_type = db.query(OperationType).filter(
        OperationType.operation_type_code.ilike(transaction_operation_type_code)
    ).first()

    if not operation_type:
        raise HTTPException(
            status_code=400,
            detail="Operation type not found",
        )

    if operation_type.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active operation types can be used",
        )

    if not clean_optional_text(entry.transaction.primary_asset_code):
        raise HTTPException(
            status_code=400,
            detail="Primary asset is missing in operation entry request",
        )

    asset = db.query(Asset).filter(
        Asset.asset_code.ilike(entry.transaction.primary_asset_code)
    ).first()

    if not asset:
        raise HTTPException(
            status_code=400,
            detail="Asset not found",
        )

    if asset.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active assets can be used for operation",
        )

    if asset.asset_type_code.lower() != operation_type.applicable_asset_type_code.lower():
        raise HTTPException(
            status_code=400,
            detail="Selected operation type is not applicable for this asset type",
        )

    if not clean_optional_text(entry.transaction.origin_location_code):
        raise HTTPException(
            status_code=400,
            detail="Origin location is missing in operation entry request",
        )

    origin_location = db.query(Location).filter(
        Location.location_code.ilike(entry.transaction.origin_location_code)
    ).first()

    if not origin_location:
        raise HTTPException(
            status_code=400,
            detail="Origin location not found",
        )

    if origin_location.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Only Active origin location can be used",
        )

    if entry.transaction.destination_location_code:
        destination_location = db.query(Location).filter(
            Location.location_code.ilike(entry.transaction.destination_location_code)
        ).first()

        if not destination_location:
            raise HTTPException(
                status_code=400,
                detail="Destination location not found",
            )

        if destination_location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active destination location can be used",
            )

    if operation_type.requires_sender_location == "Yes":
        if not entry.transaction.sender_location_code:
            raise HTTPException(
                status_code=400,
                detail="Sender location is required for this operation type",
            )

    if operation_type.requires_receiver_location == "Yes":
        if not entry.transaction.receiver_location_code:
            raise HTTPException(
                status_code=400,
                detail="Receiver location is required for this operation type",
            )

    template_fields = (
        db.query(OperationTemplateField)
        .filter(
            OperationTemplateField.template_id == template.id,
            OperationTemplateField.status == "Active",
        )
        .order_by(OperationTemplateField.sort_order, OperationTemplateField.id)
        .all()
    )

    if len(template_fields) == 0:
        raise HTTPException(
            status_code=400,
            detail="Selected operation template has no active fields",
        )

    field_map = {
        field.field_code: field
        for field in template_fields
    }

    value_map = {
        value.field_code: value.field_value
        for value in entry.values
    }

    for field in template_fields:
        if field.is_required == "Yes" and field.input_mode == "Manual":
            if field.field_code not in value_map:
                raise HTTPException(
                    status_code=400,
                    detail=f"Required field missing: {field.field_name}",
                )

            value = value_map.get(field.field_code)

            if value is None or str(value).strip() == "":
                raise HTTPException(
                    status_code=400,
                    detail=f"Required field cannot be blank: {field.field_name}",
                )

    for value in entry.values:
        if value.field_code not in field_map:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid field code: {value.field_code}",
            )

    return (
        template,
        operation_type,
        asset,
        template_fields,
        value_map,
        transaction_operation_type_code,
    )


def format_operation_date_for_ticket(operation_date):
    if operation_date is None:
        return datetime.now().strftime("%Y%m%d")

    if isinstance(operation_date, str):
        try:
            return datetime.fromisoformat(operation_date).strftime("%Y%m%d")
        except ValueError:
            return datetime.now().strftime("%Y%m%d")

    return operation_date.strftime("%Y%m%d")


def generate_operation_ticket_number(db, location_code, asset_code, operation_date):
    ticket_date = format_operation_date_for_ticket(operation_date)

    clean_location_code = str(location_code).strip().upper()
    clean_asset_code = str(asset_code).strip().upper()

    ticket_prefix = f"{clean_location_code}-{clean_asset_code}-{ticket_date}"

    existing_tickets = (
        db.query(OperationTransaction.operation_ticket_number)
        .filter(OperationTransaction.operation_ticket_number.like(f"{ticket_prefix}-%"))
        .all()
    )

    serial_numbers = []

    for row in existing_tickets:
        existing_ticket = row[0]

        if not existing_ticket:
            continue

        try:
            serial_numbers.append(int(str(existing_ticket).split("-")[-1]))
        except ValueError:
            continue

    next_serial_number = max(serial_numbers) + 1 if serial_numbers else 1

    return f"{ticket_prefix}-{next_serial_number:03d}"


def normalize_jsonb_value(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    if hasattr(value, "isoformat"):
        return value.isoformat()

    if isinstance(value, list):
        return [
            normalize_jsonb_value(item)
            for item in value
        ]

    if isinstance(value, dict):
        return {
            str(key): normalize_jsonb_value(item_value)
            for key, item_value in value.items()
        }

    return value


@app.get(
    "/operation-entries",
    response_model=list[OperationEntryResponse],
)
def get_operation_entries(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    transactions = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.operation_template_id.isnot(None))
        .filter(OperationTransaction.status.in_(["Draft", "Rejected"]))
        .order_by(OperationTransaction.id.desc())
        .all()
    )

    return [
        build_operation_entry_response(transaction, db)
        for transaction in transactions
    ]


@app.post(
    "/operation-entries",
    response_model=OperationEntryResponse,
)
def create_operation_entry(
    entry: OperationEntryCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    (
        template,
        operation_type,
        asset,
        template_fields,
        value_map,
        transaction_operation_type_code,
    ) = validate_operation_entry(entry, db)

    # ✅ Trip lock ONLY for barges (avoid blocking shuttle/tanker convoy numbers)
    trip = None
    if str(operation_type.applicable_asset_type_code or "").strip().upper() == "BARGE":
        trip = get_trip_by_convoy_or_none(db, entry.transaction.convoy_number)
        ensure_trip_not_closed(trip)

    # ✅ Shuttle lock ONLY for Shuttle Tracking templates
    if str(template.entry_layout_type or "").strip() == "Shuttle Tracking":
        voyage = get_or_create_shuttle_voyage(
            db=db,
            location_code=entry.transaction.origin_location_code,
            shuttle_number=entry.transaction.convoy_number or "",
            shuttle_asset_code=asset.asset_code,
            current_user=current_user,
        )
        ensure_shuttle_voyage_not_closed(voyage)

    ticket_number = generate_operation_ticket_number(
        db=db,
        location_code=entry.transaction.origin_location_code,
        asset_code=asset.asset_code,
        operation_date=entry.transaction.operation_date,
    )

    new_transaction = OperationTransaction(
        operation_number=generate_operation_number(db),
        operation_ticket_number=ticket_number,
        operation_type_code=transaction_operation_type_code,
        operation_template_id=template.id,
        primary_asset_code=asset.asset_code,
        primary_asset_type_code=asset.asset_type_code,
        convoy_number=clean_optional_text(entry.transaction.convoy_number),
        origin_location_code=entry.transaction.origin_location_code.strip(),
        destination_location_code=clean_optional_text(
            entry.transaction.destination_location_code
        ),
        sender_location_code=clean_optional_text(entry.transaction.sender_location_code),
        receiver_location_code=clean_optional_text(entry.transaction.receiver_location_code),
        operation_date=entry.transaction.operation_date,
        operation_start_datetime=entry.transaction.operation_start_datetime,
        operation_end_datetime=entry.transaction.operation_end_datetime,
        product_name=clean_optional_text(entry.transaction.product_name),
        created_by=(
            f"{current_user.full_name} ({current_user.username})"
            if current_user.full_name
            else current_user.username
        ),
        remarks=clean_optional_text(entry.transaction.remarks),
        status=entry.transaction.status or "Draft",
    )

    db.add(new_transaction)
    db.flush()

    for field in template_fields:
        new_value = OperationTransactionValue(
            transaction_id=new_transaction.id,
            field_code=field.field_code,
            field_name=field.field_name,
            field_group=field.field_group,
            data_type=field.data_type,
            unit=field.unit,
            input_mode=field.input_mode,
            calculation_role=field.calculation_role,
            field_value=normalize_jsonb_value(value_map.get(field.field_code)),
            sort_order=field.sort_order,
        )

        db.add(new_value)

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Create Operation Entry",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=new_transaction.id,
        entity_label=ticket_number,
        ticket_number=ticket_number,
        operation_number=new_transaction.operation_number,
        new_status=new_transaction.status,
        remarks="Operation entry created",
        request_path="/operation-entries",
        details={
            "operation_type_code": new_transaction.operation_type_code,
            "operation_template_id": new_transaction.operation_template_id,
            "primary_asset_code": new_transaction.primary_asset_code,
            "origin_location_code": new_transaction.origin_location_code,
            "operation_date": str(new_transaction.operation_date),
        },
    )

    db.commit()
    db.refresh(new_transaction)

    return build_operation_entry_response(new_transaction, db)


@app.post("/operation-transactions/backfill-ticket-numbers")
def backfill_operation_transaction_ticket_numbers(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Operation Template",
        db,
    )

    transactions = (
        db.query(OperationTransaction)
        .filter(
            (OperationTransaction.operation_ticket_number == None)
            | (OperationTransaction.operation_ticket_number == "")
        )
        .order_by(
            OperationTransaction.operation_date.asc(),
            OperationTransaction.origin_location_code.asc(),
            OperationTransaction.primary_asset_code.asc(),
            OperationTransaction.id.asc(),
        )
        .all()
    )

    total_candidates = len(transactions)
    updated_count = 0
    skipped_count = 0
    examples = []

    for transaction in transactions:
        if not transaction.origin_location_code or not transaction.primary_asset_code:
            skipped_count += 1
            continue

        old_ticket = transaction.operation_ticket_number

        ticket_number = generate_operation_ticket_number(
            db=db,
            location_code=transaction.origin_location_code,
            asset_code=transaction.primary_asset_code,
            operation_date=transaction.operation_date,
        )

        transaction.operation_ticket_number = ticket_number
        updated_count += 1

        if len(examples) < 10:
            examples.append(
                {
                    "transaction_id": transaction.id,
                    "operation_number": transaction.operation_number,
                    "old_ticket_number": old_ticket,
                    "new_ticket_number": ticket_number,
                    "origin_location_code": transaction.origin_location_code,
                    "primary_asset_code": transaction.primary_asset_code,
                    "operation_date": str(transaction.operation_date),
                }
            )

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Backfill Ticket Numbers",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=None,
        entity_label="Backfill Ticket Numbers",
        remarks="Backfilled missing operation ticket numbers",
        request_path="/operation-transactions/backfill-ticket-numbers",
        details={
            "total_candidates": total_candidates,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "examples": examples,
        },
    )

    db.commit()

    return {
        "message": "Backfill completed",
        "total_candidates": total_candidates,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
    }


@app.get("/operation-transactions/{transaction_id}")
def get_operation_transaction_detail(
    transaction_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == transaction_id)
        .first()
    )

    if transaction is None:
        raise HTTPException(status_code=404, detail="Operation transaction not found")

    operation_type = get_operation_type_by_code(transaction.operation_type_code, db)
    location = get_location_by_code(transaction.origin_location_code, db)
    primary_asset = get_asset_by_code(transaction.primary_asset_code, db)

    values = (
        db.query(OperationTransactionValue)
        .filter(OperationTransactionValue.transaction_id == transaction.id)
        .order_by(OperationTransactionValue.sort_order.asc(), OperationTransactionValue.id.asc())
        .all()
    )

    field_values = [
        {
            "id": value.id,
            "field_code": value.field_code,
            "field_name": value.field_name,
            "field_group": value.field_group,
            "data_type": value.data_type,
            "unit": value.unit,
            "input_mode": value.input_mode,
            "calculation_role": value.calculation_role,
            "field_value": value.field_value,
            "sort_order": value.sort_order,
        }
        for value in values
    ]

    return {
        "id": transaction.id,
        "operation_number": transaction.operation_number,
        "operation_ticket_number": get_transaction_ticket_number(transaction),
        "ticket_number": get_transaction_ticket_number(transaction),
        "operation_date": transaction.operation_date,
        "operation_type_code": transaction.operation_type_code,
        "operation_type_name": operation_type.operation_type_name if operation_type else "",
        "location_name": location.location_name if location else "",
        "location_code": transaction.origin_location_code,
        "primary_asset_name": primary_asset.asset_name if primary_asset else "",
        "primary_asset_code": transaction.primary_asset_code,
        "convoy_number": transaction.convoy_number,
        "status": transaction.status,
        "created_by": transaction.created_by,
        "created_at": transaction.created_at,
        "updated_at": transaction.updated_at,
        "field_values": field_values,
    }

# -------------------------
# Tanker Transaction Report APIs
# -------------------------

def parse_json_field_value(value):
    if value is None:
        return None

    if isinstance(value, dict):
        return value

    try:
        import json
        return json.loads(str(value))
    except Exception:
        return None


def get_nested_value(data: dict, path: list[str], default=None):
    current = data

    for key in path:
        if not isinstance(current, dict):
            return default

        current = current.get(key)

        if current is None:
            return default

    return current


def get_tanker_payload_for_transaction(
    db: Session,
    transaction_id: int,
):
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "tanker_payload",
        )
        .first()
    )

    if not payload_row:
        return None

    return parse_json_field_value(payload_row.field_value)


def build_tanker_transaction_report_row(
    transaction: OperationTransaction,
    tanker_payload: dict,
    db: Session,
):
    operation_type = get_operation_type_by_code(
        transaction.operation_type_code,
        db,
    )

    location = get_location_by_code(
        transaction.origin_location_code,
        db,
    )

    asset = get_asset_by_code(
        transaction.primary_asset_code,
        db,
    )

    inputs = tanker_payload.get("inputs") or {}
    calculated = tanker_payload.get("calculated") or {}

    def number_from_input(key: str):
        return safe_float(inputs.get(key))

    def number_from_calculated(*keys: str):
        for key in keys:
            value = calculated.get(key)
            if value is not None:
                return safe_float(value)
        return 0

    return {
        "transaction_id": transaction.id,
        "operation_number": transaction.operation_number,
        "ticket_number": get_transaction_ticket_number(transaction),

        "operation_date": transaction.operation_date,
        "operation_start_datetime": transaction.operation_start_datetime,
        "operation_end_datetime": transaction.operation_end_datetime,

        "operation_type_code": transaction.operation_type_code,
        "operation_type_name": operation_type.operation_type_name if operation_type else "",

        "location_code": transaction.origin_location_code,
        "location_name": location.location_name if location else "",

        "asset_code": transaction.primary_asset_code,
        "asset_name": asset.asset_name if asset else "",
        "asset_type_code": transaction.primary_asset_type_code,

        "convoy_number": transaction.convoy_number,
        "tanker_name": inputs.get("tankerName") or (asset.asset_name if asset else ""),
        "prime_mover_number": inputs.get("primeMoverNumber"),
        "chassis_number": inputs.get("chassisNumber"),

        "cargo": inputs.get("cargo") or transaction.product_name,
        "tanker_operation": inputs.get("operation"),
        "destination": inputs.get("destination"),
        "loading_bay": inputs.get("loadingBay"),
        "compartment": inputs.get("compartment"),

        "total_dip_cm": number_from_input("totalDipCm"),
        "water_dip_cm": number_from_input("waterDipCm"),
        "bsw_percent": number_from_input("bswPercent"),

        "tank_temperature": (
            safe_float(inputs.get("tankTemperature"))
            if inputs.get("tankTemperature") is not None
            else None
        ),
        "tank_temperature_unit": inputs.get("tankTemperatureUnit"),
        "sample_temperature": (
            safe_float(inputs.get("sampleTemperature"))
            if inputs.get("sampleTemperature") is not None
            else None
        ),
        "sample_temperature_unit": inputs.get("sampleTemperatureUnit"),

        "observed_input_type": inputs.get("observedInputType"),
        "observed_api": (
            safe_float(inputs.get("observedApi"))
            if inputs.get("observedApi") is not None
            else calculated.get("observedApi")
        ),
        "observed_density": (
            safe_float(inputs.get("observedDensity"))
            if inputs.get("observedDensity") is not None
            else calculated.get("observedDensity")
        ),
        "api60": calculated.get("api60"),
        "vcf": calculated.get("vcf"),

        "tov_bbl": number_from_calculated("tovBbl", "totalVolumeBbl", "total_volume_bbl"),
        "free_water_bbl": number_from_calculated("freeWaterBbl", "waterVolumeBbl", "water_volume_bbl"),
        "gov_bbl": number_from_calculated("govBbl", "gov_bbl"),
        "gsv_bbl": number_from_calculated("gsvBbl", "gsv_bbl"),
        "bsw_bbl": number_from_calculated("bswBbl", "bsw_vol_bbl", "bswVolumeBbl"),
        "nsv_bbl": number_from_calculated("nsvBbl", "nsv_bbl"),

        "lt_factor": calculated.get("ltFactor"),
        "lt": number_from_calculated("lt"),
        "mt": number_from_calculated("mt"),

        "seal_c1": inputs.get("sealC1"),
        "seal_c2": inputs.get("sealC2"),
        "seal_m1": inputs.get("sealM1"),
        "seal_m2": inputs.get("sealM2"),

        "remarks": inputs.get("remarks") or transaction.remarks,
        "status": transaction.status,
        "created_by": transaction.created_by,
        "created_at": transaction.created_at,
        "updated_at": transaction.updated_at,
    }


def get_filtered_tanker_transaction_report_rows(
    db: Session,
    date_from: str | None = None,
    date_to: str | None = None,
    location_code: str | None = None,
    asset_code: str | None = None,
    convoy_number: str | None = None,
    status: str | None = None,
    search: str | None = None,
):
    query = db.query(OperationTransaction).join(
        OperationTransactionValue,
        OperationTransactionValue.transaction_id == OperationTransaction.id,
    ).filter(
        OperationTransactionValue.field_code == "tanker_payload",
        OperationTransactionValue.field_value != None,
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value:
        query = query.filter(OperationTransaction.operation_date >= date_from_value)

    if date_to_value:
        query = query.filter(OperationTransaction.operation_date <= date_to_value)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_asset_code = clean_optional_text(asset_code)
    cleaned_convoy_number = clean_optional_text(convoy_number)
    cleaned_status = clean_optional_text(status)

    if cleaned_location_code:
        query = query.filter(
            OperationTransaction.origin_location_code.ilike(cleaned_location_code)
        )

    if cleaned_asset_code:
        query = query.filter(
            OperationTransaction.primary_asset_code.ilike(cleaned_asset_code)
        )

    if cleaned_convoy_number:
        query = query.filter(
            OperationTransaction.convoy_number.ilike(cleaned_convoy_number)
        )

    if cleaned_status:
        query = query.filter(OperationTransaction.status == cleaned_status)

    transactions = (
        query.order_by(
            OperationTransaction.operation_date.desc(),
            OperationTransaction.id.desc(),
        )
        .all()
    )

    rows = []

    for transaction in transactions:
        tanker_payload = get_tanker_payload_for_transaction(db, transaction.id)

        if not tanker_payload:
            continue

        row = build_tanker_transaction_report_row(
            transaction=transaction,
            tanker_payload=tanker_payload,
            db=db,
        )

        cleaned_search = clean_optional_text(search)

        if cleaned_search:
            search_value = cleaned_search.lower()

            searchable_text = " ".join(
                [
                    str(row.get("ticket_number") or ""),
                    str(row.get("operation_number") or ""),
                    str(row.get("operation_type_code") or ""),
                    str(row.get("operation_type_name") or ""),
                    str(row.get("location_code") or ""),
                    str(row.get("location_name") or ""),
                    str(row.get("asset_code") or ""),
                    str(row.get("asset_name") or ""),
                    str(row.get("convoy_number") or ""),
                    str(row.get("tanker_name") or ""),
                    str(row.get("prime_mover_number") or ""),
                    str(row.get("chassis_number") or ""),
                    str(row.get("destination") or ""),
                    str(row.get("cargo") or ""),
                    str(row.get("status") or ""),
                ]
            ).lower()

            if search_value not in searchable_text:
                continue

        rows.append(row)

    return rows


def build_tanker_transaction_report_totals(rows: list[dict]):
    return {
        "rows_count": len(rows),
        "total_tov_bbl": round(sum(safe_float(row.get("tov_bbl")) for row in rows), 3),
        "total_free_water_bbl": round(sum(safe_float(row.get("free_water_bbl")) for row in rows), 3),
        "total_gov_bbl": round(sum(safe_float(row.get("gov_bbl")) for row in rows), 3),
        "total_gsv_bbl": round(sum(safe_float(row.get("gsv_bbl")) for row in rows), 3),
        "total_bsw_bbl": round(sum(safe_float(row.get("bsw_bbl")) for row in rows), 3),
        "total_nsv_bbl": round(sum(safe_float(row.get("nsv_bbl")) for row in rows), 3),
        "total_lt": round(sum(safe_float(row.get("lt")) for row in rows), 3),
        "total_mt": round(sum(safe_float(row.get("mt")) for row in rows), 3),
    }


@app.get(
    "/tanker-transaction-report",
    response_model=TankerTransactionReportResponse,
)
def get_tanker_transaction_report(
    date_from: str | None = None,
    date_to: str | None = None,
    location_code: str | None = None,
    asset_code: str | None = None,
    convoy_number: str | None = None,
    status: str | None = None,
    search: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    rows = get_filtered_tanker_transaction_report_rows(
        db=db,
        date_from=date_from,
        date_to=date_to,
        location_code=location_code,
        asset_code=asset_code,
        convoy_number=convoy_number,
        status=status,
        search=search,
    )

    return {
        "rows": rows,
        "totals": build_tanker_transaction_report_totals(rows),
    }

# -------------------------
# Tanker Tracking APIs
# -------------------------

def get_payload_asset_value(payload: dict, section_names: list[str], keys: list[str]):
    for section_name in section_names:
        section = payload.get(section_name)

        if not isinstance(section, dict):
            continue

        for key in keys:
            value = section.get(key)
            if value is not None and str(value).strip() != "":
                return value

    return None


def detect_tanker_movement_role(transaction: OperationTransaction, operation_type: OperationType | None):
    text = " ".join(
        [
            str(transaction.operation_type_code or ""),
            str(operation_type.operation_type_name if operation_type else ""),
        ]
    ).upper()

    receiver_keywords = [
        "RECEIPT",
        "RECEIVE",
        "RECEIVED",
        "UNLOAD",
        "UNLOADING",
        "DISCHARGE",
        "DESTINATION",
    ]

    sender_keywords = [
        "LOAD",
        "LOADING",
        "DISPATCH",
        "SEND",
        "SENDER",
        "SOURCE",
    ]

    if any(keyword in text for keyword in receiver_keywords):
        return "RECEIVER"

    if any(keyword in text for keyword in sender_keywords):
        return "SENDER"

    return "UNKNOWN"


def build_tanker_tracking_ticket(
    transaction: OperationTransaction,
    tanker_payload: dict,
    db: Session,
):
    operation_type = get_operation_type_by_code(
        transaction.operation_type_code,
        db,
    )

    primary_asset = get_asset_by_code(
        transaction.primary_asset_code,
        db,
    )

    origin_location = get_location_by_code(
        transaction.origin_location_code,
        db,
    )

    destination_location = (
        get_location_by_code(transaction.destination_location_code, db)
        if transaction.destination_location_code
        else None
    )

    sender_location = (
        get_location_by_code(transaction.sender_location_code, db)
        if transaction.sender_location_code
        else None
    )

    receiver_location = (
        get_location_by_code(transaction.receiver_location_code, db)
        if transaction.receiver_location_code
        else None
    )

    inputs = tanker_payload.get("inputs") or {}
    calculated = tanker_payload.get("calculated") or {}

    prime_mover_asset_code = get_payload_asset_value(
        tanker_payload,
        ["prime_mover_asset"],
        ["asset_code", "assetCode", "prime_mover_asset_code"],
    )

    prime_mover_asset_name = get_payload_asset_value(
        tanker_payload,
        ["prime_mover_asset"],
        ["asset_name", "assetName", "prime_mover_asset_name"],
    )

    tanker_asset_code = get_payload_asset_value(
        tanker_payload,
        ["tanker_trailer_asset", "linked_tanker_asset", "tanker_asset", "asset"],
        ["asset_code", "assetCode", "tanker_asset_code"],
    )

    tanker_asset_name = get_payload_asset_value(
        tanker_payload,
        ["tanker_trailer_asset", "linked_tanker_asset", "tanker_asset", "asset"],
        ["asset_name", "assetName", "tanker_asset_name"],
    )

    tanker_chassis_number = get_payload_asset_value(
        tanker_payload,
        ["tanker_trailer_asset", "linked_tanker_asset", "tanker_asset", "asset"],
        ["serial_number", "serialNumber", "tanker_chassis_number"],
    )

    # Backward compatibility:
    # If tanker payload was saved before prime_mover/tanker separation,
    # use primary asset as fallback.
    if tanker_asset_code is None:
        tanker_asset_code = transaction.primary_asset_code

    if tanker_asset_name is None:
        tanker_asset_name = primary_asset.asset_name if primary_asset else ""

    if tanker_chassis_number is None:
        tanker_chassis_number = primary_asset.serial_number if primary_asset else ""

    if prime_mover_asset_code is None:
        prime_mover_asset_code = inputs.get("primeMoverNumber")

    movement_role = detect_tanker_movement_role(transaction, operation_type)

    def number_from_input(key: str):
        return safe_float(inputs.get(key))

    def number_from_calculated(*keys: str):
        for key in keys:
            value = calculated.get(key)
            if value is not None:
                return safe_float(value)
        return 0

    return {
        "transaction_id": transaction.id,
        "ticket_number": get_transaction_ticket_number(transaction),
        "operation_number": transaction.operation_number,

        "movement_role": movement_role,

        "operation_date": transaction.operation_date,
        "operation_type_code": transaction.operation_type_code,
        "operation_type_name": operation_type.operation_type_name if operation_type else "",

        "origin_location_code": transaction.origin_location_code,
        "origin_location_name": origin_location.location_name if origin_location else "",
        "destination_location_code": transaction.destination_location_code,
        "destination_location_name": destination_location.location_name if destination_location else "",
        "sender_location_code": transaction.sender_location_code,
        "sender_location_name": sender_location.location_name if sender_location else "",
        "receiver_location_code": transaction.receiver_location_code,
        "receiver_location_name": receiver_location.location_name if receiver_location else "",

        "primary_asset_code": transaction.primary_asset_code,
        "primary_asset_name": primary_asset.asset_name if primary_asset else "",
        "primary_asset_type_code": transaction.primary_asset_type_code,

        "prime_mover_asset_code": prime_mover_asset_code,
        "prime_mover_asset_name": prime_mover_asset_name,

        "tanker_asset_code": tanker_asset_code,
        "tanker_asset_name": tanker_asset_name,
        "tanker_chassis_number": tanker_chassis_number,

        "convoy_number": transaction.convoy_number,
        "product_name": transaction.product_name,

        "compartment": inputs.get("compartment"),
        "total_dip_cm": number_from_input("totalDipCm"),
        "water_dip_cm": number_from_input("waterDipCm"),
        "bsw_percent": number_from_input("bswPercent"),

        "tank_temperature": (
            safe_float(inputs.get("tankTemperature"))
            if inputs.get("tankTemperature") is not None
            else None
        ),
        "tank_temperature_unit": inputs.get("tankTemperatureUnit"),
        "sample_temperature": (
            safe_float(inputs.get("sampleTemperature"))
            if inputs.get("sampleTemperature") is not None
            else None
        ),
        "sample_temperature_unit": inputs.get("sampleTemperatureUnit"),

        "observed_input_type": inputs.get("observedInputType"),
        "observed_api": (
            safe_float(inputs.get("observedApi"))
            if inputs.get("observedApi") is not None
            else calculated.get("observedApi")
        ),
        "observed_density": (
            safe_float(inputs.get("observedDensity"))
            if inputs.get("observedDensity") is not None
            else calculated.get("observedDensity")
        ),
        "api60": calculated.get("api60"),
        "vcf": calculated.get("vcf"),

        "tov_bbl": number_from_calculated("tovBbl", "totalVolumeBbl", "total_volume_bbl"),
        "free_water_bbl": number_from_calculated("freeWaterBbl", "waterVolumeBbl", "water_volume_bbl"),
        "gov_bbl": number_from_calculated("govBbl", "gov_bbl"),
        "gsv_bbl": number_from_calculated("gsvBbl", "gsv_bbl"),
        "bsw_bbl": number_from_calculated("bswBbl", "bsw_vol_bbl", "bswVolumeBbl"),
        "nsv_bbl": number_from_calculated("nsvBbl", "nsv_bbl"),
        "lt": number_from_calculated("lt"),
        "mt": number_from_calculated("mt"),

        "seal_c1": inputs.get("sealC1"),
        "seal_c2": inputs.get("sealC2"),
        "seal_m1": inputs.get("sealM1"),
        "seal_m2": inputs.get("sealM2"),

        "remarks": inputs.get("remarks") or transaction.remarks,
        "status": transaction.status,
        "created_by": transaction.created_by,
        "created_at": transaction.created_at,
        "updated_at": transaction.updated_at,
    }


def build_tanker_seal_checks(sender_ticket: dict | None, receiver_ticket: dict | None):
    seal_fields = [
        ("C1", "seal_c1"),
        ("C2", "seal_c2"),
        ("M1", "seal_m1"),
        ("M2", "seal_m2"),
    ]

    checks = []

    for seal_name, field_name in seal_fields:
        sender_value = None
        receiver_value = None

        if sender_ticket:
            sender_value = clean_optional_text(sender_ticket.get(field_name))

        if receiver_ticket:
            receiver_value = clean_optional_text(receiver_ticket.get(field_name))

        if not sender_value and not receiver_value:
            status = "NOT_ENTERED"
        elif sender_value and not receiver_value:
            status = "RECEIVER_MISSING"
        elif not sender_value and receiver_value:
            status = "SENDER_MISSING"
        elif str(sender_value).strip().upper() == str(receiver_value).strip().upper():
            status = "MATCHED"
        else:
            status = "MISMATCH"

        checks.append(
            {
                "seal_name": seal_name,
                "sender_value": sender_value,
                "receiver_value": receiver_value,
                "status": status,
            }
        )

    return checks


def build_tanker_quantity_comparison(sender_ticket: dict | None, receiver_ticket: dict | None):
    if not sender_ticket or not receiver_ticket:
        return None

    sender_gov = safe_float(sender_ticket.get("gov_bbl"))
    receiver_gov = safe_float(receiver_ticket.get("gov_bbl"))

    sender_gsv = safe_float(sender_ticket.get("gsv_bbl"))
    receiver_gsv = safe_float(receiver_ticket.get("gsv_bbl"))

    sender_nsv = safe_float(sender_ticket.get("nsv_bbl"))
    receiver_nsv = safe_float(receiver_ticket.get("nsv_bbl"))

    sender_lt = safe_float(sender_ticket.get("lt"))
    receiver_lt = safe_float(receiver_ticket.get("lt"))

    sender_mt = safe_float(sender_ticket.get("mt"))
    receiver_mt = safe_float(receiver_ticket.get("mt"))

    nsv_variance = receiver_nsv - sender_nsv

    if sender_nsv != 0:
        nsv_variance_percent = (nsv_variance / sender_nsv) * 100
    else:
        nsv_variance_percent = 0

    return {
        "sender_transaction_id": sender_ticket.get("transaction_id"),
        "receiver_transaction_id": receiver_ticket.get("transaction_id"),

        "sender_gov_bbl": round(sender_gov, 3),
        "receiver_gov_bbl": round(receiver_gov, 3),
        "gov_variance_bbl": round(receiver_gov - sender_gov, 3),

        "sender_gsv_bbl": round(sender_gsv, 3),
        "receiver_gsv_bbl": round(receiver_gsv, 3),
        "gsv_variance_bbl": round(receiver_gsv - sender_gsv, 3),

        "sender_nsv_bbl": round(sender_nsv, 3),
        "receiver_nsv_bbl": round(receiver_nsv, 3),
        "nsv_variance_bbl": round(nsv_variance, 3),
        "nsv_variance_percent": round(nsv_variance_percent, 4),

        "sender_lt": round(sender_lt, 3),
        "receiver_lt": round(receiver_lt, 3),
        "lt_variance": round(receiver_lt - sender_lt, 3),

        "sender_mt": round(sender_mt, 3),
        "receiver_mt": round(receiver_mt, 3),
        "mt_variance": round(receiver_mt - sender_mt, 3),
    }


def get_current_user_label(current_user: User):
    full_name = str(current_user.full_name or "").strip()
    username = str(current_user.username or "").strip()

    if full_name and username:
        return f"{full_name} ({username})"

    if full_name:
        return full_name

    return username or None


def get_tanker_acknowledgement_by_sender(
    db: Session,
    sender_transaction_id: int | None,
):
    if sender_transaction_id is None:
        return None

    return (
        db.query(TankerReceiptAcknowledgement)
        .filter(
            TankerReceiptAcknowledgement.sender_transaction_id
            == sender_transaction_id,
            TankerReceiptAcknowledgement.status.in_(["Acknowledged", "Closed"]),
        )
        .first()
    )


def build_tanker_acknowledgement_response(
    acknowledgement: TankerReceiptAcknowledgement,
    db: Session,
):
    tanker_asset = None

    if acknowledgement.tanker_asset_code:
        tanker_asset = get_asset_by_code(
            acknowledgement.tanker_asset_code,
            db,
        )

    prime_mover_asset = None

    if acknowledgement.prime_mover_asset_code:
        prime_mover_asset = get_asset_by_code(
            acknowledgement.prime_mover_asset_code,
            db,
        )

    receiver_location = None

    if acknowledgement.receiver_location_code:
        receiver_location = get_location_by_code(
            acknowledgement.receiver_location_code,
            db,
        )

    return {
        "id": acknowledgement.id,
        "sender_transaction_id": acknowledgement.sender_transaction_id,
        "convoy_number": acknowledgement.convoy_number,
        "tanker_asset_code": acknowledgement.tanker_asset_code,
        "tanker_asset_name": tanker_asset.asset_name if tanker_asset else "",
        "tanker_chassis_number": tanker_asset.serial_number
        if tanker_asset
        else "",
        "prime_mover_asset_code": acknowledgement.prime_mover_asset_code,
        "prime_mover_asset_name": prime_mover_asset.asset_name
        if prime_mover_asset
        else "",
        "receiver_location_code": acknowledgement.receiver_location_code,
        "receiver_location_name": receiver_location.location_name
        if receiver_location
        else "",
        "acknowledged_by": acknowledgement.acknowledged_by,
        "acknowledged_at": acknowledgement.acknowledged_at,
        "remarks": acknowledgement.remarks,
        "status": acknowledgement.status,
        "closed_by": acknowledgement.closed_by,
        "closed_at": acknowledgement.closed_at,
        "closure_remarks": acknowledgement.closure_remarks,
        "created_at": acknowledgement.created_at,
        "updated_at": acknowledgement.updated_at,
    }


def build_tanker_acknowledgement_audit_snapshot(
    acknowledgement: TankerReceiptAcknowledgement,
    db: Session,
):
    return build_tanker_acknowledgement_response(acknowledgement, db)


def get_tanker_tracking_group_status(
    sender_ticket: dict | None,
    receiver_tickets: list[dict],
    seal_checks: list[dict],
    quantity_comparison: dict | None,
    acknowledgement: TankerReceiptAcknowledgement | None = None,
):
    if acknowledgement and acknowledgement.status == "Closed":
        return "CLOSED"

    if not sender_ticket:
        return "NO_SENDER"

    if len(receiver_tickets) == 0:
        if acknowledgement and acknowledgement.status == "Acknowledged":
            return "ACKNOWLEDGED"

        return "PENDING_RECEIPT"

    seal_mismatch = any(
        check.get("status") in ["MISMATCH", "RECEIVER_MISSING", "SENDER_MISSING"]
        for check in seal_checks
    )

    if seal_mismatch:
        return "SEAL_MISMATCH"

    if quantity_comparison:
        nsv_variance = abs(safe_float(quantity_comparison.get("nsv_variance_bbl")))

        if nsv_variance > 0:
            return "QUANTITY_VARIANCE"

        return "MATCHED"

    return "RECEIVED"


def build_tanker_tracking_groups(tickets: list[dict], db: Session):
    grouped = {}

    for ticket in tickets:
        convoy = clean_optional_text(ticket.get("convoy_number"))

        if convoy is None:
            continue

        tanker_asset_code = clean_optional_text(ticket.get("tanker_asset_code")) or "UNKNOWN_TANKER"

        group_key = f"{convoy}::{tanker_asset_code}"

        if group_key not in grouped:
            grouped[group_key] = {
                "group_key": group_key,
                "convoy_number": convoy,
                "tanker_asset_code": ticket.get("tanker_asset_code"),
                "tanker_asset_name": ticket.get("tanker_asset_name"),
                "tanker_chassis_number": ticket.get("tanker_chassis_number"),
                "prime_mover_asset_code": ticket.get("prime_mover_asset_code"),
                "prime_mover_asset_name": ticket.get("prime_mover_asset_name"),
                "product_name": ticket.get("product_name"),
                "tickets": [],
            }

        grouped[group_key]["tickets"].append(ticket)

    current_tracking_db = db

    tracking_rows = []

    for group in grouped.values():
        sorted_tickets = sorted(
            group["tickets"],
            key=lambda item: (
                item.get("operation_date") or date.min,
                item.get("transaction_id") or 0,
            ),
        )

        sender_tickets = [
            ticket
            for ticket in sorted_tickets
            if ticket.get("movement_role") == "SENDER"
        ]

        receiver_tickets = [
            ticket
            for ticket in sorted_tickets
            if ticket.get("movement_role") == "RECEIVER"
        ]

        unknown_tickets = [
            ticket
            for ticket in sorted_tickets
            if ticket.get("movement_role") == "UNKNOWN"
        ]

        warning_messages = []

        if len(sender_tickets) == 0 and len(unknown_tickets) > 0:
            # If operation names are not clear, use the first unknown ticket as sender fallback.
            sender_ticket = unknown_tickets[0]
            warning_messages.append(
                "Sender/receiver role could not be detected from operation type. First unknown ticket is treated as sender."
            )
        elif len(sender_tickets) > 0:
            sender_ticket = sender_tickets[0]
        else:
            sender_ticket = None

        if len(sender_tickets) > 1:
            warning_messages.append(
                "Multiple sender tickets found for this convoy/tanker. First sender ticket is used for comparison."
            )

        latest_receiver_ticket = receiver_tickets[-1] if receiver_tickets else None

        seal_checks = build_tanker_seal_checks(
            sender_ticket,
            latest_receiver_ticket,
        )

        quantity_comparison = build_tanker_quantity_comparison(
            sender_ticket,
            latest_receiver_ticket,
        )

        acknowledgement = None

        if sender_ticket:
            acknowledgement = get_tanker_acknowledgement_by_sender(
                db=current_tracking_db,
                sender_transaction_id=sender_ticket.get("transaction_id"),
            )

        tracking_status = get_tanker_tracking_group_status(
            sender_ticket,
            receiver_tickets,
            seal_checks,
            quantity_comparison,
            acknowledgement,
        )

        tracking_rows.append(
            {
                "group_key": group["group_key"],
                "convoy_number": group["convoy_number"],

                "tanker_asset_code": group["tanker_asset_code"],
                "tanker_asset_name": group["tanker_asset_name"],
                "tanker_chassis_number": group["tanker_chassis_number"],

                "prime_mover_asset_code": group["prime_mover_asset_code"],
                "prime_mover_asset_name": group["prime_mover_asset_name"],

                "product_name": group["product_name"],

                "sender_ticket": sender_ticket,
                "receiver_tickets": receiver_tickets,
                "latest_receiver_ticket": latest_receiver_ticket,

                "seal_checks": seal_checks,
                "quantity_comparison": quantity_comparison,

                "acknowledgement_id": acknowledgement.id
                if acknowledgement
                else None,
                "acknowledged_by": acknowledgement.acknowledged_by
                if acknowledgement
                else None,
                "acknowledged_at": acknowledgement.acknowledged_at
                if acknowledgement
                else None,
                "acknowledgement_remarks": acknowledgement.remarks
                if acknowledgement
                else None,
                "closed_by": acknowledgement.closed_by if acknowledgement else None,
                "closed_at": acknowledgement.closed_at if acknowledgement else None,
                "closure_remarks": acknowledgement.closure_remarks if acknowledgement else None,

                "tracking_status": tracking_status,
                "warning_messages": warning_messages,
            }
        )

    return sorted(
        tracking_rows,
        key=lambda item: (
            item.get("convoy_number") or "",
            item.get("tanker_asset_code") or "",
        ),
    )


def get_tanker_tracking_rows(
    db: Session,
    date_from: str | None = None,
    date_to: str | None = None,
    convoy_number: str | None = None,
    location_code: str | None = None,
    tanker_asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
):
    query = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "tanker_payload",
            OperationTransactionValue.field_value != None,
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
    )

    date_from_value = parse_date_filter(date_from, "Date From")
    date_to_value = parse_date_filter(date_to, "Date To")

    if date_from_value:
        query = query.filter(OperationTransaction.operation_date >= date_from_value)

    if date_to_value:
        query = query.filter(OperationTransaction.operation_date <= date_to_value)

    cleaned_convoy_number = clean_optional_text(convoy_number)
    cleaned_location_code = clean_optional_text(location_code)

    if cleaned_convoy_number:
        query = query.filter(OperationTransaction.convoy_number.ilike(cleaned_convoy_number))

    if cleaned_location_code:
        query = query.filter(
            (
                OperationTransaction.origin_location_code.ilike(cleaned_location_code)
            )
            | (
                OperationTransaction.destination_location_code.ilike(cleaned_location_code)
            )
            | (
                OperationTransaction.sender_location_code.ilike(cleaned_location_code)
            )
            | (
                OperationTransaction.receiver_location_code.ilike(cleaned_location_code)
            )
        )

    transactions = (
        query.order_by(
            OperationTransaction.convoy_number.asc(),
            OperationTransaction.operation_date.asc(),
            OperationTransaction.id.asc(),
        )
        .all()
    )

    tickets = []

    cleaned_tanker_asset_code = clean_optional_text(tanker_asset_code)
    cleaned_search = clean_optional_text(search)

    for transaction in transactions:
        tanker_payload = get_tanker_payload_for_transaction(db, transaction.id)

        if not tanker_payload:
            continue

        ticket = build_tanker_tracking_ticket(
            transaction=transaction,
            tanker_payload=tanker_payload,
            db=db,
        )

        if cleaned_tanker_asset_code:
            ticket_tanker_code = clean_optional_text(ticket.get("tanker_asset_code"))

            if not ticket_tanker_code or ticket_tanker_code.lower() != cleaned_tanker_asset_code.lower():
                continue

        if cleaned_search:
            searchable_text = " ".join(
                [
                    str(ticket.get("ticket_number") or ""),
                    str(ticket.get("operation_number") or ""),
                    str(ticket.get("convoy_number") or ""),
                    str(ticket.get("primary_asset_code") or ""),
                    str(ticket.get("primary_asset_name") or ""),
                    str(ticket.get("prime_mover_asset_code") or ""),
                    str(ticket.get("prime_mover_asset_name") or ""),
                    str(ticket.get("tanker_asset_code") or ""),
                    str(ticket.get("tanker_asset_name") or ""),
                    str(ticket.get("tanker_chassis_number") or ""),
                    str(ticket.get("origin_location_code") or ""),
                    str(ticket.get("destination_location_code") or ""),
                    str(ticket.get("product_name") or ""),
                    str(ticket.get("status") or ""),
                ]
            ).lower()

            if cleaned_search.lower() not in searchable_text:
                continue

        tickets.append(ticket)

    return build_tanker_tracking_groups(tickets, db)


def build_tanker_tracking_summary(rows: list[dict]):
    pending_receipts = 0
    received_groups = 0
    compared_groups = 0
    seal_mismatch_groups = 0
    quantity_variance_groups = 0

    for row in rows:
        status = row.get("tracking_status")

        if status == "PENDING_RECEIPT":
            pending_receipts += 1

        if status in [
            "ACKNOWLEDGED",
            "RECEIVED",
            "MATCHED",
            "SEAL_MISMATCH",
            "QUANTITY_VARIANCE",
        ]:
            received_groups += 1

        if row.get("quantity_comparison") is not None:
            compared_groups += 1

        if status == "SEAL_MISMATCH":
            seal_mismatch_groups += 1

        if status == "QUANTITY_VARIANCE":
            quantity_variance_groups += 1

    return {
        "total_groups": len(rows),
        "pending_receipts": pending_receipts,
        "received_groups": received_groups,
        "compared_groups": compared_groups,
        "seal_mismatch_groups": seal_mismatch_groups,
        "quantity_variance_groups": quantity_variance_groups,
    }


@app.get(
    "/tanker-tracking",
    response_model=TankerTrackingResponse,
)
def get_tanker_tracking(
    date_from: str | None = None,
    date_to: str | None = None,
    convoy_number: str | None = None,
    location_code: str | None = None,
    tanker_asset_code: str | None = None,
    status: str | None = None,
    search: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    rows = get_tanker_tracking_rows(
        db=db,
        date_from=date_from,
        date_to=date_to,
        convoy_number=convoy_number,
        location_code=location_code,
        tanker_asset_code=tanker_asset_code,
        status=status,
        search=search,
    )

    summary = build_tanker_tracking_summary(rows)

    return {
        "rows": rows,
        **summary,
    }


@app.get(
    "/tanker-tracking/by-convoy/{convoy_number}",
    response_model=TankerTrackingResponse,
)
def get_tanker_tracking_by_convoy(
    convoy_number: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    rows = get_tanker_tracking_rows(
        db=db,
        convoy_number=convoy_number,
    )

    summary = build_tanker_tracking_summary(rows)

    return {
        "rows": rows,
        **summary,
    }

@app.get(
    "/tanker-tracking/sender-reference/{sender_transaction_id}",
    response_model=TankerTrackingTicketResponse,
)
def get_tanker_sender_reference(
    sender_transaction_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    sender_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == sender_transaction_id)
        .first()
    )

    if not sender_transaction:
        raise HTTPException(
            status_code=404,
            detail="Sender tanker transaction not found",
        )

    if sender_transaction.status != APPROVED_TRANSACTION_STATUS:
        raise HTTPException(
            status_code=400,
            detail="Only Approved sender tanker transactions can be used as receiver reference",
        )

    tanker_payload = get_tanker_payload_for_transaction(
        db,
        sender_transaction.id,
    )

    if not tanker_payload:
        raise HTTPException(
            status_code=400,
            detail="Selected sender transaction does not have tanker payload",
        )

    sender_ticket = build_tanker_tracking_ticket(
        transaction=sender_transaction,
        tanker_payload=tanker_payload,
        db=db,
    )

    if sender_ticket.get("movement_role") != "SENDER":
        raise HTTPException(
            status_code=400,
            detail="Selected transaction is not detected as a sender tanker transaction",
        )

    return sender_ticket

@app.get(
    "/tanker-tracking/acknowledgements",
    response_model=list[TankerReceiptAcknowledgementResponse],
)
def get_tanker_receipt_acknowledgements(
    convoy_number: str | None = None,
    tanker_asset_code: str | None = None,
    receiver_location_code: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    query = db.query(TankerReceiptAcknowledgement)

    cleaned_convoy_number = clean_optional_text(convoy_number)

    if cleaned_convoy_number:
        query = query.filter(
            TankerReceiptAcknowledgement.convoy_number.ilike(
                cleaned_convoy_number
            )
        )

    cleaned_tanker_asset_code = clean_optional_text(tanker_asset_code)

    if cleaned_tanker_asset_code:
        query = query.filter(
            TankerReceiptAcknowledgement.tanker_asset_code.ilike(
                cleaned_tanker_asset_code
            )
        )

    cleaned_receiver_location_code = clean_optional_text(receiver_location_code)

    if cleaned_receiver_location_code:
        query = query.filter(
            TankerReceiptAcknowledgement.receiver_location_code.ilike(
                cleaned_receiver_location_code
            )
        )

    acknowledgements = (
        query.order_by(
            TankerReceiptAcknowledgement.acknowledged_at.desc(),
            TankerReceiptAcknowledgement.id.desc(),
        )
        .all()
    )

    return [
        build_tanker_acknowledgement_response(acknowledgement, db)
        for acknowledgement in acknowledgements
    ]


@app.post(
    "/tanker-tracking/acknowledge",
    response_model=TankerReceiptAcknowledgementResponse,
)
def acknowledge_tanker_receipt(
    request: TankerReceiptAcknowledgementCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    sender_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == request.sender_transaction_id)
        .first()
    )

    if not sender_transaction:
        raise HTTPException(
            status_code=404,
            detail="Sender tanker transaction not found",
        )

    if sender_transaction.status != APPROVED_TRANSACTION_STATUS:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Approved sender tanker transactions can be acknowledged "
                "by the receiver"
            ),
        )

    tanker_payload = get_tanker_payload_for_transaction(
        db,
        sender_transaction.id,
    )

    if not tanker_payload:
        raise HTTPException(
            status_code=400,
            detail="Selected sender transaction does not have tanker payload",
        )

    existing_acknowledgement = get_tanker_acknowledgement_by_sender(
        db,
        sender_transaction.id,
    )

    if existing_acknowledgement:
        raise HTTPException(
            status_code=400,
            detail="This tanker sender transaction is already acknowledged",
        )

    if clean_optional_text(sender_transaction.convoy_number) is None:
        raise HTTPException(
            status_code=400,
            detail="Sender tanker transaction does not have convoy number",
        )

    sender_ticket = build_tanker_tracking_ticket(
        transaction=sender_transaction,
        tanker_payload=tanker_payload,
        db=db,
    )

    receiver_location_code = clean_optional_text(
        request.receiver_location_code
    )

    if receiver_location_code:
        receiver_location = get_location_by_code(receiver_location_code, db)

        if not receiver_location:
            raise HTTPException(
                status_code=400,
                detail="Receiver location not found",
            )

        if receiver_location.status != "Active":
            raise HTTPException(
                status_code=400,
                detail="Only Active receiver location can acknowledge receipt",
            )

    new_acknowledgement = TankerReceiptAcknowledgement(
        sender_transaction_id=sender_transaction.id,
        convoy_number=sender_transaction.convoy_number,
        tanker_asset_code=sender_ticket.get("tanker_asset_code"),
        prime_mover_asset_code=sender_ticket.get("prime_mover_asset_code"),
        receiver_location_code=receiver_location_code,
        acknowledged_by=get_current_user_label(current_user),
        acknowledged_at=datetime.utcnow(),
        remarks=clean_optional_text(request.remarks),
        status="Acknowledged",
    )

    db.add(new_acknowledgement)
    db.flush()

    after_data = build_tanker_acknowledgement_audit_snapshot(
        new_acknowledgement,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Tanker Tracking",
        action="Acknowledge Tanker Receipt",
        current_user=current_user,
        entity_type="TankerReceiptAcknowledgement",
        entity_id=new_acknowledgement.id,
        entity_label=(
            f"{new_acknowledgement.convoy_number} - "
            f"{new_acknowledgement.tanker_asset_code or ''}"
        ),
        ticket_number=get_transaction_ticket_number(sender_transaction),
        operation_number=sender_transaction.operation_number,
        remarks="Tanker receipt acknowledged",
        request_path="/tanker-tracking/acknowledge",
        details={
            "after": after_data,
            "sender_transaction": {
                "id": sender_transaction.id,
                "ticket_number": get_transaction_ticket_number(
                    sender_transaction
                ),
                "operation_number": sender_transaction.operation_number,
                "convoy_number": sender_transaction.convoy_number,
            },
        },
    )

    db.commit()
    db.refresh(new_acknowledgement)

    return build_tanker_acknowledgement_response(
        new_acknowledgement,
        db,
    )

@app.post(
    "/tanker-tracking/acknowledgements/{acknowledgement_id}/revoke",
    response_model=TankerReceiptAcknowledgementResponse,
)
def revoke_tanker_receipt_acknowledgement(
    acknowledgement_id: int,
    remarks: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    acknowledgement = (
        db.query(TankerReceiptAcknowledgement)
        .filter(TankerReceiptAcknowledgement.id == acknowledgement_id)
        .first()
    )

    if not acknowledgement:
        raise HTTPException(
            status_code=404,
            detail="Tanker receipt acknowledgement not found",
        )

    if acknowledgement.status == "Revoked":
        raise HTTPException(
            status_code=400,
            detail="This tanker receipt acknowledgement is already revoked",
        )

    before_data = build_tanker_acknowledgement_audit_snapshot(
        acknowledgement,
        db,
    )

    acknowledgement.status = "Revoked"

    cleaned_remarks = clean_optional_text(remarks)

    if cleaned_remarks:
        acknowledgement.remarks = (
            f"{acknowledgement.remarks or ''}\nRevoke Remarks: {cleaned_remarks}"
        ).strip()

    acknowledgement.updated_at = datetime.utcnow()

    db.flush()

    after_data = build_tanker_acknowledgement_audit_snapshot(
        acknowledgement,
        db,
    )

    sender_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == acknowledgement.sender_transaction_id)
        .first()
    )

    create_audit_log(
        db=db,
        module_name="Tanker Tracking",
        action="Revoke Tanker Receipt Acknowledgement",
        current_user=current_user,
        entity_type="TankerReceiptAcknowledgement",
        entity_id=acknowledgement.id,
        entity_label=(
            f"{acknowledgement.convoy_number} - "
            f"{acknowledgement.tanker_asset_code or ''}"
        ),
        ticket_number=(
            get_transaction_ticket_number(sender_transaction)
            if sender_transaction
            else None
        ),
        operation_number=(
            sender_transaction.operation_number
            if sender_transaction
            else None
        ),
        remarks="Tanker receipt acknowledgement revoked",
        request_path=(
            f"/tanker-tracking/acknowledgements/"
            f"{acknowledgement_id}/revoke"
        ),
        details={
            "before": before_data,
            "after": after_data,
        },
    )

    db.commit()
    db.refresh(acknowledgement)

    return build_tanker_acknowledgement_response(
        acknowledgement,
        db,
    )


@app.post(
    "/tanker-tracking/close",
    response_model=TankerReceiptAcknowledgementResponse,
)
def close_tanker_tracking_movement(
    request: TankerTrackingClosureCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    acknowledgement = (
        db.query(TankerReceiptAcknowledgement)
        .filter(TankerReceiptAcknowledgement.id == request.acknowledgement_id)
        .first()
    )

    if not acknowledgement:
        raise HTTPException(
            status_code=404,
            detail="Tanker acknowledgement not found",
        )

    if acknowledgement.status == "Revoked":
        raise HTTPException(
            status_code=400,
            detail="Revoked tanker acknowledgement cannot be closed",
        )

    if acknowledgement.status == "Closed":
        raise HTTPException(
            status_code=400,
            detail="This tanker movement is already closed",
        )

    sender_transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == acknowledgement.sender_transaction_id)
        .first()
    )

    if not sender_transaction:
        raise HTTPException(
            status_code=404,
            detail="Sender transaction not found",
        )

    require_approved_transaction_for_tracking(
        sender_transaction,
        "tanker movement closure",
    )

    tracking_rows = get_tanker_tracking_rows(
        db=db,
        convoy_number=acknowledgement.convoy_number,
        tanker_asset_code=acknowledgement.tanker_asset_code,
    )

    target_row = None

    for row in tracking_rows:
        if row.get("acknowledgement_id") == acknowledgement.id:
            target_row = row
            break

    if not target_row:
        raise HTTPException(
            status_code=400,
            detail="Unable to find tanker tracking row for this acknowledgement",
        )

    if not target_row.get("latest_receiver_ticket"):
        raise HTTPException(
            status_code=400,
            detail="Cannot close tanker movement before receiver ticket is Approved",
        )

    if not target_row.get("quantity_comparison"):
        raise HTTPException(
            status_code=400,
            detail="Cannot close tanker movement before quantity comparison is available",
        )

    before_data = build_tanker_acknowledgement_audit_snapshot(
        acknowledgement,
        db,
    )

    acknowledgement.status = "Closed"
    acknowledgement.closed_by = get_current_user_label(current_user)
    acknowledgement.closed_at = datetime.utcnow()
    acknowledgement.closure_remarks = clean_optional_text(request.closure_remarks)
    acknowledgement.updated_at = datetime.utcnow()

    db.flush()

    after_data = build_tanker_acknowledgement_audit_snapshot(
        acknowledgement,
        db,
    )

    create_audit_log(
        db=db,
        module_name="Tanker Tracking",
        action="Close Tanker Movement",
        current_user=current_user,
        entity_type="TankerReceiptAcknowledgement",
        entity_id=acknowledgement.id,
        entity_label=(
            f"{acknowledgement.convoy_number} - "
            f"{acknowledgement.tanker_asset_code or ''}"
        ),
        ticket_number=get_transaction_ticket_number(sender_transaction),
        operation_number=sender_transaction.operation_number,
        remarks="Tanker movement closed after comparison",
        request_path="/tanker-tracking/close",
        details={
            "before": before_data,
            "after": after_data,
            "comparison": target_row.get("quantity_comparison"),
            "tracking_status_before_close": target_row.get("tracking_status"),
        },
    )

    db.commit()
    db.refresh(acknowledgement)

    return build_tanker_acknowledgement_response(
        acknowledgement,
        db,
    )


@app.post("/shuttle-voyages/close", response_model=ShuttleVoyageResponse)
def close_shuttle_voyage(
    request: ShuttleVoyageCloseRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Shuttle Tracking", db)

    voyage = get_or_create_shuttle_voyage(
        db=db,
        location_code=request.location_code,
        shuttle_number=request.shuttle_number,
        shuttle_asset_code=request.shuttle_asset_code,
        current_user=current_user,
    )

    ensure_shuttle_voyage_not_closed(voyage)

    voyage.status = "CLOSED"
    voyage.closed_by = get_current_user_label(current_user)
    voyage.closed_at = datetime.utcnow()
    voyage.closure_remarks = clean_optional_text(request.closure_remarks)
    voyage.updated_at = datetime.utcnow()

    create_audit_log(
        db=db,
        module_name="Shuttle Tracking",
        action="Close Shuttle Voyage",
        current_user=current_user,
        entity_type="ShuttleVoyage",
        entity_id=voyage.id,
        entity_label=f"{voyage.location_code}-{voyage.shuttle_asset_code}-{voyage.shuttle_number}",
        remarks="Shuttle voyage closed",
        request_path="/shuttle-voyages/close",
        details={
            "location_code": voyage.location_code,
            "shuttle_number": voyage.shuttle_number,
            "shuttle_asset_code": voyage.shuttle_asset_code,
        },
    )

    db.commit()
    db.refresh(voyage)
    return voyage


@app.post("/shuttle-voyages/reopen", response_model=ShuttleVoyageResponse)
def reopen_shuttle_voyage(
    request: ShuttleVoyageReopenRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage Shuttle Tracking", db)

    voyage = get_shuttle_voyage_by_key(
        db,
        request.location_code,
        request.shuttle_number,
        request.shuttle_asset_code,
    )

    if not voyage:
        raise HTTPException(status_code=404, detail="Shuttle voyage not found")

    voyage.status = "OPEN"
    voyage.closed_by = None
    voyage.closed_at = None
    voyage.closure_remarks = None
    voyage.updated_at = datetime.utcnow()

    if request.remarks:
        voyage.remarks = clean_optional_text(request.remarks)

    create_audit_log(
        db=db,
        module_name="Shuttle Tracking",
        action="Reopen Shuttle Voyage",
        current_user=current_user,
        entity_type="ShuttleVoyage",
        entity_id=voyage.id,
        entity_label=f"{voyage.location_code}-{voyage.shuttle_asset_code}-{voyage.shuttle_number}",
        remarks="Shuttle voyage reopened",
        request_path="/shuttle-voyages/reopen",
        details={
            "location_code": voyage.location_code,
            "shuttle_number": voyage.shuttle_number,
            "shuttle_asset_code": voyage.shuttle_asset_code,
        },
    )

    db.commit()
    db.refresh(voyage)
    return voyage


@app.post("/fso-voyages/close", response_model=FSOVoyageResponse)
def close_fso_voyage(
    request: FSOVoyageCloseRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage FSO Tracking", db)

    voyage = get_or_create_fso_voyage(
        db=db,
        location_code=request.location_code,
        shuttle_number=request.shuttle_number,
        fso_asset_code=request.fso_asset_code,
        current_user=current_user,
    )

    ensure_fso_voyage_not_closed(voyage)

    voyage.status = "CLOSED"
    voyage.closed_by = get_current_user_label(current_user)
    voyage.closed_at = datetime.utcnow()
    voyage.closure_remarks = clean_optional_text(request.closure_remarks)
    voyage.updated_at = datetime.utcnow()

    create_audit_log(
        db=db,
        module_name="FSO Tracking",
        action="Close FSO Voyage",
        current_user=current_user,
        entity_type="FSOVoyage",
        entity_id=voyage.id,
        entity_label=f"{voyage.location_code}-{voyage.fso_asset_code}-{voyage.shuttle_number}",
        remarks="FSO voyage closed",
        request_path="/fso-voyages/close",
        details={
            "location_code": voyage.location_code,
            "shuttle_number": voyage.shuttle_number,
            "fso_asset_code": voyage.fso_asset_code,
        },
    )

    db.commit()
    db.refresh(voyage)
    return voyage


@app.post("/fso-voyages/reopen", response_model=FSOVoyageResponse)
def reopen_fso_voyage(
    request: FSOVoyageReopenRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Manage FSO Tracking", db)

    voyage = get_fso_voyage_by_key(
        db,
        request.location_code,
        request.shuttle_number,
        request.fso_asset_code,
    )

    if not voyage:
        raise HTTPException(status_code=404, detail="FSO voyage not found")

    voyage.status = "OPEN"
    voyage.closed_by = None
    voyage.closed_at = None
    voyage.closure_remarks = None
    voyage.updated_at = datetime.utcnow()

    if request.remarks:
        voyage.remarks = clean_optional_text(request.remarks)

    create_audit_log(
        db=db,
        module_name="FSO Tracking",
        action="Reopen FSO Voyage",
        current_user=current_user,
        entity_type="FSOVoyage",
        entity_id=voyage.id,
        entity_label=f"{voyage.location_code}-{voyage.fso_asset_code}-{voyage.shuttle_number}",
        remarks="FSO voyage reopened",
        request_path="/fso-voyages/reopen",
        details={
            "location_code": voyage.location_code,
            "shuttle_number": voyage.shuttle_number,
            "fso_asset_code": voyage.fso_asset_code,
        },
    )

    db.commit()
    db.refresh(voyage)
    return voyage


@app.get("/shuttle-tracking", response_model=ShuttleTrackingResponse)
def get_shuttle_tracking(
    date_from: str | None = None,
    date_to: str | None = None,
    location_code: str | None = None,
    shuttle_number: str | None = None,
    shuttle_asset_code: str | None = None,
    # ✅ NEW enterprise params
    tab: str | None = None,          # OPEN / CLOSED
    search: str | None = None,       # keyword search
    page: int = 1,
    page_size: int = 20,
    include_tickets: bool = False,   # list-only by default
    group_key: str | None = None,    # lazy-load one voyage
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Shuttle Tracking", db)

    def _norm(v):
        return str(v or "").strip().upper()

    def _abs_qty(net_stock, net_water):
        try:
            return abs(float(net_stock or 0.0)) + abs(float(net_water or 0.0))
        except Exception:
            return 0.0

    def _op_code(meta):
        return _norm(meta.get("vessel_operation_code"))

    def _op_label(meta):
        return _norm(meta.get("vessel_operation_label"))

    def _is_loading(meta):
        code = _op_code(meta)
        if code == "LOADING":
            return True
        label = _op_label(meta)
        return ("LOADING" in label) and ("UNLOADING" not in label)

    def _is_sts_in(meta):
        code = _op_code(meta)
        if code == "STS_IN":
            return True
        label = _op_label(meta)
        return "STS IN" in label or "STS_IN" in label

    def _is_sts_out(meta):
        code = _op_code(meta)
        if code == "STS_OUT":
            return True
        label = _op_label(meta)
        return "STS OUT" in label or "STS_OUT" in label

    def _is_unloading(meta):
        code = _op_code(meta)
        if code == "UNLOADING":
            return True
        label = _op_label(meta)
        return ("UNLOADING" in label) or ("UNLOAD" in label)

    def _is_top_up(meta):
        code = _op_code(meta)
        if code == "TOP_UP":
            return True
        label = _op_label(meta)
        return ("TOP UP" in label) or ("TOP-UP" in label) or ("TOP_UP" in label)

    df = parse_date_filter(date_from, "date_from")
    dt = parse_date_filter(date_to, "date_to")

    lc = clean_optional_text(location_code)
    sn = clean_optional_text(shuttle_number)
    ac = clean_optional_text(shuttle_asset_code)

    tab_norm = (clean_optional_text(tab) or "OPEN").upper()
    search_norm = (clean_optional_text(search) or "").strip()

    page = 1 if page is None or page < 1 else page
    page_size = 20 if page_size is None or page_size < 1 else min(int(page_size), 200)
    offset = (page - 1) * page_size

    # -----------------------------
    # Base query (Approved + has shuttle_payload)
    # -----------------------------
    base_q = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "shuttle_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
    )

    if df:
        base_q = base_q.filter(OperationTransaction.operation_date >= df)
    if dt:
        base_q = base_q.filter(OperationTransaction.operation_date <= dt)
    if lc:
        base_q = base_q.filter(OperationTransaction.origin_location_code.ilike(lc))
    if sn:
        base_q = base_q.filter(OperationTransaction.convoy_number.ilike(sn))
    if ac:
        base_q = base_q.filter(OperationTransaction.primary_asset_code.ilike(ac))

    # Optional keyword search (basic)
    if search_norm:
        like = f"%{search_norm}%"
        base_q = base_q.filter(
            or_(
                OperationTransaction.origin_location_code.ilike(like),
                OperationTransaction.convoy_number.ilike(like),
                OperationTransaction.primary_asset_code.ilike(like),
                OperationTransaction.operation_number.ilike(like),
                OperationTransaction.operation_ticket_number.ilike(like),
                OperationTransaction.product_name.ilike(like),
            )
        )

    # -----------------------------
    # If lazy-loading a single voyage, parse group_key => filter base_q
    # group_key format: location|shuttle_number|asset_code
    # -----------------------------
    if group_key:
        parts = [p.strip() for p in str(group_key).split("|")]
        if len(parts) == 3:
            g_loc, g_shuttle, g_asset = parts
            base_q = base_q.filter(
                OperationTransaction.origin_location_code == g_loc,
                OperationTransaction.convoy_number == g_shuttle,
                OperationTransaction.primary_asset_code == g_asset,
            )
            include_tickets = True
        else:
            raise HTTPException(status_code=400, detail="Invalid group_key format")

    # -----------------------------
    # Get distinct voyage keys with paging
    # Join ShuttleVoyage for OPEN/CLOSED filtering (enterprise)
    # -----------------------------
    voyage_status_expr = func.coalesce(ShuttleVoyage.status, literal("OPEN"))

    key_loc = OperationTransaction.origin_location_code
    key_shuttle = OperationTransaction.convoy_number
    key_asset = OperationTransaction.primary_asset_code

    group_q = (
        base_q.with_entities(
            key_loc.label("location_code"),
            key_shuttle.label("shuttle_number"),
            key_asset.label("shuttle_asset_code"),
            func.min(OperationTransaction.operation_date).label("first_date"),
            func.max(OperationTransaction.operation_date).label("last_date"),
        )
        .outerjoin(
            ShuttleVoyage,
            and_(
                ShuttleVoyage.location_code == key_loc,
                ShuttleVoyage.shuttle_number == key_shuttle,
                ShuttleVoyage.shuttle_asset_code == key_asset,
            ),
        )
        .group_by(key_loc, key_shuttle, key_asset, voyage_status_expr)
    )

    if tab_norm == "CLOSED":
        group_q = group_q.filter(voyage_status_expr == "CLOSED")
    else:
        group_q = group_q.filter(voyage_status_expr != "CLOSED")

    total_groups = group_q.count()

    group_rows = (
        group_q.order_by(func.max(OperationTransaction.operation_date).desc(), key_loc.asc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # No voyages on this page
    if not group_rows:
        return {
            "rows": [],
            "total_groups": total_groups,
            "page": page,
            "page_size": page_size,
            "has_more": total_groups > offset + page_size,
        }

    # Build keys list for fetching tickets only for paged voyages
    keys = [
        (r.location_code, r.shuttle_number, r.shuttle_asset_code)
        for r in group_rows
    ]

    # If list-only mode, return voyages without tickets (fast)
    if not include_tickets:
        # ✅ Compute totals for voyages on this page WITHOUT returning ticket list
        # Pull only shuttle_payload values for these keys and sum net fields.
        key_filters = []
        for (loc_code, sh_num, asset_code) in keys:
            key_filters.append(
                and_(
                    OperationTransaction.origin_location_code == loc_code,
                    OperationTransaction.convoy_number == sh_num,
                    OperationTransaction.primary_asset_code == asset_code,
                )
            )

        payload_rows = (
            db.query(
                OperationTransaction.origin_location_code,
                OperationTransaction.convoy_number,
                OperationTransaction.primary_asset_code,
                OperationTransactionValue.field_value,
            )
            .join(
                OperationTransactionValue,
                OperationTransactionValue.transaction_id == OperationTransaction.id,
            )
            .filter(
                OperationTransactionValue.field_code == "shuttle_payload",
                OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            )
            .filter(or_(*key_filters))
            .all()
        )

        totals_map = {}  # key -> {net_receipt_bbl, net_discharge_bbl}
        for (loc_code, sh_num, asset_code, field_value) in payload_rows:
            k = f"{loc_code}|{sh_num}|{asset_code}"
            if k not in totals_map:
                totals_map[k] = {"net_receipt_bbl": 0.0, "net_discharge_bbl": 0.0}

            if isinstance(field_value, dict):
                fv = field_value
                meta = fv.get("meta") or {}
                net = ((fv.get("calculated") or {}).get("net") or {})
                net_stock = float(safe_float(net.get("net_stock_bbl")))
                net_water = float(safe_float(net.get("net_water_bbl")))
                qty_bbl = _abs_qty(net_stock, net_water)

                if (_is_loading(meta) or _is_sts_in(meta) or _is_top_up(meta)) and (not _is_unloading(meta)):
                    totals_map[k]["net_receipt_bbl"] += qty_bbl
                if _is_unloading(meta) and not _is_sts_out(meta):
                    totals_map[k]["net_discharge_bbl"] += qty_bbl

        rows = []
        for (loc_code, sh_num, asset_code) in keys:
            asset = get_asset_by_code(asset_code, db)
            loc = get_location_by_code(loc_code, db)
            voyage = get_shuttle_voyage_by_key(db, loc_code, sh_num or "", asset_code)

            k = f"{loc_code}|{sh_num}|{asset_code}"
            t = totals_map.get(k, {"net_receipt_bbl": 0.0, "net_discharge_bbl": 0.0})

            rows.append(
                {
                    "group_key": k,
                    "location_code": loc_code,
                    "location_name": loc.location_name if loc else "",
                    "shuttle_number": sh_num or "",
                    "shuttle_asset_code": asset_code,
                    "shuttle_asset_name": asset.asset_name if asset else "",
                    "voyage_status": voyage.status if voyage else "OPEN",
                    "closed_by": voyage.closed_by if voyage else None,
                    "closed_at": voyage.closed_at if voyage else None,
                    "closure_remarks": voyage.closure_remarks if voyage else None,
                    "total_tov_bbl": 0.0,
                    "total_free_water_bbl": 0.0,
                    "total_nsv_bbl": 0.0,
                    "net_receipt_bbl": float(t["net_receipt_bbl"]),
                    "net_discharge_bbl": float(t["net_discharge_bbl"]),
                    "tickets": [],
                }
            )

        return {
            "rows": rows,
            "total_groups": total_groups,
            "page": page,
            "page_size": page_size,
            "has_more": total_groups > offset + page_size,
        }

    # -----------------------------
    # Include tickets: fetch only transactions for these keys
    # -----------------------------
    tx_q = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "shuttle_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
    )

    # Reapply the same filters
    if df:
        tx_q = tx_q.filter(OperationTransaction.operation_date >= df)
    if dt:
        tx_q = tx_q.filter(OperationTransaction.operation_date <= dt)
    if search_norm:
        like = f"%{search_norm}%"
        tx_q = tx_q.filter(
            or_(
                OperationTransaction.origin_location_code.ilike(like),
                OperationTransaction.convoy_number.ilike(like),
                OperationTransaction.primary_asset_code.ilike(like),
                OperationTransaction.operation_number.ilike(like),
                OperationTransaction.operation_ticket_number.ilike(like),
                OperationTransaction.product_name.ilike(like),
            )
        )

    # Key filter: only paged voyages
    key_filters = []
    for (loc_code, sh_num, asset_code) in keys:
        key_filters.append(
            and_(
                OperationTransaction.origin_location_code == loc_code,
                OperationTransaction.convoy_number == sh_num,
                OperationTransaction.primary_asset_code == asset_code,
            )
        )
    tx_q = tx_q.filter(or_(*key_filters))

    txs = (
        tx_q.order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
        .all()
    )

    grouped = {}

    for tx in txs:
        payload = get_shuttle_payload_for_transaction(db, tx.id) or {}
        meta = (payload.get("meta") or {}) if isinstance(payload, dict) else {}
        net = ((payload.get("calculated") or {}).get("net") or {}) if isinstance(payload, dict) else {}
        inputs = (payload.get("inputs") or {}) if isinstance(payload, dict) else {}

        # Raw values from payload
        tov_raw = float(safe_float(net.get("TOV")))
        fw_raw = float(safe_float(net.get("FW")))
        nsv_raw = float(safe_float(net.get("NSV")))

        # Apply movement sign (enterprise-safe)
        sign = str(meta.get("vessel_operation_sign") or "").strip().upper()

        multiplier = 1.0
        if sign in ("OUT", "-"):
            multiplier = -1.0
        elif sign in ("SET", "NEUTRAL", "0"):
            multiplier = 0.0

        tov = tov_raw * multiplier
        fw = fw_raw * multiplier
        nsv = nsv_raw * multiplier

        event_time = inputs.get("event_time")
        opening_stock = float(safe_float(inputs.get("opening_stock_bbl")))
        opening_water = float(safe_float(inputs.get("opening_water_bbl")))
        closing_stock = float(safe_float(inputs.get("closing_stock_bbl")))
        closing_water = float(safe_float(inputs.get("closing_water_bbl")))
        net_stock = float(safe_float(net.get("net_stock_bbl")))
        net_water = float(safe_float(net.get("net_water_bbl")))

        barge_reference = inputs.get("barge_reference")
        remarks = inputs.get("remarks")

        key = f"{tx.origin_location_code}|{tx.convoy_number}|{tx.primary_asset_code}"

        asset = get_asset_by_code(tx.primary_asset_code, db)
        loc = get_location_by_code(tx.origin_location_code, db)

        if key not in grouped:
            voyage = get_shuttle_voyage_by_key(
                db,
                tx.origin_location_code,
                tx.convoy_number or "",
                tx.primary_asset_code,
            )

            grouped[key] = {
                "group_key": key,
                "location_code": tx.origin_location_code,
                "location_name": loc.location_name if loc else "",
                "shuttle_number": tx.convoy_number or "",
                "shuttle_asset_code": tx.primary_asset_code,
                "shuttle_asset_name": asset.asset_name if asset else "",
                "voyage_status": voyage.status if voyage else "OPEN",
                "closed_by": voyage.closed_by if voyage else None,
                "closed_at": voyage.closed_at if voyage else None,
                "closure_remarks": voyage.closure_remarks if voyage else None,
                "tickets": [],
                "total_tov_bbl": 0.0,
                "total_free_water_bbl": 0.0,
                "total_nsv_bbl": 0.0,
                "net_receipt_bbl": 0.0,
                "net_discharge_bbl": 0.0,
            }

        qty_bbl = _abs_qty(net_stock, net_water)
        if (_is_loading(meta) or _is_sts_in(meta) or _is_top_up(meta)) and (not _is_unloading(meta)):
            grouped[key]["net_receipt_bbl"] += qty_bbl
        if _is_unloading(meta) and not _is_sts_out(meta):
            grouped[key]["net_discharge_bbl"] += qty_bbl

        grouped[key]["tickets"].append(
            {
                "transaction_id": tx.id,
                "ticket_number": get_transaction_ticket_number(tx),
                "operation_number": tx.operation_number,
                "location_code": tx.origin_location_code,
                "location_name": loc.location_name if loc else "",
                "shuttle_number": tx.convoy_number or "",
                "shuttle_asset_code": tx.primary_asset_code,
                "shuttle_asset_name": asset.asset_name if asset else "",
                "product_name": tx.product_name,
                "operation_date": tx.operation_date,
                "event_time": event_time,
                "opening_stock_bbl": opening_stock,
                "opening_water_bbl": opening_water,
                "closing_stock_bbl": closing_stock,
                "closing_water_bbl": closing_water,
                "net_stock_bbl": net_stock,
                "net_water_bbl": net_water,
                "barge_reference": barge_reference,
                "remarks": remarks,
                "vessel_operation_code": meta.get("vessel_operation_code"),
                "vessel_operation_label": meta.get("vessel_operation_label"),
                "vessel_operation_category": meta.get("vessel_operation_category"),
                "vessel_operation_sign": meta.get("vessel_operation_sign"),
                "tov_bbl": tov,
                "free_water_bbl": fw,
                "nsv_bbl": nsv,
                "status": tx.status,
                "created_by": tx.created_by,
                "created_at": tx.created_at,
                "updated_at": tx.updated_at,
            }
        )

        grouped[key]["total_tov_bbl"] += tov
        grouped[key]["total_free_water_bbl"] += fw
        grouped[key]["total_nsv_bbl"] += nsv

    # Return rows in the same order as group_rows (paged ordering)
    rows = []
    for (loc_code, sh_num, asset_code) in keys:
        k = f"{loc_code}|{sh_num}|{asset_code}"
        if k in grouped:
            rows.append(grouped[k])
        else:
            # In case of no tickets (rare)
            asset = get_asset_by_code(asset_code, db)
            loc = get_location_by_code(loc_code, db)
            voyage = get_shuttle_voyage_by_key(db, loc_code, sh_num or "", asset_code)
            rows.append(
                {
                    "group_key": k,
                    "location_code": loc_code,
                    "location_name": loc.location_name if loc else "",
                    "shuttle_number": sh_num or "",
                    "shuttle_asset_code": asset_code,
                    "shuttle_asset_name": asset.asset_name if asset else "",
                    "voyage_status": voyage.status if voyage else "OPEN",
                    "closed_by": voyage.closed_by if voyage else None,
                    "closed_at": voyage.closed_at if voyage else None,
                    "closure_remarks": voyage.closure_remarks if voyage else None,
                    "tickets": [],
                    "total_tov_bbl": 0.0,
                    "total_free_water_bbl": 0.0,
                    "total_nsv_bbl": 0.0,
                    "net_receipt_bbl": 0.0,
                    "net_discharge_bbl": 0.0,
                }
            )

    return {
        "rows": rows,
        "total_groups": total_groups,
        "page": page,
        "page_size": page_size,
        "has_more": total_groups > offset + page_size,
    }


@app.get("/shuttle-tracking/export/xlsx")
def export_shuttle_voyage_xlsx(
    group_key: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View Shuttle Tracking", db)

    parts = [p.strip() for p in str(group_key).split("|")]
    if len(parts) != 3:
        raise HTTPException(
            status_code=400,
            detail="Invalid group_key format. Expected location|shuttle_number|asset_code",
        )

    loc_code, shuttle_no, asset_code = parts

    txs = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "shuttle_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransaction.origin_location_code == loc_code,
            OperationTransaction.convoy_number == shuttle_no,
            OperationTransaction.primary_asset_code == asset_code,
        )
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
        .all()
    )

    if not txs:
        raise HTTPException(status_code=404, detail="No approved shuttle tickets found for this voyage")

    loc = get_location_by_code(loc_code, db)
    asset = get_asset_by_code(asset_code, db)

    def _sf(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _abs_qty(net_stock, net_water):
        return abs(_sf(net_stock)) + abs(_sf(net_water))

    receipt_total = 0.0
    discharge_total = 0.0

    last_closing_stock = 0.0
    last_closing_water = 0.0

    rows = []

    for tx in txs:
        payload = get_shuttle_payload_for_transaction(db, tx.id) or {}
        meta = (payload.get("meta") or {}) if isinstance(payload, dict) else {}
        inputs = (payload.get("inputs") or {}) if isinstance(payload, dict) else {}
        net = ((payload.get("calculated") or {}).get("net") or {}) if isinstance(payload, dict) else {}

        op_code = str(meta.get("vessel_operation_code") or "").strip().upper()
        op_label = str(meta.get("vessel_operation_label") or "").strip()
        op_sign = str(meta.get("vessel_operation_sign") or "").strip().upper()

        event_time = inputs.get("event_time") or ""
        closing_stock = _sf(inputs.get("closing_stock_bbl"))
        closing_water = _sf(inputs.get("closing_water_bbl"))

        net_stock = _sf(net.get("net_stock_bbl"))
        net_water = _sf(net.get("net_water_bbl"))
        qty = _abs_qty(net_stock, net_water)

        if op_code in ("LOADING", "STS_IN", "TOP_UP"):
            receipt_total += qty
        if op_code == "UNLOADING":
            discharge_total += qty

        last_closing_stock = closing_stock
        last_closing_water = closing_water

        rows.append(
            {
                "date": str(tx.operation_date),
                "time": event_time,
                "operation": op_label or op_code,
                "sign": op_sign,
                "net_stock": net_stock,
                "net_water": net_water,
                "qty": qty,
                "ticket": get_transaction_ticket_number(tx),
            }
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shuttle MTR"

    ws.append(["SHUTTLE VOYAGE MTR"])
    ws.append([f"Location: {loc_code} - {(loc.location_name if loc else '')}"])
    ws.append([f"Shuttle: {asset_code} - {(asset.asset_name if asset else '')}"])
    ws.append([f"Shuttle Number: {shuttle_no}"])
    ws.append([""])

    ws.append(
        [
            "Status",
            "Tickets",
            "Net Receipt (BBL)",
            "Net Discharge (BBL)",
            "Last Closing Stock",
            "Last Closing Water",
        ]
    )
    ws.append(
        [
            "Approved",
            len(rows),
            round(receipt_total, 3),
            round(discharge_total, 3),
            round(last_closing_stock, 3),
            round(last_closing_water, 3),
        ]
    )
    ws.append([""])

    ws.append(["Date", "Time", "Operation", "Sign", "Net Stock", "Net Water", "Qty (Abs S+W)", "Ticket"])
    for r in rows:
        ws.append(
            [
                r["date"],
                r["time"],
                r["operation"],
                r["sign"],
                round(r["net_stock"], 3),
                round(r["net_water"], 3),
                round(r["qty"], 3),
                r["ticket"],
            ]
        )

    _xlsx_autofit(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"shuttle_mtr_{loc_code}_{asset_code}_{shuttle_no}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/fso-tracking", response_model=FSOTrackingResponse)
def get_fso_tracking(
    tab: str | None = "OPEN",
    search: str | None = None,
    page: int = 1,
    page_size: int = 20,
    include_tickets: bool = False,
    group_key: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "View FSO Tracking", db)

    def _sf(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _abs_qty(net_stock, net_water):
        try:
            return abs(float(net_stock or 0.0)) + abs(float(net_water or 0.0))
        except Exception:
            return 0.0

    def _norm(v):
        return str(v or "").strip().upper()

    def _sh_op_code(meta):
        return _norm((meta or {}).get("vessel_operation_code"))

    def _sh_op_label(meta):
        return _norm((meta or {}).get("vessel_operation_label"))

    def _sh_is_unloading(meta):
        code = _sh_op_code(meta)
        if code == "UNLOADING":
            return True
        label = _sh_op_label(meta)
        return ("UNLOADING" in label) or ("UNLOAD" in label)

    def _sh_is_sts_out(meta):
        code = _sh_op_code(meta)
        if code == "STS_OUT":
            return True
        label = _sh_op_label(meta)
        return "STS OUT" in label or "STS_OUT" in label

    def _build_shuttle_discharge_fallback(pairs):
        if not pairs:
            return {}

        voyage_filters = []
        for (loc_code, sh_num) in pairs:
            voyage_filters.append(
                and_(
                    ShuttleVoyage.location_code == loc_code,
                    ShuttleVoyage.shuttle_number == sh_num,
                )
            )

        voyages = (
            db.query(ShuttleVoyage)
            .filter(ShuttleVoyage.status == "CLOSED")
            .filter(or_(*voyage_filters))
            .all()
        )

        latest_voyage = {}
        for v in voyages:
            key = f"{v.location_code}|{v.shuttle_number}"
            cur = latest_voyage.get(key)
            if not cur:
                latest_voyage[key] = v
                continue
            cur_dt = cur.closed_at or datetime.min
            v_dt = v.closed_at or datetime.min
            if v_dt > cur_dt:
                latest_voyage[key] = v

        tx_filters = []
        for _, v in latest_voyage.items():
            tx_filters.append(
                and_(
                    OperationTransaction.origin_location_code == v.location_code,
                    OperationTransaction.convoy_number == v.shuttle_number,
                    OperationTransaction.primary_asset_code == v.shuttle_asset_code,
                )
            )

        if not tx_filters:
            return {}

        payloads = (
            db.query(
                OperationTransaction.origin_location_code,
                OperationTransaction.convoy_number,
                OperationTransactionValue.field_value,
            )
            .join(
                OperationTransactionValue,
                OperationTransactionValue.transaction_id == OperationTransaction.id,
            )
            .filter(
                OperationTransactionValue.field_code == "shuttle_payload",
                OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            )
            .filter(or_(*tx_filters))
            .all()
        )

        out = {}
        for (loc_code, sh_num, fv) in payloads:
            if not isinstance(fv, dict):
                continue
            meta = fv.get("meta") or {}
            net = ((fv.get("calculated") or {}).get("net") or {})
            net_stock = _sf(net.get("net_stock_bbl"))
            net_water = _sf(net.get("net_water_bbl"))
            qty = _abs_qty(net_stock, net_water)

            if _sh_is_unloading(meta) and (not _sh_is_sts_out(meta)):
                key = f"{loc_code}|{sh_num}"
                out[key] = float(out.get(key, 0.0)) + qty

        return out

    tab_norm = str(tab or "OPEN").strip().upper()
    page = max(int(page or 1), 1)
    page_size = min(max(int(page_size or 20), 1), 200)
    offset = (page - 1) * page_size

    # Approved-only, and only tickets that have fso_payload
    base_q = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransactionValue.field_code == "fso_payload",
        )
    )

    # Optional: enforce primary asset type is FSO (recommended)
    base_q = base_q.filter(OperationTransaction.primary_asset_type_code.ilike("FSO"))

    # Lazy-load a single voyage by group_key: location|shuttle_number|fso_asset_code
    if group_key:
        parts = [p.strip() for p in str(group_key).split("|")]
        if len(parts) != 3:
            raise HTTPException(status_code=400, detail="Invalid group_key format for FSO tracking")
        g_loc, g_shuttle, g_fso = parts
        base_q = base_q.filter(
            OperationTransaction.origin_location_code == g_loc,
            OperationTransaction.convoy_number == g_shuttle,
            OperationTransaction.primary_asset_code == g_fso,
        )
        include_tickets = True

    if search and str(search).strip():
        s = str(search).strip()
        base_q = base_q.filter(
            or_(
                OperationTransaction.origin_location_code.ilike(f"%{s}%"),
                OperationTransaction.primary_asset_code.ilike(f"%{s}%"),
                OperationTransaction.convoy_number.ilike(f"%{s}%"),
                OperationTransaction.ticket_number.ilike(f"%{s}%"),
            )
        )

    key_loc = OperationTransaction.origin_location_code
    key_shuttle = OperationTransaction.convoy_number
    key_asset = OperationTransaction.primary_asset_code

    voyage_status_expr = func.coalesce(FSOVoyage.status, literal("OPEN")).label("voyage_status")

    group_q = (
        base_q.with_entities(
            key_loc.label("location_code"),
            key_shuttle.label("shuttle_number"),
            key_asset.label("fso_asset_code"),
            func.min(OperationTransaction.operation_date).label("first_date"),
            func.max(OperationTransaction.operation_date).label("last_date"),
            voyage_status_expr,
        )
        .outerjoin(
            FSOVoyage,
            and_(
                FSOVoyage.location_code == key_loc,
                FSOVoyage.shuttle_number == key_shuttle,
                FSOVoyage.fso_asset_code == key_asset,
            ),
        )
        .group_by(key_loc, key_shuttle, key_asset, voyage_status_expr)
    )

    if tab_norm == "CLOSED":
        group_q = group_q.filter(voyage_status_expr == "CLOSED")
    else:
        group_q = group_q.filter(voyage_status_expr != "CLOSED")

    total_groups = group_q.count()

    group_rows = (
        group_q.order_by(func.max(OperationTransaction.operation_date).desc(), key_loc.asc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    if not group_rows:
        return {
            "rows": [],
            "total_groups": total_groups,
            "page": page,
            "page_size": page_size,
            "has_more": total_groups > offset + page_size,
        }

    keys = [(r.location_code, r.shuttle_number, r.fso_asset_code) for r in group_rows]

    if not include_tickets:
        # Compute totals for the voyages on this page WITHOUT returning full tickets
        key_filters = []
        for (loc_code, sh_num, fso_code) in keys:
            key_filters.append(
                and_(
                    OperationTransaction.origin_location_code == loc_code,
                    OperationTransaction.convoy_number == sh_num,
                    OperationTransaction.primary_asset_code == fso_code,
                )
            )

        payload_rows = (
            db.query(
                OperationTransaction.origin_location_code,
                OperationTransaction.convoy_number,
                OperationTransaction.primary_asset_code,
                OperationTransactionValue.field_value,
            )
            .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
            .filter(
                OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
                OperationTransactionValue.field_code == "fso_payload",
            )
            .filter(or_(*key_filters))
            .all()
        )

        def _norm_l(v):
            return str(v or "").strip().lower()

        totals_map = {}  # group_key -> totals
        for (loc_code, sh_num, fso_code, fv) in payload_rows:
            k = f"{loc_code}|{sh_num}|{fso_code}"
            if k not in totals_map:
                totals_map[k] = {
                    "receipts": 0.0,
                    "exports": 0.0,
                    "water_in": 0.0,
                    "water_out": 0.0,
                    "loss_gain": 0.0,
                    "variance": 0.0,
                    "shuttle_discharge_meta": 0.0,
                    "fso_receipt_bbl": 0.0,
                }

            if not isinstance(fv, dict):
                continue

            meta = fv.get("meta") or {}
            op_label = meta.get("operation_label") or meta.get("operation") or ""
            op_norm = _norm_l(op_label)

            inputs = fv.get("inputs") or {}
            net = ((fv.get("calculated") or {}).get("net") or {})

            op_sign = _norm(meta.get("operation_sign"))

            net_stock = _sf(net.get("net_stock_bbl"))
            net_water = _sf(net.get("net_water_bbl"))
            vessel_qty = _sf(inputs.get("vessel_quantity_bbl") or meta.get("vessel_quantity_bbl"))
            variance = _sf(net.get("variance_bbl") or meta.get("variance_bbl"))

            qty_bbl = _abs_qty(net_stock, net_water)
            if op_sign == "IN":
                totals_map[k]["fso_receipt_bbl"] += qty_bbl
                totals_map[k]["shuttle_discharge_meta"] += _sf(
                    meta.get("source_shuttle_discharge_bbl")
                )

            if op_norm == "receipt":
                totals_map[k]["receipts"] += max(net_stock, 0.0)
            elif op_norm == "export":
                totals_map[k]["exports"] += abs(net_stock)
            elif op_norm == "stock opening":
                totals_map[k]["loss_gain"] += net_stock

            if net_water > 0:
                totals_map[k]["water_in"] += net_water
            elif net_water < 0:
                totals_map[k]["water_out"] += abs(net_water)

            if op_norm != "export":
                if abs(variance) > 0:
                    totals_map[k]["variance"] += variance
                else:
                    totals_map[k]["variance"] += (abs(net_stock) - vessel_qty)

        rows = []
        shuttle_fallback = _build_shuttle_discharge_fallback(
            list({(loc_code, sh_num) for (loc_code, sh_num, _) in keys})
        )
        for r in group_rows:
            loc = db.query(Location).filter(Location.location_code.ilike(r.location_code)).first()
            fso_asset = db.query(Asset).filter(Asset.asset_code.ilike(r.fso_asset_code)).first()
            voyage = get_fso_voyage_by_key(db, r.location_code, r.shuttle_number, r.fso_asset_code)

            k = f"{r.location_code}|{r.shuttle_number}|{r.fso_asset_code}"
            t = totals_map.get(
                k,
                {
                    "receipts": 0,
                    "exports": 0,
                    "water_in": 0,
                    "water_out": 0,
                    "loss_gain": 0,
                    "variance": 0,
                    "shuttle_discharge_meta": 0,
                    "fso_receipt_bbl": 0,
                },
            )
            net_water = float(t["water_in"]) - float(t["water_out"])
            shuttle_discharge_meta = float(t.get("shuttle_discharge_meta") or 0.0)
            fallback = float(
                shuttle_fallback.get(f"{r.location_code}|{r.shuttle_number}", 0.0)
            )
            shuttle_discharge_bbl = shuttle_discharge_meta if shuttle_discharge_meta > 0 else fallback
            fso_receipt_bbl = float(t.get("fso_receipt_bbl") or 0.0)
            variance_bbl = fso_receipt_bbl - shuttle_discharge_bbl

            rows.append(
                {
                    "group_key": k,
                    "location_code": r.location_code,
                    "location_name": loc.location_name if loc else "",
                    "shuttle_number": r.shuttle_number or "",
                    "fso_asset_code": r.fso_asset_code or "",
                    "fso_asset_name": fso_asset.asset_name if fso_asset else "",
                    "voyage_status": (voyage.status if voyage else "OPEN"),
                    "closed_by": (voyage.closed_by if voyage else None),
                    "closed_at": (voyage.closed_at if voyage else None),
                    "closure_remarks": (voyage.closure_remarks if voyage else None),
                    "total_receipts_bbl": float(t["receipts"]),
                    "total_exports_bbl": float(t["exports"]),
                    "total_water_in_bbl": float(t["water_in"]),
                    "total_water_out_bbl": float(t["water_out"]),
                    "net_water_bbl": net_water,
                    "loss_gain_bbl": float(t["loss_gain"]),
                    "total_variance_bbl": float(t["variance"]),
                    "shuttle_discharge_bbl": float(shuttle_discharge_bbl),
                    "fso_receipt_bbl": float(fso_receipt_bbl),
                    "variance_bbl": float(variance_bbl),
                    "tickets": [],
                }
            )

        return {
            "rows": rows,
            "total_groups": total_groups,
            "page": page,
            "page_size": page_size,
            "has_more": total_groups > offset + page_size,
        }

    # Fetch tickets only for the paged groups
    key_filters = []
    for (loc_code, sh_num, fso_code) in keys:
        key_filters.append(
            and_(
                OperationTransaction.origin_location_code == loc_code,
                OperationTransaction.convoy_number == sh_num,
                OperationTransaction.primary_asset_code == fso_code,
            )
        )

    tx_rows = (
        db.query(OperationTransaction)
        .join(OperationTransactionValue, OperationTransactionValue.transaction_id == OperationTransaction.id)
        .filter(
            OperationTransactionValue.field_code == "fso_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
        .filter(or_(*key_filters))
        .order_by(OperationTransaction.operation_date.asc(), OperationTransaction.id.asc())
        .all()
    )

    # Build group map
    groups = {}
    for r in group_rows:
        k = f"{r.location_code}|{r.shuttle_number}|{r.fso_asset_code}"
        loc = db.query(Location).filter(Location.location_code.ilike(r.location_code)).first()
        fso_asset = db.query(Asset).filter(Asset.asset_code.ilike(r.fso_asset_code)).first()
        voyage = get_fso_voyage_by_key(db, r.location_code, r.shuttle_number, r.fso_asset_code)

        groups[k] = {
            "group_key": k,
            "location_code": r.location_code,
            "location_name": loc.location_name if loc else "",
            "shuttle_number": r.shuttle_number or "",
            "fso_asset_code": r.fso_asset_code or "",
            "fso_asset_name": fso_asset.asset_name if fso_asset else "",
            "voyage_status": (voyage.status if voyage else "OPEN"),
            "closed_by": (voyage.closed_by if voyage else None),
            "closed_at": (voyage.closed_at if voyage else None),
            "closure_remarks": (voyage.closure_remarks if voyage else None),
            "tickets": [],
            "total_receipts_bbl": 0.0,
            "total_exports_bbl": 0.0,
            "total_water_in_bbl": 0.0,
            "total_water_out_bbl": 0.0,
            "net_water_bbl": 0.0,
            "loss_gain_bbl": 0.0,
            "total_variance_bbl": 0.0,
            "shuttle_discharge_bbl": 0.0,
            "fso_receipt_bbl": 0.0,
            "variance_bbl": 0.0,
        }

    def _safe_float(v):
        try:
            return float(v or 0.0)
        except Exception:
            return 0.0

    def _norm_op(v):
        return str(v or "").strip().lower()

    # Build tickets + OTR summary using payload fields (same logic as old fso_operations.py)
    for tx in tx_rows:
        payload = get_fso_payload_for_transaction(db, tx.id) or {}
        meta = payload.get("meta") or {}
        calcs = payload.get("calculated") or {}
        net_calc = (calcs.get("net") or {}) if isinstance(calcs, dict) else {}

        loc_code = tx.origin_location_code or ""
        sh_num = tx.convoy_number or ""
        fso_code = tx.primary_asset_code or ""
        k = f"{loc_code}|{sh_num}|{fso_code}"

        if k not in groups:
            continue

        op_label = meta.get("operation_label") or meta.get("operation") or ""
        op_norm = _norm_op(op_label)

        opening_stock = _safe_float(meta.get("opening_stock_bbl"))
        opening_water = _safe_float(meta.get("opening_water_bbl"))
        closing_stock = _safe_float(meta.get("closing_stock_bbl"))
        closing_water = _safe_float(meta.get("closing_water_bbl"))
        net_stock = _safe_float(meta.get("net_stock_bbl"))
        net_water = _safe_float(meta.get("net_water_bbl"))
        vessel_qty = _safe_float(meta.get("vessel_quantity_bbl"))
        variance = _safe_float(meta.get("variance_bbl"))

        ticket = {
            "transaction_id": tx.id,
            "ticket_number": get_transaction_ticket_number(tx),
            "operation_number": tx.operation_ticket_number or tx.operation_number,
            "location_code": loc_code,
            "location_name": groups[k]["location_name"],
            "shuttle_number": sh_num,
            "fso_asset_code": fso_code,
            "fso_asset_name": groups[k]["fso_asset_name"],
            "product_name": tx.product_name,
            "operation_date": tx.operation_date,
            "event_time": meta.get("event_time"),
            "operation_label": op_label,
            "vessel_name": meta.get("vessel_name"),
            "vessel_quantity_bbl": vessel_qty,
            "opening_stock_bbl": opening_stock,
            "opening_water_bbl": opening_water,
            "closing_stock_bbl": closing_stock,
            "closing_water_bbl": closing_water,
            "net_stock_bbl": net_stock,
            "net_water_bbl": net_water,
            "variance_bbl": variance,
            "remarks": meta.get("remarks") or tx.remarks,
            "status": tx.status,
            "created_by": tx.created_by,
            "created_at": tx.created_at,
            "updated_at": tx.updated_at,
        }

        groups[k]["tickets"].append(ticket)

        op_sign = _norm(meta.get("operation_sign"))
        qty_bbl = _abs_qty(_safe_float(net_calc.get("net_stock_bbl")), _safe_float(net_calc.get("net_water_bbl")))
        if op_sign == "IN":
            groups[k]["fso_receipt_bbl"] += qty_bbl
            groups[k]["shuttle_discharge_bbl"] += _safe_float(meta.get("source_shuttle_discharge_bbl"))

        # Summary metrics (old OTR rules)
        if op_norm == "receipt":
            groups[k]["total_receipts_bbl"] += max(net_stock, 0.0)
        elif op_norm == "export":
            groups[k]["total_exports_bbl"] += abs(net_stock)
        elif op_norm == "stock opening":
            groups[k]["loss_gain_bbl"] += net_stock

        if net_water > 0:
            groups[k]["total_water_in_bbl"] += net_water
        elif net_water < 0:
            groups[k]["total_water_out_bbl"] += abs(net_water)

        # Total variance: skip export rows (same style as old)
        if op_norm != "export":
            if abs(variance) > 0:
                groups[k]["total_variance_bbl"] += variance
            else:
                groups[k]["total_variance_bbl"] += (abs(net_stock) - vessel_qty)

    # finalize net water
    for k in groups:
        groups[k]["net_water_bbl"] = groups[k]["total_water_in_bbl"] - groups[k]["total_water_out_bbl"]

    shuttle_fallback = _build_shuttle_discharge_fallback(
        list({(r.location_code, r.shuttle_number) for r in group_rows})
    )
    for k in groups:
        loc_code = str(groups[k].get("location_code") or "").strip()
        sh_num = str(groups[k].get("shuttle_number") or "").strip()
        if float(groups[k].get("shuttle_discharge_bbl") or 0.0) <= 0.0:
            groups[k]["shuttle_discharge_bbl"] = float(shuttle_fallback.get(f"{loc_code}|{sh_num}", 0.0))
        groups[k]["variance_bbl"] = float(groups[k]["fso_receipt_bbl"]) - float(groups[k]["shuttle_discharge_bbl"])

    return {
        "rows": list(groups.values()),
        "total_groups": total_groups,
        "page": page,
        "page_size": page_size,
        "has_more": total_groups > offset + page_size,
    }

# -------------------------
# Barge / Trip Tracking APIs
# -------------------------

@app.get("/convoy-tracker", response_model=ConvoyTrackerResponse)
def get_convoy_tracker(
    convoy_number: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    # View-only: reuse existing permission
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    convoy = clean_optional_text(convoy_number)

    if convoy is None:
        raise HTTPException(
            status_code=400,
            detail="convoy_number is required",
        )

    transactions = (
        db.query(OperationTransaction)
        .filter(
            OperationTransaction.convoy_number.ilike(convoy),
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
        .order_by(
            OperationTransaction.operation_date.asc(),
            OperationTransaction.id.asc(),
        )
        .all()
    )

    # Group by asset (barge)
    asset_map = {}

    for tx in transactions:
        asset_code = tx.primary_asset_code
        asset = get_asset_by_code(asset_code, db)

        if asset_code not in asset_map:
            asset_map[asset_code] = {
                "asset_code": asset_code,
                "asset_name": asset.asset_name if asset else "",
                "tickets": [],
            }

        op_type = get_operation_type_by_code(tx.operation_type_code, db)

        asset_map[asset_code]["tickets"].append(
            {
                "transaction_id": tx.id,
                "ticket_number": get_transaction_ticket_number(tx),
                "operation_type_code": tx.operation_type_code,
                "operation_type_name": op_type.operation_type_name if op_type else "",
                "operation_date": tx.operation_date,
                "origin_location_code": tx.origin_location_code,
                "origin_location_name": get_location_name_by_code(tx.origin_location_code, db),
                "destination_location_code": tx.destination_location_code,
                "destination_location_name": get_location_name_by_code(tx.destination_location_code, db),
                "status": tx.status,
            }
        )

    return {
        "convoy_number": convoy,
        "total_tickets": len(transactions),
        "assets": list(asset_map.values()),
    }


@app.get("/barge-tracking", response_model=ConvoyTrackerResponse)
def get_barge_tracking(
    convoy_number: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    return get_convoy_tracker(
        convoy_number=convoy_number,
        current_user=current_user,
        db=db,
    )

def ensure_barge_unload_comparison(
    db: Session,
    trip: Trip,
    asset_code: str,
    unload_tx: OperationTransaction,
    current_user: User,
    remarks: str | None = None,
):
    if not trip or not unload_tx:
        return None

    require_approved_transaction_for_tracking(unload_tx, "barge comparison")

    asset = str(asset_code or "").strip()
    if not asset:
        return None

    comparison_type = "LOAD_AFTER_vs_UNLOAD_BEFORE"

    # Latest LOAD for same trip+barge
    latest_load_event = (
        db.query(TripEvent)
        .filter(
            TripEvent.trip_id == trip.id,
            TripEvent.asset_code == asset,
            TripEvent.event_type.in_(["LOAD_1", "LOAD_2_TOPUP"]),
            TripEvent.operation_transaction_id.isnot(None),
        )
        .order_by(TripEvent.sequence_no.desc(), TripEvent.id.desc())
        .first()
    )

    if not latest_load_event or not latest_load_event.operation_transaction_id:
        return None

    left_tx = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == latest_load_event.operation_transaction_id)
        .first()
    )

    require_approved_transaction_for_tracking(left_tx, "barge comparison")

    existing = (
        db.query(TripComparison)
        .filter(
            TripComparison.trip_id == trip.id,
            TripComparison.comparison_type == comparison_type,
            TripComparison.left_transaction_id == left_tx.id,
            TripComparison.right_transaction_id == unload_tx.id,
        )
        .first()
    )
    if existing:
        return existing

    left_payload = load_multi_tank_payload(db, left_tx.id)
    right_payload = load_multi_tank_payload(db, unload_tx.id)

    # If these are not Multi-Tank tickets, we cannot auto-build the comparison JSON
    if not left_payload or not right_payload:
        return None

    summary_json, per_tank_json = build_multitank_comparison_json(
        left_tx=left_tx,
        right_tx=unload_tx,
        comparison_type=comparison_type,
        left_payload=left_payload,
        right_payload=right_payload,
    )

    created_by_display = get_current_user_display_name(current_user)

    new_cmp = TripComparison(
        trip_id=trip.id,
        comparison_type=comparison_type,
        left_transaction_id=left_tx.id,
        right_transaction_id=unload_tx.id,
        summary_json=summary_json,
        per_tank_json=per_tank_json,
        created_by=created_by_display,
        remarks=clean_optional_text(remarks) or "Auto-created on UNLOAD event tagging",
    )

    db.add(new_cmp)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Barge Tracking",
        action="Auto Create Barge Comparison",
        current_user=current_user,
        entity_type="TripComparison",
        entity_id=new_cmp.id,
        entity_label=f"{trip.convoy_number} | {asset} | {comparison_type}",
        ticket_number=get_transaction_ticket_number(left_tx),
        operation_number=left_tx.operation_number,
        remarks="Auto-created from trip event tagging",
        request_path="/trip-events",
        details={
            "convoy_number": trip.convoy_number,
            "trip_id": trip.id,
            "asset_code": asset,
            "comparison_type": comparison_type,
            "left_transaction_id": left_tx.id,
            "right_transaction_id": unload_tx.id,
        },
    )

    return new_cmp

@app.post("/trip-events", response_model=TripEventResponse)
def create_trip_event(
    request: TripEventCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    convoy = clean_optional_text(request.convoy_number)
    if convoy is None:
        raise HTTPException(status_code=400, detail="convoy_number is required")

    asset_code = clean_optional_text(request.asset_code)
    if asset_code is None:
        raise HTTPException(status_code=400, detail="asset_code is required")

    tx = None
    if request.operation_transaction_id is not None:
        tx = (
            db.query(OperationTransaction)
            .filter(OperationTransaction.id == request.operation_transaction_id)
            .first()
        )
        if not tx:
            raise HTTPException(status_code=404, detail="Operation transaction not found")

        if str(tx.primary_asset_code or "").strip().lower() != asset_code.lower():
            raise HTTPException(
                status_code=400,
                detail="asset_code does not match the operation ticket primary_asset_code",
            )

        # ✅ Approved-only rule for any event linked to a ticket
        require_approved_transaction_for_tracking(
            tx,
            "barge timeline event",
        )

        # Align convoy if needed
        if clean_optional_text(tx.convoy_number) is None:
            tx.convoy_number = convoy
            db.flush()
        elif str(tx.convoy_number).strip().lower() != convoy.lower():
            raise HTTPException(
                status_code=400,
                detail="Ticket convoy_number does not match request convoy_number",
            )

        # ✅ Idempotency: if the ticket already has an event, UPDATE it instead of INSERT
        existing_event_for_ticket = (
            db.query(TripEvent)
            .filter(TripEvent.operation_transaction_id == tx.id)
            .first()
        )
        if existing_event_for_ticket:
            existing_event_for_ticket.event_type = (
                clean_optional_text(request.event_type)
                or existing_event_for_ticket.event_type
            )
            existing_event_for_ticket.location_code = (
                clean_optional_text(request.location_code)
                or existing_event_for_ticket.location_code
            )
            existing_event_for_ticket.asset_code = asset_code
            existing_event_for_ticket.event_datetime = (
                request.event_datetime
                or existing_event_for_ticket.event_datetime
            )

            cleaned_remarks = clean_optional_text(request.remarks)
            if cleaned_remarks:
                existing_event_for_ticket.remarks = cleaned_remarks

            existing_event_for_ticket.updated_at = datetime.now()
            db.commit()
            db.refresh(existing_event_for_ticket)

            # ✅ If timeline is being corrected to UNLOAD, backfill comparison now
            if tx and str(existing_event_for_ticket.event_type or "").strip().upper() == "UNLOAD":
                trip_for_cmp = db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()
                if trip_for_cmp:
                    ensure_trip_not_closed(trip_for_cmp)
                    ensure_barge_unload_comparison(
                        db=db,
                        trip=trip_for_cmp,
                        asset_code=asset_code,
                        unload_tx=tx,
                        current_user=current_user,
                        remarks="Backfilled from Fix Timeline",
                    )
                    db.commit()

            return {
                "id": existing_event_for_ticket.id,
                "trip_id": existing_event_for_ticket.trip_id,
                "convoy_number": convoy,
                "event_type": existing_event_for_ticket.event_type,
                "location_code": existing_event_for_ticket.location_code,
                "asset_code": existing_event_for_ticket.asset_code,
                "operation_transaction_id": existing_event_for_ticket.operation_transaction_id,
                "sequence_no": existing_event_for_ticket.sequence_no,
                "event_datetime": existing_event_for_ticket.event_datetime,
                "created_by": existing_event_for_ticket.created_by,
                "remarks": existing_event_for_ticket.remarks,
                "created_at": existing_event_for_ticket.created_at,
                "updated_at": existing_event_for_ticket.updated_at,
            }

    # Ensure Trip exists
    trip = db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()
    created_by_display = get_current_user_display_name(current_user)

    if not trip:
        trip = Trip(
            convoy_number=convoy,
            primary_barge_asset_code=asset_code,
            status="OPEN",
            created_by=created_by_display,
            remarks=None,
        )
        db.add(trip)
        db.flush()

    ensure_trip_not_closed(trip)

    # Auto sequence if missing
    if request.sequence_no is None:
        max_seq = (
            db.query(func.max(TripEvent.sequence_no))
            .filter(TripEvent.trip_id == trip.id)
            .scalar()
        )
        sequence_no = (max_seq or 0) + 1
    else:
        sequence_no = int(request.sequence_no)

    event_type = clean_optional_text(request.event_type)
    if event_type is None:
        raise HTTPException(status_code=400, detail="event_type is required")

    # location_code must exist for ACK events (no ticket)
    location_code = clean_optional_text(request.location_code)
    if location_code is None and tx is not None:
        location_code = clean_optional_text(tx.origin_location_code)

    if location_code is None:
        raise HTTPException(
            status_code=400,
            detail="location_code is required when operation_transaction_id is not provided",
        )

    event_datetime = (
        request.event_datetime
        or (tx.operation_start_datetime if tx else None)
        or datetime.now()
    )

    op_tx_id = tx.id if tx else None

    new_event = TripEvent(
        trip_id=trip.id,
        event_type=event_type.upper(),
        location_code=location_code,
        asset_code=asset_code,
        operation_transaction_id=op_tx_id,
        sequence_no=sequence_no,
        event_datetime=event_datetime,
        created_by=created_by_display,
        remarks=clean_optional_text(request.remarks),
    )

    try:
        db.add(new_event)
        db.flush()
    except IntegrityError:
        # ✅ Safety net (race/duplicate click): fetch existing + update instead of 500
        db.rollback()

        if op_tx_id is not None:
            existing = (
                db.query(TripEvent)
                .filter(TripEvent.operation_transaction_id == op_tx_id)
                .first()
            )
            if existing:
                existing.event_type = event_type.upper()
                existing.location_code = location_code
                existing.asset_code = asset_code
                existing.event_datetime = event_datetime
                cleaned_remarks = clean_optional_text(request.remarks)
                if cleaned_remarks:
                    existing.remarks = cleaned_remarks
                existing.updated_at = datetime.now()
                db.commit()
                db.refresh(existing)

                return {
                    "id": existing.id,
                    "trip_id": existing.trip_id,
                    "convoy_number": convoy,
                    "event_type": existing.event_type,
                    "location_code": existing.location_code,
                    "asset_code": existing.asset_code,
                    "operation_transaction_id": existing.operation_transaction_id,
                    "sequence_no": existing.sequence_no,
                    "event_datetime": existing.event_datetime,
                    "created_by": existing.created_by,
                    "remarks": existing.remarks,
                    "created_at": existing.created_at,
                    "updated_at": existing.updated_at,
                }

        raise HTTPException(
            status_code=500,
            detail="Failed to create/update trip event due to duplicate operation_transaction_id",
        )

    create_audit_log(
        db=db,
        module_name="Barge Tracking",
        action="Create Trip Event",
        current_user=current_user,
        entity_type="TripEvent",
        entity_id=new_event.id,
        entity_label=f"{convoy} | {new_event.event_type} | {asset_code}",
        ticket_number=(get_transaction_ticket_number(tx) if tx else None),
        operation_number=(tx.operation_number if tx else None),
        remarks="Trip event created",
        request_path="/trip-events",
        details={
            "convoy_number": convoy,
            "trip_id": trip.id,
            "event_type": new_event.event_type,
            "asset_code": asset_code,
            "location_code": location_code,
            "operation_transaction_id": op_tx_id,
            "sequence_no": sequence_no,
        },
    )

    # ✅ If this newly created event is UNLOAD and linked to an approved ticket, create comparison
    if tx and str(new_event.event_type or "").strip().upper() == "UNLOAD":
        ensure_barge_unload_comparison(
            db=db,
            trip=trip,
            asset_code=asset_code,
            unload_tx=tx,
            current_user=current_user,
            remarks="Auto-created from trip event creation",
        )

    db.commit()
    db.refresh(new_event)

    return {
        "id": new_event.id,
        "trip_id": new_event.trip_id,
        "convoy_number": convoy,
        "event_type": new_event.event_type,
        "location_code": new_event.location_code,
        "asset_code": new_event.asset_code,
        "operation_transaction_id": new_event.operation_transaction_id,
        "sequence_no": new_event.sequence_no,
        "event_datetime": new_event.event_datetime,
        "created_by": new_event.created_by,
        "remarks": new_event.remarks,
        "created_at": new_event.created_at,
        "updated_at": new_event.updated_at,
    }

@app.get("/trips/by-convoy/{convoy_number}")
def get_trip_timeline_by_convoy(
    convoy_number: str,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    convoy = clean_optional_text(convoy_number)

    if convoy is None:
        raise HTTPException(
            status_code=400,
            detail="convoy_number is required",
        )

    trip = db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()

    if not trip:
        raise HTTPException(
            status_code=404,
            detail="Trip not found for this convoy number",
        )

    events = (
        db.query(TripEvent)
        .filter(TripEvent.trip_id == trip.id)
        .order_by(TripEvent.sequence_no.asc(), TripEvent.id.asc())
        .all()
    )

    comparisons = (
        db.query(TripComparison)
        .filter(TripComparison.trip_id == trip.id)
        .order_by(TripComparison.id.asc())
        .all()
    )

    # Enrich events with ticket numbers for UI convenience
    event_rows = []

    for ev in events:
        tx = db.query(OperationTransaction).filter(OperationTransaction.id == ev.operation_transaction_id).first()
        asset = get_asset_by_code(ev.asset_code, db)

        event_rows.append(
            {
                "id": ev.id,
                "trip_id": ev.trip_id,
                "convoy_number": convoy,
                "event_type": ev.event_type,
                "sequence_no": ev.sequence_no,
                "event_datetime": ev.event_datetime,
                "location_code": ev.location_code,
                "location_name": get_location_name_by_code(ev.location_code, db),
                "asset_code": ev.asset_code,
                "asset_name": asset.asset_name if asset else "",
                "operation_transaction_id": ev.operation_transaction_id,
                "ticket_number": get_transaction_ticket_number(tx) if tx else "",
                "ticket_status": tx.status if tx else "",
            }
        )

    comparison_rows = []
    did_backfill = False

    for cmp in comparisons:
        left_tx = (
            db.query(OperationTransaction)
            .filter(OperationTransaction.id == cmp.left_transaction_id)
            .first()
        )
        right_tx = (
            db.query(OperationTransaction)
            .filter(OperationTransaction.id == cmp.right_transaction_id)
            .first()
        )

        # ✅ Backfill missing JSON for older comparisons (fixes blank reports)
        if (
            cmp.summary_json is None or cmp.per_tank_json is None
        ) and left_tx and right_tx:
            left_payload = load_multi_tank_payload(db, left_tx.id)
            right_payload = load_multi_tank_payload(db, right_tx.id)

            if left_payload and right_payload:
                auto_summary, auto_per_tank = build_multitank_comparison_json(
                    left_tx=left_tx,
                    right_tx=right_tx,
                    comparison_type=cmp.comparison_type,
                    left_payload=left_payload,
                    right_payload=right_payload,
                )
                if cmp.summary_json is None:
                    cmp.summary_json = auto_summary
                if cmp.per_tank_json is None:
                    cmp.per_tank_json = auto_per_tank
                did_backfill = True

        asset_code = (left_tx.primary_asset_code if left_tx else "") or (
            right_tx.primary_asset_code if right_tx else ""
        )
        asset = get_asset_by_code(asset_code, db) if asset_code else None

        comparison_rows.append(
            {
                "id": cmp.id,
                "trip_id": cmp.trip_id,
                "convoy_number": convoy,
                "comparison_type": cmp.comparison_type,
                "asset_code": asset_code,
                "asset_name": asset.asset_name if asset else "",
                "left_transaction_id": cmp.left_transaction_id,
                "left_ticket_number": get_transaction_ticket_number(left_tx)
                if left_tx
                else "",
                "right_transaction_id": cmp.right_transaction_id,
                "right_ticket_number": get_transaction_ticket_number(right_tx)
                if right_tx
                else "",
                "summary_json": cmp.summary_json,
                "per_tank_json": cmp.per_tank_json,
                "created_by": cmp.created_by,
                "remarks": cmp.remarks,
                "created_at": cmp.created_at,
                "updated_at": cmp.updated_at,
            }
        )

    if did_backfill:
        db.commit()

    return {
        "trip": {
            "id": trip.id,
            "convoy_number": trip.convoy_number,
            "primary_barge_asset_code": trip.primary_barge_asset_code,
            "status": trip.status,
            "created_by": trip.created_by,
            "remarks": trip.remarks,
            "created_at": trip.created_at,
            "updated_at": trip.updated_at,
        },
        "events": event_rows,
        "comparisons": comparison_rows,
    }


import json


def load_multi_tank_payload(db: Session, transaction_id: int):
    row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "multi_tank_payload",
        )
        .first()
    )

    if not row or row.field_value is None:
        return None

    if isinstance(row.field_value, dict):
        return row.field_value

    try:
        return json.loads(str(row.field_value))
    except Exception:
        return None


def resolve_comparison_stages(comparison_type: str):
    """
    Decide which snapshot to compare from each ticket payload.
    MultiTank payload contains:
      payload["inputs"]["before"/"after"]
      payload["perTank"]["before"/"after"]
      payload["calculated"]["before"/"after"]
    """
    t = (comparison_type or "").upper()

    # Defaults
    left_stage = "after"
    right_stage = "before"

    if "UNLOAD_BEFORE_VS_UNLOAD_AFTER" in t:
        left_stage = "before"
        right_stage = "after"

    if "LOAD_PREV_VS_LOAD_CURRENT" in t:
        left_stage = "after"
        right_stage = "before"

    # Main case
    if "LOAD_AFTER_VS_UNLOAD_BEFORE" in t:
        left_stage = "after"
        right_stage = "before"

    return left_stage, right_stage


def get_payload_stage(payload: dict, stage_key: str):
    inputs = (payload.get("inputs") or {}).get(stage_key) or {}
    per_tank = (payload.get("perTank") or {}).get(stage_key) or {}
    totals = (payload.get("calculated") or {}).get(stage_key) or {}

    return {
        "inputs": inputs,
        "per_tank": per_tank,
        "totals": totals,
    }


def build_multitank_seal_checks(left_payload: dict, right_payload: dict):
    """
    Multi-tank seal checks for barge comparison.
    We compare:
      - Sender: seals.after.temporary
      - Receiver: seals.before.temporary

    Required keys (standard): C1, C2, M1, M2
    """

    def norm(v):
        return str(v or "").strip()

    left_temp = (
        (((left_payload.get("seals") or {}).get("after") or {}).get("temporary") or {})
    )
    right_temp = (
        (((right_payload.get("seals") or {}).get("before") or {}).get("temporary") or {})
    )

    seal_fields = [
        ("C1", "sealC1"),
        ("C2", "sealC2"),
        ("M1", "sealM1"),
        ("M2", "sealM2"),
    ]

    checks = []
    for seal_name, key in seal_fields:
        sender_val = norm(left_temp.get(key))
        receiver_val = norm(right_temp.get(key))

        status = "MATCH"
        if sender_val == "" and receiver_val == "":
            status = "MISSING_BOTH"
        elif sender_val == "":
            status = "MISSING_SENDER"
        elif receiver_val == "":
            status = "MISSING_RECEIVER"
        elif sender_val != receiver_val:
            status = "MISMATCH"

        checks.append(
            {
                "seal_name": seal_name,
                "sender": sender_val,
                "receiver": receiver_val,
                "status": status,
            }
        )

    seal_mismatch = any(
        c["status"] in ("MISMATCH", "MISSING_SENDER", "MISSING_RECEIVER")
        for c in checks
    )
    return checks, seal_mismatch


def build_multitank_comparison_json(
    left_tx: OperationTransaction,
    right_tx: OperationTransaction,
    comparison_type: str,
    left_payload: dict,
    right_payload: dict,
):
    left_stage, right_stage = resolve_comparison_stages(comparison_type)

    l = get_payload_stage(left_payload, left_stage)
    r = get_payload_stage(right_payload, right_stage)

    # Tank list (union)
    tank_ids = set()
    tank_ids.update((left_payload.get("meta") or {}).get("tankIds") or [])
    tank_ids.update((right_payload.get("meta") or {}).get("tankIds") or [])
    tank_ids.update(list((l["per_tank"] or {}).keys()))
    tank_ids.update(list((r["per_tank"] or {}).keys()))
    tank_ids = [str(x) for x in tank_ids if str(x).strip()]
    tank_ids.sort()

    per_tank_rows = []
    for tid in tank_ids:
        lp = (l["per_tank"] or {}).get(tid) or {}
        rp = (r["per_tank"] or {}).get(tid) or {}

        per_tank_rows.append(
            {
                "tank_id": tid,
                "left": {
                    "total_dip": lp.get("totalDip", 0),
                    "water_dip": lp.get("waterDip", 0),
                    "tov": lp.get("tovCorrected", 0),
                    "fw": lp.get("fwCorrected", 0),
                },
                "right": {
                    "total_dip": rp.get("totalDip", 0),
                    "water_dip": rp.get("waterDip", 0),
                    "tov": rp.get("tovCorrected", 0),
                    "fw": rp.get("fwCorrected", 0),
                },
                "delta": {
                    "tov": (lp.get("tovCorrected", 0) or 0)
                    - (rp.get("tovCorrected", 0) or 0),
                    "fw": (lp.get("fwCorrected", 0) or 0)
                    - (rp.get("fwCorrected", 0) or 0),
                },
            }
        )

    def pick_totals(obj: dict):
        # totals already computed by frontend and stored in payload
        keys = [
            "TOV",
            "FW",
            "GOV",
            "GSV",
            "BSW",
            "NSV",
            "LT",
            "MT",
            "API60",
            "VCF",
            "ltFactor",
            "table11Method",
        ]
        return {k: obj.get(k) for k in keys if k in obj}

    left_totals = pick_totals(l["totals"] or {})
    right_totals = pick_totals(r["totals"] or {})

    def n(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    delta_totals = {}
    for k in ["TOV", "FW", "GOV", "GSV", "BSW", "NSV", "LT", "MT"]:
        delta_totals[k] = n(left_totals.get(k)) - n(right_totals.get(k))

    seal_checks, seal_mismatch = build_multitank_seal_checks(
        left_payload, right_payload
    )

    summary_json = {
        "comparison_type": comparison_type,
        "asset_code": left_tx.primary_asset_code,
        "seal_checks": seal_checks,
        "seal_mismatch": seal_mismatch,
        "left": {
            "transaction_id": left_tx.id,
            "ticket_number": get_transaction_ticket_number(left_tx),
            "stage": left_stage,
            "operation_date": (
                str(left_tx.operation_date) if left_tx.operation_date else ""
            ),
            "location_code": left_tx.origin_location_code or "",
            "inputs": l["inputs"],
            "totals": left_totals,
        },
        "right": {
            "transaction_id": right_tx.id,
            "ticket_number": get_transaction_ticket_number(right_tx),
            "stage": right_stage,
            "operation_date": (
                str(right_tx.operation_date) if right_tx.operation_date else ""
            ),
            "location_code": right_tx.origin_location_code or "",
            "inputs": r["inputs"],
            "totals": right_totals,
        },
        "delta": {"totals": delta_totals},
        "units": {
            "dip": ((left_payload.get("meta") or {}).get("inputXUnit") or "mm"),
            "volume": ((left_payload.get("meta") or {}).get("outputUnit") or ""),
        },
    }

    per_tank_json = {"tanks": per_tank_rows}

    return summary_json, per_tank_json


@app.post("/trip-comparisons", response_model=TripComparisonResponse)
def create_trip_comparison(
    request: TripComparisonCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    # For now, treat as operational action
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    convoy = clean_optional_text(request.convoy_number)

    if convoy is None:
        raise HTTPException(status_code=400, detail="convoy_number is required")

    trip = db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()

    created_by_display = get_current_user_display_name(current_user)

    if not trip:
        trip = Trip(
            convoy_number=convoy,
            primary_barge_asset_code=None,
            status="OPEN",
            created_by=created_by_display,
            remarks=None,
        )
        db.add(trip)
        db.flush()

    ensure_trip_not_closed(trip)

    left_tx = db.query(OperationTransaction).filter(OperationTransaction.id == request.left_transaction_id).first()
    right_tx = db.query(OperationTransaction).filter(OperationTransaction.id == request.right_transaction_id).first()

    if not left_tx or not right_tx:
        raise HTTPException(status_code=404, detail="Left or Right transaction not found")

    require_approved_transaction_for_tracking(
        left_tx,
        "barge sender/receiver comparison",
    )

    require_approved_transaction_for_tracking(
        right_tx,
        "barge sender/receiver comparison",
    )

    # Align ticket convoy if missing
    if clean_optional_text(left_tx.convoy_number) is None:
        left_tx.convoy_number = convoy
    if clean_optional_text(right_tx.convoy_number) is None:
        right_tx.convoy_number = convoy

    # Reject mismatch
    if str(left_tx.convoy_number).strip().lower() != convoy.lower() or str(right_tx.convoy_number).strip().lower() != convoy.lower():
        raise HTTPException(status_code=400, detail="Both tickets must belong to the same convoy_number")

    comparison_type = clean_optional_text(request.comparison_type)
    if comparison_type is None:
        raise HTTPException(status_code=400, detail="comparison_type is required")

    # Auto-build comparison JSON if missing
    summary_json = request.summary_json
    per_tank_json = request.per_tank_json

    left_payload = None
    right_payload = None

    if summary_json is None or per_tank_json is None:
        left_payload = load_multi_tank_payload(db, left_tx.id)
        right_payload = load_multi_tank_payload(db, right_tx.id)

        if left_payload and right_payload:
            auto_summary, auto_per_tank = build_multitank_comparison_json(
                left_tx=left_tx,
                right_tx=right_tx,
                comparison_type=comparison_type,
                left_payload=left_payload,
                right_payload=right_payload,
            )
            if summary_json is None:
                summary_json = auto_summary
            if per_tank_json is None:
                per_tank_json = auto_per_tank

    # ✅ Prevent empty comparison records (this is why your report shows blanks)
    if summary_json is None or per_tank_json is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Unable to auto-build comparison data. "
                "Ensure BOTH tickets are Multi-Tank tickets and contain field_code 'multi_tank_payload'. "
                f"left_ticket_id={left_tx.id} has_payload={bool(left_payload)} | "
                f"right_ticket_id={right_tx.id} has_payload={bool(right_payload)}"
            ),
        )

    new_cmp = TripComparison(
        trip_id=trip.id,
        comparison_type=comparison_type,
        left_transaction_id=left_tx.id,
        right_transaction_id=right_tx.id,
        summary_json=summary_json,
        per_tank_json=per_tank_json,
        created_by=created_by_display,
        remarks=clean_optional_text(request.remarks),
    )

    db.add(new_cmp)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Barge Tracking",
        action="Create Barge Comparison",
        current_user=current_user,
        entity_type="TripComparison",
        entity_id=new_cmp.id,
        entity_label=f"{convoy} | {comparison_type}",
        ticket_number=get_transaction_ticket_number(left_tx),
        operation_number=left_tx.operation_number,
        remarks="Barge comparison created",
        request_path="/trip-comparisons",
        details={
            "convoy_number": convoy,
            "trip_id": trip.id,
            "comparison_type": comparison_type,
            "left_transaction_id": left_tx.id,
            "left_ticket_number": get_transaction_ticket_number(left_tx),
            "right_transaction_id": right_tx.id,
            "right_ticket_number": get_transaction_ticket_number(right_tx),
        },
    )

    db.commit()
    db.refresh(new_cmp)

    return {
        "id": new_cmp.id,
        "trip_id": new_cmp.trip_id,
        "convoy_number": convoy,
        "comparison_type": new_cmp.comparison_type,
        "left_transaction_id": new_cmp.left_transaction_id,
        "right_transaction_id": new_cmp.right_transaction_id,
        "summary_json": new_cmp.summary_json,
        "per_tank_json": new_cmp.per_tank_json,
        "created_by": new_cmp.created_by,
        "remarks": new_cmp.remarks,
        "created_at": new_cmp.created_at,
        "updated_at": new_cmp.updated_at,
    }

# -------------------------
# Trip Close / Reopen + Lock enforcement
# -------------------------

class TripStatusUpdateRequest(BaseModel):
    remarks: str | None = None


def get_trip_by_convoy_or_none(db: Session, convoy_number: str | None):
    convoy = clean_optional_text(convoy_number)
    if convoy is None:
        return None
    return db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()


def ensure_trip_not_closed(trip: Trip | None):
    if not trip:
        return
    if str(trip.status or "").strip().upper() == "CLOSED":
        raise HTTPException(
            status_code=400,
            detail="Trip is CLOSED for this convoy. Reopen the trip to continue.",
        )

class ShuttleVoyageStatusUpdateRequest(BaseModel):
    location_code: str
    shuttle_number: str
    shuttle_asset_code: str
    remarks: str | None = None


def get_shuttle_voyage_by_key(db: Session, location_code: str, shuttle_number: str, shuttle_asset_code: str):
    lc = clean_optional_text(location_code)
    sn = clean_optional_text(shuttle_number)
    ac = clean_optional_text(shuttle_asset_code)

    if not lc or not sn or not ac:
        return None

    return (
        db.query(ShuttleVoyage)
        .filter(
            ShuttleVoyage.location_code.ilike(lc),
            ShuttleVoyage.shuttle_number.ilike(sn),
            ShuttleVoyage.shuttle_asset_code.ilike(ac),
        )
        .first()
    )


def ensure_shuttle_voyage_not_closed(voyage: ShuttleVoyage | None):
    if not voyage:
        return
    if str(voyage.status or "").strip().upper() == "CLOSED":
        raise HTTPException(
            status_code=400,
            detail="Shuttle voyage is CLOSED for this key. Reopen the voyage to continue.",
        )


def get_or_create_shuttle_voyage(
    db: Session,
    location_code: str,
    shuttle_number: str,
    shuttle_asset_code: str,
    current_user: User,
):
    voyage = get_shuttle_voyage_by_key(db, location_code, shuttle_number, shuttle_asset_code)

    if voyage:
        return voyage

    created_by_display = get_current_user_label(current_user)

    voyage = ShuttleVoyage(
        location_code=str(location_code).strip(),
        shuttle_number=str(shuttle_number).strip(),
        shuttle_asset_code=str(shuttle_asset_code).strip(),
        status="OPEN",
        created_by=created_by_display,
        remarks=None,
    )
    db.add(voyage)
    db.flush()
    return voyage


def get_shuttle_payload_for_transaction(db: Session, transaction_id: int):
    row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "shuttle_payload",
        )
        .first()
    )
    if not row:
        return None
    if isinstance(row.field_value, dict):
        return row.field_value
    return None


def get_fso_voyage_by_key(db: Session, location_code: str, shuttle_number: str, fso_asset_code: str):
    lc = clean_optional_text(location_code)
    sn = clean_optional_text(shuttle_number)
    ac = clean_optional_text(fso_asset_code)

    if not lc or not sn or not ac:
        return None

    return (
        db.query(FSOVoyage)
        .filter(
            FSOVoyage.location_code.ilike(lc),
            FSOVoyage.shuttle_number.ilike(sn),
            FSOVoyage.fso_asset_code.ilike(ac),
        )
        .first()
    )


def ensure_fso_voyage_not_closed(voyage: FSOVoyage | None):
    if not voyage:
        return
    if str(voyage.status or "").strip().upper() == "CLOSED":
        raise HTTPException(
            status_code=400,
            detail="FSO voyage is CLOSED for this key. Reopen the voyage to continue.",
        )


def get_or_create_fso_voyage(
    db: Session,
    location_code: str,
    shuttle_number: str,
    fso_asset_code: str,
    current_user: User,
):
    voyage = get_fso_voyage_by_key(db, location_code, shuttle_number, fso_asset_code)
    if voyage:
        return voyage

    created_by_display = get_current_user_label(current_user)

    voyage = FSOVoyage(
        location_code=str(location_code).strip(),
        shuttle_number=str(shuttle_number).strip(),
        fso_asset_code=str(fso_asset_code).strip(),
        status="OPEN",
        created_by=created_by_display,
        remarks=None,
    )
    db.add(voyage)
    db.flush()
    return voyage


def get_fso_payload_for_transaction(db: Session, transaction_id: int):
    row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "fso_payload",
        )
        .first()
    )
    if not row:
        return None
    if isinstance(row.field_value, dict):
        return row.field_value
    return None

def require_barge_tracking_ready_for_closure(
    trip: Trip,
    db: Session,
):
    approved_transactions = (
        db.query(OperationTransaction)
        .filter(
            OperationTransaction.convoy_number.ilike(trip.convoy_number),
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
            OperationTransaction.primary_asset_type_code.ilike("BARGE"),
        )
        .all()
    )

    approved_asset_codes = {
        str(tx.primary_asset_code or "").strip()
        for tx in approved_transactions
        if str(tx.primary_asset_code or "").strip()
    }

    if len(approved_asset_codes) == 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot close barge movement because no Approved barge tickets were found.",
        )

    if len(approved_transactions) < 2:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot close barge movement before both sender and receiver "
                "transactions are Approved."
            ),
        )

    comparisons = (
        db.query(TripComparison)
        .filter(TripComparison.trip_id == trip.id)
        .all()
    )

    compared_asset_codes = set()

    for comparison in comparisons:
        # Only sender/receiver comparison should qualify for closure
        if str(comparison.comparison_type or "").strip() != "LOAD_AFTER_vs_UNLOAD_BEFORE":
            continue

        left_tx = (
            db.query(OperationTransaction)
            .filter(OperationTransaction.id == comparison.left_transaction_id)
            .first()
        )

        right_tx = (
            db.query(OperationTransaction)
            .filter(OperationTransaction.id == comparison.right_transaction_id)
            .first()
        )

        if not left_tx or not right_tx:
            continue

        if left_tx.status != APPROVED_TRANSACTION_STATUS:
            continue

        if right_tx.status != APPROVED_TRANSACTION_STATUS:
            continue

        if str(left_tx.primary_asset_code or "").strip().lower() != str(
            right_tx.primary_asset_code or ""
        ).strip().lower():
            continue

        asset_code = str(left_tx.primary_asset_code or "").strip()

        if asset_code:
            compared_asset_codes.add(asset_code)

    pending_asset_codes = sorted(
        list(approved_asset_codes - compared_asset_codes)
    )

    if pending_asset_codes:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot close convoy because comparison is pending for barge(s): "
                + ", ".join(pending_asset_codes)
            ),
        )

@app.post("/trips/{trip_id}/close")
def close_trip(
    trip_id: int,
    request: TripStatusUpdateRequest | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Create Operation Entry", db)

    trip = db.query(Trip).filter(Trip.id == trip_id).first()

    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")

    if str(trip.status or "").upper() == "CLOSED":
        return {
            "message": "Barge movement already CLOSED",
            "trip_id": trip.id,
            "status": trip.status,
        }

    require_barge_tracking_ready_for_closure(trip, db)

    before_status = trip.status
    trip.status = "CLOSED"
    trip.updated_at = datetime.now()

    closure_remarks = clean_optional_text(request.remarks) if request else None

    if closure_remarks:
        trip.remarks = (
            f"{trip.remarks or ''}\n"
            f"[Barge Movement Closed] {closure_remarks}"
        ).strip()

    create_audit_log(
        db=db,
        module_name="Barge Tracking",
        action="Close Barge Movement",
        current_user=current_user,
        entity_type="Trip",
        entity_id=trip.id,
        entity_label=trip.convoy_number,
        remarks="Barge movement closed after comparison review",
        request_path=f"/trips/{trip_id}/close",
        details={
            "convoy_number": trip.convoy_number,
            "before_status": before_status,
            "after_status": trip.status,
            "closure_remarks": closure_remarks,
        },
    )

    db.commit()
    db.refresh(trip)

    return {
        "message": "Barge movement CLOSED",
        "trip_id": trip.id,
        "status": trip.status,
    }

@app.post("/trips/{trip_id}/reopen")
def reopen_trip(
    trip_id: int,
    request: TripStatusUpdateRequest | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(current_user, "Create Operation Entry", db)

    trip = db.query(Trip).filter(Trip.id == trip_id).first()

    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")

    if str(trip.status or "").upper() == "OPEN":
        return {
            "message": "Barge movement already OPEN",
            "trip_id": trip.id,
            "status": trip.status,
        }

    before_status = trip.status
    trip.status = "OPEN"
    trip.updated_at = datetime.now()

    reopen_remarks = clean_optional_text(request.remarks) if request else None

    if reopen_remarks:
        trip.remarks = (
            f"{trip.remarks or ''}\n"
            f"[Barge Movement Reopened] {reopen_remarks}"
        ).strip()

    create_audit_log(
        db=db,
        module_name="Barge Tracking",
        action="Reopen Barge Movement",
        current_user=current_user,
        entity_type="Trip",
        entity_id=trip.id,
        entity_label=trip.convoy_number,
        remarks="Barge movement reopened manually",
        request_path=f"/trips/{trip_id}/reopen",
        details={
            "convoy_number": trip.convoy_number,
            "before_status": before_status,
            "after_status": trip.status,
            "reopen_remarks": reopen_remarks,
        },
    )

    db.commit()
    db.refresh(trip)

    return {
        "message": "Barge movement OPEN",
        "trip_id": trip.id,
        "status": trip.status,
    }

# -------------------------
# Tank Stock Ledger Creation Helpers
# -------------------------

def safe_float(value, default_value: float = 0):
    try:
        if value is None:
            return default_value

        if str(value).strip() == "":
            return default_value

        return float(value)
    except (TypeError, ValueError):
        return default_value


def get_tank_gauging_payload_for_transaction(
    db: Session,
    transaction_id: int,
):
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == "tank_gauging_payload",
        )
        .first()
    )

    if payload_row is None or payload_row.field_value is None:
        return None

    if not isinstance(payload_row.field_value, dict):
        return None

    return payload_row.field_value


def parse_payload_gauging_datetime(payload: dict):
    inputs = payload.get("inputs") or {}

    gauging_date = clean_optional_text(inputs.get("gaugingDate"))
    gauging_time = clean_optional_text(inputs.get("gaugingTime"))

    if not gauging_date or not gauging_time:
        return None

    try:
        return datetime.fromisoformat(f"{gauging_date}T{gauging_time}")
    except ValueError:
        return None


def resolve_transaction_datetime_for_accounting_day(
    transaction: OperationTransaction,
    payload: dict,
):
    if transaction.operation_start_datetime is not None:
        return transaction.operation_start_datetime

    payload_datetime = parse_payload_gauging_datetime(payload)

    if payload_datetime is not None:
        return payload_datetime

    raise HTTPException(
        status_code=400,
        detail=(
            "Operation Start Date/Time or Tank Gauging Date/Time is required "
            "to calculate the Location Accounting Day."
        ),
    )


def calculate_accounting_window_from_setting(
    setting: LocationAccountingDaySetting,
    transaction_datetime: datetime,
):
    transaction_date = transaction_datetime.date()
    transaction_time = transaction_datetime.time()

    start_time = setting.day_start_time
    end_time = setting.day_end_time

    # Most hydrocarbon operational days are overnight:
    # Example 06:01 today to 06:00 next day.
    is_overnight_window = end_time < start_time

    if is_overnight_window:
        if transaction_time >= start_time:
            accounting_date = transaction_date
        else:
            accounting_date = transaction_date - timedelta(days=1)

        accounting_day_start = datetime.combine(accounting_date, start_time)
        accounting_day_end = datetime.combine(
            accounting_date + timedelta(days=1),
            end_time,
        )

    else:
        # Supports same-calendar-day windows if ever required.
        if transaction_time >= start_time:
            accounting_date = transaction_date
        else:
            accounting_date = transaction_date - timedelta(days=1)

        accounting_day_start = datetime.combine(accounting_date, start_time)
        accounting_day_end = datetime.combine(accounting_date, end_time)

    return {
        "accounting_date": accounting_date,
        "accounting_day_start": accounting_day_start,
        "accounting_day_end": accounting_day_end,
    }


def get_location_accounting_day_for_transaction(
    db: Session,
    location_code: str,
    transaction_datetime: datetime,
):
    cleaned_location_code = clean_optional_text(location_code)

    if not cleaned_location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required to calculate accounting day",
        )

    active_settings = (
        db.query(LocationAccountingDaySetting)
        .filter(
            LocationAccountingDaySetting.location_code.ilike(cleaned_location_code),
            LocationAccountingDaySetting.status == "Active",
        )
        .order_by(
            LocationAccountingDaySetting.effective_from.desc(),
            LocationAccountingDaySetting.id.desc(),
        )
        .all()
    )

    if len(active_settings) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No Active Location Accounting Day Setting found for "
                f"{cleaned_location_code}. Configure it before approving "
                "Tank Gauging tickets."
            ),
        )

    matching_options = []

    for setting in active_settings:
        window = calculate_accounting_window_from_setting(
            setting=setting,
            transaction_datetime=transaction_datetime,
        )

        accounting_date = window["accounting_date"]

        effective_to = setting.effective_to or date(9999, 12, 31)

        if setting.effective_from <= accounting_date <= effective_to:
            if (
                window["accounting_day_start"]
                <= transaction_datetime
                <= window["accounting_day_end"]
            ):
                matching_options.append(
                    {
                        "setting": setting,
                        "window": window,
                    }
                )

    if len(matching_options) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No effective Location Accounting Day Setting matched this "
                "transaction date/time. Check Effective From/To settings."
            ),
        )

    selected = matching_options[0]
    selected_setting = selected["setting"]
    selected_window = selected["window"]

    return {
        "setting_id": selected_setting.id,
        "accounting_date": selected_window["accounting_date"],
        "accounting_day_start": selected_window["accounting_day_start"],
        "accounting_day_end": selected_window["accounting_day_end"],
    }

def get_ledger_sort_datetime(ledger: TankStockLedger):
    if ledger.accounting_day_start is not None:
        return ledger.accounting_day_start

    if ledger.operation_date is not None:
        return datetime.combine(ledger.operation_date, datetime_time(0, 0))

    return datetime.min


def get_previous_active_ledger_row(
    db: Session,
    location_code: str,
    tank_asset_code: str,
    product_name: str | None,
    transaction_datetime: datetime,
    exclude_ledger_id: int | None = None,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.status == "Active",
        TankStockLedger.location_code.ilike(location_code),
        TankStockLedger.tank_asset_code.ilike(tank_asset_code),
    )

    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_product_name:
        query = query.filter(TankStockLedger.product_name.ilike(cleaned_product_name))
    else:
        query = query.filter(TankStockLedger.product_name == None)

    if exclude_ledger_id is not None:
        query = query.filter(TankStockLedger.id != exclude_ledger_id)

    candidate_rows = query.all()

    previous_rows = []

    for row in candidate_rows:
        row_datetime = row.accounting_day_start

        # Prefer actual operation timestamp if stored in source payload.
        try:
            payload = row.source_payload or {}
            payload_inputs = payload.get("inputs") or {}
            gauging_date = clean_optional_text(payload_inputs.get("gaugingDate"))
            gauging_time = clean_optional_text(payload_inputs.get("gaugingTime"))

            if gauging_date and gauging_time:
                row_datetime = datetime.fromisoformat(
                    f"{gauging_date}T{gauging_time}"
                )
        except Exception:
            row_datetime = None

        if row_datetime is None:
            row_datetime = datetime.combine(row.operation_date, datetime_time(0, 0))

        if row_datetime < transaction_datetime:
            previous_rows.append((row_datetime, row.id, row))

    if not previous_rows:
        return None

    previous_rows.sort(key=lambda item: (item[0], item[1]))

    return previous_rows[-1][2]


def calculate_stock_movement_from_snapshot(
    operation_sign: str,
    current_gsv_bbl: float,
    current_nsv_bbl: float,
    current_lt: float,
    current_mt: float,
    previous_ledger: TankStockLedger | None,
):
    sign = str(operation_sign or "").upper()

    previous_gsv_bbl = 0
    previous_nsv_bbl = 0
    previous_lt = 0
    previous_mt = 0

    if previous_ledger is not None:
        previous_gsv_bbl = safe_float(
            previous_ledger.stock_gsv_bbl
            if previous_ledger.stock_gsv_bbl is not None
            else previous_ledger.running_balance_gsv_bbl
        )
        previous_nsv_bbl = safe_float(
            previous_ledger.stock_nsv_bbl
            if previous_ledger.stock_nsv_bbl is not None
            else previous_ledger.running_balance_nsv_bbl
        )
        previous_lt = safe_float(
            previous_ledger.stock_lt
            if previous_ledger.stock_lt is not None
            else previous_ledger.running_balance_lt
        )
        previous_mt = safe_float(
            previous_ledger.stock_mt
            if previous_ledger.stock_mt is not None
            else previous_ledger.running_balance_mt
        )

    if sign == "SET":
        movement_gsv_bbl = current_gsv_bbl
        movement_nsv_bbl = current_nsv_bbl
        movement_lt = current_lt
        movement_mt = current_mt

    elif sign == "IN":
        movement_gsv_bbl = max(current_gsv_bbl - previous_gsv_bbl, 0)
        movement_nsv_bbl = max(current_nsv_bbl - previous_nsv_bbl, 0)
        movement_lt = max(current_lt - previous_lt, 0)
        movement_mt = max(current_mt - previous_mt, 0)

    elif sign == "OUT":
        movement_gsv_bbl = max(previous_gsv_bbl - current_gsv_bbl, 0)
        movement_nsv_bbl = max(previous_nsv_bbl - current_nsv_bbl, 0)
        movement_lt = max(previous_lt - current_lt, 0)
        movement_mt = max(previous_mt - current_mt, 0)

    elif sign == "NEUTRAL":
        movement_gsv_bbl = 0
        movement_nsv_bbl = 0
        movement_lt = 0
        movement_mt = 0

    else:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Tank Operation Sign: {operation_sign}",
        )

    return {
        "previous_gsv_bbl": previous_gsv_bbl,
        "previous_nsv_bbl": previous_nsv_bbl,
        "previous_lt": previous_lt,
        "previous_mt": previous_mt,
        "movement_gsv_bbl": movement_gsv_bbl,
        "movement_nsv_bbl": movement_nsv_bbl,
        "movement_lt": movement_lt,
        "movement_mt": movement_mt,
    }

def is_tank_gauging_transaction(
    db: Session,
    transaction: OperationTransaction,
):
    if transaction.operation_template_id is None:
        return False

    template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.id == transaction.operation_template_id)
        .first()
    )

    if not template:
        return False

    entry_layout_type = str(template.entry_layout_type or "").strip()
    calculation_engine = str(template.calculation_engine or "").strip()

    if entry_layout_type == "Tank Gauging":
        return True

    if calculation_engine == "Tank Quantity":
        return True

    payload = get_tank_gauging_payload_for_transaction(
        db=db,
        transaction_id=transaction.id,
    )

    return payload is not None


def rebuild_tank_stock_running_balances(
    db: Session,
    location_code: str,
    tank_asset_code: str,
    product_name: str | None,
):
    query = db.query(TankStockLedger).filter(
        TankStockLedger.location_code.ilike(location_code),
        TankStockLedger.tank_asset_code.ilike(tank_asset_code),
        TankStockLedger.status == "Active",
    )

    cleaned_product_name = clean_optional_text(product_name)

    if cleaned_product_name:
        query = query.filter(
            TankStockLedger.product_name.ilike(cleaned_product_name)
        )
    else:
        query = query.filter(TankStockLedger.product_name == None)

    ledger_rows = query.all()

    sortable_rows = []

    for row in ledger_rows:
        row_datetime = row.accounting_day_start

        try:
            payload = row.source_payload or {}
            payload_inputs = payload.get("inputs") or {}
            gauging_date = clean_optional_text(payload_inputs.get("gaugingDate"))
            gauging_time = clean_optional_text(payload_inputs.get("gaugingTime"))

            if gauging_date and gauging_time:
                row_datetime = datetime.fromisoformat(
                    f"{gauging_date}T{gauging_time}"
                )
        except Exception:
            row_datetime = None

        if row_datetime is None:
            row_datetime = datetime.combine(row.operation_date, datetime_time(0, 0))

        sortable_rows.append((row_datetime, row.id, row))

    sortable_rows.sort(key=lambda item: (item[0], item[1]))

    previous_row = None

    for row_datetime, _row_id, row in sortable_rows:
        # Backfill accounting day fields for old ledger rows created before
        # Location Accounting Day Settings were connected to the ledger.
        if (
            row.accounting_date is None
            or row.accounting_day_start is None
            or row.accounting_day_end is None
            or row.accounting_day_setting_id is None
        ):
            try:
                payload = row.source_payload or {}

                transaction_datetime = resolve_transaction_datetime_for_accounting_day(
                    transaction=db.query(OperationTransaction)
                    .filter(OperationTransaction.id == row.transaction_id)
                    .first(),
                    payload=payload,
                )

                accounting_day = get_location_accounting_day_for_transaction(
                    db=db,
                    location_code=row.location_code,
                    transaction_datetime=transaction_datetime,
                )

                row.accounting_date = accounting_day["accounting_date"]
                row.accounting_day_start = accounting_day["accounting_day_start"]
                row.accounting_day_end = accounting_day["accounting_day_end"]
                row.accounting_day_setting_id = accounting_day["setting_id"]

            except Exception:
                # Keep rebuild running for other rows.
                # Validation will still show rows that could not be backfilled.
                pass
        current_gsv_bbl = safe_float(row.stock_gsv_bbl)
        current_nsv_bbl = safe_float(row.stock_nsv_bbl)
        current_lt = safe_float(row.stock_lt)
        current_mt = safe_float(row.stock_mt)

        # Backward compatibility for old rows before stock_* columns existed.
        if current_gsv_bbl == 0 and current_nsv_bbl == 0:
            current_gsv_bbl = safe_float(row.running_balance_gsv_bbl)
            current_nsv_bbl = safe_float(row.running_balance_nsv_bbl)
            current_lt = safe_float(row.running_balance_lt)
            current_mt = safe_float(row.running_balance_mt)

            row.stock_gsv_bbl = current_gsv_bbl
            row.stock_nsv_bbl = current_nsv_bbl
            row.stock_lt = current_lt
            row.stock_mt = current_mt

        movement = calculate_stock_movement_from_snapshot(
            operation_sign=row.tank_operation_sign,
            current_gsv_bbl=current_gsv_bbl,
            current_nsv_bbl=current_nsv_bbl,
            current_lt=current_lt,
            current_mt=current_mt,
            previous_ledger=previous_row,
        )

        row.previous_stock_gsv_bbl = movement["previous_gsv_bbl"]
        row.previous_stock_nsv_bbl = movement["previous_nsv_bbl"]
        row.previous_stock_lt = movement["previous_lt"]
        row.previous_stock_mt = movement["previous_mt"]

        row.movement_gsv_bbl = movement["movement_gsv_bbl"]
        row.movement_nsv_bbl = movement["movement_nsv_bbl"]
        row.movement_lt = movement["movement_lt"]
        row.movement_mt = movement["movement_mt"]

        # Running balance is the current stock snapshot after the operation.
        row.running_balance_gsv_bbl = current_gsv_bbl
        row.running_balance_nsv_bbl = current_nsv_bbl
        row.running_balance_lt = current_lt
        row.running_balance_mt = current_mt

        row.updated_at = datetime.now()

        previous_row = row

    db.flush()

def create_tank_stock_ledger_from_approved_transaction(
    db: Session,
    transaction: OperationTransaction,
    current_user: User,
):
    if transaction.status != "Approved":
        return None

    if not is_tank_gauging_transaction(db, transaction):
        return None

    existing_ledger = (
        db.query(TankStockLedger)
        .filter(TankStockLedger.transaction_id == transaction.id)
        .first()
    )

    if existing_ledger:
        return existing_ledger

    payload = get_tank_gauging_payload_for_transaction(
        db=db,
        transaction_id=transaction.id,
    )

    if payload is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tank Gauging payload is missing. Open Operation Entry, "
                "save the tank gauging ticket, then approve again."
            ),
        )

    inputs = payload.get("inputs") or {}
    calculated = payload.get("calculated") or {}
    payload_asset = payload.get("asset") or {}

    transaction_datetime = resolve_transaction_datetime_for_accounting_day(
        transaction=transaction,
        payload=payload,
    )

    accounting_day = get_location_accounting_day_for_transaction(
        db=db,
        location_code=transaction.origin_location_code,
        transaction_datetime=transaction_datetime,
    )

    tank_operation_code = clean_optional_text(
        inputs.get("tankOperationCode")
    )
    tank_operation_label = clean_optional_text(
        inputs.get("tankOperationLabel")
    )
    tank_operation_category = clean_optional_text(
        inputs.get("tankOperationCategory")
    )
    tank_operation_sign = clean_optional_text(
        inputs.get("tankOperationSign")
    )

    if not tank_operation_code:
        raise HTTPException(
            status_code=400,
            detail=(
                "Tank Operation is missing in Tank Gauging payload. "
                "Open the ticket, select Tank Operation, save, then approve."
            ),
        )

    if not tank_operation_label:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Label is missing in Tank Gauging payload.",
        )

    if not tank_operation_category:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Category is missing in Tank Gauging payload.",
        )

    if not tank_operation_sign:
        raise HTTPException(
            status_code=400,
            detail="Tank Operation Sign is missing in Tank Gauging payload.",
        )

    current_stock_gsv_bbl = safe_float(calculated.get("gsvBbl"))
    current_stock_nsv_bbl = safe_float(calculated.get("nsvBbl"))
    current_stock_lt = safe_float(calculated.get("lt"))
    current_stock_mt = safe_float(calculated.get("mt"))

    if current_stock_nsv_bbl == 0 and current_stock_gsv_bbl == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "Calculated tank quantity is missing or zero. "
                "Open the ticket, verify Tank Gauging calculations, save, then approve."
            ),
        )

    tank_asset = get_asset_by_code(transaction.primary_asset_code, db)

    tank_asset_name = ""

    if tank_asset:
        tank_asset_name = tank_asset.asset_name
    else:
        tank_asset_name = clean_optional_text(
            payload_asset.get("asset_name")
        ) or ""

    created_by_display = get_current_user_display_name(current_user)

    new_ledger = TankStockLedger(
        transaction_id=transaction.id,
        ticket_number=get_transaction_ticket_number(transaction),
        operation_number=transaction.operation_number,
        location_code=transaction.origin_location_code,
        tank_asset_code=transaction.primary_asset_code,
        tank_asset_name=tank_asset_name,
        operation_date=transaction.operation_date,
        product_name=clean_optional_text(transaction.product_name),
        accounting_date=accounting_day["accounting_date"],
        accounting_day_start=accounting_day["accounting_day_start"],
        accounting_day_end=accounting_day["accounting_day_end"],
        accounting_day_setting_id=accounting_day["setting_id"],
        tank_operation_code=tank_operation_code,
        tank_operation_label=tank_operation_label,
        tank_operation_category=tank_operation_category,
        tank_operation_sign=tank_operation_sign,
        movement_gsv_bbl=0,
        movement_nsv_bbl=0,
        movement_lt=0,
        movement_mt=0,
        stock_gsv_bbl=current_stock_gsv_bbl,
        stock_nsv_bbl=current_stock_nsv_bbl,
        stock_lt=current_stock_lt,
        stock_mt=current_stock_mt,
        previous_stock_gsv_bbl=0,
        previous_stock_nsv_bbl=0,
        previous_stock_lt=0,
        previous_stock_mt=0,
        running_balance_gsv_bbl=current_stock_gsv_bbl,
        running_balance_nsv_bbl=current_stock_nsv_bbl,
        running_balance_lt=current_stock_lt,
        running_balance_mt=current_stock_mt,
        source_payload=normalize_jsonb_value(payload),
        status="Active",
        created_by=created_by_display,
        remarks="Auto-created when Tank Gauging ticket was approved",
    )

    db.add(new_ledger)
    db.flush()

    rebuild_tank_stock_running_balances(
        db=db,
        location_code=new_ledger.location_code,
        tank_asset_code=new_ledger.tank_asset_code,
        product_name=new_ledger.product_name,
    )

    db.flush()

    create_audit_log(
        db=db,
        module_name="Tank Stock Ledger",
        action="Create Tank Stock Ledger Entry",
        current_user=current_user,
        entity_type="TankStockLedger",
        entity_id=new_ledger.id,
        entity_label=(
            f"{new_ledger.ticket_number} | "
            f"{new_ledger.tank_asset_code} | "
            f"{new_ledger.tank_operation_label}"
        ),
        ticket_number=new_ledger.ticket_number,
        operation_number=new_ledger.operation_number,
        remarks="Auto-created on Tank Gauging approval",
        request_path="/operation-transactions/{id}/status",
        details={
            "transaction_id": transaction.id,
            "location_code": new_ledger.location_code,
            "tank_asset_code": new_ledger.tank_asset_code,
            "operation_date": str(new_ledger.operation_date),
            "transaction_datetime": transaction_datetime.isoformat(),
            "accounting_date": str(new_ledger.accounting_date),
            "accounting_day_start": (
                new_ledger.accounting_day_start.isoformat()
                if new_ledger.accounting_day_start
                else None
            ),
            "accounting_day_end": (
                new_ledger.accounting_day_end.isoformat()
                if new_ledger.accounting_day_end
                else None
            ),
            "accounting_day_setting_id": new_ledger.accounting_day_setting_id,
            "product_name": new_ledger.product_name,
            "tank_operation_code": new_ledger.tank_operation_code,
            "tank_operation_label": new_ledger.tank_operation_label,
            "tank_operation_category": new_ledger.tank_operation_category,
            "tank_operation_sign": new_ledger.tank_operation_sign,
            "stock_gsv_bbl": new_ledger.stock_gsv_bbl,
            "stock_nsv_bbl": new_ledger.stock_nsv_bbl,
            "stock_lt": new_ledger.stock_lt,
            "stock_mt": new_ledger.stock_mt,
            "previous_stock_gsv_bbl": new_ledger.previous_stock_gsv_bbl,
            "previous_stock_nsv_bbl": new_ledger.previous_stock_nsv_bbl,
            "previous_stock_lt": new_ledger.previous_stock_lt,
            "previous_stock_mt": new_ledger.previous_stock_mt,
            "movement_gsv_bbl": new_ledger.movement_gsv_bbl,
            "movement_nsv_bbl": new_ledger.movement_nsv_bbl,
            "movement_lt": new_ledger.movement_lt,
            "movement_mt": new_ledger.movement_mt,
            "running_balance_gsv_bbl": new_ledger.running_balance_gsv_bbl,
            "running_balance_nsv_bbl": new_ledger.running_balance_nsv_bbl,
            "running_balance_lt": new_ledger.running_balance_lt,
            "running_balance_mt": new_ledger.running_balance_mt,
        },
    )

    return new_ledger

def validate_operation_status_transition(current_status, next_status):
    allowed_transitions = {
        "Draft": ["Submitted", "Cancelled"],
        "Submitted": ["Approved", "Rejected", "Draft"],
        "Rejected": ["Submitted", "Cancelled"],
        "Approved": [],
        "Cancelled": [],
    }

    if current_status not in allowed_transitions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid current status: {current_status}",
        )

    if next_status not in allowed_transitions[current_status]:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot change status from {current_status} to {next_status}",
        )

def validate_multi_tank_seals_before_submit(
    db: Session,
    transaction: OperationTransaction,
    submit_remarks: str | None,
):
    # Check template layout type
    template = (
        db.query(OperationTemplate)
        .filter(OperationTemplate.id == transaction.operation_template_id)
        .first()
    )

    if not template or (template.entry_layout_type or "") != "Multi-Tank Before/After":
        return None  # Not a multi-tank ticket, no seal validation needed

    # Load multi_tank_payload
    payload_row = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction.id,
            OperationTransactionValue.field_code == "multi_tank_payload",
        )
        .first()
    )

    if payload_row is None or payload_row.field_value is None:
        raise HTTPException(
            status_code=400,
            detail="Multi-Tank payload is missing. Open Operation Entry and save the ticket before submitting.",
        )

    payload = payload_row.field_value if isinstance(payload_row.field_value, dict) else {}

    seals_after = (((payload.get("seals") or {}).get("after")) or {})
    temporary = (seals_after.get("temporary") or {})

    # Required temporary seals (AFTER only)
    required_temp_keys = [
        "portManifoldSeal",
        "stbdManifoldSeal",
        "pumproomSeal",
    ]

    missing = []
    for k in required_temp_keys:
        if not str(temporary.get(k) or "").strip():
            missing.append(k)

    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Seal details are incomplete. Please enter AFTER temporary seals: "
                + ", ".join(missing)
            ),
        )

    # If mismatch exists, remarks required
    tank_seals = seals_after.get("tankSeals") or {}
    mismatch_count = 0

    if isinstance(tank_seals, dict):
        for _, positions in tank_seals.items():
            if not isinstance(positions, dict):
                continue
            for _, cell in positions.items():
                if not isinstance(cell, dict):
                    continue
                master = str(cell.get("master") or "").strip()
                observed = str(cell.get("observed") or "").strip()
                if master and observed and master != observed:
                    mismatch_count += 1

    if mismatch_count > 0 and not str(submit_remarks or "").strip():
        raise HTTPException(
            status_code=400,
            detail=(
                f"Seal mismatch detected ({mismatch_count} mismatch). "
                "Please add remarks before submitting."
            ),
        )

    return {
        "required_temp_seals": required_temp_keys,
        "missing_temp_seals": missing,
        "mismatch_count": mismatch_count,
    }

def get_transaction_value_text(db: Session, transaction_id: int, field_code: str):
    v = (
        db.query(OperationTransactionValue)
        .filter(
            OperationTransactionValue.transaction_id == transaction_id,
            OperationTransactionValue.field_code == field_code,
        )
        .first()
    )
    if not v:
        return None
    if v.field_value is None:
        return None
    return str(v.field_value).strip()


def resolve_barge_event_type_from_ticket(db: Session, transaction: OperationTransaction):
    # First preference: stored stage
    stage = get_transaction_value_text(db, transaction.id, "barge_event_type")
    if stage:
        stage_u = stage.strip().upper()
        if stage_u in ["LOAD_1", "LOAD_2_TOPUP", "UNLOAD", "STS"]:
            return stage_u

    # Fallback: old logic (minimal)
    code_u = str(transaction.operation_type_code or "").upper()
    if any(k in code_u for k in ["UNLOAD", "DISCHARGE", "RECEIPT", "RECEIVE"]):
        return "UNLOAD"

    return None


def auto_create_trip_event_on_submit(
    db: Session,
    transaction: OperationTransaction,
    current_user: User,
):
    return None, None


def auto_create_barge_tracking_on_approval(
    db: Session,
    transaction: OperationTransaction,
    current_user: User,
):
    convoy = clean_optional_text(transaction.convoy_number)
    if convoy is None:
        return None, None, None

    if str(transaction.primary_asset_type_code or "").strip().upper() != "BARGE":
        return None, None, None

    if transaction.status != "Approved":
        return None, None, None

    asset_code = str(transaction.primary_asset_code or "").strip()
    if not asset_code:
        return None, None, None

    created_by_display = get_current_user_display_name(current_user)

    trip = db.query(Trip).filter(Trip.convoy_number.ilike(convoy)).first()
    if not trip:
        trip = Trip(
            convoy_number=convoy,
            primary_barge_asset_code=asset_code,
            status="OPEN",
            created_by=created_by_display,
            remarks=None,
        )
        db.add(trip)
        db.flush()

    ensure_trip_not_closed(trip)

    # Decide event type
    chosen = resolve_barge_event_type_from_ticket(db, transaction)

    if chosen is None:
        # fallback: LOAD_1 or LOAD_2_TOPUP based on previous load events
        prev_load = (
            db.query(TripEvent)
            .filter(
                TripEvent.trip_id == trip.id,
                TripEvent.asset_code == asset_code,
                TripEvent.event_type.in_(["LOAD_1", "LOAD_2_TOPUP"]),
            )
            .order_by(TripEvent.sequence_no.desc(), TripEvent.id.desc())
            .first()
        )
        chosen = "LOAD_1" if not prev_load else "LOAD_2_TOPUP"

    existing = (
        db.query(TripEvent)
        .filter(TripEvent.operation_transaction_id == transaction.id)
        .first()
    )

    if existing:
        existing.event_type = chosen
        existing.location_code = clean_optional_text(transaction.origin_location_code) or existing.location_code
        existing.asset_code = asset_code
        existing.event_datetime = transaction.operation_start_datetime or existing.event_datetime
        existing.updated_at = datetime.now()
        new_event = existing
    else:
        max_seq = (
            db.query(func.max(TripEvent.sequence_no))
            .filter(TripEvent.trip_id == trip.id)
            .scalar()
        )
        seq = (max_seq or 0) + 1

        new_event = TripEvent(
            trip_id=trip.id,
            event_type=chosen,
            location_code=clean_optional_text(transaction.origin_location_code),
            asset_code=asset_code,
            operation_transaction_id=transaction.id,
            sequence_no=seq,
            event_datetime=transaction.operation_start_datetime or datetime.now(),
            created_by=created_by_display,
            remarks="Auto-created on Approval",
        )
        db.add(new_event)
        db.flush()

    new_cmp = None

    # Only UNLOAD creates the main comparison
    if chosen == "UNLOAD":
        latest_load = (
            db.query(TripEvent)
            .filter(
                TripEvent.trip_id == trip.id,
                TripEvent.asset_code == asset_code,
                TripEvent.event_type.in_(["LOAD_1", "LOAD_2_TOPUP"]),
                TripEvent.operation_transaction_id.isnot(None),
            )
            .order_by(TripEvent.sequence_no.desc(), TripEvent.id.desc())
            .first()
        )

        if latest_load and latest_load.operation_transaction_id:
            left_tx = (
                db.query(OperationTransaction)
                .filter(OperationTransaction.id == latest_load.operation_transaction_id)
                .first()
            )

            if left_tx and left_tx.status == "Approved":
                existing_cmp = (
                    db.query(TripComparison)
                    .filter(
                        TripComparison.trip_id == trip.id,
                        TripComparison.comparison_type == "LOAD_AFTER_vs_UNLOAD_BEFORE",
                        TripComparison.left_transaction_id == left_tx.id,
                        TripComparison.right_transaction_id == transaction.id,
                    )
                    .first()
                )

                if not existing_cmp:
                    left_payload = load_multi_tank_payload(db, left_tx.id)
                    right_payload = load_multi_tank_payload(db, transaction.id)

                    if left_payload and right_payload:
                        summary_json, per_tank_json = build_multitank_comparison_json(
                            left_tx=left_tx,
                            right_tx=transaction,
                            comparison_type="LOAD_AFTER_vs_UNLOAD_BEFORE",
                            left_payload=left_payload,
                            right_payload=right_payload,
                        )

                        new_cmp = TripComparison(
                            trip_id=trip.id,
                            comparison_type="LOAD_AFTER_vs_UNLOAD_BEFORE",
                            left_transaction_id=left_tx.id,
                            right_transaction_id=transaction.id,
                            summary_json=summary_json,
                            per_tank_json=per_tank_json,
                            created_by=created_by_display,
                            remarks="Auto-created on UNLOAD Approval",
                        )
                        db.add(new_cmp)
                        db.flush()

    return trip, new_event, new_cmp

@app.patch("/operation-transactions/{transaction_id}/status")
def update_operation_transaction_status(
    transaction_id: int,
    status_update: OperationTransactionStatusUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == transaction_id)
        .first()
    )

    if transaction is None:
        raise HTTPException(
            status_code=404,
            detail="Operation transaction not found",
        )

    trip = None
    if str(transaction.primary_asset_type_code or "").strip().upper() == "BARGE":
        trip = get_trip_by_convoy_or_none(db, transaction.convoy_number)
        ensure_trip_not_closed(trip)

    next_status = clean_optional_text(status_update.status)

    if next_status is None:
        raise HTTPException(
            status_code=400,
            detail="Status is required",
        )

    allowed_statuses = ["Draft", "Submitted", "Approved", "Rejected", "Cancelled"]

    if next_status not in allowed_statuses:
        raise HTTPException(
            status_code=400,
            detail="Invalid transaction status",
        )

    required_permission = get_required_permission_for_status_change(next_status)

    if required_permission:
        require_user_permission(current_user, required_permission, db)

    validate_operation_status_transition(transaction.status, next_status)

    old_status = transaction.status
    changed_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )
    
    status_remarks = clean_optional_text(status_update.remarks)

    review_confirmed = bool(getattr(status_update, "review_confirmed", False))

    if next_status in ["Submitted", "Approved"] and not review_confirmed:
        raise HTTPException(
            status_code=400,
            detail="Review confirmation is required for Submit/Approve.",
        )

    seal_validation_details = None
    if next_status == "Submitted":
        seal_validation_details = validate_multi_tank_seals_before_submit(
            db=db,
            transaction=transaction,
            submit_remarks=status_remarks,
        )

    if next_status in ["Submitted", "Approved"] and review_confirmed:
        status_remarks = (status_remarks or "").strip()
        status_remarks = (status_remarks + "\n[REVIEW CONFIRMED]").strip()

    transaction.status = next_status
    transaction.updated_at = datetime.now()

    # Movement tracking must start only after Approval.
    # Draft / Submitted tickets must not move into Barge Tracking.
    if next_status == "Approved":
        template = None
        if transaction.operation_template_id:
            template = (
                db.query(OperationTemplate)
                .filter(OperationTemplate.id == transaction.operation_template_id)
                .first()
            )

        if template and str(template.entry_layout_type or "").strip() == "Shuttle Tracking":
            voyage = get_or_create_shuttle_voyage(
                db=db,
                location_code=transaction.origin_location_code,
                shuttle_number=transaction.convoy_number or "",
                shuttle_asset_code=transaction.primary_asset_code,
                current_user=current_user,
            )
            ensure_shuttle_voyage_not_closed(voyage)

        auto_create_barge_tracking_on_approval(
            db=db,
            transaction=transaction,
            current_user=current_user,
        )

        create_tank_stock_ledger_from_approved_transaction(
            db=db,
            transaction=transaction,
            current_user=current_user,
        )

        create_or_update_vessel_stock_ledger_from_approved_transaction(
            db=db,
            transaction=transaction,
            current_user=current_user,
        )
    if status_remarks:
        existing_remarks = transaction.remarks or ""
        transaction.remarks = (
            f"{existing_remarks}\n"
            f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
            f"{changed_by} changed status from {old_status} to {next_status}: "
            f"{status_remarks}"
        ).strip()

    history = OperationTransactionStatusHistory(
        transaction_id=transaction.id,
        old_status=old_status,
        new_status=next_status,
        changed_by=changed_by,
        remarks=status_remarks,
        changed_at=datetime.now(),
    )

    db.add(history)

    action_name = f"Change Status to {next_status}"

    if next_status == "Submitted":
        action_name = "Submit Operation Transaction"
    elif next_status == "Approved":
        action_name = "Approve Operation Transaction"
    elif next_status == "Rejected":
        action_name = "Reject Operation Transaction"
    elif next_status == "Draft":
        action_name = "Recall Operation Transaction"
    elif next_status == "Cancelled":
        action_name = "Cancel Operation Transaction"

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action=action_name,
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=transaction.id,
        entity_label=get_transaction_ticket_number(transaction),
        ticket_number=get_transaction_ticket_number(transaction),
        operation_number=transaction.operation_number,
        old_status=old_status,
        new_status=next_status,
        remarks=status_remarks or "",
        request_path=f"/operation-transactions/{transaction_id}/status",
        details={
            "operation_type_code": transaction.operation_type_code,
            "operation_template_id": transaction.operation_template_id,
            "primary_asset_code": transaction.primary_asset_code,
            "origin_location_code": transaction.origin_location_code,
            "operation_date": str(transaction.operation_date),
            "seal_validation": seal_validation_details,
            "review_confirmed": review_confirmed,
            "reviewed_by": changed_by if review_confirmed else None,
            "reviewed_at": datetime.now().isoformat() if review_confirmed else None,
        },
    )

    db.commit()
    db.refresh(transaction)

    return {
        "message": f"Transaction status changed to {next_status}",
        "transaction": build_operation_transaction_response(transaction, db),
    }


@app.get("/operation-transactions/{transaction_id}/status-history")
def get_operation_transaction_status_history(
    transaction_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Operation Transaction",
        db,
    )

    transaction = (
        db.query(OperationTransaction)
        .filter(OperationTransaction.id == transaction_id)
        .first()
    )

    if transaction is None:
        raise HTTPException(
            status_code=404,
            detail="Operation transaction not found",
        )

    history = (
        db.query(OperationTransactionStatusHistory)
        .filter(OperationTransactionStatusHistory.transaction_id == transaction_id)
        .order_by(OperationTransactionStatusHistory.changed_at.asc())
        .all()
    )

    return [
        {
            "id": item.id,
            "transaction_id": item.transaction_id,
            "old_status": item.old_status,
            "new_status": item.new_status,
            "changed_by": item.changed_by,
            "remarks": item.remarks,
            "changed_at": item.changed_at,
        }
        for item in history
    ]


@app.put(
    "/operation-entries/{transaction_id}",
    response_model=OperationEntryResponse,
)
def update_operation_entry(
    transaction_id: int,
    entry: OperationEntryCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Create Operation Entry",
        db,
    )

    existing_transaction = db.query(OperationTransaction).filter(
        OperationTransaction.id == transaction_id
    ).first()

    if not existing_transaction:
        raise HTTPException(
            status_code=404,
            detail="Operation entry not found",
        )

    if existing_transaction.status not in ["Draft", "Rejected"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Rejected operation entries can be edited. "
                "Recall Submitted tickets to Draft before editing."
            ),
        )
    
    convoy_to_check = clean_optional_text(entry.transaction.convoy_number) or clean_optional_text(existing_transaction.convoy_number)
    trip = get_trip_by_convoy_or_none(db, convoy_to_check)
    ensure_trip_not_closed(trip)

    (
        template,
        operation_type,
        asset,
        template_fields,
        value_map,
        transaction_operation_type_code,
    ) = validate_operation_entry(entry, db)

    existing_transaction.operation_type_code = transaction_operation_type_code
    existing_transaction.operation_template_id = template.id
    existing_transaction.primary_asset_code = asset.asset_code
    existing_transaction.primary_asset_type_code = asset.asset_type_code
    existing_transaction.convoy_number = clean_optional_text(entry.transaction.convoy_number)
    existing_transaction.origin_location_code = entry.transaction.origin_location_code.strip()
    existing_transaction.destination_location_code = clean_optional_text(
        entry.transaction.destination_location_code
    )
    existing_transaction.sender_location_code = clean_optional_text(
        entry.transaction.sender_location_code
    )
    existing_transaction.receiver_location_code = clean_optional_text(
        entry.transaction.receiver_location_code
    )
    existing_transaction.operation_date = entry.transaction.operation_date
    existing_transaction.operation_start_datetime = entry.transaction.operation_start_datetime
    existing_transaction.operation_end_datetime = entry.transaction.operation_end_datetime
    existing_transaction.product_name = clean_optional_text(entry.transaction.product_name)
    existing_transaction.remarks = clean_optional_text(entry.transaction.remarks)
    existing_transaction.updated_at = datetime.now()

    db.query(OperationTransactionValue).filter(
        OperationTransactionValue.transaction_id == transaction_id
    ).delete()

    for field in template_fields:
        new_value = OperationTransactionValue(
            transaction_id=transaction_id,
            field_code=field.field_code,
            field_name=field.field_name,
            field_group=field.field_group,
            data_type=field.data_type,
            unit=field.unit,
            input_mode=field.input_mode,
            calculation_role=field.calculation_role,
            field_value=normalize_jsonb_value(value_map.get(field.field_code)),
            sort_order=field.sort_order,
        )

        db.add(new_value)

    changed_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )

    existing_remarks = existing_transaction.remarks or ""
    existing_transaction.remarks = (
        f"{existing_remarks}\n"
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
        f"Edited by {changed_by}"
    ).strip()

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Update Operation Entry",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=existing_transaction.id,
        entity_label=get_transaction_ticket_number(existing_transaction),
        ticket_number=get_transaction_ticket_number(existing_transaction),
        operation_number=existing_transaction.operation_number,
        old_status=existing_transaction.status,
        new_status=existing_transaction.status,
        remarks="Operation entry edited",
        request_path=f"/operation-entries/{transaction_id}",
        details={
            "operation_type_code": existing_transaction.operation_type_code,
            "operation_template_id": existing_transaction.operation_template_id,
            "primary_asset_code": existing_transaction.primary_asset_code,
            "origin_location_code": existing_transaction.origin_location_code,
            "operation_date": str(existing_transaction.operation_date),
            "field_count": len(template_fields),
        },
    )

    db.commit()
    db.refresh(existing_transaction)

    return build_operation_entry_response(existing_transaction, db)


@app.delete("/operation-entries/{transaction_id}")
def delete_operation_entry(
    transaction_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Cancel Operation Transaction",
        db,
    )

    existing_transaction = db.query(OperationTransaction).filter(
        OperationTransaction.id == transaction_id
    ).first()

    if not existing_transaction:
        raise HTTPException(
            status_code=404,
            detail="Operation entry not found",
        )

    if existing_transaction.status not in ["Draft", "Rejected"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "Only Draft or Rejected operation entries can be cancelled. "
                "Submitted tickets must be recalled first. Approved and Cancelled tickets are locked."
            ),
        )

    old_status = existing_transaction.status

    changed_by = (
        f"{current_user.full_name} ({current_user.username})"
        if current_user.full_name
        else current_user.username
    )

    existing_transaction.status = "Cancelled"
    existing_transaction.updated_at = datetime.now()

    existing_remarks = existing_transaction.remarks or ""
    existing_transaction.remarks = (
        f"{existing_remarks}\n"
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
        f"Cancelled by {changed_by}"
    ).strip()

    history = OperationTransactionStatusHistory(
        transaction_id=existing_transaction.id,
        old_status=old_status,
        new_status="Cancelled",
        changed_by=changed_by,
        remarks="Cancelled from Operation Entry editable list",
        changed_at=datetime.now(),
    )

    db.add(history)

    field_count = (
        db.query(OperationTransactionValue)
        .filter(OperationTransactionValue.transaction_id == existing_transaction.id)
        .count()
    )

    create_audit_log(
        db=db,
        module_name="Operation Transaction",
        action="Cancel Operation Entry",
        current_user=current_user,
        entity_type="OperationTransaction",
        entity_id=existing_transaction.id,
        entity_label=get_transaction_ticket_number(existing_transaction),
        ticket_number=get_transaction_ticket_number(existing_transaction),
        operation_number=existing_transaction.operation_number,
        old_status=old_status,
        new_status="Cancelled",
        remarks="Cancelled from Operation Entry editable list",
        request_path=f"/operation-entries/{transaction_id}",
        details={
            "operation_type_code": existing_transaction.operation_type_code,
            "operation_template_id": existing_transaction.operation_template_id,
            "primary_asset_code": existing_transaction.primary_asset_code,
            "origin_location_code": existing_transaction.origin_location_code,
            "operation_date": str(existing_transaction.operation_date),
            "field_count": field_count,
        },
    )

    db.commit()
    db.refresh(existing_transaction)

    return {
        "message": "Operation entry cancelled successfully"
    }

# -------------------------
# Common Table 11 Factor APIs
# -------------------------

def build_table11_factor_response(row: Table11Factor):
    return {
        "id": row.id,
        "api60": float(row.api60),
        "lt_factor": float(row.lt_factor),
        "created_at": row.created_at,
        "updated_at": row.updated_at,
    }


def interpolate_table11_factor(api60: float, db: Session):
    if api60 is None:
        raise HTTPException(
            status_code=400,
            detail="API @ 60°F is required",
        )

    api_value = float(api60)

    rows = (
        db.query(Table11Factor)
        .order_by(Table11Factor.api60.asc())
        .all()
    )

    if len(rows) == 0:
        raise HTTPException(
            status_code=400,
            detail="Table 11 factor master is empty. Please upload API@60 and LT factor data first.",
        )

    exact_row = next(
        (
            row
            for row in rows
            if float(row.api60) == api_value
        ),
        None,
    )

    if exact_row:
        return {
            "api60": api_value,
            "lower_api60": float(exact_row.api60),
            "upper_api60": float(exact_row.api60),
            "lt_factor": float(exact_row.lt_factor),
            "lookup_method": "Exact match",
        }

    lower_row = None
    upper_row = None

    for row in rows:
        row_api = float(row.api60)

        if row_api < api_value:
            lower_row = row

        if row_api > api_value:
            upper_row = row
            break

    if lower_row is None:
        first_row = rows[0]

        return {
            "api60": api_value,
            "lower_api60": float(first_row.api60),
            "upper_api60": float(first_row.api60),
            "lt_factor": float(first_row.lt_factor),
            "lookup_method": "Below range - nearest factor used",
        }

    if upper_row is None:
        last_row = rows[-1]

        return {
            "api60": api_value,
            "lower_api60": float(last_row.api60),
            "upper_api60": float(last_row.api60),
            "lt_factor": float(last_row.lt_factor),
            "lookup_method": "Above range - nearest factor used",
        }

    lower_api = float(lower_row.api60)
    upper_api = float(upper_row.api60)
    lower_factor = float(lower_row.lt_factor)
    upper_factor = float(upper_row.lt_factor)

    if upper_api == lower_api:
        interpolated_factor = lower_factor
    else:
        ratio = (api_value - lower_api) / (upper_api - lower_api)
        interpolated_factor = lower_factor + ratio * (upper_factor - lower_factor)

    return {
        "api60": api_value,
        "lower_api60": lower_api,
        "upper_api60": upper_api,
        "lt_factor": round(interpolated_factor, 10),
        "lookup_method": "Linear interpolation",
    }


@app.get(
    "/table11-factors",
    response_model=list[Table11FactorResponse],
)
def get_table11_factors(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset Calibration",
        db,
    )

    rows = (
        db.query(Table11Factor)
        .order_by(Table11Factor.api60.asc())
        .all()
    )

    return [
        build_table11_factor_response(row)
        for row in rows
    ]


@app.get(
    "/table11-factors/lookup",
    response_model=Table11LookupResponse,
)
def lookup_table11_factor(
    api60: float,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Asset Calibration",
        db,
    )

    return interpolate_table11_factor(api60, db)

def build_table11_audit_snapshot(db: Session, preview_limit: int = 20):
    rows = db.query(Table11Factor).order_by(Table11Factor.api60.asc()).all()

    count = len(rows)

    min_api = float(rows[0].api60) if count > 0 else None
    max_api = float(rows[-1].api60) if count > 0 else None

    preview_rows = rows[:preview_limit]

    return {
        "count": count,
        "min_api60": min_api,
        "max_api60": max_api,
        "preview_limit": preview_limit,
        "preview_rows": [
            {
                "api60": float(r.api60),
                "lt_factor": float(r.lt_factor),
            }
            for r in preview_rows
        ],
    }

@app.post(
    "/table11-factors/bulk",
    response_model=list[Table11FactorResponse],
)
def bulk_save_table11_factors(
    request: Table11FactorBulkCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Calibration",
        db,
    )

    if len(request.rows) == 0:
        raise HTTPException(
            status_code=400,
            detail="Please provide at least one Table 11 row",
        )

    api_values = [float(row.api60) for row in request.rows]

    if len(api_values) != len(set(api_values)):
        raise HTTPException(
            status_code=400,
            detail="Duplicate API @ 60°F values are not allowed",
        )

    for row in request.rows:
        if row.api60 <= 0:
            raise HTTPException(
                status_code=400,
                detail="API @ 60°F must be greater than zero",
            )
        if row.lt_factor <= 0:
            raise HTTPException(
                status_code=400,
                detail=f"LT factor must be greater than zero for API @ 60°F {row.api60}",
            )

    before_snapshot = build_table11_audit_snapshot(db, preview_limit=20)

    # Replace all
    db.query(Table11Factor).delete()

    for row in request.rows:
        db.add(
            Table11Factor(
                api60=float(row.api60),
                lt_factor=float(row.lt_factor),
            )
        )

    db.flush()

    after_snapshot = build_table11_audit_snapshot(db, preview_limit=20)

    create_audit_log(
        db=db,
        module_name="Table 11 Factor Master",
        action="Bulk Save Table 11 Factors",
        current_user=current_user,
        entity_type="Table11Factor",
        entity_id=None,
        entity_label="Table 11 Factor Master",
        remarks="Replaced Table 11 factor master rows",
        request_path="/table11-factors/bulk",
        details={
            "before": before_snapshot,
            "after": after_snapshot,
            "input_row_count": len(request.rows),
        },
    )

    db.commit()

    saved_rows = db.query(Table11Factor).order_by(Table11Factor.api60.asc()).all()

    return [build_table11_factor_response(row) for row in saved_rows]


@app.delete("/table11-factors")
def clear_table11_factors(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Asset Calibration",
        db,
    )

    before_snapshot = build_table11_audit_snapshot(db, preview_limit=20)

    deleted_count = db.query(Table11Factor).delete()
    db.flush()

    after_snapshot = build_table11_audit_snapshot(db, preview_limit=20)

    create_audit_log(
        db=db,
        module_name="Table 11 Factor Master",
        action="Clear Table 11 Factors",
        current_user=current_user,
        entity_type="Table11Factor",
        entity_id=None,
        entity_label="Table 11 Factor Master",
        remarks="Cleared all Table 11 factor rows",
        request_path="/table11-factors",
        details={
            "before": before_snapshot,
            "after": after_snapshot,
            "deleted_count": deleted_count,
        },
    )

    db.commit()

    return {
        "message": "Table 11 factors cleared successfully",
        "deleted_count": deleted_count,
    }

# -------------------------
# Material Balance Template Configuration Helpers
# -------------------------

VALID_MATERIAL_BALANCE_COLUMN_TYPES = {
    "OPENING",
    "MOVEMENT",
    "BOOK_CLOSING",
    "ACTUAL_CLOSING",
    "LOSS_GAIN",
    "FORMULA",
    "INFO",
}

VALID_MATERIAL_BALANCE_DIRECTIONS = {
    "IN",
    "OUT",
    "NEUTRAL",
}


def normalize_column_key(value: str):
    cleaned_value = clean_optional_text(value)

    if not cleaned_value:
        return ""

    normalized = cleaned_value.strip().lower()
    normalized = normalized.replace(" ", "_")
    normalized = normalized.replace("-", "_")
    normalized = normalized.replace("/", "_")

    while "__" in normalized:
        normalized = normalized.replace("__", "_")

    return normalized.upper()


def validate_yes_no(value: str | None, field_name: str):
    cleaned_value = clean_optional_text(value) or "No"
    cleaned_value = cleaned_value.strip().title()

    if cleaned_value not in ["Yes", "No"]:
        raise HTTPException(
            status_code=400,
            detail=f"{field_name} must be Yes or No",
        )

    return cleaned_value


def validate_material_balance_template_column_payload(column):
    column_type = clean_optional_text(column.column_type).upper()

    if column_type not in VALID_MATERIAL_BALANCE_COLUMN_TYPES:
        raise HTTPException(
            status_code=400,
            detail=(
                "Column Type must be one of: "
                + ", ".join(sorted(VALID_MATERIAL_BALANCE_COLUMN_TYPES))
            ),
        )

    movement_direction = clean_optional_text(column.movement_direction)

    if column_type == "MOVEMENT":
        if not movement_direction:
            raise HTTPException(
                status_code=400,
                detail="Movement Direction is required for MOVEMENT columns",
            )

        movement_direction = movement_direction.upper()

        if movement_direction not in VALID_MATERIAL_BALANCE_DIRECTIONS:
            raise HTTPException(
                status_code=400,
                detail="Movement Direction must be IN, OUT, or NEUTRAL",
            )

        if len(column.mapped_operation_codes or []) == 0:
            raise HTTPException(
                status_code=400,
                detail="At least one Tank Operation must be mapped for MOVEMENT columns",
            )
    else:
        movement_direction = None

    include_in_material_balance = validate_yes_no(
        column.include_in_material_balance,
        "Include in Material Balance",
    )

    include_in_book_closing = validate_yes_no(
        column.include_in_book_closing,
        "Include in Book Closing",
    )

    is_internal_transfer = validate_yes_no(
        column.is_internal_transfer,
        "Is Internal Transfer",
    )

    # Safety rule:
    # Internal transfer columns should not affect Material Balance or Book Closing.
    if is_internal_transfer == "Yes":
        include_in_material_balance = "No"
        include_in_book_closing = "No"

    column_key = normalize_column_key(column.column_key or column.column_label)

    if not column_key:
        raise HTTPException(
            status_code=400,
            detail="Column Key is required",
        )

    mapped_operation_codes = [
        normalize_column_key(item)
        for item in (column.mapped_operation_codes or [])
        if clean_optional_text(item)
    ]

    excluded_operation_codes = [
        normalize_column_key(item)
        for item in (column.excluded_operation_codes or [])
        if clean_optional_text(item)
    ]

    return {
        "column_key": column_key,
        "column_type": column_type,
        "movement_direction": movement_direction,
        "mapped_operation_codes": mapped_operation_codes,
        "excluded_operation_codes": excluded_operation_codes,
        "include_in_material_balance": include_in_material_balance,
        "include_in_book_closing": include_in_book_closing,
        "is_internal_transfer": is_internal_transfer,
    }


def build_material_balance_template_response(
    template: MaterialBalanceTemplate,
):
    return {
        "id": template.id,
        "location_code": template.location_code,
        "template_name": template.template_name,
        "description": template.description,
        "status": template.status,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


def build_material_balance_template_column_response(
    column: MaterialBalanceTemplateColumn,
):
    return {
        "id": column.id,
        "template_id": column.template_id,
        "column_label": column.column_label,
        "column_key": column.column_key,
        "column_order": column.column_order,
        "column_type": column.column_type,
        "movement_direction": column.movement_direction,
        "mapped_operation_codes": column.mapped_operation_codes or [],
        "excluded_operation_codes": column.excluded_operation_codes or [],
        "include_in_material_balance": column.include_in_material_balance,
        "include_in_book_closing": column.include_in_book_closing,
        "is_internal_transfer": column.is_internal_transfer,
        "formula_json": column.formula_json,
        "remarks": column.remarks,
        "status": column.status,
        "created_at": column.created_at,
        "updated_at": column.updated_at,
    }


def build_material_balance_template_detail_response(
    template: MaterialBalanceTemplate,
    db: Session,
):
    columns = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(MaterialBalanceTemplateColumn.template_id == template.id)
        .order_by(
            MaterialBalanceTemplateColumn.column_order.asc(),
            MaterialBalanceTemplateColumn.id.asc(),
        )
        .all()
    )

    response = build_material_balance_template_response(template)
    response["columns"] = [
        build_material_balance_template_column_response(column)
        for column in columns
    ]

    return response

# -------------------------
# Material Balance Template Configuration APIs
# -------------------------

@app.get(
    "/material-balance-templates",
    response_model=list[MaterialBalanceTemplateResponse],
)
def get_material_balance_templates(
    location_code: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Material Balance Template",
        db,
    )

    query = db.query(MaterialBalanceTemplate)

    cleaned_location_code = clean_optional_text(location_code)
    cleaned_status = clean_optional_text(status)

    if cleaned_location_code:
        query = query.filter(
            MaterialBalanceTemplate.location_code.ilike(cleaned_location_code)
        )

    if cleaned_status:
        query = query.filter(MaterialBalanceTemplate.status == cleaned_status)

    templates = (
        query.order_by(
            MaterialBalanceTemplate.location_code.asc(),
            MaterialBalanceTemplate.template_name.asc(),
        )
        .all()
    )

    return [
        build_material_balance_template_response(template)
        for template in templates
    ]


@app.get(
    "/material-balance-templates/{template_id}",
    response_model=MaterialBalanceTemplateDetailResponse,
)
def get_material_balance_template_detail(
    template_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Material Balance Template",
        db,
    )

    template = (
        db.query(MaterialBalanceTemplate)
        .filter(MaterialBalanceTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template not found",
        )

    return build_material_balance_template_detail_response(template, db)


@app.post(
    "/material-balance-templates",
    response_model=MaterialBalanceTemplateResponse,
)
def create_material_balance_template(
    template_data: MaterialBalanceTemplateCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    location_code = clean_optional_text(template_data.location_code)

    if not location_code:
        raise HTTPException(
            status_code=400,
            detail="Location is required",
        )

    location = get_location_by_code(location_code, db)

    if not location:
        raise HTTPException(
            status_code=404,
            detail=f"Location {location_code} not found",
        )

    template_name = clean_optional_text(template_data.template_name)

    if not template_name:
        raise HTTPException(
            status_code=400,
            detail="Template Name is required",
        )

    existing_template = (
        db.query(MaterialBalanceTemplate)
        .filter(
            MaterialBalanceTemplate.location_code.ilike(location_code),
            MaterialBalanceTemplate.template_name.ilike(template_name),
        )
        .first()
    )

    if existing_template:
        raise HTTPException(
            status_code=400,
            detail="Material Balance Template already exists for this location",
        )

    new_template = MaterialBalanceTemplate(
        location_code=location_code.upper(),
        template_name=template_name,
        description=clean_optional_text(template_data.description),
        status=template_data.status or "Active",
    )

    db.add(new_template)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Create Material Balance Template",
        current_user=current_user,
        entity_type="MaterialBalanceTemplate",
        entity_id=new_template.id,
        entity_label=new_template.template_name,
        remarks="Created Material Balance Template",
        request_path="/material-balance-templates",
        details=build_material_balance_template_response(new_template),
    )

    db.commit()
    db.refresh(new_template)

    return build_material_balance_template_response(new_template)


@app.put(
    "/material-balance-templates/{template_id}",
    response_model=MaterialBalanceTemplateResponse,
)
def update_material_balance_template(
    template_id: int,
    template_data: MaterialBalanceTemplateUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    template = (
        db.query(MaterialBalanceTemplate)
        .filter(MaterialBalanceTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template not found",
        )

    location_code = clean_optional_text(template_data.location_code)
    template_name = clean_optional_text(template_data.template_name)

    if not location_code:
        raise HTTPException(status_code=400, detail="Location is required")

    if not template_name:
        raise HTTPException(status_code=400, detail="Template Name is required")

    duplicate_template = (
        db.query(MaterialBalanceTemplate)
        .filter(
            MaterialBalanceTemplate.id != template_id,
            MaterialBalanceTemplate.location_code.ilike(location_code),
            MaterialBalanceTemplate.template_name.ilike(template_name),
        )
        .first()
    )

    if duplicate_template:
        raise HTTPException(
            status_code=400,
            detail="Another Material Balance Template already exists for this location",
        )

    old_details = build_material_balance_template_response(template)

    template.location_code = location_code.upper()
    template.template_name = template_name
    template.description = clean_optional_text(template_data.description)
    template.status = template_data.status or "Active"
    template.updated_at = datetime.now()

    db.flush()

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Update Material Balance Template",
        current_user=current_user,
        entity_type="MaterialBalanceTemplate",
        entity_id=template.id,
        entity_label=template.template_name,
        remarks="Updated Material Balance Template",
        request_path=f"/material-balance-templates/{template_id}",
        details={
            "old": old_details,
            "new": build_material_balance_template_response(template),
        },
    )

    db.commit()
    db.refresh(template)

    return build_material_balance_template_response(template)


@app.delete("/material-balance-templates/{template_id}")
def delete_material_balance_template(
    template_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    template = (
        db.query(MaterialBalanceTemplate)
        .filter(MaterialBalanceTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template not found",
        )

    old_details = build_material_balance_template_detail_response(template, db)

    db.delete(template)

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Delete Material Balance Template",
        current_user=current_user,
        entity_type="MaterialBalanceTemplate",
        entity_id=template_id,
        entity_label=template.template_name,
        remarks="Deleted Material Balance Template",
        request_path=f"/material-balance-templates/{template_id}",
        details=old_details,
    )

    db.commit()

    return {"message": "Material Balance Template deleted successfully"}


@app.post(
    "/material-balance-templates/{template_id}/columns",
    response_model=MaterialBalanceTemplateColumnResponse,
)
def create_material_balance_template_column(
    template_id: int,
    column_data: MaterialBalanceTemplateColumnCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    template = (
        db.query(MaterialBalanceTemplate)
        .filter(MaterialBalanceTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template not found",
        )

    validated = validate_material_balance_template_column_payload(column_data)

    duplicate_column = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(
            MaterialBalanceTemplateColumn.template_id == template_id,
            MaterialBalanceTemplateColumn.column_key == validated["column_key"],
        )
        .first()
    )

    if duplicate_column:
        raise HTTPException(
            status_code=400,
            detail="Column Key already exists in this template",
        )

    new_column = MaterialBalanceTemplateColumn(
        template_id=template_id,
        column_label=clean_optional_text(column_data.column_label),
        column_key=validated["column_key"],
        column_order=column_data.column_order or 1,
        column_type=validated["column_type"],
        movement_direction=validated["movement_direction"],
        mapped_operation_codes=validated["mapped_operation_codes"],
        excluded_operation_codes=validated["excluded_operation_codes"],
        include_in_material_balance=validated["include_in_material_balance"],
        include_in_book_closing=validated["include_in_book_closing"],
        is_internal_transfer=validated["is_internal_transfer"],
        formula_json=column_data.formula_json,
        remarks=clean_optional_text(column_data.remarks),
        status=column_data.status or "Active",
    )

    db.add(new_column)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Create Material Balance Template Column",
        current_user=current_user,
        entity_type="MaterialBalanceTemplateColumn",
        entity_id=new_column.id,
        entity_label=new_column.column_label,
        remarks="Created Material Balance Template Column",
        request_path=f"/material-balance-templates/{template_id}/columns",
        details=build_material_balance_template_column_response(new_column),
    )

    db.commit()
    db.refresh(new_column)

    return build_material_balance_template_column_response(new_column)


@app.put(
    "/material-balance-template-columns/{column_id}",
    response_model=MaterialBalanceTemplateColumnResponse,
)
def update_material_balance_template_column(
    column_id: int,
    column_data: MaterialBalanceTemplateColumnUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    column = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(MaterialBalanceTemplateColumn.id == column_id)
        .first()
    )

    if not column:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template Column not found",
        )

    validated = validate_material_balance_template_column_payload(column_data)

    duplicate_column = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(
            MaterialBalanceTemplateColumn.id != column_id,
            MaterialBalanceTemplateColumn.template_id == column.template_id,
            MaterialBalanceTemplateColumn.column_key == validated["column_key"],
        )
        .first()
    )

    if duplicate_column:
        raise HTTPException(
            status_code=400,
            detail="Column Key already exists in this template",
        )

    old_details = build_material_balance_template_column_response(column)

    column.column_label = clean_optional_text(column_data.column_label)
    column.column_key = validated["column_key"]
    column.column_order = column_data.column_order or 1
    column.column_type = validated["column_type"]
    column.movement_direction = validated["movement_direction"]
    column.mapped_operation_codes = validated["mapped_operation_codes"]
    column.excluded_operation_codes = validated["excluded_operation_codes"]
    column.include_in_material_balance = validated["include_in_material_balance"]
    column.include_in_book_closing = validated["include_in_book_closing"]
    column.is_internal_transfer = validated["is_internal_transfer"]
    column.formula_json = column_data.formula_json
    column.remarks = clean_optional_text(column_data.remarks)
    column.status = column_data.status or "Active"
    column.updated_at = datetime.now()

    db.flush()

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Update Material Balance Template Column",
        current_user=current_user,
        entity_type="MaterialBalanceTemplateColumn",
        entity_id=column.id,
        entity_label=column.column_label,
        remarks="Updated Material Balance Template Column",
        request_path=f"/material-balance-template-columns/{column_id}",
        details={
            "old": old_details,
            "new": build_material_balance_template_column_response(column),
        },
    )

    db.commit()
    db.refresh(column)

    return build_material_balance_template_column_response(column)


@app.delete("/material-balance-template-columns/{column_id}")
def delete_material_balance_template_column(
    column_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Material Balance Template",
        db,
    )

    column = (
        db.query(MaterialBalanceTemplateColumn)
        .filter(MaterialBalanceTemplateColumn.id == column_id)
        .first()
    )

    if not column:
        raise HTTPException(
            status_code=404,
            detail="Material Balance Template Column not found",
        )

    old_details = build_material_balance_template_column_response(column)

    db.delete(column)

    create_audit_log(
        db=db,
        module_name="Material Balance Template",
        action="Delete Material Balance Template Column",
        current_user=current_user,
        entity_type="MaterialBalanceTemplateColumn",
        entity_id=column_id,
        entity_label=column.column_label,
        remarks="Deleted Material Balance Template Column",
        request_path=f"/material-balance-template-columns/{column_id}",
        details=old_details,
    )

    db.commit()

    return {"message": "Material Balance Template Column deleted successfully"}


VALID_DASHBOARD_SCOPE_TYPES = {"GLOBAL", "LOCATION"}
VALID_DASHBOARD_STATUSES = {"Draft", "Published", "Archived"}


def normalize_dashboard_scope_type(value: str | None):
    cleaned = str(value or "").strip().upper()
    if cleaned not in VALID_DASHBOARD_SCOPE_TYPES:
        raise HTTPException(
            status_code=400,
            detail="scope_type must be GLOBAL or LOCATION",
        )
    return cleaned


def normalize_dashboard_status(value: str | None):
    cleaned = str(value or "").strip().lower()
    mapping = {
        "draft": "Draft",
        "published": "Published",
        "archived": "Archived",
    }
    out = mapping.get(cleaned)
    if out not in VALID_DASHBOARD_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="status must be Draft, Published, or Archived",
        )
    return out


def build_dashboard_config_response(config: DashboardConfig):
    return {
        "id": config.id,
        "name": config.name,
        "scope_type": config.scope_type,
        "location_code": config.location_code,
        "status": config.status,
        "active_version_id": config.active_version_id,
        "created_by": config.created_by,
        "remarks": config.remarks,
        "created_at": config.created_at,
        "updated_at": config.updated_at,
    }


def build_dashboard_version_response(version: DashboardVersion):
    return {
        "id": version.id,
        "config_id": version.config_id,
        "version_number": version.version_number,
        "config_json": version.config_json,
        "change_note": version.change_note,
        "created_by": version.created_by,
        "created_at": version.created_at,
    }


@app.get("/dashboard-configs", response_model=list[DashboardConfigResponse])
def get_dashboard_configs(
    scope_type: str | None = None,
    location_code: str | None = None,
    status: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    query = db.query(DashboardConfig)

    if scope_type:
        query = query.filter(
            DashboardConfig.scope_type == normalize_dashboard_scope_type(scope_type)
        )

    if location_code is not None and str(location_code).strip() != "":
        query = query.filter(
            DashboardConfig.location_code.ilike(str(location_code).strip())
        )

    if status:
        query = query.filter(
            DashboardConfig.status == normalize_dashboard_status(status)
        )

    configs = (
        query.order_by(
            DashboardConfig.scope_type.asc(),
            DashboardConfig.location_code.asc().nullsfirst(),
            DashboardConfig.name.asc(),
        )
        .all()
    )

    return [
        build_dashboard_config_response(config)
        for config in configs
    ]


@app.get("/dashboard-configs/{config_id}", response_model=DashboardConfigResponse)
def get_dashboard_config_by_id(
    config_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    config = (
        db.query(DashboardConfig)
        .filter(DashboardConfig.id == config_id)
        .first()
    )

    if not config:
        raise HTTPException(
            status_code=404,
            detail="Dashboard config not found",
        )

    return build_dashboard_config_response(config)


@app.post("/dashboard-configs", response_model=DashboardConfigResponse)
def create_dashboard_config(
    request: DashboardConfigCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    name = str(request.name or "").strip()
    if name == "":
        raise HTTPException(
            status_code=400,
            detail="name is required",
        )

    scope_type = normalize_dashboard_scope_type(request.scope_type)
    location_code = clean_optional_text(request.location_code)

    if scope_type == "GLOBAL":
        location_code = None
    else:
        if not location_code:
            raise HTTPException(
                status_code=400,
                detail="location_code is required when scope_type is LOCATION",
            )

    new_config = DashboardConfig(
        name=name,
        scope_type=scope_type,
        location_code=location_code,
        status="Draft",
        active_version_id=None,
        created_by=get_current_user_display_name(current_user),
        remarks=clean_optional_text(request.remarks),
    )

    db.add(new_config)

    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Dashboard config already exists for this scope",
        )

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Create DashboardConfig",
        current_user=current_user,
        entity_type="DashboardConfig",
        entity_id=new_config.id,
        entity_label=new_config.name,
        remarks="Dashboard config created",
        request_path="/dashboard-configs",
        details=build_dashboard_config_response(new_config),
    )

    db.commit()
    db.refresh(new_config)

    return build_dashboard_config_response(new_config)


@app.put("/dashboard-configs/{config_id}", response_model=DashboardConfigResponse)
def update_dashboard_config(
    config_id: int,
    request: DashboardConfigUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    config = (
        db.query(DashboardConfig)
        .filter(DashboardConfig.id == config_id)
        .first()
    )

    if not config:
        raise HTTPException(
            status_code=404,
            detail="Dashboard config not found",
        )

    old_details = build_dashboard_config_response(config)
    payload = request.model_dump(exclude_unset=True)

    if "name" in payload:
        name = str(payload.get("name") or "").strip()
        if name == "":
            raise HTTPException(
                status_code=400,
                detail="name cannot be empty",
            )
        config.name = name

    if "scope_type" in payload:
        config.scope_type = normalize_dashboard_scope_type(payload.get("scope_type"))

    if "location_code" in payload:
        config.location_code = clean_optional_text(payload.get("location_code"))

    if "status" in payload:
        config.status = normalize_dashboard_status(payload.get("status"))

    if "remarks" in payload:
        config.remarks = clean_optional_text(payload.get("remarks"))

    if config.scope_type == "GLOBAL":
        config.location_code = None
    else:
        if not clean_optional_text(config.location_code):
            raise HTTPException(
                status_code=400,
                detail="location_code is required when scope_type is LOCATION",
            )

    config.updated_at = datetime.now()

    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Dashboard config already exists for this scope",
        )

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Update DashboardConfig",
        current_user=current_user,
        entity_type="DashboardConfig",
        entity_id=config.id,
        entity_label=config.name,
        remarks="Dashboard config updated",
        request_path=f"/dashboard-configs/{config_id}",
        details={
            "old": old_details,
            "new": build_dashboard_config_response(config),
        },
    )

    db.commit()
    db.refresh(config)

    return build_dashboard_config_response(config)


@app.get(
    "/dashboard-configs/{config_id}/versions",
    response_model=list[DashboardVersionResponse],
)
def get_dashboard_versions_for_config(
    config_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    config = (
        db.query(DashboardConfig)
        .filter(DashboardConfig.id == config_id)
        .first()
    )

    if not config:
        raise HTTPException(
            status_code=404,
            detail="Dashboard config not found",
        )

    versions = (
        db.query(DashboardVersion)
        .filter(DashboardVersion.config_id == config_id)
        .order_by(DashboardVersion.version_number.desc())
        .all()
    )

    return [
        build_dashboard_version_response(v)
        for v in versions
    ]


@app.get("/dashboard-versions/{version_id}", response_model=DashboardVersionResponse)
def get_dashboard_version_by_id(
    version_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    version = (
        db.query(DashboardVersion)
        .filter(DashboardVersion.id == version_id)
        .first()
    )

    if not version:
        raise HTTPException(
            status_code=404,
            detail="Dashboard version not found",
        )

    return build_dashboard_version_response(version)


@app.post("/dashboard-configs/{config_id}/publish", response_model=DashboardConfigResponse)
def publish_dashboard_config(
    config_id: int,
    request: DashboardPublishRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    config = (
        db.query(DashboardConfig)
        .filter(DashboardConfig.id == config_id)
        .first()
    )

    if not config:
        raise HTTPException(
            status_code=404,
            detail="Dashboard config not found",
        )

    max_version = (
        db.query(func.max(DashboardVersion.version_number))
        .filter(DashboardVersion.config_id == config_id)
        .scalar()
        or 0
    )

    next_version_number = int(max_version) + 1

    new_version = DashboardVersion(
        config_id=config_id,
        version_number=next_version_number,
        config_json=jsonable_encoder(request.config_json),
        change_note=clean_optional_text(request.change_note),
        created_by=get_current_user_display_name(current_user),
    )

    db.add(new_version)
    db.flush()

    old_details = build_dashboard_config_response(config)

    config.status = "Published"
    config.active_version_id = new_version.id
    config.updated_at = datetime.now()

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Publish Dashboard",
        current_user=current_user,
        entity_type="DashboardConfig",
        entity_id=config.id,
        entity_label=config.name,
        remarks="Dashboard published",
        request_path=f"/dashboard-configs/{config_id}/publish",
        details={
            "config_id": config.id,
            "version_id": new_version.id,
            "version_number": new_version.version_number,
            "change_note": new_version.change_note,
            "old": old_details,
            "new": build_dashboard_config_response(config),
        },
    )

    db.commit()
    db.refresh(config)

    return build_dashboard_config_response(config)


@app.post("/dashboard-configs/{config_id}/revert", response_model=DashboardConfigResponse)
def revert_dashboard_config(
    config_id: int,
    request: DashboardRevertRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    config = (
        db.query(DashboardConfig)
        .filter(DashboardConfig.id == config_id)
        .first()
    )

    if not config:
        raise HTTPException(
            status_code=404,
            detail="Dashboard config not found",
        )

    version = (
        db.query(DashboardVersion)
        .filter(
            DashboardVersion.id == request.version_id,
            DashboardVersion.config_id == config_id,
        )
        .first()
    )

    if not version:
        raise HTTPException(
            status_code=400,
            detail="version_id does not belong to this config",
        )

    from_version_id = config.active_version_id
    old_details = build_dashboard_config_response(config)

    config.active_version_id = version.id
    config.updated_at = datetime.now()

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Revert Dashboard",
        current_user=current_user,
        entity_type="DashboardConfig",
        entity_id=config.id,
        entity_label=config.name,
        remarks="Dashboard reverted",
        request_path=f"/dashboard-configs/{config_id}/revert",
        details={
            "from_version_id": from_version_id,
            "to_version_id": version.id,
            "change_note": clean_optional_text(request.change_note),
            "old": old_details,
            "new": build_dashboard_config_response(config),
        },
    )

    db.commit()
    db.refresh(config)

    return build_dashboard_config_response(config)


VALID_DASHBOARD_DATA_SOURCE_STATUSES = {"Active", "Inactive"}
VALID_DASHBOARD_HANDLER_KEYS = {
    "FSO_OTR",
    "FSO_MATERIAL_BALANCE",
    "FSO_OUTTURN",
    "SHUTTLE_SUMMARY",
    "TANK_STOCK_SNAPSHOT",
    # Generic (Any Asset) dashboard handlers
    "ASSET_LIST",
    "OP_TRANSACTIONS",
    "OP_STATUS_COUNTS",
}


def normalize_dashboard_data_source_code(value: str | None):
    cleaned = str(value or "").strip().upper()
    if cleaned == "":
        raise HTTPException(
            status_code=400,
            detail="data_source_code is required",
        )
    return cleaned


def normalize_dashboard_data_source_status(value: str | None):
    cleaned = str(value or "").strip()
    if cleaned == "":
        cleaned = "Active"
    if cleaned not in VALID_DASHBOARD_DATA_SOURCE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="status must be Active or Inactive",
        )
    return cleaned


def normalize_dashboard_handler_key(value: str | None):
    cleaned = str(value or "").strip().upper()
    if cleaned == "":
        raise HTTPException(
            status_code=400,
            detail="handler_key is required",
        )
    return cleaned


def normalize_allowed_params_json(value):
    if not isinstance(value, dict):
        raise HTTPException(
            status_code=400,
            detail="allowed_params_json must be an object",
        )
    allowed = value.get("allowed")
    if not isinstance(allowed, list):
        raise HTTPException(
            status_code=400,
            detail='allowed_params_json must contain "allowed" as a list',
        )

    normalized_allowed = []
    seen_keys = set()
    for item in allowed:
        if not isinstance(item, dict):
            raise HTTPException(
                status_code=400,
                detail='allowed_params_json.allowed items must be objects',
            )

        key = str(item.get("key") or "").strip()
        if key == "":
            raise HTTPException(
                status_code=400,
                detail='allowed_params_json.allowed[].key is required',
            )
        if key in seen_keys:
            raise HTTPException(
                status_code=400,
                detail=f"Duplicate allowed param key: {key}",
            )
        seen_keys.add(key)

        typ = str(item.get("type") or "").strip().lower()
        if typ not in ("str", "int", "float", "bool", "date"):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid allowed param type for {key}",
            )

        required = bool(item.get("required"))
        normalized_allowed.append(
            {
                "key": key,
                "type": typ,
                "required": required,
            }
        )

    return {"allowed": normalized_allowed}


def build_dashboard_data_source_response(obj: DashboardDataSource):
    return {
        "id": obj.id,
        "data_source_code": obj.data_source_code,
        "data_source_name": obj.data_source_name,
        "description": obj.description,
        "handler_key": obj.handler_key,
        "allowed_params_json": obj.allowed_params_json,
        "status": obj.status,
        "created_by": obj.created_by,
        "remarks": obj.remarks,
        "created_at": obj.created_at,
        "updated_at": obj.updated_at,
    }


@app.get(
    "/dashboard-data-sources",
    response_model=list[DashboardDataSourceResponse],
)
def get_dashboard_data_sources(
    status: str | None = None,
    q: str | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    query = db.query(DashboardDataSource)

    cleaned_status = clean_optional_text(status)
    if cleaned_status:
        query = query.filter(
            DashboardDataSource.status == normalize_dashboard_data_source_status(cleaned_status)
        )

    cleaned_q = clean_optional_text(q)
    if cleaned_q:
        like = f"%{cleaned_q}%"
        query = query.filter(
            or_(
                DashboardDataSource.data_source_code.ilike(like),
                DashboardDataSource.data_source_name.ilike(like),
                DashboardDataSource.handler_key.ilike(like),
            )
        )

    rows = query.order_by(DashboardDataSource.data_source_code.asc()).all()

    return [
        build_dashboard_data_source_response(r)
        for r in rows
    ]


@app.get(
    "/dashboard-data-sources/{data_source_id}",
    response_model=DashboardDataSourceResponse,
)
def get_dashboard_data_source_by_id(
    data_source_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    obj = (
        db.query(DashboardDataSource)
        .filter(DashboardDataSource.id == data_source_id)
        .first()
    )

    if not obj:
        raise HTTPException(
            status_code=404,
            detail="Dashboard data source not found",
        )

    return build_dashboard_data_source_response(obj)


@app.post(
    "/dashboard-data-sources",
    response_model=DashboardDataSourceResponse,
)
def create_dashboard_data_source(
    request: DashboardDataSourceCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    code = normalize_dashboard_data_source_code(request.data_source_code)
    name = str(request.data_source_name or "").strip()
    if name == "":
        raise HTTPException(
            status_code=400,
            detail="data_source_name is required",
        )

    handler_key = normalize_dashboard_handler_key(request.handler_key)
    allowed_params_json = normalize_allowed_params_json(request.allowed_params_json)
    status_value = normalize_dashboard_data_source_status(request.status)

    obj = DashboardDataSource(
        data_source_code=code,
        data_source_name=name,
        description=clean_optional_text(request.description),
        handler_key=handler_key,
        allowed_params_json=jsonable_encoder(allowed_params_json),
        status=status_value,
        created_by=get_current_user_display_name(current_user),
        remarks=clean_optional_text(request.remarks),
    )

    db.add(obj)

    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="data_source_code already exists",
        )

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Create DashboardDataSource",
        current_user=current_user,
        entity_type="DashboardDataSource",
        entity_id=obj.id,
        entity_label=obj.data_source_code,
        remarks="Dashboard data source created",
        request_path="/dashboard-data-sources",
        details=build_dashboard_data_source_response(obj),
    )

    db.commit()
    db.refresh(obj)

    return build_dashboard_data_source_response(obj)


@app.put(
    "/dashboard-data-sources/{data_source_id}",
    response_model=DashboardDataSourceResponse,
)
def update_dashboard_data_source(
    data_source_id: int,
    request: DashboardDataSourceUpdate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    obj = (
        db.query(DashboardDataSource)
        .filter(DashboardDataSource.id == data_source_id)
        .first()
    )

    if not obj:
        raise HTTPException(
            status_code=404,
            detail="Dashboard data source not found",
        )

    old_details = build_dashboard_data_source_response(obj)
    payload = request.model_dump(exclude_unset=True)

    if "data_source_code" in payload:
        obj.data_source_code = normalize_dashboard_data_source_code(payload.get("data_source_code"))

    if "data_source_name" in payload:
        name = str(payload.get("data_source_name") or "").strip()
        if name == "":
            raise HTTPException(
                status_code=400,
                detail="data_source_name cannot be empty",
            )
        obj.data_source_name = name

    if "description" in payload:
        obj.description = clean_optional_text(payload.get("description"))

    if "handler_key" in payload:
        obj.handler_key = normalize_dashboard_handler_key(payload.get("handler_key"))

    if "allowed_params_json" in payload:
        obj.allowed_params_json = jsonable_encoder(
            normalize_allowed_params_json(payload.get("allowed_params_json"))
        )

    if "status" in payload:
        obj.status = normalize_dashboard_data_source_status(payload.get("status"))

    if "remarks" in payload:
        obj.remarks = clean_optional_text(payload.get("remarks"))

    obj.updated_at = datetime.now()

    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="data_source_code already exists",
        )

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Update DashboardDataSource",
        current_user=current_user,
        entity_type="DashboardDataSource",
        entity_id=obj.id,
        entity_label=obj.data_source_code,
        remarks="Dashboard data source updated",
        request_path=f"/dashboard-data-sources/{data_source_id}",
        details={
            "old": old_details,
            "new": build_dashboard_data_source_response(obj),
        },
    )

    db.commit()
    db.refresh(obj)

    return build_dashboard_data_source_response(obj)


@app.delete("/dashboard-data-sources/{data_source_id}")
def delete_dashboard_data_source(
    data_source_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    obj = (
        db.query(DashboardDataSource)
        .filter(DashboardDataSource.id == data_source_id)
        .first()
    )

    if not obj:
        raise HTTPException(
            status_code=404,
            detail="Dashboard data source not found",
        )

    old_details = build_dashboard_data_source_response(obj)

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Delete DashboardDataSource",
        current_user=current_user,
        entity_type="DashboardDataSource",
        entity_id=obj.id,
        entity_label=obj.data_source_code,
        remarks="Dashboard data source deleted",
        request_path=f"/dashboard-data-sources/{data_source_id}",
        details=old_details,
    )

    db.delete(obj)
    db.commit()

    return {"message": "Dashboard data source deleted successfully"}


def seed_dashboard_data_sources(db: Session, current_user_label: str | None):
    defs = [
        {
            "data_source_code": "FSO_OTR",
            "data_source_name": "FSO OTR",
            "description": "FSO operations ticket register (OTR)",
            "handler_key": "FSO_OTR",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": True},
                    {"key": "fso_asset_code", "type": "str", "required": True},
                    {"key": "date_from", "type": "date", "required": True},
                    {"key": "date_to", "type": "date", "required": True},
                    {"key": "shuttle_number", "type": "str", "required": False},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "FSO_MATERIAL_BALANCE",
            "data_source_name": "FSO Material Balance",
            "description": "FSO daily material balance",
            "handler_key": "FSO_MATERIAL_BALANCE",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": True},
                    {"key": "fso_asset_code", "type": "str", "required": True},
                    {"key": "date_from", "type": "date", "required": True},
                    {"key": "date_to", "type": "date", "required": True},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "FSO_OUTTURN",
            "data_source_name": "FSO Outturn",
            "description": "FSO outturn (receipt vs shuttle discharge)",
            "handler_key": "FSO_OUTTURN",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": True},
                    {"key": "fso_asset_code", "type": "str", "required": True},
                    {"key": "date_from", "type": "date", "required": True},
                    {"key": "date_to", "type": "date", "required": True},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "SHUTTLE_SUMMARY",
            "data_source_name": "Shuttle Summary",
            "description": "Shuttle voyage summary (group list)",
            "handler_key": "SHUTTLE_SUMMARY",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": False},
                    {"key": "shuttle_number", "type": "str", "required": False},
                    {"key": "shuttle_asset_code", "type": "str", "required": False},
                    {"key": "tab", "type": "str", "required": False},
                    {"key": "search", "type": "str", "required": False},
                    {"key": "date_from", "type": "date", "required": False},
                    {"key": "date_to", "type": "date", "required": False},
                    {"key": "page", "type": "int", "required": False},
                    {"key": "page_size", "type": "int", "required": False},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "ASSET_LIST",
            "data_source_name": "Asset Registry (Any Asset)",
            "description": "List of assets filtered by location/type/status.",
            "handler_key": "ASSET_LIST",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": False},
                    {"key": "asset_type_code", "type": "str", "required": False},
                    {"key": "status", "type": "str", "required": False},
                    {"key": "limit", "type": "int", "required": False},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "OP_TRANSACTIONS",
            "data_source_name": "Operation Tickets (Any Asset)",
            "description": "Operation transactions list (filter by location/asset/type/status/date).",
            "handler_key": "OP_TRANSACTIONS",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": False},
                    {"key": "asset_type_code", "type": "str", "required": False},
                    {"key": "asset_code", "type": "str", "required": False},
                    {"key": "operation_type_code", "type": "str", "required": False},
                    {"key": "status", "type": "str", "required": False},
                    {"key": "date_from", "type": "date", "required": False},
                    {"key": "date_to", "type": "date", "required": False},
                    {"key": "limit", "type": "int", "required": False},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "OP_STATUS_COUNTS",
            "data_source_name": "Operation Ticket Status Counts",
            "description": "Counts grouped by status (filter by location/asset/type/date).",
            "handler_key": "OP_STATUS_COUNTS",
            "allowed_params_json": {
                "allowed": [
                    {"key": "location_code", "type": "str", "required": False},
                    {"key": "asset_type_code", "type": "str", "required": False},
                    {"key": "asset_code", "type": "str", "required": False},
                    {"key": "operation_type_code", "type": "str", "required": False},
                    {"key": "date_from", "type": "date", "required": False},
                    {"key": "date_to", "type": "date", "required": False},
                ]
            },
            "status": "Active",
        },
        {
            "data_source_code": "TANK_STOCK_SNAPSHOT",
            "data_source_name": "Tank Stock Snapshot",
            "description": "Tank stock snapshot (placeholder)",
            "handler_key": "TANK_STOCK_SNAPSHOT",
            "allowed_params_json": {"allowed": []},
            "status": "Active",
        },
    ]

    created = 0
    updated = 0

    for d in defs:
        code = normalize_dashboard_data_source_code(d.get("data_source_code"))
        existing = (
            db.query(DashboardDataSource)
            .filter(DashboardDataSource.data_source_code == code)
            .first()
        )

        normalized_allowed = normalize_allowed_params_json(d.get("allowed_params_json"))
        handler_key = normalize_dashboard_handler_key(d.get("handler_key"))
        status_value = normalize_dashboard_data_source_status(d.get("status"))

        if not existing:
            obj = DashboardDataSource(
                data_source_code=code,
                data_source_name=str(d.get("data_source_name") or "").strip(),
                description=clean_optional_text(d.get("description")),
                handler_key=handler_key,
                allowed_params_json=jsonable_encoder(normalized_allowed),
                status=status_value,
                created_by=current_user_label,
                remarks=None,
            )
            db.add(obj)
            created += 1
        else:
            existing.data_source_name = str(d.get("data_source_name") or "").strip()
            existing.description = clean_optional_text(d.get("description"))
            existing.handler_key = handler_key
            existing.allowed_params_json = jsonable_encoder(normalized_allowed)
            existing.status = status_value
            existing.updated_at = datetime.now()
            updated += 1

    return {"created": created, "updated": updated, "total": created + updated}


@app.post("/dashboard-data-sources/seed")
def seed_dashboard_data_sources_api(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Dashboard",
        db,
    )

    current_user_label = get_current_user_display_name(current_user)
    result = seed_dashboard_data_sources(db, current_user_label)

    create_audit_log(
        db=db,
        module_name="Dashboard",
        action="Seed DashboardDataSources",
        current_user=current_user,
        entity_type="DashboardDataSource",
        entity_id=None,
        entity_label="dashboard_data_sources",
        remarks="Dashboard data sources seeded",
        request_path="/dashboard-data-sources/seed",
        details=result,
    )

    db.commit()
    return result


def coerce_dashboard_param_value(param_type: str, raw_value, field_name: str):
    t = str(param_type or "").strip().lower()

    if t == "date":
        if raw_value is None:
            return None
        return parse_date_filter(str(raw_value), field_name)

    if t == "int":
        if raw_value is None:
            return None
        try:
            return int(raw_value)
        except Exception:
            raise HTTPException(status_code=400, detail=f"{field_name} must be an integer")

    if t == "float":
        if raw_value is None:
            return None
        try:
            return float(raw_value)
        except Exception:
            raise HTTPException(status_code=400, detail=f"{field_name} must be a number")

    if t == "bool":
        if raw_value is None:
            return None
        if isinstance(raw_value, bool):
            return raw_value
        v = str(raw_value or "").strip().lower()
        if v in ("true", "1", "yes", "y", "on"):
            return True
        if v in ("false", "0", "no", "n", "off"):
            return False
        raise HTTPException(status_code=400, detail=f"{field_name} must be a boolean")

    if raw_value is None:
        return None
    return str(raw_value).strip()


def build_shuttle_summary_rows(
    db: Session,
    date_from: date | None = None,
    date_to: date | None = None,
    location_code: str | None = None,
    shuttle_number: str | None = None,
    shuttle_asset_code: str | None = None,
    tab: str | None = None,
    search: str | None = None,
    page: int = 1,
    page_size: int = 20,
):
    def _norm(v):
        return str(v or "").strip().upper()

    def _abs_qty(net_stock, net_water):
        try:
            return abs(float(net_stock or 0.0)) + abs(float(net_water or 0.0))
        except Exception:
            return 0.0

    def _op_code(meta):
        return _norm(meta.get("vessel_operation_code"))

    def _op_label(meta):
        return _norm(meta.get("vessel_operation_label"))

    def _is_loading(meta):
        code = _op_code(meta)
        if code == "LOADING":
            return True
        label = _op_label(meta)
        return ("LOADING" in label) and ("UNLOADING" not in label)

    def _is_sts_in(meta):
        code = _op_code(meta)
        if code == "STS_IN":
            return True
        label = _op_label(meta)
        return "STS IN" in label or "STS_IN" in label

    def _is_sts_out(meta):
        code = _op_code(meta)
        if code == "STS_OUT":
            return True
        label = _op_label(meta)
        return "STS OUT" in label or "STS_OUT" in label

    def _is_unloading(meta):
        code = _op_code(meta)
        if code == "UNLOADING":
            return True
        label = _op_label(meta)
        return ("UNLOADING" in label) or ("UNLOAD" in label)

    def _is_top_up(meta):
        code = _op_code(meta)
        if code == "TOP_UP":
            return True
        label = _op_label(meta)
        return ("TOP UP" in label) or ("TOP-UP" in label) or ("TOP_UP" in label)

    lc = clean_optional_text(location_code)
    sn = clean_optional_text(shuttle_number)
    ac = clean_optional_text(shuttle_asset_code)

    tab_norm = (clean_optional_text(tab) or "OPEN").upper()
    search_norm = (clean_optional_text(search) or "").strip()

    page = 1 if page is None or page < 1 else page
    page_size = 20 if page_size is None or page_size < 1 else min(int(page_size), 200)
    offset = (page - 1) * page_size

    base_q = (
        db.query(OperationTransaction)
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "shuttle_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
    )

    if date_from:
        base_q = base_q.filter(OperationTransaction.operation_date >= date_from)
    if date_to:
        base_q = base_q.filter(OperationTransaction.operation_date <= date_to)
    if lc:
        base_q = base_q.filter(OperationTransaction.origin_location_code.ilike(lc))
    if sn:
        base_q = base_q.filter(OperationTransaction.convoy_number.ilike(sn))
    if ac:
        base_q = base_q.filter(OperationTransaction.primary_asset_code.ilike(ac))

    if search_norm:
        like = f"%{search_norm}%"
        base_q = base_q.filter(
            or_(
                OperationTransaction.origin_location_code.ilike(like),
                OperationTransaction.convoy_number.ilike(like),
                OperationTransaction.primary_asset_code.ilike(like),
                OperationTransaction.operation_number.ilike(like),
                OperationTransaction.operation_ticket_number.ilike(like),
                OperationTransaction.product_name.ilike(like),
            )
        )

    voyage_status_expr = func.coalesce(ShuttleVoyage.status, literal("OPEN"))

    key_loc = OperationTransaction.origin_location_code
    key_shuttle = OperationTransaction.convoy_number
    key_asset = OperationTransaction.primary_asset_code

    group_q = (
        base_q.with_entities(
            key_loc.label("location_code"),
            key_shuttle.label("shuttle_number"),
            key_asset.label("shuttle_asset_code"),
            func.min(OperationTransaction.operation_date).label("first_date"),
            func.max(OperationTransaction.operation_date).label("last_date"),
        )
        .outerjoin(
            ShuttleVoyage,
            and_(
                ShuttleVoyage.location_code == key_loc,
                ShuttleVoyage.shuttle_number == key_shuttle,
                ShuttleVoyage.shuttle_asset_code == key_asset,
            ),
        )
        .group_by(key_loc, key_shuttle, key_asset, voyage_status_expr)
    )

    if tab_norm == "CLOSED":
        group_q = group_q.filter(voyage_status_expr == "CLOSED")
    else:
        group_q = group_q.filter(voyage_status_expr != "CLOSED")

    total_groups = group_q.count()

    group_rows = (
        group_q.order_by(func.max(OperationTransaction.operation_date).desc(), key_loc.asc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    if not group_rows:
        return {
            "rows": [],
            "total_groups": total_groups,
            "page": page,
            "page_size": page_size,
            "has_more": total_groups > offset + page_size,
        }

    keys = [(r.location_code, r.shuttle_number, r.shuttle_asset_code) for r in group_rows]

    key_filters = []
    for (loc_code, sh_num, asset_code) in keys:
        key_filters.append(
            and_(
                OperationTransaction.origin_location_code == loc_code,
                OperationTransaction.convoy_number == sh_num,
                OperationTransaction.primary_asset_code == asset_code,
            )
        )

    payload_rows = (
        db.query(
            OperationTransaction.origin_location_code,
            OperationTransaction.convoy_number,
            OperationTransaction.primary_asset_code,
            OperationTransactionValue.field_value,
        )
        .join(
            OperationTransactionValue,
            OperationTransactionValue.transaction_id == OperationTransaction.id,
        )
        .filter(
            OperationTransactionValue.field_code == "shuttle_payload",
            OperationTransaction.status == APPROVED_TRANSACTION_STATUS,
        )
        .filter(or_(*key_filters))
        .all()
    )

    totals_map = {}
    for (loc_code, sh_num, asset_code, field_value) in payload_rows:
        k = f"{loc_code}|{sh_num}|{asset_code}"
        if k not in totals_map:
            totals_map[k] = {"net_receipt_bbl": 0.0, "net_discharge_bbl": 0.0}

        if isinstance(field_value, dict):
            fv = field_value
            meta = fv.get("meta") or {}
            net = ((fv.get("calculated") or {}).get("net") or {})
            net_stock = float(safe_float(net.get("net_stock_bbl")))
            net_water = float(safe_float(net.get("net_water_bbl")))
            qty_bbl = _abs_qty(net_stock, net_water)

            if (_is_loading(meta) or _is_sts_in(meta) or _is_top_up(meta)) and (not _is_unloading(meta)):
                totals_map[k]["net_receipt_bbl"] += qty_bbl
            if _is_unloading(meta) and not _is_sts_out(meta):
                totals_map[k]["net_discharge_bbl"] += qty_bbl

    rows = []
    for (loc_code, sh_num, asset_code) in keys:
        asset = get_asset_by_code(asset_code, db)
        loc = get_location_by_code(loc_code, db)
        voyage = get_shuttle_voyage_by_key(db, loc_code, sh_num or "", asset_code)

        k = f"{loc_code}|{sh_num}|{asset_code}"
        t = totals_map.get(k, {"net_receipt_bbl": 0.0, "net_discharge_bbl": 0.0})

        rows.append(
            {
                "group_key": k,
                "location_code": loc_code,
                "location_name": loc.location_name if loc else "",
                "shuttle_number": sh_num or "",
                "shuttle_asset_code": asset_code,
                "shuttle_asset_name": asset.asset_name if asset else "",
                "voyage_status": voyage.status if voyage else "OPEN",
                "closed_by": voyage.closed_by if voyage else None,
                "closed_at": voyage.closed_at if voyage else None,
                "closure_remarks": voyage.closure_remarks if voyage else None,
                "net_receipt_bbl": float(t["net_receipt_bbl"]),
                "net_discharge_bbl": float(t["net_discharge_bbl"]),
            }
        )

    return {
        "rows": rows,
        "total_groups": total_groups,
        "page": page,
        "page_size": page_size,
        "has_more": total_groups > offset + page_size,
    }


@app.post("/dashboard/data", response_model=DashboardDataResponse)
def dashboard_data_gateway(
    request: DashboardDataRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Dashboard",
        db,
    )

    code = normalize_dashboard_data_source_code(request.data_source_code)
    data_source = (
        db.query(DashboardDataSource)
        .filter(DashboardDataSource.data_source_code == code)
        .first()
    )

    if not data_source:
        raise HTTPException(
            status_code=404,
            detail="Dashboard data source not found",
        )

    if data_source.status != "Active":
        raise HTTPException(
            status_code=400,
            detail="Dashboard data source is not Active",
        )

    allowed_spec = data_source.allowed_params_json if isinstance(data_source.allowed_params_json, dict) else {}
    allowed = allowed_spec.get("allowed") if isinstance(allowed_spec, dict) else None
    if not isinstance(allowed, list):
        raise HTTPException(
            status_code=400,
            detail="Dashboard data source allowed_params_json is invalid",
        )

    params = request.params if isinstance(request.params, dict) else None
    if params is None:
        raise HTTPException(
            status_code=400,
            detail="params must be an object",
        )

    allowed_keys = [
        str(i.get("key")).strip()
        for i in allowed
        if isinstance(i, dict) and i.get("key") is not None and str(i.get("key")).strip() != ""
    ]
    allowed_key_set = set(allowed_keys)

    extra_keys = [k for k in params.keys() if k not in allowed_key_set]
    if extra_keys:
        raise HTTPException(
            status_code=400,
            detail=f"Unexpected params: {', '.join(sorted(extra_keys))}",
        )

    missing_required = []
    resolved = {}
    for item in allowed:
        if not isinstance(item, dict):
            continue
        k = str(item.get("key") or "").strip()
        if k == "":
            continue
        required = bool(item.get("required"))
        if k in params:
            resolved[k] = coerce_dashboard_param_value(item.get("type"), params.get(k), k)
        if required and (k not in resolved or resolved.get(k) in (None, "")):
            missing_required.append(k)

    if missing_required:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required params: {', '.join(missing_required)}",
        )

    handler_key = str(data_source.handler_key or "").strip().upper()
    if handler_key not in VALID_DASHBOARD_HANDLER_KEYS:
        raise HTTPException(
            status_code=400,
            detail="Unsupported handler_key",
        )

    if handler_key == "FSO_OTR":
        rows, totals = build_fso_otr_report(
            db=db,
            location_code=resolved["location_code"],
            fso_asset_code=resolved["fso_asset_code"],
            date_from=resolved["date_from"],
            date_to=resolved["date_to"],
            shuttle_number=resolved.get("shuttle_number"),
        )
        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {
                "totals": totals,
            },
        }

    if handler_key == "FSO_MATERIAL_BALANCE":
        rows = build_fso_material_balance(
            db=db,
            location_code=resolved["location_code"],
            fso_asset_code=resolved["fso_asset_code"],
            date_from=resolved["date_from"],
            date_to=resolved["date_to"],
        )
        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {},
        }

    if handler_key == "FSO_OUTTURN":
        rows, totals, totals_pct = build_fso_outturn_report(
            db=db,
            location_code=resolved["location_code"],
            fso_asset_code=resolved["fso_asset_code"],
            date_from=resolved["date_from"],
            date_to=resolved["date_to"],
        )
        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {
                "totals": totals,
                "totals_pct": totals_pct,
            },
        }

    if handler_key == "SHUTTLE_SUMMARY":
        summary = build_shuttle_summary_rows(
            db=db,
            date_from=resolved.get("date_from"),
            date_to=resolved.get("date_to"),
            location_code=resolved.get("location_code"),
            shuttle_number=resolved.get("shuttle_number"),
            shuttle_asset_code=resolved.get("shuttle_asset_code"),
            tab=resolved.get("tab"),
            search=resolved.get("search"),
            page=resolved.get("page") or 1,
            page_size=resolved.get("page_size") or 20,
        )
        return {
            "data_source_code": code,
            "rows": summary.get("rows") or [],
            "meta": {
                "total_groups": summary.get("total_groups", 0),
                "page": summary.get("page", 1),
                "page_size": summary.get("page_size", 20),
                "has_more": summary.get("has_more", False),
            },
        }

    if handler_key == "ASSET_LIST":
        q = db.query(Asset)

        if resolved.get("location_code"):
            q = q.filter(Asset.location_code == resolved["location_code"])

        if resolved.get("asset_type_code"):
            q = q.filter(Asset.asset_type_code == resolved["asset_type_code"])

        if resolved.get("status"):
            q = q.filter(Asset.status == resolved["status"])

        total_rows = q.count()

        limit = resolved.get("limit") or 500
        if limit < 1:
            limit = 1
        if limit > 5000:
            limit = 5000

        items = q.order_by(Asset.asset_code.asc()).limit(limit).all()

        rows = []
        for a in items:
            rows.append(
                {
                    "asset_code": a.asset_code,
                    "asset_name": a.asset_name,
                    "asset_type_code": a.asset_type_code,
                    "asset_scope": a.asset_scope,
                    "location_code": a.location_code,
                    "status": a.status,
                }
            )

        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {"total_rows": total_rows, "limit": limit},
        }

    if handler_key == "OP_TRANSACTIONS":
        q = db.query(OperationTransaction)

        if resolved.get("location_code"):
            q = q.filter(
                OperationTransaction.origin_location_code == resolved["location_code"]
            )

        if resolved.get("asset_type_code"):
            q = q.filter(
                OperationTransaction.primary_asset_type_code == resolved["asset_type_code"]
            )

        if resolved.get("asset_code"):
            q = q.filter(OperationTransaction.primary_asset_code == resolved["asset_code"])

        if resolved.get("operation_type_code"):
            q = q.filter(
                OperationTransaction.operation_type_code == resolved["operation_type_code"]
            )

        if resolved.get("status"):
            q = q.filter(OperationTransaction.status == resolved["status"])

        if resolved.get("date_from"):
            q = q.filter(OperationTransaction.operation_date >= resolved["date_from"])

        if resolved.get("date_to"):
            q = q.filter(OperationTransaction.operation_date <= resolved["date_to"])

        total_rows = q.count()

        limit = resolved.get("limit") or 200
        if limit < 1:
            limit = 1
        if limit > 2000:
            limit = 2000

        items = (
            q.order_by(
                OperationTransaction.operation_date.desc(),
                OperationTransaction.id.desc(),
            )
            .limit(limit)
            .all()
        )

        rows = []
        for t in items:
            rows.append(
                {
                    "transaction_id": t.id,
                    "ticket_number": t.operation_ticket_number or t.operation_number,
                    "operation_type_code": t.operation_type_code,
                    "asset_code": t.primary_asset_code,
                    "asset_type_code": t.primary_asset_type_code,
                    "location_code": t.origin_location_code,
                    "operation_date": str(t.operation_date) if t.operation_date else None,
                    "status": t.status,
                    "product_name": t.product_name,
                    "remarks": t.remarks,
                }
            )

        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {"total_rows": total_rows, "limit": limit},
        }

    if handler_key == "OP_STATUS_COUNTS":
        q = db.query(
            OperationTransaction.status.label("status"),
            func.count(OperationTransaction.id).label("count"),
        )

        if resolved.get("location_code"):
            q = q.filter(
                OperationTransaction.origin_location_code == resolved["location_code"]
            )

        if resolved.get("asset_type_code"):
            q = q.filter(
                OperationTransaction.primary_asset_type_code == resolved["asset_type_code"]
            )

        if resolved.get("asset_code"):
            q = q.filter(OperationTransaction.primary_asset_code == resolved["asset_code"])

        if resolved.get("operation_type_code"):
            q = q.filter(
                OperationTransaction.operation_type_code == resolved["operation_type_code"]
            )

        if resolved.get("date_from"):
            q = q.filter(OperationTransaction.operation_date >= resolved["date_from"])

        if resolved.get("date_to"):
            q = q.filter(OperationTransaction.operation_date <= resolved["date_to"])

        q = q.group_by(OperationTransaction.status).order_by(
            func.count(OperationTransaction.id).desc()
        )

        rows = [{"status": r.status, "count": int(r.count)} for r in q.all()]

        return {
            "data_source_code": code,
            "rows": rows,
            "meta": {"total_rows": len(rows)},
        }

    return {
        "data_source_code": code,
        "rows": [],
        "meta": {"note": "not implemented"},
    }

# -------------------------
# Company Report Profile APIs
# -------------------------

VALID_COMPANY_REPORT_PROFILE_STATUSES = [
    "Active",
    "Inactive",
    "Blocked",
]


def build_company_report_profile_response(profile: CompanyReportProfile):
    return {
        "id": profile.id,
        "profile_name": profile.profile_name,
        "company_name": profile.company_name,
        "system_name": profile.system_name,
        "report_subtitle": profile.report_subtitle,
        "logo_data_url": profile.logo_data_url,
        "logo_text": profile.logo_text,
        "footer_formula": profile.footer_formula,
        "footer_note": profile.footer_note,
        "status": profile.status,
        "created_at": profile.created_at,
        "updated_at": profile.updated_at,
    }


def validate_company_report_profile(
    profile: CompanyReportProfileCreate,
):
    if profile.profile_name.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Profile Name is required",
        )

    if profile.company_name.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Company Name is required",
        )

    if profile.system_name.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="System Name is required",
        )

    if profile.report_subtitle.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Report Subtitle is required",
        )

    if profile.logo_text.strip() == "":
        raise HTTPException(
            status_code=400,
            detail="Logo placeholder text is required",
        )

    if profile.status not in VALID_COMPANY_REPORT_PROFILE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="Status must be Active, Inactive, or Blocked",
        )

    if profile.logo_data_url:
        logo_value = profile.logo_data_url.strip()

        if not (
            logo_value.startswith("data:image/png;base64,")
            or logo_value.startswith("data:image/jpeg;base64,")
            or logo_value.startswith("data:image/jpg;base64,")
        ):
            raise HTTPException(
                status_code=400,
                detail="Logo must be a PNG, JPG, or JPEG data URL",
            )

        max_logo_length = 2_000_000

        if len(logo_value) > max_logo_length:
            raise HTTPException(
                status_code=400,
                detail="Logo image is too large. Please upload a smaller PNG/JPG/JPEG file.",
            )

# -------------------------
# Audit Log APIs
# -------------------------

@app.get("/audit-logs", response_model=list[AuditLogResponse])
def get_audit_logs(
    module_name: str | None = None,
    action: str | None = None,
    entity_type: str | None = None,
    ticket_number: str | None = None,
    operation_number: str | None = None,
    performed_by: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    limit: int = 200,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Audit Log",
        db,
    )

    query = db.query(AuditLog)

    if module_name:
        query = query.filter(AuditLog.module_name.ilike(f"%{module_name.strip()}%"))

    if action:
        query = query.filter(AuditLog.action.ilike(f"%{action.strip()}%"))

    if entity_type:
        query = query.filter(AuditLog.entity_type.ilike(f"%{entity_type.strip()}%"))

    if ticket_number:
        query = query.filter(AuditLog.ticket_number.ilike(f"%{ticket_number.strip()}%"))

    if operation_number:
        query = query.filter(
            AuditLog.operation_number.ilike(f"%{operation_number.strip()}%")
        )

    if performed_by:
        query = query.filter(AuditLog.performed_by.ilike(f"%{performed_by.strip()}%"))

    if date_from:
        try:
            parsed_date_from = datetime.fromisoformat(date_from)
            query = query.filter(AuditLog.created_at >= parsed_date_from)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="date_from must be in ISO format, for example 2026-05-13",
            )

    if date_to:
        try:
            parsed_date_to = datetime.fromisoformat(date_to)
            query = query.filter(AuditLog.created_at <= parsed_date_to)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="date_to must be in ISO format, for example 2026-05-13",
            )

    safe_limit = min(max(limit, 1), 1000)

    audit_logs = (
        query.order_by(AuditLog.id.desc())
        .limit(safe_limit)
        .all()
    )

    return [
        build_audit_log_response(audit_log)
        for audit_log in audit_logs
    ]

@app.get("/audit-logs/{audit_log_id}", response_model=AuditLogResponse)
def get_audit_log_by_id(
    audit_log_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Audit Log",
        db,
    )

    audit_log = db.query(AuditLog).filter(AuditLog.id == audit_log_id).first()

    if not audit_log:
        raise HTTPException(
            status_code=404,
            detail="Audit log not found",
        )

    return build_audit_log_response(audit_log)

@app.get(
    "/company-report-profiles",
    response_model=list[CompanyReportProfileResponse],
)
def get_company_report_profiles(
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Company Report Profile",
        db,
    )

    profiles = (
        db.query(CompanyReportProfile)
        .order_by(CompanyReportProfile.profile_name.asc())
        .all()
    )

    return [
        build_company_report_profile_response(profile)
        for profile in profiles
    ]


@app.post(
    "/company-report-profiles",
    response_model=CompanyReportProfileResponse,
)
def create_company_report_profile(
    profile: CompanyReportProfileCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Company Report Profile",
        db,
    )

    validate_company_report_profile(profile)

    existing_profile = (
        db.query(CompanyReportProfile)
        .filter(CompanyReportProfile.profile_name.ilike(profile.profile_name.strip()))
        .first()
    )

    if existing_profile:
        raise HTTPException(
            status_code=400,
            detail="Report profile name already exists",
        )

    new_profile = CompanyReportProfile(
        profile_name=profile.profile_name.strip(),
        company_name=profile.company_name.strip(),
        system_name=profile.system_name.strip(),
        report_subtitle=profile.report_subtitle.strip(),
        logo_data_url=clean_optional_text(profile.logo_data_url),
        logo_text=profile.logo_text.strip(),
        footer_formula=clean_optional_text(profile.footer_formula),
        footer_note=clean_optional_text(profile.footer_note),
        status=profile.status,
    )

    db.add(new_profile)
    db.flush()

    create_audit_log(
        db=db,
        module_name="Company Report Profile",
        action="Create Company Report Profile",
        current_user=current_user,
        entity_type="CompanyReportProfile",
        entity_id=new_profile.id,
        entity_label=new_profile.profile_name,
        remarks="Company report profile created",
        request_path="/company-report-profiles",
        details={
            "profile_name": new_profile.profile_name,
            "company_name": new_profile.company_name,
            "system_name": new_profile.system_name,
            "report_subtitle": new_profile.report_subtitle,
            "logo_uploaded": bool(new_profile.logo_data_url),
            "logo_text": new_profile.logo_text,
            "footer_formula_available": bool(new_profile.footer_formula),
            "footer_note_available": bool(new_profile.footer_note),
            "status": new_profile.status,
        },
    )

    db.commit()
    db.refresh(new_profile)

    return build_company_report_profile_response(new_profile)


@app.put(
    "/company-report-profiles/{profile_id}",
    response_model=CompanyReportProfileResponse,
)
def update_company_report_profile(
    profile_id: int,
    profile: CompanyReportProfileCreate,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Company Report Profile",
        db,
    )

    existing_profile = (
        db.query(CompanyReportProfile)
        .filter(CompanyReportProfile.id == profile_id)
        .first()
    )

    if not existing_profile:
        raise HTTPException(
            status_code=404,
            detail="Company report profile not found",
        )

    validate_company_report_profile(profile)

    duplicate_profile = (
        db.query(CompanyReportProfile)
        .filter(
            CompanyReportProfile.profile_name.ilike(profile.profile_name.strip()),
            CompanyReportProfile.id != profile_id,
        )
        .first()
    )

    if duplicate_profile:
        raise HTTPException(
            status_code=400,
            detail="Report profile name already exists",
        )

    old_profile_data = {
        "profile_name": existing_profile.profile_name,
        "company_name": existing_profile.company_name,
        "system_name": existing_profile.system_name,
        "report_subtitle": existing_profile.report_subtitle,
        "logo_uploaded": bool(existing_profile.logo_data_url),
        "logo_text": existing_profile.logo_text,
        "footer_formula_available": bool(existing_profile.footer_formula),
        "footer_note_available": bool(existing_profile.footer_note),
        "status": existing_profile.status,
    }

    existing_profile.profile_name = profile.profile_name.strip()
    existing_profile.company_name = profile.company_name.strip()
    existing_profile.system_name = profile.system_name.strip()
    existing_profile.report_subtitle = profile.report_subtitle.strip()
    existing_profile.logo_data_url = clean_optional_text(profile.logo_data_url)
    existing_profile.logo_text = profile.logo_text.strip()
    existing_profile.footer_formula = clean_optional_text(profile.footer_formula)
    existing_profile.footer_note = clean_optional_text(profile.footer_note)
    existing_profile.status = profile.status
    existing_profile.updated_at = datetime.now()

    new_profile_data = {
        "profile_name": existing_profile.profile_name,
        "company_name": existing_profile.company_name,
        "system_name": existing_profile.system_name,
        "report_subtitle": existing_profile.report_subtitle,
        "logo_uploaded": bool(existing_profile.logo_data_url),
        "logo_text": existing_profile.logo_text,
        "footer_formula_available": bool(existing_profile.footer_formula),
        "footer_note_available": bool(existing_profile.footer_note),
        "status": existing_profile.status,
    }

    create_audit_log(
        db=db,
        module_name="Company Report Profile",
        action="Update Company Report Profile",
        current_user=current_user,
        entity_type="CompanyReportProfile",
        entity_id=existing_profile.id,
        entity_label=existing_profile.profile_name,
        remarks="Company report profile updated",
        request_path=f"/company-report-profiles/{profile_id}",
        details={
            "before": old_profile_data,
            "after": new_profile_data,
        },
    )

    db.commit()
    db.refresh(existing_profile)

    return build_company_report_profile_response(existing_profile)


@app.delete("/company-report-profiles/{profile_id}")
def delete_company_report_profile(
    profile_id: int,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Company Report Profile",
        db,
    )

    existing_profile = (
        db.query(CompanyReportProfile)
        .filter(CompanyReportProfile.id == profile_id)
        .first()
    )

    if not existing_profile:
        raise HTTPException(
            status_code=404,
            detail="Company report profile not found",
        )

    deleted_profile_data = {
        "profile_name": existing_profile.profile_name,
        "company_name": existing_profile.company_name,
        "system_name": existing_profile.system_name,
        "report_subtitle": existing_profile.report_subtitle,
        "logo_uploaded": bool(existing_profile.logo_data_url),
        "logo_text": existing_profile.logo_text,
        "footer_formula_available": bool(existing_profile.footer_formula),
        "footer_note_available": bool(existing_profile.footer_note),
        "status": existing_profile.status,
    }

    create_audit_log(
        db=db,
        module_name="Company Report Profile",
        action="Delete Company Report Profile",
        current_user=current_user,
        entity_type="CompanyReportProfile",
        entity_id=existing_profile.id,
        entity_label=existing_profile.profile_name,
        remarks="Company report profile deleted",
        request_path=f"/company-report-profiles/{profile_id}",
        details={
            "deleted_profile": deleted_profile_data,
        },
    )

    db.delete(existing_profile)
    db.commit()

    return {
        "message": "Company report profile deleted successfully",
    }

# -------------------------
# Barge Seal Master APIs
# -------------------------

@app.get("/barge-seal-master", response_model=list[BargeSealMasterResponse])
def get_barge_seal_master(
    asset_code: str,
    effective_date: date | None = None,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "View Barge Seal Master",
        db,
    )

    asset_code_clean = (asset_code or "").strip()
    if asset_code_clean == "":
        raise HTTPException(status_code=400, detail="asset_code is required")

    query = db.query(BargeSealMaster).filter(
        BargeSealMaster.asset_code.ilike(asset_code_clean)
    )

    if effective_date is None:
        query = query.filter(BargeSealMaster.effective_date.is_(None))
    else:
        query = query.filter(BargeSealMaster.effective_date == effective_date)

    rows = (
        query.order_by(
            BargeSealMaster.tank_id.asc(),
            BargeSealMaster.seal_position.asc(),
        ).all()
    )

    return rows


@app.post("/barge-seal-master/bulk", response_model=list[BargeSealMasterResponse])
def bulk_save_barge_seal_master(
    request: BargeSealMasterBulkSaveRequest,
    current_user: User = Depends(get_current_user_from_token),
    db: Session = Depends(get_db),
):
    require_user_permission(
        current_user,
        "Manage Barge Seal Master",
        db,
    )

    asset_code = (request.asset_code or "").strip()
    if asset_code == "":
        raise HTTPException(status_code=400, detail="asset_code is required")

    asset = db.query(Asset).filter(Asset.asset_code.ilike(asset_code)).first()
    if not asset:
        raise HTTPException(status_code=400, detail="Asset not found")

    if request.rows is None or len(request.rows) == 0:
        raise HTTPException(status_code=400, detail="Please provide at least one seal row")

    def norm(s: str) -> str:
        return str(s or "").strip()

    def norm_pos(s: str) -> str:
        return str(s or "").strip().upper()

    # Build request map + validate duplicates
    req_map = {}
    duplicate_keys = []

    for row in request.rows:
        tank_id = norm(row.tank_id)
        seal_position = norm_pos(row.seal_position)
        seal_number = norm(row.seal_number)

        if tank_id == "":
            raise HTTPException(status_code=400, detail="tank_id is required in rows")
        if seal_position == "":
            raise HTTPException(status_code=400, detail="seal_position is required in rows")
        if seal_number == "":
            raise HTTPException(status_code=400, detail="seal_number is required in rows")

        key = (tank_id, seal_position)
        if key in req_map:
            duplicate_keys.append(f"{tank_id}:{seal_position}")
            continue

        req_map[key] = {
            "tank_id": tank_id,
            "seal_position": seal_position,
            "seal_number": seal_number,
            "remarks": clean_optional_text(row.remarks),
            "status": row.status or "Active",
        }

    if duplicate_keys:
        raise HTTPException(
            status_code=400,
            detail=f"Duplicate seal keys in request: {', '.join(duplicate_keys)}",
        )

    # Load existing master rows for this asset + effective_date
    existing_q = db.query(BargeSealMaster).filter(
        BargeSealMaster.asset_code.ilike(asset_code)
    )

    if request.effective_date is None:
        existing_q = existing_q.filter(BargeSealMaster.effective_date.is_(None))
    else:
        existing_q = existing_q.filter(BargeSealMaster.effective_date == request.effective_date)

    existing_rows = existing_q.all()

    def existing_key(obj: BargeSealMaster):
        return (norm(obj.tank_id), norm_pos(obj.seal_position))

    existing_map = {existing_key(r): r for r in existing_rows}

    before_count = len(existing_rows)

    added = []
    updated = []
    removed = []

    # Removed: exists in DB but not in request
    for key, obj in existing_map.items():
        if key not in req_map:
            removed.append({
                "tank_id": obj.tank_id,
                "seal_position": obj.seal_position,
                "seal_number": obj.seal_number,
                "status": obj.status,
            })
            db.delete(obj)

    # Added/Updated
    for key, incoming in req_map.items():
        if key in existing_map:
            obj = existing_map[key]

            changed = (
                (obj.seal_number or "") != (incoming["seal_number"] or "")
                or (obj.status or "") != (incoming["status"] or "")
                or (obj.remarks or "") != (incoming["remarks"] or "")
                or obj.effective_date != request.effective_date
            )

            if changed:
                updated.append({
                    "tank_id": obj.tank_id,
                    "seal_position": obj.seal_position,
                    "before_seal_number": obj.seal_number,
                    "after_seal_number": incoming["seal_number"],
                    "before_status": obj.status,
                    "after_status": incoming["status"],
                })

                obj.seal_number = incoming["seal_number"]
                obj.status = incoming["status"]
                obj.remarks = incoming["remarks"]
                obj.effective_date = request.effective_date
                obj.updated_at = datetime.now()
        else:
            new_row = BargeSealMaster(
                asset_code=asset_code,
                tank_id=incoming["tank_id"],
                seal_position=incoming["seal_position"],
                seal_number=incoming["seal_number"],
                effective_date=request.effective_date,
                remarks=incoming["remarks"],
                status=incoming["status"],
            )
            db.add(new_row)

            added.append({
                "tank_id": incoming["tank_id"],
                "seal_position": incoming["seal_position"],
                "seal_number": incoming["seal_number"],
                "status": incoming["status"],
            })

    db.flush()

    after_count = before_count - len(removed) + len(added)

    # Audit (store counts + small samples only)
    create_audit_log(
        db=db,
        module_name="Barge Seal Master",
        action="Bulk Save Barge Seals",
        current_user=current_user,
        entity_type="BargeSealMaster",
        entity_id=None,
        entity_label=asset_code,
        remarks="Barge seal master bulk saved",
        request_path="/barge-seal-master/bulk",
        details={
            "asset_code": asset_code,
            "effective_date": str(request.effective_date) if request.effective_date else None,
            "before_count": before_count,
            "after_count": after_count,
            "added_count": len(added),
            "updated_count": len(updated),
            "removed_count": len(removed),
            "added_sample": added[:20],
            "updated_sample": updated[:20],
            "removed_sample": removed[:20],
        },
    )

    db.commit()

    # Return saved rows
    out_q = db.query(BargeSealMaster).filter(
        BargeSealMaster.asset_code.ilike(asset_code)
    )

    if request.effective_date is None:
        out_q = out_q.filter(BargeSealMaster.effective_date.is_(None))
    else:
        out_q = out_q.filter(BargeSealMaster.effective_date == request.effective_date)

    return out_q.order_by(
        BargeSealMaster.tank_id.asc(),
        BargeSealMaster.seal_position.asc(),
    ).all()
